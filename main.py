# -------------------- IMPORTS --------------------
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os
from datetime import datetime, date, timedelta
from calendar import month_name
import uuid
import json
import io
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font 
from flask import make_response

# -------------------- FLASK APP SETUP --------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "super_secret_key_change_later")

# -------------------- DATABASE CONFIGURATION --------------------
DATABASE_PATH = os.environ.get("DATABASE_PATH", "hospital.db")

def get_db_connection():
    """Establish SQLite database connection with error handling."""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row  # This enables column access by name
        return conn
    except Exception as e:
        app.logger.error(f"Database connection error: {e}")
        return None

# -------------------- DATABASE INITIALIZATION --------------------
def update_prescriptions_table():
    """Update prescriptions table to allow NULL patient_id"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Check if prescriptions table exists
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='prescriptions'")
        if cur.fetchone():
            # Get current table info
            cur.execute("PRAGMA table_info(prescriptions)")
            columns = [col[1] for col in cur.fetchall()]
            
            # Check if patient_id column is set to NOT NULL
            cur.execute("PRAGMA table_info(prescriptions)")
            for col in cur.fetchall():
                if col[1] == 'patient_id' and col[3] == 1:  # notnull = 1
                    # Need to recreate table without NOT NULL constraint
                    print("Updating prescriptions table to allow NULL patient_id...")
                    
                    # Create temporary table
                    cur.execute("""
                        CREATE TABLE prescriptions_temp (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            prescription_no VARCHAR(50) UNIQUE NOT NULL,
                            patient_id INTEGER,
                            patient_name VARCHAR(200) NOT NULL,
                            doctor_id INTEGER NOT NULL,
                            doctor_name VARCHAR(100) NOT NULL,
                            drug_id INTEGER,
                            drug_name VARCHAR(100) NOT NULL,
                            strength VARCHAR(50),
                            dosage VARCHAR(100) NOT NULL,
                            frequency VARCHAR(100) NOT NULL,
                            duration VARCHAR(50) NOT NULL,
                            quantity INTEGER NOT NULL,
                            instructions TEXT,
                            refills INTEGER DEFAULT 0,
                            is_controlled BOOLEAN DEFAULT 0,
                            status VARCHAR(20) DEFAULT 'Active',
                            prescribed_date DATE NOT NULL,
                            prescribed_time TIME NOT NULL,
                            expires_date DATE,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        )
                    """)
                    
                    # Copy data
                    cur.execute("""
                        INSERT INTO prescriptions_temp SELECT * FROM prescriptions
                    """)
                    
                    # Drop old table
                    cur.execute("DROP TABLE prescriptions")
                    
                    # Rename temp table
                    cur.execute("ALTER TABLE prescriptions_temp RENAME TO prescriptions")
                    
                    conn.commit()
                    print("Prescriptions table updated successfully!")
                    break
            
    except Exception as e:
        app.logger.error(f"Error updating prescriptions table: {e}")
    finally:
        cur.close()
        conn.close()
        
        
def create_tables():
    """Create all necessary tables if they don't exist."""
    queries = {
                
        "patients": """
            CREATE TABLE IF NOT EXISTS patients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_id VARCHAR(50) UNIQUE NOT NULL,
                hospital_no VARCHAR(50) UNIQUE NOT NULL,
                card_number VARCHAR(50) UNIQUE,
                first_name VARCHAR(100) NOT NULL,
                last_name VARCHAR(100) NOT NULL,
                date_of_birth DATE NOT NULL,
                gender VARCHAR(10) NOT NULL,
                blood_group VARCHAR(5),
                genotype VARCHAR(5),
                phone VARCHAR(20),
                email VARCHAR(100),
                address TEXT,
                emergency_contact_name VARCHAR(100),
                emergency_contact_phone VARCHAR(20),
                occupation VARCHAR(100),
                marital_status VARCHAR(20),
                nationality VARCHAR(50),
                religion VARCHAR(50),
                next_of_kin VARCHAR(100),
                next_of_kin_phone VARCHAR(20),
                registration_date DATE NOT NULL,
                registration_time TIME NOT NULL,
                status VARCHAR(20) DEFAULT 'Active',
                profile_photo TEXT,
                notes TEXT,
                created_by INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "appointments": """
            CREATE TABLE IF NOT EXISTS appointments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                appointment_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                doctor_name VARCHAR(100),
                department VARCHAR(100),
                appointment_date DATE NOT NULL,
                appointment_time TIME NOT NULL,
                purpose TEXT,
                status VARCHAR(20) DEFAULT 'Scheduled',
                priority VARCHAR(20) DEFAULT 'Normal',
                notes TEXT,
                created_by INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "queue": """
            CREATE TABLE IF NOT EXISTS queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                token_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                service_type VARCHAR(100) NOT NULL,
                priority VARCHAR(20) DEFAULT 'Normal',
                status VARCHAR(20) DEFAULT 'Waiting',
                queue_position INTEGER,
                assigned_to VARCHAR(100),
                called_at TIMESTAMP,
                called_by VARCHAR(100),
                completed_at TIMESTAMP,
                waiting_time INTEGER,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "patient_services": """
            CREATE TABLE IF NOT EXISTS patient_services (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                service_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                service_type VARCHAR(100) NOT NULL,
                check_in_time TIMESTAMP NOT NULL,
                check_out_time TIMESTAMP,
                status VARCHAR(20) DEFAULT 'Active',
                ward_requested BOOLEAN DEFAULT 0,
                ward_type VARCHAR(50),
                is_emergency BOOLEAN DEFAULT 0,
                notes TEXT,
                created_by INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "patient_notes": """
            CREATE TABLE IF NOT EXISTS patient_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                note_type VARCHAR(50),
                note TEXT NOT NULL,
                created_by INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "patient_communication": """
            CREATE TABLE IF NOT EXISTS patient_communication (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                communication_type VARCHAR(50),
                subject VARCHAR(200),
                message TEXT,
                status VARCHAR(20) DEFAULT 'Sent',
                sent_by INTEGER NOT NULL,
                sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "patient_users": """
            CREATE TABLE IF NOT EXISTS patient_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100) NOT NULL,
                role VARCHAR(50) DEFAULT 'Receptionist',
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "pharmacists": """
            CREATE TABLE IF NOT EXISTS pharmacists (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "drugs": """
            CREATE TABLE IF NOT EXISTS drugs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(100) NOT NULL,
                strength VARCHAR(50) NOT NULL,
                unit_price DECIMAL(10, 2) NOT NULL,
                stock_quantity INT NOT NULL,
                expiry_date DATE NOT NULL,
                low_stock_threshold INT DEFAULT 20,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "drug_sales": """
            CREATE TABLE IF NOT EXISTS drug_sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receipt_no VARCHAR(50) UNIQUE NOT NULL,
                patient_name VARCHAR(100),
                patient_id VARCHAR(50),
                items JSON NOT NULL,
                subtotal DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0.00,
                tax DECIMAL(10, 2) DEFAULT 0.00,
                grand_total DECIMAL(10, 2) NOT NULL,
                pharmacist VARCHAR(50) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "receipts": """
            CREATE TABLE IF NOT EXISTS receipts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_name VARCHAR(100),
                patient_id VARCHAR(50),
                subtotal DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0.00,
                tax DECIMAL(10, 2) DEFAULT 0.00,
                total_amount DECIMAL(10, 2) NOT NULL,
                grand_total DECIMAL(10, 2) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "receipt_items": """
            CREATE TABLE IF NOT EXISTS receipt_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receipt_id INT NOT NULL REFERENCES receipts(id),
                drug_name VARCHAR(100) NOT NULL,
                strength VARCHAR(50) NOT NULL,
                quantity INT NOT NULL,
                unit_price DECIMAL(10, 2) NOT NULL
            );
        """,
        "stock_movements": """
            CREATE TABLE IF NOT EXISTS stock_movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                drug_id INT NOT NULL REFERENCES drugs(id),
                movement_type VARCHAR(20) NOT NULL,
                quantity INT NOT NULL,
                user_id INT NOT NULL,
                note TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "billing_users": """
            CREATE TABLE IF NOT EXISTS billing_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "billing_invoice": """
            CREATE TABLE IF NOT EXISTS billing_invoice (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_name VARCHAR(100) NOT NULL,
                service_type VARCHAR(100) NOT NULL,
                amount DECIMAL(10, 2) NOT NULL,
                status VARCHAR(20) DEFAULT 'UNPAID',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "billing_receipt": """
            CREATE TABLE IF NOT EXISTS billing_receipt (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INT NOT NULL REFERENCES billing_invoice(id),
                amount_paid DECIMAL(10, 2) NOT NULL,
                payment_method VARCHAR(50) NOT NULL,
                received_by VARCHAR(50) NOT NULL,
                payment_date TIMESTAMP NOT NULL
            );
        """,
        "payments": """
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_name VARCHAR(100) NOT NULL,
                service_type VARCHAR(100) NOT NULL,
                subtotal DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0.00,
                tax DECIMAL(10, 2) DEFAULT 0.00,
                grand_total DECIMAL(10, 2) NOT NULL,
                amount_paid DECIMAL(10, 2) NOT NULL,
                balance DECIMAL(10, 2) NOT NULL,
                payment_method VARCHAR(50) NOT NULL,
                status VARCHAR(20) NOT NULL,
                payment_date DATE NOT NULL,
                recorded_by INT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "users": """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "hr_users": """
            CREATE TABLE IF NOT EXISTS hr_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100) NOT NULL,
                email VARCHAR(100),
                role VARCHAR(50) DEFAULT 'HR Staff',
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "departments": """
            CREATE TABLE IF NOT EXISTS departments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(100) NOT NULL,
                code VARCHAR(20) UNIQUE NOT NULL,
                description TEXT,
                head_of_dept VARCHAR(100),
                status VARCHAR(20) DEFAULT 'Active',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "staff": """
            CREATE TABLE IF NOT EXISTS staff (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id VARCHAR(50) UNIQUE NOT NULL,
                first_name VARCHAR(100) NOT NULL,
                last_name VARCHAR(100) NOT NULL,
                department_id INTEGER REFERENCES departments(id),
                position VARCHAR(100) NOT NULL,
                employment_type VARCHAR(50),
                email VARCHAR(100),
                phone VARCHAR(20),
                hire_date DATE NOT NULL,
                salary DECIMAL(12, 2),
                status VARCHAR(20) DEFAULT 'Active',
                emergency_contact VARCHAR(100),
                address TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "attendance": """
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id),
                date DATE NOT NULL,
                check_in TIME,
                check_out TIME,
                status VARCHAR(20),
                remarks TEXT,
                recorded_by INTEGER REFERENCES hr_users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "leaves": """
            CREATE TABLE IF NOT EXISTS leaves (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id),
                leave_type VARCHAR(50) NOT NULL,
                start_date DATE NOT NULL,
                end_date DATE NOT NULL,
                days_requested INTEGER NOT NULL,
                reason TEXT,
                status VARCHAR(20) DEFAULT 'Pending',
                approved_by INTEGER REFERENCES hr_users(id),
                approved_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "schedules": """
            CREATE TABLE IF NOT EXISTS schedules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id),
                schedule_date DATE NOT NULL,
                shift_type VARCHAR(50),
                start_time TIME NOT NULL,
                end_time TIME NOT NULL,
                location VARCHAR(100),
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "payroll": """
            CREATE TABLE IF NOT EXISTS payroll (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id),
                pay_period VARCHAR(50),
                basic_salary DECIMAL(12, 2),
                allowances DECIMAL(12, 2),
                deductions DECIMAL(12, 2),
                net_salary DECIMAL(12, 2),
                status VARCHAR(20) DEFAULT 'Pending',
                payment_date DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "documents": """
            CREATE TABLE IF NOT EXISTS documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id),
                document_type VARCHAR(50),
                document_name VARCHAR(255),
                file_path VARCHAR(500),
                uploaded_by INTEGER REFERENCES hr_users(id),
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "shift_swap_requests": """
            CREATE TABLE IF NOT EXISTS shift_swap_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                schedule_id INTEGER NOT NULL REFERENCES schedules(id) ON DELETE CASCADE,
                from_staff_id INTEGER NOT NULL REFERENCES staff(id) ON DELETE CASCADE,
                to_staff_id INTEGER NOT NULL REFERENCES staff(id) ON DELETE CASCADE,
                reason TEXT,
                status VARCHAR(20) DEFAULT 'Pending',
                requested_by INTEGER REFERENCES hr_users(id),
                requested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                approved_by INTEGER REFERENCES hr_users(id),
                approved_at TIMESTAMP,
                reviewed_by INTEGER REFERENCES hr_users(id),
                reviewed_at TIMESTAMP,
                rejection_reason TEXT,
                notes TEXT
            );
        """,
        "lab_tests": """
            CREATE TABLE IF NOT EXISTS lab_tests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                test_id VARCHAR(50) UNIQUE NOT NULL,
                test_name VARCHAR(100) NOT NULL,
                category VARCHAR(50),
                price DECIMAL(10, 2) NOT NULL,
                description TEXT,
                preparation_instructions TEXT,
                normal_range VARCHAR(100),
                turnaround_time VARCHAR(50),
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "lab_requests": """
            CREATE TABLE IF NOT EXISTS lab_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_no VARCHAR(50) UNIQUE NOT NULL,
                patient_name VARCHAR(100) NOT NULL,
                patient_id VARCHAR(50),
                patient_age INTEGER,
                patient_gender VARCHAR(10),
                doctor_name VARCHAR(100),
                test_ids TEXT,
                test_names TEXT,
                total_amount DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0,
                grand_total DECIMAL(10, 2) NOT NULL,
                amount_paid DECIMAL(10, 2) DEFAULT 0,
                balance DECIMAL(10, 2) DEFAULT 0,
                status VARCHAR(20) DEFAULT 'Pending',
                priority VARCHAR(20) DEFAULT 'Normal',
                requested_date DATE NOT NULL,
                collected_date DATE,
                collected_by VARCHAR(100),
                completed_date DATE,
                completed_by VARCHAR(100),
                approved_by VARCHAR(100),
                notes TEXT,
                created_by INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "lab_results": """
            CREATE TABLE IF NOT EXISTS lab_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL REFERENCES lab_requests(id),
                test_id INTEGER NOT NULL REFERENCES lab_tests(id),
                test_name VARCHAR(100) NOT NULL,
                result_value VARCHAR(200),
                normal_range VARCHAR(100),
                unit VARCHAR(50),
                interpretation VARCHAR(20),
                remarks TEXT,
                performed_by VARCHAR(100),
                performed_date DATE,
                verified_by VARCHAR(100),
                verified_date DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "lab_users": """
            CREATE TABLE IF NOT EXISTS lab_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100) NOT NULL,
                role VARCHAR(50) DEFAULT 'Lab Technician',
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
       

        "consultations": """
            CREATE TABLE IF NOT EXISTS consultations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                consultation_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                doctor_id INTEGER NOT NULL,
                doctor_name VARCHAR(100) NOT NULL,
                consultation_date DATE NOT NULL,
                consultation_time TIME NOT NULL,
                status VARCHAR(20) DEFAULT 'In Progress',
                chief_complaint TEXT,
                history_of_present_illness TEXT,
                past_medical_history TEXT,
                family_history TEXT,
                social_history TEXT,
                review_of_systems TEXT,
                physical_exam TEXT,
                diagnosis TEXT,
                treatment_plan TEXT,
                follow_up_date DATE,
                follow_up_instructions TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                completed_at TIMESTAMP,
                FOREIGN KEY (patient_id) REFERENCES patients(id)
            );
        """,

        "clinical_notes": """
            CREATE TABLE IF NOT EXISTS clinical_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                note_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                doctor_id INTEGER NOT NULL,
                doctor_name VARCHAR(100) NOT NULL,
                note_type VARCHAR(50), -- SOAP, Progress, Discharge, Procedure
                subjective TEXT,
                objective TEXT,
                assessment TEXT,
                plan TEXT,
                note_date DATE NOT NULL,
                note_time TIME NOT NULL,
                status VARCHAR(20) DEFAULT 'Draft',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "prescriptions": """
            CREATE TABLE IF NOT EXISTS prescriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                prescription_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                doctor_id INTEGER NOT NULL,
                doctor_name VARCHAR(100) NOT NULL,
                drug_id INTEGER REFERENCES drugs(id),
                drug_name VARCHAR(100) NOT NULL,
                strength VARCHAR(50),
                dosage VARCHAR(100) NOT NULL,
                frequency VARCHAR(100) NOT NULL,
                duration VARCHAR(50) NOT NULL,
                quantity INTEGER NOT NULL,
                instructions TEXT,
                refills INTEGER DEFAULT 0,
                is_controlled BOOLEAN DEFAULT 0,
                status VARCHAR(20) DEFAULT 'Active',
                prescribed_date DATE NOT NULL,
                prescribed_time TIME NOT NULL,
                expires_date DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "doctor_queue": """
            CREATE TABLE IF NOT EXISTS doctor_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                queue_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                doctor_id INTEGER NOT NULL,
                doctor_name VARCHAR(100) NOT NULL,
                priority VARCHAR(20) DEFAULT 'Normal',
                status VARCHAR(20) DEFAULT 'Waiting',
                wait_time_minutes INTEGER,
                consultation_type VARCHAR(50), -- New, Follow-up, Review, Emergency
                token_no VARCHAR(50),
                called_at TIMESTAMP,
                started_at TIMESTAMP,
                completed_at TIMESTAMP,
                vitals_recorded BOOLEAN DEFAULT 0,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "lab_orders": """
            CREATE TABLE IF NOT EXISTS lab_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                doctor_id INTEGER NOT NULL,
                doctor_name VARCHAR(100) NOT NULL,
                test_name VARCHAR(100) NOT NULL,
                urgency VARCHAR(20) DEFAULT 'Routine',
                status VARCHAR(20) DEFAULT 'Pending', -- Pending, Collected, Processing, Completed
                ordered_date DATE NOT NULL,
                ordered_time TIME NOT NULL,
                collected_date DATE,
                collected_by VARCHAR(100),
                result_date DATE,
                result_value TEXT,
                normal_range VARCHAR(100),
                interpretation VARCHAR(50),
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "imaging_orders": """
            CREATE TABLE IF NOT EXISTS imaging_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                doctor_id INTEGER NOT NULL,
                doctor_name VARCHAR(100) NOT NULL,
                imaging_type VARCHAR(50) NOT NULL, -- X-Ray, CT, MRI, Ultrasound, etc.
                body_part VARCHAR(100) NOT NULL,
                urgency VARCHAR(20) DEFAULT 'Routine',
                clinical_indication TEXT,
                status VARCHAR(20) DEFAULT 'Pending',
                ordered_date DATE NOT NULL,
                ordered_time TIME NOT NULL,
                performed_date DATE,
                report_date DATE,
                report_finding TEXT,
                report_impression TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "referrals": """
            CREATE TABLE IF NOT EXISTS referrals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                referral_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                referring_doctor_id INTEGER NOT NULL,
                referring_doctor_name VARCHAR(100) NOT NULL,
                referred_to VARCHAR(100) NOT NULL,
                referred_to_doctor VARCHAR(100),
                specialization VARCHAR(100) NOT NULL,
                reason TEXT NOT NULL,
                urgency VARCHAR(20) DEFAULT 'Normal',
                status VARCHAR(20) DEFAULT 'Pending',
                referral_date DATE NOT NULL,
                referral_time TIME NOT NULL,
                acceptance_date DATE,
                rejection_reason TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "vitals": """
            CREATE TABLE IF NOT EXISTS vitals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vitals_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER NOT NULL REFERENCES patients(id),
                patient_name VARCHAR(200) NOT NULL,
                recorded_by VARCHAR(100) NOT NULL,
                recorded_date DATE NOT NULL,
                recorded_time TIME NOT NULL,
                temperature DECIMAL(4,1),
                blood_pressure_systolic INTEGER,
                blood_pressure_diastolic INTEGER,
                heart_rate INTEGER,
                respiratory_rate INTEGER,
                oxygen_saturation INTEGER,
                weight DECIMAL(5,2),
                height DECIMAL(5,2),
                bmi DECIMAL(4,1),
                pain_score INTEGER,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "clinical_users": """
            CREATE TABLE IF NOT EXISTS clinical_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100) NOT NULL,
                email VARCHAR(100),
                specialization VARCHAR(100),
                license_number VARCHAR(50),
                role VARCHAR(50) DEFAULT 'Doctor',
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "radiology_users": """
            CREATE TABLE IF NOT EXISTS radiology_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100) NOT NULL,
                role VARCHAR(50) DEFAULT 'Technologist',
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,

        "imaging_orders": """
            CREATE TABLE IF NOT EXISTS imaging_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no VARCHAR(50) UNIQUE NOT NULL,
                patient_id INTEGER,
                patient_name VARCHAR(200) NOT NULL,
                doctor_id INTEGER NOT NULL,
                doctor_name VARCHAR(100) NOT NULL,
                imaging_type VARCHAR(100) NOT NULL,
                body_part VARCHAR(100) NOT NULL,
                urgency VARCHAR(20) DEFAULT 'Routine',
                clinical_indication TEXT,
                status VARCHAR(20) DEFAULT 'Pending',
                ordered_date DATE NOT NULL,
                ordered_time TIME NOT NULL,
                performed_date DATE,
                performed_by VARCHAR(100),
                report TEXT,
                report_date DATE,
                report_approved_by VARCHAR(100),
                images_stored TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "management_users": """
            CREATE TABLE IF NOT EXISTS management_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100) NOT NULL,
                role VARCHAR(50) DEFAULT 'Manager',
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """
    }

    conn = get_db_connection()
    if not conn:
        return

    cursor = conn.cursor()
    for table, query in queries.items():
        try:
            cursor.execute(query)
        except Exception as e:
            app.logger.error(f"Error creating table {table}: {e}")
    
    # Create indexes for better performance
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_staff_department ON staff(department_id);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_staff_date ON attendance(staff_id, date);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_staff_status ON leaves(staff_id, status);",
        "CREATE INDEX IF NOT EXISTS idx_schedules_staff_date ON schedules(staff_id, schedule_date);",
        "CREATE INDEX IF NOT EXISTS idx_payroll_staff_period ON payroll(staff_id, pay_period);",
        "CREATE INDEX IF NOT EXISTS idx_staff_status ON staff(status);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(date);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_status ON leaves(status);",
        "CREATE INDEX IF NOT EXISTS idx_shift_swap_status ON shift_swap_requests(status);",
        "CREATE INDEX IF NOT EXISTS idx_shift_swap_schedule ON shift_swap_requests(schedule_id);",
        "CREATE INDEX IF NOT EXISTS idx_lab_requests_status ON lab_requests(status);",
        "CREATE INDEX IF NOT EXISTS idx_lab_requests_patient ON lab_requests(patient_name);",
        "CREATE INDEX IF NOT EXISTS idx_lab_results_request ON lab_results(request_id);"
    ]
    
    for index_query in indexes:
        try:
            cursor.execute(index_query)
        except Exception as e:
            app.logger.warning(f"Could not create index: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()

def create_default_users():
    """Create default users for all modules."""
    default_users = {
        "pharmacists": ("pharmacist1", "pharma123"),
        "billing_users": ("billing1", "billing123"),
        "lab_users": ("labtech1", "lab123"),
        "patient_users": ("reception1", "reception123")
    }
    # Note: clinical_users is handled separately in create_default_clinical_user()

    conn = get_db_connection()
    if not conn:
        return

    cursor = conn.cursor()
    for table, (username, password) in default_users.items():
        hashed_pw = generate_password_hash(password)
        try:
            if table == "lab_users":
                cursor.execute(f"""
                    INSERT OR REPLACE INTO {table} (username, password, full_name, role)
                    VALUES (?, ?, ?, ?)
                """, (username, hashed_pw, 'Lab Technician', 'Lab Technician'))
            elif table == "patient_users":
                cursor.execute(f"""
                    INSERT OR REPLACE INTO {table} (username, password, full_name, role, is_active)
                    VALUES (?, ?, ?, ?, 1)
                """, (username, hashed_pw, 'Receptionist', 'Receptionist'))
            else:
                cursor.execute(f"""
                    INSERT OR REPLACE INTO {table} (username, password)
                    VALUES (?, ?)
                """, (username, hashed_pw))
        except Exception as e:
            app.logger.error(f"Error creating default user {username}: {e}")
    conn.commit()
    cursor.close()
    conn.close()
    
        
# -------------------- HELPER FUNCTIONS --------------------
def format_currency(amount):
    """Format amount as Nigerian Naira currency."""
    return f"₦{amount:,.2f}"

@app.route("/hr/leave/calendar")
def leave_calendar():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    month = request.args.get("month", date.today().month, type=int)
    year = request.args.get("year", date.today().year, type=int)
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT 
                l.start_date,
                l.end_date,
                l.leave_type,
                l.status,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department
            FROM leaves l
            JOIN staff s ON l.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE l.status IN ('Approved', 'Pending')
                AND (l.start_date BETWEEN ? AND ? OR l.end_date BETWEEN ? AND ?)
            ORDER BY l.start_date
        """, (start_date.isoformat(), end_date.isoformat(), start_date.isoformat(), end_date.isoformat()))
        
        leaves = cur.fetchall()
        
        # Group leaves by date for calendar
        calendar_data = {}
        for leave in leaves:
            # Parse dates if they are strings
            start = leave[0]
            end = leave[1]
            
            if isinstance(start, str):
                start = datetime.strptime(start, '%Y-%m-%d').date()
            if isinstance(end, str):
                end = datetime.strptime(end, '%Y-%m-%d').date()
            
            current = max(start, start_date)
            end_date_limit = min(end, end_date)
            
            while current <= end_date_limit:
                date_str = current.strftime("%Y-%m-%d")
                if date_str not in calendar_data:
                    calendar_data[date_str] = []
                
                calendar_data[date_str].append({
                    'type': leave[2],
                    'status': leave[3],
                    'staff_name': f"{leave[4]} {leave[5]}",
                    'position': leave[6],
                    'department': leave[7] or 'N/A'
                })
                current += timedelta(days=1)
        
        # Get months for navigation
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        
    except Exception as e:
        app.logger.error(f"Error generating leave calendar: {e}")
        calendar_data = {}
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "leave_calendar.html",
        calendar_data=calendar_data,
        month=month,
        year=year,
        months=months,
        years=years,
        start_date=start_date,
        end_date=end_date,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )
    
    
@app.route("/hr/departments/add", methods=["GET", "POST"])
def add_department():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        name = request.form.get("name")
        code = request.form.get("code")
        description = request.form.get("description")
        head_of_dept = request.form.get("head_of_dept")
        status = request.form.get("status", "Active")
        
        # Validate required fields
        if not all([name, code]):
            flash("Department name and code are required", "danger")
            return redirect(url_for("add_department"))
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Check if code already exists
            cur.execute("SELECT id FROM departments WHERE code = ?", (code.upper(),))
            if cur.fetchone():
                flash(f"Department code '{code}' already exists", "danger")
                return redirect(url_for("add_department"))
            
            # Insert new department
            cur.execute("""
                INSERT INTO departments (name, code, description, head_of_dept, status)
                VALUES (?, ?, ?, ?, ?)
            """, (name, code.upper(), description, head_of_dept, status))
            
            conn.commit()
            flash(f"Department '{name}' added successfully!", "success")
            return redirect(url_for("departments"))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error adding department: {e}")
            flash(f"Error adding department: {str(e)}", "danger")
            return redirect(url_for("add_department"))
            
        finally:
            cur.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get potential department heads (staff members)
        cur.execute("""
            SELECT id, first_name, last_name, position 
            FROM staff 
            WHERE status = 'Active'
            ORDER BY first_name, last_name
        """)
        potential_heads = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error loading form data: {e}")
        potential_heads = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "add_department.html",
        potential_heads=potential_heads,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )



@app.route("/hr/departments")
def departments():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get all departments with detailed statistics
        cur.execute("""
            SELECT 
                d.id,
                d.name,
                d.code,
                d.description,
                d.head_of_dept,
                d.status,
                d.created_at,
                COUNT(DISTINCT s.id) as staff_count,
                COUNT(DISTINCT CASE WHEN s.status = 'Active' THEN s.id END) as active_staff
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            GROUP BY d.id, d.name, d.code, d.description, d.head_of_dept, d.status, d.created_at
            ORDER BY d.name
        """)
        
        departments_list = cur.fetchall()
        
        # Get staff count by department for chart
        cur.execute("""
            SELECT 
                d.name,
                COUNT(s.id) as staff_count
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            WHERE d.status = 'Active'
            GROUP BY d.name
            ORDER BY staff_count DESC
        """)
        chart_data = cur.fetchall()
        
        # Statistics
        cur.execute("SELECT COUNT(*) FROM departments WHERE status = 'Active'")
        active_depts = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM staff WHERE department_id IS NOT NULL")
        staff_with_dept = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COALESCE(AVG(staff_count), 0)
            FROM (
                SELECT COUNT(*) as staff_count 
                FROM staff 
                WHERE department_id IS NOT NULL 
                GROUP BY department_id
            ) as dept_stats
        """)
        avg_staff_per_dept = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error fetching departments: {e}")
        departments_list = []
        chart_data = []
        active_depts = 0
        staff_with_dept = 0
        avg_staff_per_dept = 0
    
    finally:
        cur.close()
        conn.close()
    
    # Format departments for template - FIXED DATE CONVERSION
    formatted_depts = []
    for dept in departments_list:
        # Convert created_at to datetime if it's a string
        created_at = dept[6]
        if created_at:
            if isinstance(created_at, str):
                try:
                    created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                except:
                    try:
                        created_at = datetime.strptime(created_at, '%Y-%m-%d')
                    except:
                        # If all fails, use current date as fallback
                        created_at = date.today()
            elif isinstance(created_at, (date, datetime)):
                # It's already a date/datetime object
                pass
            else:
                created_at = date.today()
        else:
            created_at = date.today()
        
        formatted_depts.append({
            'id': dept[0],
            'name': dept[1],
            'code': dept[2],
            'description': dept[3] or 'No description',
            'head_of_dept': dept[4] or 'Not assigned',
            'status': dept[5],
            'created_at': created_at,
            'staff_count': dept[7] or 0,
            'active_staff': dept[8] or 0
        })
    
    return render_template(
        "departments.html",
        departments=formatted_depts,
        chart_data=chart_data,
        active_depts=active_depts,
        staff_with_dept=staff_with_dept,
        avg_staff_per_dept=round(avg_staff_per_dept, 1),
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )    

@app.route("/hr/departments/<int:department_id>")
def view_department(department_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get department details
        cur.execute("""
            SELECT 
                d.id,
                d.name,
                d.code,
                d.description,
                d.head_of_dept,
                d.status,
                d.created_at,
                COUNT(DISTINCT s.id) as total_staff,
                COUNT(DISTINCT CASE WHEN s.status = 'Active' THEN s.id END) as active_staff
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            WHERE d.id = ?
            GROUP BY d.id, d.name, d.code, d.description, d.head_of_dept, d.status, d.created_at
        """, (department_id,))
        
        department = cur.fetchone()
        
        if not department:
            flash("Department not found", "danger")
            return redirect(url_for("departments"))
        
        # Get staff in this department
        cur.execute("""
            SELECT 
                s.id,
                s.staff_id,
                s.first_name,
                s.last_name,
                s.position,
                s.email,
                s.phone,
                s.status,
                s.hire_date,
                s.salary
            FROM staff s
            WHERE s.department_id = ?
            ORDER BY s.first_name, s.last_name
        """, (department_id,))
        
        staff_list = cur.fetchall()
        
        # Convert created_at to datetime
        created_at = department[6]
        if isinstance(created_at, str):
            try:
                created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
            except:
                try:
                    created_at = datetime.strptime(created_at, '%Y-%m-%d')
                except:
                    created_at = date.today()
        
        dept_dict = {
            'id': department[0],
            'name': department[1],
            'code': department[2],
            'description': department[3] or 'No description provided',
            'head_of_dept': department[4] or 'Not assigned',
            'status': department[5],
            'created_at': created_at,
            'total_staff': department[7] or 0,
            'active_staff': department[8] or 0
        }
        
        # Format staff list
        formatted_staff = []
        for staff in staff_list:
            formatted_staff.append({
                'id': staff[0],
                'staff_id': staff[1],
                'first_name': staff[2],
                'last_name': staff[3],
                'name': f"{staff[2]} {staff[3]}",
                'position': staff[4] or 'N/A',
                'email': staff[5] or 'N/A',
                'phone': staff[6] or 'N/A',
                'status': staff[7],
                'hire_date': staff[8],
                'salary': float(staff[9]) if staff[9] else 0
            })
        
    except Exception as e:
        app.logger.error(f"Error fetching department details: {e}")
        flash(f"Error loading department details: {str(e)}", "danger")
        return redirect(url_for("departments"))
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "view_department.html",
        department=dept_dict,
        staff_list=formatted_staff,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )

@app.route("/hr/departments/edit/<int:department_id>", methods=["GET", "POST"])
def edit_department(department_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == "POST":
        name = request.form.get("name")
        code = request.form.get("code")
        description = request.form.get("description")
        head_of_dept = request.form.get("head_of_dept")
        status = request.form.get("status")
        
        if not all([name, code]):
            flash("Department name and code are required", "danger")
            return redirect(url_for("edit_department", department_id=department_id))
        
        try:
            # Check if code exists for other departments
            cur.execute("""
                SELECT id FROM departments 
                WHERE code = ? AND id != ?
            """, (code.upper(), department_id))
            
            if cur.fetchone():
                flash(f"Department code '{code}' already exists", "danger")
                return redirect(url_for("edit_department", department_id=department_id))
            
            # Update department
            cur.execute("""
                UPDATE departments 
                SET name = ?,
                    code = ?,
                    description = ?,
                    head_of_dept = ?,
                    status = ?
                WHERE id = ?
            """, (name, code.upper(), description, head_of_dept, status, department_id))
            
            conn.commit()
            flash("Department updated successfully!", "success")
            return redirect(url_for("departments"))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error updating department: {e}")
            flash(f"Error updating department: {str(e)}", "danger")
            return redirect(url_for("edit_department", department_id=department_id))
    
    # GET request - load department data
    try:
        cur.execute("SELECT * FROM departments WHERE id = ?", (department_id,))
        department = cur.fetchone()
        
        if not department:
            flash("Department not found", "danger")
            return redirect(url_for("departments"))
        
        # Get potential department heads (active staff)
        cur.execute("""
            SELECT 
                s.id,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department_name
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.status = 'Active'
            ORDER BY s.first_name, s.last_name
        """)
        potential_heads = cur.fetchall()
        
        # Format potential heads for template
        formatted_heads = []
        for head in potential_heads:
            formatted_heads.append({
                'id': head[0],
                'name': f"{head[1]} {head[2]}",
                'position': head[3],
                'department': head[4] or 'No Department',
                'full_title': f"{head[1]} {head[2]} - {head[3]} ({head[4] or 'No Dept'})"
            })
        
    except Exception as e:
        app.logger.error(f"Error loading department for edit: {e}")
        flash("Error loading department details", "danger")
        return redirect(url_for("departments"))
    
    finally:
        cur.close()
        conn.close()
    
    # Format department data
    dept_dict = {
        'id': department[0],
        'name': department[1],
        'code': department[2],
        'description': department[3] or '',
        'head_of_dept': department[4] or '',
        'status': department[5],
        'created_at': department[6]
    }
    
    return render_template(
        "edit_department.html",
        department=dept_dict,
        potential_heads=formatted_heads,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )
@app.route("/hr/leave/export")
def export_leave_report():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get all leave data for export
        cur.execute("""
            SELECT 
                l.id,
                s.staff_id as employee_id,
                s.first_name || ' ' || s.last_name as staff_name,
                d.name as department,
                s.position,
                l.leave_type,
                l.start_date,
                l.end_date,
                l.days_requested,
                l.reason,
                l.status,
                l.created_at,
                l.approved_at,
                hu.username as approved_by
            FROM leaves l
            JOIN staff s ON l.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN hr_users hu ON l.approved_by = hu.id
            ORDER BY l.created_at DESC
        """)
        
        leaves = cur.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leave Report"
        
        # Add headers
        headers = [
            "Leave ID", "Employee ID", "Staff Name", "Department", "Position",
            "Leave Type", "Start Date", "End Date", "Days Requested",
            "Reason", "Status", "Requested Date", "Approved Date", "Approved By"
        ]
        ws.append(headers)
        
        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for leave in leaves:
            ws.append([
                leave[0],  # id
                leave[1],  # employee_id
                leave[2],  # staff_name
                leave[3] or 'N/A',  # department
                leave[4] or 'N/A',  # position
                leave[5],  # leave_type
                leave[6],  # start_date
                leave[7],  # end_date
                leave[8],  # days_requested
                leave[9] or '',  # reason
                leave[10],  # status
                leave[11] if leave[11] else '',  # created_at
                leave[12] if leave[12] else '',  # approved_at
                leave[13] or ''  # approved_by
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        # Generate filename
        filename = f"leave_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error exporting leave report: {e}")
        flash(f"Error exporting report: {str(e)}", "danger")
        return redirect(url_for("leave_management"))
        
    finally:
        cur.close()
        conn.close()
         
@app.route("/hr/attendance/report")
def attendance_report():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get report parameters
    report_type = request.args.get("type", "daily")
    month = request.args.get("month", date.today().month, type=int)
    year = request.args.get("year", date.today().year, type=int)
    department_id = request.args.get("department_id", "")
    staff_id = request.args.get("staff_id", "")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        if report_type == "daily":
            # Daily report for selected date
            report_date = request.args.get("date", date.today().strftime("%Y-%m-%d"))
            start_date = end_date = datetime.strptime(report_date, "%Y-%m-%d").date()
            
        elif report_type == "weekly":
            # Weekly report
            week_start = request.args.get("week_start")
            if week_start:
                start_date = datetime.strptime(week_start, "%Y-%m-%d").date()
            else:
                today = date.today()
                start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            
        else:  # monthly
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # Build query
        query = """
            SELECT 
                a.date,
                s.first_name || ' ' || s.last_name as staff_name,
                s.position,
                d.name as department,
                a.check_in,
                a.check_out,
                a.status,
                a.remarks,
                ROUND((julianday(a.check_out) - julianday(a.check_in)) * 24, 1) as hours_worked
            FROM attendance a
            JOIN staff s ON a.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE a.date BETWEEN ? AND ?
        """
        params = [start_date.isoformat(), end_date.isoformat()]
        
        if department_id:
            query += " AND s.department_id = ?"
            params.append(department_id)
        
        if staff_id:
            query += " AND a.staff_id = ?"
            params.append(staff_id)
        
        query += " ORDER BY a.date DESC, d.name, s.first_name"
        
        cur.execute(query, params)
        report_data = cur.fetchall()
        
        # Calculate summary statistics
        cur.execute("""
            SELECT 
                COUNT(*) as total_records,
                SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) as present,
                SUM(CASE WHEN status = 'Late' THEN 1 ELSE 0 END) as late,
                SUM(CASE WHEN status = 'Absent' THEN 1 ELSE 0 END) as absent,
                SUM(CASE WHEN status = 'Half Day' THEN 1 ELSE 0 END) as half_day,
                SUM(CASE WHEN status = 'Holiday' THEN 1 ELSE 0 END) as holiday
            FROM attendance
            WHERE date BETWEEN ? AND ?
        """, (start_date.isoformat(), end_date.isoformat()))
        
        summary = cur.fetchone()
        
        # Get departments for filter
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get staff for filter
        cur.execute("""
            SELECT id, first_name, last_name 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        # Get months for dropdown
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        
    except Exception as e:
        app.logger.error(f"Error generating attendance report: {e}")
        report_data = []
        summary = (0, 0, 0, 0, 0, 0)
        departments = []
        staff_list = []
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        start_date = date.today()
        end_date = date.today()
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "attendance_report.html",
        report_data=report_data,
        summary=summary,
        report_type=report_type,
        departments=departments,
        staff_list=staff_list,
        months=months,
        years=years,
        start_date=start_date,
        end_date=end_date,
        selected_month=month,
        selected_year=year,
        selected_department=department_id,
        selected_staff=staff_id,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )
    
@app.route("/hr/attendance/summary")
def attendance_summary():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get parameters
    period = request.args.get("period", "month")
    month = request.args.get("month", date.today().month, type=int)
    year = request.args.get("year", date.today().year, type=int)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        if period == "month":
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
        else:  # year
            start_date = date(year, 1, 1)
            end_date = date(year, 12, 31)
        
        # Get attendance summary by staff
        cur.execute("""
            SELECT 
                s.id,
                s.first_name || ' ' || s.last_name as staff_name,
                s.position,
                d.name as department,
                COUNT(CASE WHEN a.status = 'Present' THEN 1 END) as present_days,
                COUNT(CASE WHEN a.status = 'Late' THEN 1 END) as late_days,
                COUNT(CASE WHEN a.status = 'Absent' THEN 1 END) as absent_days,
                COUNT(CASE WHEN a.status = 'Half Day' THEN 1 END) as half_days,
                COUNT(CASE WHEN a.status = 'Holiday' THEN 1 END) as holiday_days,
                COUNT(a.id) as total_days,
                ROUND(AVG((julianday(a.check_out) - julianday(a.check_in)) * 24), 1) as avg_hours
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN attendance a ON s.id = a.staff_id 
                AND a.date BETWEEN ? AND ?
            WHERE s.status = 'Active'
            GROUP BY s.id, s.first_name, s.last_name, s.position, d.name
            ORDER BY d.name, s.first_name
        """, (start_date.isoformat(), end_date.isoformat()))
        
        staff_summary = cur.fetchall()
        
        # Get daily summary
        cur.execute("""
            SELECT 
                a.date,
                COUNT(DISTINCT a.staff_id) as staff_present,
                COUNT(CASE WHEN a.status = 'Present' THEN 1 END) as present,
                COUNT(CASE WHEN a.status = 'Late' THEN 1 END) as late,
                COUNT(CASE WHEN a.status = 'Absent' THEN 1 END) as absent,
                COUNT(CASE WHEN a.status = 'Half Day' THEN 1 END) as half_day
            FROM attendance a
            WHERE a.date BETWEEN ? AND ?
            GROUP BY a.date
            ORDER BY a.date
        """, (start_date.isoformat(), end_date.isoformat()))
        
        daily_summary = cur.fetchall()
        
        # Get months for dropdown
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        
    except Exception as e:
        app.logger.error(f"Error generating attendance summary: {e}")
        staff_summary = []
        daily_summary = []
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "attendance_summary.html",
        staff_summary=staff_summary,
        daily_summary=daily_summary,
        period=period,
        months=months,
        years=years,
        selected_month=month,
        selected_year=year,
        start_date=start_date,
        end_date=end_date,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )
         
@app.route("/hr/departments/export")
def export_departments():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get all departments with statistics
        cur.execute("""
            SELECT 
                d.name,
                d.code,
                d.head_of_dept,
                d.status,
                d.created_at,
                COUNT(s.id) as total_staff,
                SUM(CASE WHEN s.status = 'Active' THEN 1 ELSE 0 END) as active_staff,
                COALESCE(AVG(s.salary), 0) as avg_salary
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            GROUP BY d.id, d.name, d.code, d.head_of_dept, d.status, d.created_at
            ORDER BY d.name
        """)
        
        departments = cur.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Departments Report"
        
        # Add headers
        headers = [
            "Department Name", "Code", "Head of Department", "Status",
            "Created Date", "Total Staff", "Active Staff", "Average Salary (₦)"
        ]
        ws.append(headers)
        
        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for dept in departments:
            ws.append([
                dept[0],
                dept[1],
                dept[2] or 'Not assigned',
                dept[3],
                dept[4] if dept[4] else '',
                dept[5] or 0,
                dept[6] or 0,
                round(dept[7] or 0, 2)
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        # Generate filename
        filename = f"departments_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error exporting departments: {e}")
        flash(f"Error exporting report: {str(e)}", "danger")
        return redirect(url_for("departments"))
        
    finally:
        cur.close()
        conn.close()
@app.route("/hr/reports")
def hr_reports():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get staff list with complete data for staff report
        cur.execute("""
            SELECT 
                s.id,
                s.staff_id,
                s.first_name,
                s.last_name,
                s.position,
                s.employment_type,
                s.email,
                s.phone,
                s.hire_date,
                s.salary,
                s.status,
                s.emergency_contact,
                s.address,
                d.name as department_name,
                d.id as department_id
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.status = 'Active'
            ORDER BY s.first_name, s.last_name
        """)
        staff_list = cur.fetchall()
        
        # Get departments for filters
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get attendance data for today
        today = date.today().isoformat()
        cur.execute("""
            SELECT 
                COUNT(DISTINCT s.id) as total_staff,
                COUNT(CASE WHEN a.status IN ('Present', 'Late') THEN 1 END) as present_count,
                COUNT(CASE WHEN a.status = 'Absent' THEN 1 END) as absent_count,
                COUNT(CASE WHEN a.status = 'Late' THEN 1 END) as late_count
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
            WHERE s.status = 'Active'
        """, (today,))
        attendance_today = cur.fetchone()
        
        # Get leave data for current year
        current_year = date.today().year
        cur.execute("""
            SELECT 
                l.id,
                s.first_name || ' ' || s.last_name as staff_name,
                d.name as department,
                l.leave_type,
                l.start_date,
                l.end_date,
                l.days_requested,
                l.status,
                l.approved_by,
                hu.username as approved_by_name
            FROM leaves l
            JOIN staff s ON l.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN hr_users hu ON l.approved_by = hu.id
            WHERE strftime('%Y', l.start_date) = ?
            ORDER BY l.start_date DESC
            LIMIT 100
        """, (str(current_year),))
        leave_data = cur.fetchall()
        
        # Get payroll data
        cur.execute("""
            SELECT 
                s.staff_id,
                s.first_name || ' ' || s.last_name as staff_name,
                d.name as department,
                s.position,
                s.salary as basic_salary,
                COALESCE(SUM(CASE WHEN a.status IN ('Present', 'Late') THEN 1 ELSE 0 END), 0) as days_worked,
                COUNT(DISTINCT a.date) as working_days
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN attendance a ON s.id = a.staff_id 
                AND strftime('%Y-%m', a.date) = strftime('%Y-%m', 'now')
            WHERE s.status = 'Active'
            GROUP BY s.id, s.staff_id, s.first_name, s.last_name, d.name, s.position, s.salary
            ORDER BY d.name, s.first_name
        """)
        payroll_data = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error loading HR reports: {e}")
        staff_list = []
        departments = []
        attendance_today = (0, 0, 0, 0)
        leave_data = []
        payroll_data = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "hr_reports.html",
        staff_list=staff_list,
        departments=departments,
        attendance_today=attendance_today,
        leave_data=leave_data,
        payroll_data=payroll_data,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )
    
def build_stock_snapshot(rows, today):
    stock = []
    for r in rows:
        # Convert row to dict if it's a Row object
        if hasattr(r, 'keys'):
            drug_dict = dict(r)
            expiry_date = drug_dict.get('expiry_date')
            quantity = drug_dict.get('stock_quantity')
            threshold = drug_dict.get('low_stock_threshold') or 20
            drug_id = drug_dict.get('id')
            name = drug_dict.get('name')
            strength = drug_dict.get('strength')
            unit_price = drug_dict.get('unit_price')
        else:
            # Handle tuple
            drug_id, name, strength, quantity, unit_price, expiry_date, threshold = r
            if threshold is None:
                threshold = 20

        if expiry_date:
            expiry_date_obj = datetime.strptime(expiry_date, '%Y-%m-%d').date() if isinstance(expiry_date, str) else expiry_date
            days_left = (expiry_date_obj - today).days
            status = "EXPIRED" if days_left < 0 else "EXPIRING_SOON" if days_left <= 30 else "VALID"
        else:
            days_left = None
            status = "UNKNOWN"

        stock.append({
            "id": drug_id,
            "name": name,
            "strength": strength,
            "quantity": quantity,
            "unit_price": float(unit_price),
            "expiry_date": expiry_date,
            "days_left": days_left,
            "status": status,
            "low_stock_threshold": threshold,
            "total_value": quantity * float(unit_price)
        })
    return stock

def apply_stock_filter(stock, filter_type):
    if filter_type == "expired":
        return [d for d in stock if d["status"] == "EXPIRED"]
    elif filter_type in ("expiring", "expiring_soon"):
        return [d for d in stock if d["status"] == "EXPIRING_SOON"]
    elif filter_type in ("low", "low_stock"):
        return [d for d in stock if d["quantity"] <= d["low_stock_threshold"]]
    return stock

# -------------------- ROUTES: LANDING & MODULES --------------------
@app.route('/')
def landing_page():
    hospital_name = "John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    modules = [
        "System Admin", "Patient Services", "Clinical Services",
        "Pharmacy", "Laboratory", "Radiology", "Billing and Revenue",
        "Human Resources", "Management and Reports"
    ]
    return render_template("dashboard.html", hospital_name=hospital_name, modules=modules)

@app.route('/<module_name>')
def module_placeholder(module_name):
    display_name = module_name.replace('_', ' ').title()
    if module_name.lower() == "pharmacy":
        return redirect(url_for('pharmacy_login'))
    return render_template("module_placeholder.html", module_name=display_name)

# -------------------- ROUTES: PHARMACY MODULE --------------------
@app.route('/pharmacy/login', methods=['GET', 'POST'])
def pharmacy_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return render_template("pharmacy_login.html")

        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, username, password FROM pharmacists WHERE username=? AND is_active=1",
            (username,)
        )
        pharmacist = cursor.fetchone()
        cursor.close()
        conn.close()

        if pharmacist and check_password_hash(pharmacist[2], password):
            session['pharmacist_id'] = pharmacist[0]
            session['pharmacist_username'] = pharmacist[1]
            return redirect(url_for('pharmacy_dashboard'))
        else:
            flash("Invalid username or password", "danger")

    return render_template("pharmacy_login.html")

@app.route('/pharmacy/dashboard')
def pharmacy_dashboard():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))
    return render_template(
        "pharmacy_dashboard.html",
        pharmacist_name=session.get('pharmacist_username')
    )

@app.route('/pharmacy/logout')
def pharmacy_logout():
    session.clear()
    return redirect(url_for('pharmacy_login'))

@app.route('/pharmacy/drug_sales')
def drug_sales():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))
    return render_template(
        "drug_sales_dashboard.html",
        pharmacist_name=session.get('pharmacist_username'),
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    )

@app.route('/pharmacy/add-stock', methods=['GET', 'POST'])
def add_stock():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    if request.method == 'POST':
        drug_name = request.form.get('drug_name', '').strip()
        strength = request.form.get('strength', '').strip()
        unit_price = request.form.get('unit_price', '').strip()
        quantity = request.form.get('quantity', '').strip()
        expiry_date = request.form.get('expiry_date', '').strip()

        if not all([drug_name, strength, unit_price, quantity, expiry_date]):
            flash("All fields including expiry date are required.", "danger")
            return redirect(url_for('add_stock'))

        try:
            unit_price = float(unit_price)
            quantity = int(quantity)
            expiry_date_obj = datetime.strptime(expiry_date, "%Y-%m-%d").date()
        except ValueError:
            flash("Invalid input data.", "danger")
            return redirect(url_for('add_stock'))

        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('add_stock'))

        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, stock_quantity FROM drugs
            WHERE name = ? AND strength = ? AND expiry_date = ?
        """, (drug_name, strength, expiry_date_obj.isoformat()))

        existing = cursor.fetchone()

        if existing:
            cursor.execute("""
                UPDATE drugs
                SET stock_quantity = stock_quantity + ?,
                    unit_price = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (quantity, unit_price, existing[0]))
        else:
            cursor.execute("""
                INSERT INTO drugs (name, strength, unit_price, stock_quantity, expiry_date)
                VALUES (?, ?, ?, ?, ?)
            """, (drug_name, strength, unit_price, quantity, expiry_date_obj.isoformat()))

        conn.commit()
        cursor.close()
        conn.close()

        flash("Stock added successfully.", "success")
        return redirect(url_for('add_stock'))

    return render_template("add_stock.html")

@app.route('/api/drugs')
def api_drugs():
    if 'pharmacist_id' not in session:
        return jsonify([])

    search = request.args.get('q', '').strip()
    conn = get_db_connection()
    if not conn:
        return jsonify([])

    cur = conn.cursor()
    query = """
        SELECT id, name, strength, unit_price, stock_quantity
        FROM drugs
        WHERE stock_quantity > 0
    """
    params = []
    
    if search:
        query += " AND LOWER(name) LIKE LOWER(?)"
        params.append(f"{search}%")
    
    query += " ORDER BY name ASC"
    cur.execute(query, params)
    
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify([{
        "id": r[0], "name": r[1], "strength": r[2],
        "unit_price": float(r[3]), "stock_quantity": r[4]
    } for r in rows])

@app.route('/pharmacy/receipt', methods=['POST'])
def pharmacy_receipt():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    data = request.json
    receipt_no = f"RX-{uuid.uuid4().hex[:8].upper()}"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO drug_sales
        (receipt_no, items, subtotal, discount, tax, grand_total, pharmacist)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        receipt_no,
        json.dumps(data["items"]),
        data["subtotal"],
        data["discount"],
        data["tax"],
        data["grand_total"],
        session.get('pharmacist_username')
    ))

    conn.commit()
    cur.close()
    conn.close()

    data["receipt_no"] = receipt_no
    return render_template(
        "receipt.html",
        receipt=data,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        pharmacist_name=session.get('pharmacist_username')
    )

@app.route('/pharmacy/save-patient', methods=['POST'])
def save_patient_info():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    receipt_no = request.form['receipt_no']
    patient_name = request.form['patient_name']
    patient_id = request.form.get('patient_id')

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE drug_sales
        SET patient_name=?, patient_id=?
        WHERE receipt_no=?
    """, (patient_name, patient_id, receipt_no))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for('reprint_receipt', receipt_no=receipt_no))

@app.route('/pharmacy/receipt/<receipt_no>')
def reprint_receipt(receipt_no):
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT receipt_no, patient_name, patient_id, items,
               subtotal, discount, tax, grand_total, pharmacist, created_at
        FROM drug_sales
        WHERE receipt_no = ?
    """, (receipt_no,))

    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        flash("Receipt not found", "danger")
        return redirect(url_for('pharmacy_dashboard'))

    receipt = {
        "receipt_no": row[0], "patient_name": row[1], "patient_id": row[2],
        "items": json.loads(row[3]) if row[3] else [],
        "subtotal": float(row[4]), "discount": float(row[5]),
        "tax": float(row[6]), "grand_total": float(row[7]),
        "pharmacist": row[8], "date": row[9]
    }

    return render_template("receipt.html", receipt=receipt, hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

@app.route("/pharmacy/confirm-payment", methods=["POST"])
def confirm_payment():
    if "pharmacist_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json()
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO receipts (
                patient_name, patient_id, subtotal, discount, tax,
                total_amount, grand_total, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data.get("patient_name"), data.get("patient_id"),
            data["subtotal"], data["discount"], data["tax"],
            data["grand_total"], data["grand_total"], datetime.now().isoformat()
        ))

        receipt_id = cur.lastrowid

        for item in data["items"]:
            cur.execute("""
                INSERT INTO receipt_items (
                    receipt_id, drug_name, strength, quantity, unit_price
                )
                VALUES (?, ?, ?, ?, ?);
            """, (
                receipt_id, item["drug_name"], item["strength"],
                item["quantity"], item["unit_price"]
            ))

            cur.execute("""
                UPDATE drugs
                SET stock_quantity = stock_quantity - ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE name = ? AND strength = ?;
            """, (
                item["quantity"], item["drug_name"], item["strength"]
            ))

        conn.commit()
        return jsonify({"success": True, "receipt_id": receipt_id})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        cur.close()
        conn.close()

@app.route("/pharmacy/stock-report")
def stock_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    filter_type = request.args.get("filter", "all")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, strength, stock_quantity, unit_price, 
               expiry_date, low_stock_threshold
        FROM drugs
        ORDER BY expiry_date ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    stock = build_stock_snapshot(rows, date.today())
    stock = apply_stock_filter(stock, filter_type)

    return render_template(
        "stock_report.html",
        stock=stock,
        current_filter=filter_type,
        expired_count=sum(1 for d in stock if d["status"] == "EXPIRED"),
        expiring_soon_count=sum(1 for d in stock if d["status"] == "EXPIRING_SOON"),
        low_stock_count=sum(1 for d in stock if d["quantity"] <= d["low_stock_threshold"]),
        total_stock_value=sum(d["total_value"] for d in stock)
    )

@app.route("/pharmacy/stock-report/export")
def export_stock_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    filter_type = request.args.get("filter", "all")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, strength, stock_quantity, unit_price, 
               expiry_date, low_stock_threshold
        FROM drugs
        ORDER BY expiry_date ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    stock = build_stock_snapshot(rows, date.today())
    stock = apply_stock_filter(stock, filter_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    headers = [
        "Drug Name", "Strength", "Quantity", "Unit Price (₦)",
        "Expiry Date", "Days Left", "Status", "Total Value (₦)", "Low Stock Threshold"
    ]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    fills = {
        "EXPIRED": PatternFill("solid", fgColor="FF9999"),
        "EXPIRING_SOON": PatternFill("solid", fgColor="FFFF99"),
        "LOW": PatternFill("solid", fgColor="ADD8E6")
    }

    for item in stock:
        ws.append([
            item["name"], item["strength"], item["quantity"],
            float(item["unit_price"]), item["expiry_date"],
            item["days_left"], item["status"],
            float(item["total_value"]), item["low_stock_threshold"]
        ])

        row_idx = ws.max_row
        if item["status"] in fills:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fills[item["status"]]
        elif item["quantity"] <= item["low_stock_threshold"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fills["LOW"]

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=f"pharmacy_stock_report_{date.today()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/pharmacy/stock-movements')
def stock_movements():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT sm.id, d.name, d.strength, sm.movement_type, 
               sm.quantity, u.username, sm.created_at, sm.note
        FROM stock_movements sm
        JOIN drugs d ON sm.drug_id = d.id
        JOIN users u ON sm.user_id = u.id
        ORDER BY sm.created_at DESC
    """)
    movements = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('stock_movements.html', movements=movements)

@app.route("/pharmacy/revenue-report", methods=["GET", "POST"])
def revenue_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    report_type = request.form.get("period", "daily")
    selected_day = request.form.get("day")
    selected_month = request.form.get("month")
    selected_year = request.form.get("year")
    today = date.today()

    if report_type == "daily":
        start_date = end_date = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today
    elif report_type == "weekly":
        d = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today
        start_date = d - timedelta(days=d.weekday())
        end_date = start_date + timedelta(days=6)
    elif report_type == "monthly":
        month = int(selected_month) if selected_month else today.month
        year = int(selected_year) if selected_year else today.year
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    else:
        flash("Invalid report period", "danger")
        return redirect(url_for("pharmacy_dashboard"))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, patient_name, patient_id, grand_total, created_at
        FROM receipts
        WHERE DATE(created_at) BETWEEN ? AND ?
        ORDER BY created_at ASC;
    """, (start_date.isoformat(), end_date.isoformat()))
    
    sales_rows = cur.fetchall()
    cur.close()
    conn.close()

    # Process sales to ensure created_at is a datetime object
    sales = []
    for row in sales_rows:
        # row[4] is created_at
        created_at = row[4]
        
        # Convert string to datetime if needed
        if isinstance(created_at, str):
            try:
                created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
            except:
                try:
                    created_at = datetime.strptime(created_at, '%Y-%m-%d')
                except:
                    pass
        
        # Append as tuple to maintain index access in template
        sales.append((
            row[0],  # id
            row[1] or 'N/A',  # patient_name
            row[2] or 'N/A',  # patient_id
            float(row[3]) if row[3] else 0,  # grand_total
            created_at  # created_at as datetime object
        ))

    total_revenue = sum(sale[3] for sale in sales)
    months = [(i, month_name[i]) for i in range(1, 13)]
    years = range(2024, today.year + 1)

    return render_template(
        "revenue_report.html",
        sales=sales,
        total_revenue=total_revenue,
        period=report_type,
        start_date=start_date,
        end_date=end_date,
        selected_day=selected_day,
        selected_month=int(selected_month) if selected_month else today.month,
        selected_year=int(selected_year) if selected_year else today.year,
        months=months,
        years=years
    )
    
@app.route("/receipt/<int:receipt_id>")
def receipt(receipt_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM receipts WHERE id = ?;", (receipt_id,))
    receipt = cur.fetchone()

    cur.execute("""
        SELECT drug_name, strength, quantity, unit_price
        FROM receipt_items
        WHERE receipt_id = ?;
    """, (receipt_id,))
    items = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "receipt.html",
        receipt=receipt,
        items=items
    )

@app.route("/pharmacy/receipt/<int:receipt_id>")
def view_receipt(receipt_id):
    if "pharmacist_id" not in session:
        return redirect(url_for("pharmacy_login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, patient_name, patient_id, subtotal, discount, tax, grand_total, created_at
        FROM receipts
        WHERE id = ?
    """, (receipt_id,))

    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        flash("Receipt not found", "danger")
        return redirect(url_for("pharmacy_dashboard"))

    receipt = {
        "id": row[0],
        "patient_name": row[1],
        "patient_id": row[2],
        "subtotal": float(row[3]),
        "discount": float(row[4]),
        "tax": float(row[5]),
        "grand_total": float(row[6]),
        "date": row[7]
    }

    cur.execute("""
        SELECT drug_name, strength, quantity, unit_price
        FROM receipt_items
        WHERE receipt_id = ?
    """, (receipt_id,))

    items_rows = cur.fetchall()
    items = []
    for i in items_rows:
        items.append({
            "drug_name": i[0],
            "strength": i[1],
            "quantity": i[2],
            "unit_price": float(i[3])
        })

    cur.close()
    conn.close()

    return render_template(
        "receipt.html",
        receipt=receipt,
        items=items,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    )

# -------------------- ROUTES: BILLING MODULE --------------------
@app.route("/billing/login", methods=["GET", "POST"])
def billing_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, password
            FROM billing_users
            WHERE username = ?
        """, (username,))

        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user[1], password):
            session["billing_user_id"] = user[0]
            session["billing_username"] = username
            return redirect(url_for("billing_dashboard"))
        else:
            flash("Invalid login credentials", "danger")

    return render_template("billing_login.html")

@app.route("/billing/dashboard")
def billing_dashboard():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))

    return render_template("billing_dashboard.html")

@app.route("/billing/logout")
def billing_logout():
    session.pop("billing_user_id", None)
    session.pop("billing_username", None)
    flash("Logged out successfully", "success")
    return redirect(url_for("billing_login"))

@app.route("/billing/confirm-payment", methods=["POST"])
def billing_confirm_payment():
    if "billing_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    patient_name = request.form.get("patient_name")
    service_type = request.form.get("service_type")
    receipt_date = request.form.get("receipt_date")
    payment_method = request.form.get("payment_method")
    amount_paid = float(request.form.get("amount_paid", 0))
    vat_percent = float(request.form.get("vat", 0))
    discount = float(request.form.get("discount", 0))

    subtotal = amount_paid
    vat_amount = (subtotal * vat_percent) / 100
    grand_total = subtotal + vat_amount - discount
    balance = 0 if amount_paid >= grand_total else grand_total - amount_paid
    status = "Paid" if balance <= 0 else "Partial"

    payment_date = datetime.strptime(receipt_date, "%Y-%m-%d").date()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO payments (
                patient_name, service_type, subtotal, discount, tax,
                grand_total, amount_paid, balance, payment_method,
                status, payment_date, recorded_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            patient_name, service_type, subtotal, discount, vat_amount,
            grand_total, amount_paid, balance, payment_method,
            status, payment_date.isoformat(), session["billing_user_id"]
        ))

        payment_id = cur.lastrowid
        conn.commit()
        flash(f"Payment recorded successfully. Receipt No: {payment_id}", "success")

        # Redirect to the receipt page
        return redirect(url_for("view_payment_receipt", payment_id=payment_id))

    except Exception as e:
        conn.rollback()
        flash(f"Payment error: {e}", "danger")
        return redirect(url_for("accept_payment_page"))

    finally:
        cur.close()
        conn.close()
        
@app.route("/billing/accept-payment", methods=["GET"])
def accept_payment_page():
    return render_template("accept_payment.html")

@app.route("/billing/receipt/<int:payment_id>")
def view_payment_receipt(payment_id):
    from datetime import datetime
    
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM payments WHERE id=?", (payment_id,))
    payment = cur.fetchone()
    cur.close()
    conn.close()

    if not payment:
        flash("Payment not found", "danger")
        return redirect(url_for("billing_dashboard"))

    return render_template("payment_receipt.html", payment=payment, datetime=datetime)



# -------------------- ROUTES: BILLING MODULE - PAYMENT HISTORY --------------------

@app.route("/billing/payment-history")
def payment_history():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))

    # Get filter parameters
    patient_name = request.args.get("patient_name", "").strip()
    service_type = request.args.get("service_type", "")
    payment_method = request.args.get("payment_method", "")
    status = request.args.get("status", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    
    # Pagination
    page = request.args.get("page", 1, type=int)
    per_page = 20
    
    # Build query
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Base query
    query = "SELECT * FROM payments WHERE 1=1"
    count_query = "SELECT COUNT(*) FROM payments WHERE 1=1"
    params = []
    
    # Apply filters
    if patient_name:
        query += " AND LOWER(patient_name) LIKE LOWER(?)"
        count_query += " AND LOWER(patient_name) LIKE LOWER(?)"
        params.append(f"%{patient_name}%")
    
    if service_type:
        query += " AND service_type = ?"
        count_query += " AND service_type = ?"
        params.append(service_type)
    
    if payment_method:
        query += " AND payment_method = ?"
        count_query += " AND payment_method = ?"
        params.append(payment_method)
    
    if status:
        query += " AND status = ?"
        count_query += " AND status = ?"
        params.append(status)
    
    if start_date:
        query += " AND payment_date >= ?"
        count_query += " AND payment_date >= ?"
        params.append(start_date)
    
    if end_date:
        query += " AND payment_date <= ?"
        count_query += " AND payment_date <= ?"
        params.append(end_date)
    
    # Get total count
    cur.execute(count_query, params)
    total_items = cur.fetchone()[0]
    
    # Apply ordering and pagination
    query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
    offset = (page - 1) * per_page
    params.extend([per_page, offset])
    
    # Execute main query
    cur.execute(query, params)
    payments = cur.fetchall()
    
    # Get unique service types for dropdown
    cur.execute("SELECT DISTINCT service_type FROM payments WHERE service_type IS NOT NULL ORDER BY service_type")
    service_types = [row[0] for row in cur.fetchall()]
    
    # Calculate total amount and convert dates
    total_amount = 0
    formatted_payments = []
    for payment in payments:
        # Convert payment_date to datetime if it's a string
        payment_date = payment[11]
        if isinstance(payment_date, str):
            try:
                payment_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
            except:
                try:
                    payment_date = datetime.strptime(payment_date, '%Y-%m-%d %H:%M:%S').date()
                except:
                    payment_date = payment_date
        
        # Convert created_at to datetime if it's a string
        created_at = payment[13]
        if isinstance(created_at, str):
            try:
                created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
            except:
                created_at = created_at
        
        payment_dict = {
            "id": payment[0],
            "patient_name": payment[1],
            "service_type": payment[2],
            "subtotal": float(payment[3]),
            "discount": float(payment[4]),
            "tax": float(payment[5]),
            "grand_total": float(payment[6]),
            "amount_paid": float(payment[7]),
            "balance": float(payment[8]),
            "payment_method": payment[9],
            "status": payment[10],
            "payment_date": payment_date,
            "created_at": created_at
        }
        formatted_payments.append(payment_dict)
        total_amount += payment_dict["amount_paid"]
    
    cur.close()
    conn.close()
    
    # Calculate pagination
    total_pages = (total_items + per_page - 1) // per_page
    
    return render_template(
        "billing_payment_history.html",
        payments=formatted_payments,
        service_types=service_types,
        total_items=total_items,
        total_amount=total_amount,
        page=page,
        total_pages=total_pages,
        current_filters=request.args
    )


@app.route('/billing/api/today-collection')
def api_today_collection():
    """Return today's total collections and transaction count."""
    if 'billing_user_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    today = date.today().isoformat()
    
    # Sum of amount_paid for today
    cur.execute("SELECT COALESCE(SUM(amount_paid), 0) FROM payments WHERE payment_date = ?", (today,))
    total = cur.fetchone()[0]
    
    # Count of transactions today
    cur.execute("SELECT COUNT(*) FROM payments WHERE payment_date = ?", (today,))
    count = cur.fetchone()[0]
    
    cur.close()
    conn.close()
    
    return jsonify({
        "success": True,
        "total_collected": float(total),
        "transaction_count": count
    })

@app.route("/billing/payment-history/export")
def export_payment_history():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    # Get filter parameters (same as payment_history)
    patient_name = request.args.get("patient_name", "").strip()
    service_type = request.args.get("service_type", "")
    payment_method = request.args.get("payment_method", "")
    status = request.args.get("status", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Build query without pagination
    query = "SELECT * FROM payments WHERE 1=1"
    params = []
    
    if patient_name:
        query += " AND LOWER(patient_name) LIKE LOWER(?)"
        params.append(f"%{patient_name}%")
    
    if service_type:
        query += " AND service_type = ?"
        params.append(service_type)
    
    if payment_method:
        query += " AND payment_method = ?"
        params.append(payment_method)
    
    if status:
        query += " AND status = ?"
        params.append(status)
    
    if start_date:
        query += " AND payment_date >= ?"
        params.append(start_date)
    
    if end_date:
        query += " AND payment_date <= ?"
        params.append(end_date)
    
    query += " ORDER BY created_at DESC"
    cur.execute(query, params)
    payments = cur.fetchall()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment History"
    
    # Add headers
    headers = [
        "Receipt No", "Patient Name", "Service Type", 
        "Subtotal (₦)", "Discount (₦)", "Tax (₦)", "Grand Total (₦)",
        "Amount Paid (₦)", "Balance (₦)", "Payment Method",
        "Status", "Payment Date", "Created At", "Recorded By"
    ]
    ws.append(headers)
    
    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for payment in payments:
        # Format dates for display
        payment_date = payment[11]
        if isinstance(payment_date, str):
            payment_date = payment_date[:10] if len(payment_date) >= 10 else payment_date
        
        created_at = payment[13]
        if isinstance(created_at, str):
            created_at = created_at[:19] if len(created_at) >= 19 else created_at
        
        ws.append([
            payment[0],  # id
            payment[1],  # patient_name
            payment[2],  # service_type
            float(payment[3]),  # subtotal
            float(payment[4]),  # discount
            float(payment[5]),  # tax
            float(payment[6]),  # grand_total
            float(payment[7]),  # amount_paid
            float(payment[8]),  # balance
            payment[9],  # payment_method
            payment[10],  # status
            payment_date,  # payment_date
            created_at,  # created_at
            payment[12]   # recorded_by
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary row
    ws.append([])
    ws.append(["SUMMARY", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    
    if payments:
        total_amount = sum(float(p[7]) for p in payments)
        total_balance = sum(float(p[8]) for p in payments)
        
        summary_headers = ["Total Payments", "Total Amount", "Total Balance"]
        summary_values = [len(payments), total_amount, total_balance]
        
        for i, (header, value) in enumerate(zip(summary_headers, summary_values)):
            ws.append([header, value])
        
        # Style summary
        for i in range(len(summary_headers)):
            ws.cell(row=ws.max_row - len(summary_headers) + i, column=1).font = Font(bold=True)
    
    # Save to BytesIO
    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    cur.close()
    conn.close()
    
    # Generate filename
    filename = f"payment_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -------------------- ROUTES: TODAY'S COLLECTION --------------------

@app.route("/billing/todays-collection")
def todays_collection():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    today = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get today's payments
    cur.execute("""
        SELECT id, patient_name, service_type, amount_paid, 
               payment_method, status, created_at
        FROM payments 
        WHERE DATE(payment_date) = ?
        ORDER BY created_at DESC
    """, (today.isoformat(),))
    
    today_payments = cur.fetchall()
    
    # Calculate totals by payment method
    payment_methods_data = {
        'Cash': {'amount': 0, 'count': 0},
        'Card': {'amount': 0, 'count': 0},
        'Transfer': {'amount': 0, 'count': 0},
        'POS': {'amount': 0, 'count': 0},
        'Insurance': {'amount': 0, 'count': 0},
        'Other': {'amount': 0, 'count': 0}
    }
    
    # Process payments
    total_transactions = len(today_payments)
    grand_total = 0
    amounts = []
    
    recent_transactions = []
    for payment in today_payments:
        amount_paid = float(payment[3])
        payment_method = payment[4]
        
        # Convert created_at string to datetime if needed
        created_at = payment[6]
        if isinstance(created_at, str):
            try:
                created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
            except:
                try:
                    created_at = datetime.strptime(created_at, '%Y-%m-%d')
                except:
                    created_at = created_at
        
        # Add to grand total
        grand_total += amount_paid
        amounts.append(amount_paid)
        
        # Add to payment method totals
        if payment_method in payment_methods_data:
            payment_methods_data[payment_method]['amount'] += amount_paid
            payment_methods_data[payment_method]['count'] += 1
        else:
            payment_methods_data['Other']['amount'] += amount_paid
            payment_methods_data['Other']['count'] += 1
        
        # Prepare recent transactions data
        recent_transactions.append({
            'id': payment[0],
            'patient_name': payment[1],
            'service_type': payment[2],
            'amount_paid': amount_paid,
            'payment_method': payment_method,
            'status': payment[5],
            'created_at': created_at  # Now a datetime object
        })
    
    # Calculate additional statistics
    average_transaction = grand_total / total_transactions if total_transactions > 0 else 0
    highest_transaction = max(amounts) if amounts else 0
    lowest_transaction = min(amounts) if amounts else 0
    
    # Calculate totals for time periods
    morning_total = 0  # 6AM - 12PM
    afternoon_total = 0  # 12PM - 4PM
    evening_total = 0  # 4PM - 10PM
    
    for transaction in recent_transactions:
        created_at = transaction['created_at']
        if created_at and not isinstance(created_at, str):
            hour = created_at.hour
            amount = transaction['amount_paid']
            
            if 6 <= hour < 12:
                morning_total += amount
            elif 12 <= hour < 16:
                afternoon_total += amount
            elif 16 <= hour < 22:
                evening_total += amount
    
    # Prepare payment methods for template
    payment_methods = []
    for method_name, data in payment_methods_data.items():
        if data['count'] > 0:  # Only include methods with transactions
            percentage = (data['amount'] / grand_total * 100) if grand_total > 0 else 0
            payment_methods.append({
                'name': method_name,
                'amount': data['amount'],
                'count': data['count'],
                'percentage': round(percentage, 1)
            })
    
    # Set daily target (you can make this configurable)
    daily_target = 500000.00  # ₦500,000 daily target
    
    cur.close()
    conn.close()
    
    # Format date for display
    today_date = today.strftime("%A, %B %d, %Y")
    
    return render_template(
        "todays_collection.html",
        today_date=today_date,
        grand_total=grand_total,
        cash_total=payment_methods_data['Cash']['amount'],
        card_total=payment_methods_data['Card']['amount'],
        transfer_total=payment_methods_data['Transfer']['amount'],
        pos_total=payment_methods_data['POS']['amount'],
        insurance_total=payment_methods_data['Insurance']['amount'],
        other_total=payment_methods_data['Other']['amount'],
        payment_methods=payment_methods,
        recent_transactions=recent_transactions,
        total_transactions=total_transactions,
        average_transaction=average_transaction,
        highest_transaction=highest_transaction,
        lowest_transaction=lowest_transaction,
        morning_total=morning_total,
        afternoon_total=afternoon_total,
        evening_total=evening_total,
        daily_target=daily_target
    )

@app.route("/billing/todays-collection/export")
def export_todays_collection():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    today = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get today's payments
    cur.execute("""
        SELECT id, patient_name, service_type, subtotal, discount, tax,
               grand_total, amount_paid, balance, payment_method, 
               status, payment_date, created_at
        FROM payments 
        WHERE DATE(payment_date) = ?
        ORDER BY created_at DESC
    """, (today.isoformat(),))
    
    today_payments = cur.fetchall()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Today's Collection - {today}"
    
    # Add title
    ws.append([f"Today's Collection Report - {today.strftime('%B %d, %Y')}"])
    ws.append([])
    
    # Add summary section
    ws.append(["SUMMARY"])
    ws.append([])
    
    # Calculate totals by payment method
    payment_methods_data = {
        'Cash': {'amount': 0, 'count': 0},
        'Card': {'amount': 0, 'count': 0},
        'Transfer': {'amount': 0, 'count': 0},
        'POS': {'amount': 0, 'count': 0},
        'Insurance': {'amount': 0, 'count': 0},
        'Other': {'amount': 0, 'count': 0}
    }
    
    grand_total = 0
    total_transactions = len(today_payments)
    
    for payment in today_payments:
        amount_paid = float(payment[7])
        payment_method = payment[9]
        grand_total += amount_paid
        
        if payment_method in payment_methods_data:
            payment_methods_data[payment_method]['amount'] += amount_paid
            payment_methods_data[payment_method]['count'] += 1
        else:
            payment_methods_data['Other']['amount'] += amount_paid
            payment_methods_data['Other']['count'] += 1
    
    # Write summary
    ws.append(["Total Transactions:", total_transactions])
    ws.append(["Grand Total:", grand_total])
    ws.append([])
    ws.append(["Payment Method Breakdown"])
    ws.append(["Method", "Count", "Amount", "Percentage"])
    
    for method_name, data in payment_methods_data.items():
        if data['count'] > 0:
            percentage = (data['amount'] / grand_total * 100) if grand_total > 0 else 0
            ws.append([
                method_name,
                data['count'],
                data['amount'],
                f"{percentage:.1f}%"
            ])
    
    ws.append([])
    ws.append([])
    
    # Add detailed transactions
    ws.append(["DETAILED TRANSACTIONS"])
    ws.append([])
    
    headers = [
        "Receipt No", "Patient Name", "Service Type", 
        "Subtotal", "Discount", "Tax", "Grand Total",
        "Amount Paid", "Balance", "Payment Method",
        "Status", "Payment Date", "Time"
    ]
    ws.append(headers)
    
    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))
    
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Add data rows
    for payment in today_payments:
        # Parse created_at time
        created_at = payment[12]
        if created_at:
            if isinstance(created_at, str):
                created_at_dt = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
            else:
                created_at_dt = created_at
            time_str = created_at_dt.strftime('%H:%M:%S')
        else:
            time_str = ''
        
        ws.append([
            payment[0],  # id
            payment[1],  # patient_name
            payment[2],  # service_type
            float(payment[3]),  # subtotal
            float(payment[4]),  # discount
            float(payment[5]),  # tax
            float(payment[6]),  # grand_total
            float(payment[7]),  # amount_paid
            float(payment[8]),  # balance
            payment[9],  # payment_method
            payment[10],  # status
            payment[11],  # payment_date
            time_str
        ])
    
    # Apply borders to data rows
    for row in ws.iter_rows(min_row=ws.max_row - len(today_payments) + 1, max_row=ws.max_row):
        for cell in row:
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    cur.close()
    conn.close()
    
    # Save to BytesIO
    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # Generate filename
    filename = f"todays_collection_{today.strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Custom filter for currency formatting
@app.template_filter('currency')
def currency_filter(amount):
    """Format amount as Nigerian Naira currency."""
    if amount is None:
        return "₦0.00"
    return f"₦{float(amount):,.2f}"

# -------------------- ROUTES: HR MODULE --------------------
@app.route("/hr/login", methods=["GET", "POST"])
def hr_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return render_template("hr_login.html", 
                                 hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Enugu State")

        cur = conn.cursor()

        try:
            cur.execute("""
                SELECT id, username, password, full_name, role
                FROM hr_users
                WHERE username = ? AND is_active = 1
            """, (username,))

            user = cur.fetchone()
            cur.close()
            conn.close()

            if user:
                stored_password = user[2]
                
                # Check if password is hashed or plain text
                try:
                    # Try bcrypt check first
                    if check_password_hash(stored_password, password):
                        session["hr_user_id"] = user[0]
                        session["hr_username"] = user[1]
                        session["hr_full_name"] = user[3]
                        session["hr_role"] = user[4]
                        flash(f"Welcome, {user[3]}!", "success")
                        return redirect(url_for("hr_dashboard"))
                    else:
                        flash("Invalid password", "danger")
                except ValueError:
                    # If check_password_hash fails, try direct comparison (for plain text)
                    if stored_password == password:
                        # Upgrade to hashed password
                        hashed = generate_password_hash(password)
                        conn2 = get_db_connection()
                        cur2 = conn2.cursor()
                        cur2.execute("UPDATE hr_users SET password = ? WHERE id = ?", (hashed, user[0]))
                        conn2.commit()
                        cur2.close()
                        conn2.close()
                        
                        session["hr_user_id"] = user[0]
                        session["hr_username"] = user[1]
                        session["hr_full_name"] = user[3]
                        session["hr_role"] = user[4]
                        flash(f"Welcome, {user[3]}!", "success")
                        return redirect(url_for("hr_dashboard"))
                    else:
                        flash("Invalid password", "danger")
            else:
                flash("Invalid username", "danger")

        except Exception as e:
            app.logger.error(f"HR login error: {e}")
            flash("Login error. Please try again.", "danger")

    return render_template("hr_login.html", 
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")
    
@app.route("/hr/dashboard")
def hr_dashboard():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get HR statistics
    conn = get_db_connection()
    if not conn:
        flash("Database connection error", "danger")
        return redirect(url_for("hr_login"))
    
    cur = conn.cursor()
    
    try:
        # Total staff
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active'")
        total_staff = cur.fetchone()[0] or 0
        
        # Staff active today (attendance)
        today = date.today().isoformat()
        cur.execute("""
            SELECT COUNT(DISTINCT staff_id) 
            FROM attendance 
            WHERE date = ? AND status IN ('Present', 'Late')
        """, (today,))
        active_staff = cur.fetchone()[0] or 0
        
        # Staff on leave today
        cur.execute("""
            SELECT COUNT(*) 
            FROM leaves 
            WHERE ? BETWEEN start_date AND end_date 
            AND status = 'Approved'
        """, (today,))
        on_leave = cur.fetchone()[0] or 0
        
        # Departments count
        cur.execute("SELECT COUNT(*) FROM departments WHERE status = 'Active'")
        departments_count = cur.fetchone()[0] or 0
        
        # Pending leave requests
        cur.execute("SELECT COUNT(*) FROM leaves WHERE status = 'Pending'")
        pending_leave = cur.fetchone()[0] or 0
        
        # Upcoming shifts (next 7 days)
        next_week = (date.today() + timedelta(days=7)).isoformat()
        cur.execute("""
            SELECT COUNT(*) 
            FROM schedules 
            WHERE schedule_date BETWEEN ? AND ?
        """, (today, next_week))
        upcoming_shifts = cur.fetchone()[0] or 0
        
        # Pending updates (staff with missing info)
        cur.execute("""
            SELECT COUNT(*) 
            FROM staff 
            WHERE emergency_contact IS NULL OR address IS NULL
        """)
        pending_updates = cur.fetchone()[0] or 0
        
        # Late arrivals today
        cur.execute("""
            SELECT COUNT(*) 
            FROM attendance 
            WHERE date = ? AND status = 'Late'
        """, (today,))
        late_arrivals = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error fetching HR stats: {e}")
        # Set default values on error
        total_staff = active_staff = on_leave = departments_count = 0
        pending_leave = upcoming_shifts = pending_updates = late_arrivals = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "hr_dashboard.html",
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        total_staff=total_staff,
        active_staff=active_staff,
        on_leave=on_leave,
        departments_count=departments_count,
        pending_leave=pending_leave,
        upcoming_shifts=upcoming_shifts,
        pending_updates=pending_updates,
        late_arrivals=late_arrivals,
        current_year=date.today().year
    )

@app.route("/hr/logout")
def hr_logout():
    session.pop("hr_user_id", None)
    session.pop("hr_username", None)
    session.pop("hr_full_name", None)
    session.pop("hr_role", None)
    flash("Logged out successfully", "success")
    return redirect(url_for("hr_login"))

@app.route("/hr/staff-management")
def staff_management():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    if not conn:
        flash("Database connection error", "danger")
        return redirect(url_for("hr_login"))
    
    cur = conn.cursor()
    
    try:
        # Get staff list with department names
        cur.execute("""
            SELECT s.id, s.staff_id, s.first_name, s.last_name, 
                   s.position, s.employment_type, s.email, s.phone,
                   s.hire_date, s.salary, s.status, s.emergency_contact,
                   s.address, d.name as department_name
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            ORDER BY s.id DESC
            LIMIT 100
        """)
        staff_list = cur.fetchall()
        
        # Get statistics
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active'")
        total_staff = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active' AND employment_type = 'Full-Time'")
        active_staff = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM staff WHERE employment_type = 'Contract'")
        on_contract = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(DISTINCT department_id) FROM staff")
        departments_count = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error fetching staff data: {e}")
        staff_list = []
        total_staff = active_staff = on_contract = departments_count = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("staff_management.html", 
                         module_name="Staff Management",
                         description="Manage staff profiles, positions, and employment details",
                         staff_list=staff_list,
                         total_staff=total_staff,
                         active_staff=active_staff,
                         on_contract=on_contract,
                         departments_count=departments_count,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
                         current_year=date.today().year)

# View Staff Details
@app.route("/hr/staff/<int:staff_id>")
def view_staff(staff_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get staff details with department
        cur.execute("""
            SELECT s.*, d.name as department_name, d.code as department_code
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.id = ?
        """, (staff_id,))
        
        staff = cur.fetchone()
        
        if not staff:
            flash("Staff member not found", "danger")
            return redirect(url_for("staff_management"))
        
        # Convert to dictionary for easier template access
        staff_dict = {
            'id': staff[0],
            'staff_id': staff[1],
            'first_name': staff[2],
            'last_name': staff[3],
            'department_id': staff[4],
            'position': staff[5],
            'employment_type': staff[6],
            'email': staff[7],
            'phone': staff[8],
            'hire_date': staff[9],
            'salary': float(staff[10]) if staff[10] else 0,
            'status': staff[11],
            'emergency_contact': staff[12],
            'address': staff[13],
            'department_name': staff[14],
            'department_code': staff[15]
        }
        
        # Calculate employment duration
        today = date.today()
        hire_date = staff[9]
        if isinstance(hire_date, str):
            hire_date = datetime.strptime(hire_date, '%Y-%m-%d').date()
        
        years = today.year - hire_date.year
        months = today.month - hire_date.month
        
        if months < 0:
            years -= 1
            months += 12
        
        employment_duration = f"{years} year(s), {months} month(s)"
        
    except Exception as e:
        app.logger.error(f"Error fetching staff details: {e}")
        flash("Error loading staff details", "danger")
        return redirect(url_for("staff_management"))
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("view_staff.html", 
                         staff=staff_dict,
                         today=today,
                         employment_duration=employment_duration,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

# Add New Staff
@app.route("/hr/staff/add", methods=["GET", "POST"])
def add_staff():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        # Get form data
        staff_id = request.form.get('staff_id')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        department_id = request.form.get('department_id')
        position = request.form.get('position')
        employment_type = request.form.get('employment_type')
        email = request.form.get('email')
        phone = request.form.get('phone')
        hire_date = request.form.get('hire_date')
        salary = request.form.get('salary')
        emergency_contact = request.form.get('emergency_contact')
        address = request.form.get('address')
        
        # Validate required fields
        if not all([staff_id, first_name, last_name, department_id, position, hire_date]):
            flash("Please fill in all required fields", "danger")
            return redirect(url_for("add_staff"))
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return redirect(url_for("add_staff"))
        
        cur = conn.cursor()
        
        try:
            # Check if staff ID already exists
            cur.execute("SELECT id FROM staff WHERE staff_id = ?", (staff_id,))
            if cur.fetchone():
                flash(f"Staff ID '{staff_id}' already exists. Please use a different ID.", "danger")
                return redirect(url_for("add_staff"))
            
            # Convert salary to decimal or set to 0
            try:
                salary_decimal = float(salary) if salary else 0.00
            except ValueError:
                salary_decimal = 0.00
            
            # Insert new staff
            cur.execute("""
                INSERT INTO staff (
                    staff_id, first_name, last_name, department_id, 
                    position, employment_type, email, phone, 
                    hire_date, salary, emergency_contact, address, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active')
            """, (
                staff_id, first_name, last_name, department_id,
                position, employment_type, email, phone,
                hire_date, salary_decimal, emergency_contact, address
            ))
            
            conn.commit()
            flash(f"Staff member {first_name} {last_name} (ID: {staff_id}) added successfully!", "success")
            
            # Redirect to staff management or view the new staff
            cur.execute("SELECT id FROM staff WHERE staff_id = ?", (staff_id,))
            new_staff_id = cur.fetchone()[0]
            return redirect(url_for("view_staff", staff_id=new_staff_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error adding staff: {e}")
            flash(f"Error adding staff: {str(e)}", "danger")
            return redirect(url_for("add_staff"))
            
        finally:
            cur.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    if not conn:
        flash("Database connection error", "danger")
        return redirect(url_for("staff_management"))
    
    cur = conn.cursor()
    
    try:
        # Get departments for dropdown
        cur.execute("SELECT id, name, code FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get next staff ID suggestion
        cur.execute("""
            SELECT MAX(staff_id) FROM staff 
            WHERE staff_id LIKE 'EMP%'
        """)
        last_staff_id = cur.fetchone()[0]
        
        if last_staff_id:
            # Extract number and increment
            import re
            match = re.search(r'EMP(\d+)', last_staff_id)
            if match:
                next_num = int(match.group(1)) + 1
                suggested_id = f"EMP{next_num:03d}"
            else:
                suggested_id = "EMP001"
        else:
            suggested_id = "EMP001"
            
        # Get current date for hire date default
        today = date.today().strftime("%Y-%m-%d")
        
    except Exception as e:
        app.logger.error(f"Error loading form data: {e}")
        departments = []
        suggested_id = "EMP001"
        today = date.today().strftime("%Y-%m-%d")
        
    finally:
        cur.close()
        conn.close()
    
    return render_template("add_staff.html",
                         departments=departments,
                         suggested_id=suggested_id,
                         today=today,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

# Edit Staff
@app.route("/hr/staff/edit/<int:staff_id>", methods=["GET", "POST"])
def edit_staff(staff_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == "POST":
        # Update staff
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        department_id = request.form.get('department_id')
        position = request.form.get('position')
        employment_type = request.form.get('employment_type')
        email = request.form.get('email')
        phone = request.form.get('phone')
        salary = request.form.get('salary')
        status = request.form.get('status')
        emergency_contact = request.form.get('emergency_contact')
        address = request.form.get('address')
        
        try:
            cur.execute("""
                UPDATE staff SET
                    first_name = ?,
                    last_name = ?,
                    department_id = ?,
                    position = ?,
                    employment_type = ?,
                    email = ?,
                    phone = ?,
                    salary = ?,
                    status = ?,
                    emergency_contact = ?,
                    address = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                first_name, last_name, department_id,
                position, employment_type, email, phone,
                salary, status, emergency_contact, address,
                staff_id
            ))
            
            conn.commit()
            flash("Staff details updated successfully!", "success")
            return redirect(url_for("view_staff", staff_id=staff_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error updating staff: {e}")
            flash("Error updating staff details", "danger")
    
    # GET request - load staff data
    try:
        cur.execute("SELECT * FROM staff WHERE id = ?", (staff_id,))
        staff = cur.fetchone()
        
        if not staff:
            flash("Staff member not found", "danger")
            return redirect(url_for("staff_management"))
        
        # Get departments
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error loading staff for edit: {e}")
        flash("Error loading staff details", "danger")
        return redirect(url_for("staff_management"))
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("edit_staff.html",
                         staff=staff,
                         departments=departments,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

# ==================== ROUTES: SCHEDULING MODULE ====================

@app.route("/hr/scheduling")
def scheduling():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get current month schedules
        current_month = date.today().replace(day=1)
        if current_month.month == 12:
            next_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            next_month = current_month.replace(month=current_month.month + 1)
        
        cur.execute("""
            SELECT s.*, st.first_name, st.last_name, st.position, d.name as department_name
            FROM schedules s
            JOIN staff st ON s.staff_id = st.id
            LEFT JOIN departments d ON st.department_id = d.id
            WHERE s.schedule_date >= ? AND s.schedule_date < ?
            ORDER BY s.schedule_date, s.start_time
        """, (current_month.isoformat(), next_month.isoformat()))
        
        schedules = cur.fetchall()
        
        # Get statistics
        cur.execute("SELECT COUNT(*) FROM schedules WHERE schedule_date >= CURRENT_DATE")
        upcoming_shifts = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COUNT(DISTINCT staff_id) 
            FROM schedules 
            WHERE schedule_date >= CURRENT_DATE
        """)
        staff_scheduled = cur.fetchone()[0] or 0
        
        # Get departments for filter
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get staff for filter
        cur.execute("""
            SELECT id, first_name, last_name, position 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error fetching scheduling data: {e}")
        schedules = []
        upcoming_shifts = 0
        staff_scheduled = 0
        departments = []
        staff_list = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("scheduling_dashboard.html",
                         schedules=schedules,
                         upcoming_shifts=upcoming_shifts,
                         staff_scheduled=staff_scheduled,
                         departments=departments,
                         staff_list=staff_list,
                         current_month=current_month.strftime("%B %Y"),
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

@app.route("/hr/scheduling/create", methods=["GET", "POST"])
def create_schedule():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        staff_id = request.form.get("staff_id")
        schedule_date = request.form.get("schedule_date")
        shift_type = request.form.get("shift_type")
        start_time = request.form.get("start_time")
        end_time = request.form.get("end_time")
        location = request.form.get("location")
        notes = request.form.get("notes")
        
        if not all([staff_id, schedule_date, start_time, end_time]):
            flash("Please fill in all required fields", "danger")
            return redirect(url_for("create_schedule"))
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Check for existing schedule for same staff on same date
            cur.execute("""
                SELECT id FROM schedules 
                WHERE staff_id = ? AND schedule_date = ?
            """, (staff_id, schedule_date))
            
            if cur.fetchone():
                flash("This staff already has a schedule for this date", "warning")
                return redirect(url_for("create_schedule"))
            
            # Insert new schedule
            cur.execute("""
                INSERT INTO schedules (
                    staff_id, schedule_date, shift_type, 
                    start_time, end_time, location, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (staff_id, schedule_date, shift_type, start_time, end_time, location, notes))
            
            conn.commit()
            flash("Schedule created successfully!", "success")
            return redirect(url_for("scheduling"))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating schedule: {e}")
            flash(f"Error creating schedule: {str(e)}", "danger")
            return redirect(url_for("create_schedule"))
            
        finally:
            cur.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get active staff
        cur.execute("""
            SELECT id, first_name, last_name, position, 
                   (SELECT name FROM departments WHERE id = staff.department_id) as department
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        # Get default tomorrow's date
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        
    except Exception as e:
        app.logger.error(f"Error loading schedule form data: {e}")
        staff_list = []
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        
    finally:
        cur.close()
        conn.close()
    
    return render_template("create_schedule.html",
                         staff_list=staff_list,
                         tomorrow=tomorrow,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

@app.route("/hr/scheduling/roster")
def view_roster():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get filter parameters
    department_id = request.args.get("department_id", "")
    staff_id = request.args.get("staff_id", "")
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    
    # Default to current week if no dates specified
    if not start_date:
        today = date.today()
        start_date = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
    
    if not end_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = (start_date_obj + timedelta(days=6)).strftime("%Y-%m-%d")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Build query with filters
        query = """
            SELECT 
                s.id as schedule_id,
                s.schedule_date,
                s.shift_type,
                s.start_time,
                s.end_time,
                s.location,
                s.notes,
                st.id as staff_id,
                st.first_name,
                st.last_name,
                st.position,
                d.name as department_name,
                d.id as department_id
            FROM schedules s
            JOIN staff st ON s.staff_id = st.id
            LEFT JOIN departments d ON st.department_id = d.id
            WHERE s.schedule_date BETWEEN ? AND ?
        """
        params = [start_date, end_date]
        
        if department_id:
            query += " AND st.department_id = ?"
            params.append(department_id)
        
        if staff_id:
            query += " AND s.staff_id = ?"
            params.append(staff_id)
        
        query += " ORDER BY s.schedule_date, d.name, st.first_name, s.start_time"
        
        cur.execute(query, params)
        schedules = cur.fetchall()
        
        # Get departments for filter dropdown
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get staff for filter dropdown
        cur.execute("""
            SELECT id, first_name, last_name 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        # Group schedules by date for calendar view
        schedule_dict = {}
        for schedule in schedules:
            schedule_date = schedule[1] if isinstance(schedule[1], str) else schedule[1].strftime("%Y-%m-%d")
            if schedule_date not in schedule_dict:
                schedule_dict[schedule_date] = []
            
            schedule_dict[schedule_date].append({
                'id': schedule[0],
                'date': schedule[1],
                'shift_type': schedule[2],
                'start_time': schedule[3],
                'end_time': schedule[4],
                'location': schedule[5],
                'notes': schedule[6],
                'staff_id': schedule[7],
                'first_name': schedule[8],
                'last_name': schedule[9],
                'position': schedule[10],
                'department': schedule[11]
            })
        
        # Calculate statistics
        total_shifts = len(schedules)
        unique_staff = len(set([s[7] for s in schedules]))
        unique_departments = len(set([s[11] for s in schedules if s[11]]))
        
    except Exception as e:
        app.logger.error(f"Error fetching roster: {e}")
        schedules = []
        departments = []
        staff_list = []
        schedule_dict = {}
        total_shifts = 0
        unique_staff = 0
        unique_departments = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("view_roster.html",
                         schedules=schedules,
                         schedule_dict=schedule_dict,
                         departments=departments,
                         staff_list=staff_list,
                         start_date=start_date,
                         end_date=end_date,
                         selected_department=department_id,
                         selected_staff=staff_id,
                         total_shifts=total_shifts,
                         unique_staff=unique_staff,
                         unique_departments=unique_departments,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

# Route to delete schedule
@app.route("/hr/scheduling/delete/<int:schedule_id>", methods=["POST"])
def delete_schedule(schedule_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("DELETE FROM schedules WHERE id = ?", (schedule_id,))
        conn.commit()
        return jsonify({"success": True, "message": "Schedule deleted successfully"})
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error deleting schedule: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()

# ==================== ROUTES: LEAVE MANAGEMENT MODULE ====================

@app.route("/hr/leave-management")
def leave_management():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get all leave requests with staff details
        cur.execute("""
            SELECT 
                l.id,
                l.leave_type,
                l.start_date,
                l.end_date,
                l.days_requested,
                l.reason,
                l.status,
                l.created_at,
                l.approved_at,
                l.approved_by,
                s.id as staff_id,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department,
                s.staff_id as employee_id,
                hu.username as approved_by_name
            FROM leaves l
            JOIN staff s ON l.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN hr_users hu ON l.approved_by = hu.id
            ORDER BY 
                CASE 
                    WHEN l.status = 'Pending' THEN 1
                    WHEN l.status = 'Approved' THEN 2
                    ELSE 3
                END,
                l.start_date DESC
        """)
        
        leaves = cur.fetchall()
        
        # Separate pending and history leaves
        pending_leaves = []
        approved_leaves = []
        rejected_leaves = []
        
        for leave in leaves:
            leave_dict = {
                'id': leave[0],
                'leave_type': leave[1],
                'start_date': leave[2],
                'end_date': leave[3],
                'days_requested': leave[4],
                'reason': leave[5],
                'status': leave[6],
                'created_at': leave[7],
                'approved_at': leave[8],
                'approved_by': leave[9],
                'staff_id': leave[10],
                'first_name': leave[11],
                'last_name': leave[12],
                'staff_name': f"{leave[11]} {leave[12]}",
                'position': leave[13],
                'department': leave[14] or 'N/A',
                'employee_id': leave[15],
                'approved_by_name': leave[16]
            }
            
            if leave_dict['status'] == 'Pending':
                pending_leaves.append(leave_dict)
            elif leave_dict['status'] == 'Approved':
                approved_leaves.append(leave_dict)
            elif leave_dict['status'] == 'Rejected':
                rejected_leaves.append(leave_dict)
        
        # Get leave statistics
        cur.execute("""
            SELECT 
                COUNT(*) as total_requests,
                SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending_count,
                SUM(CASE WHEN status = 'Approved' THEN 1 ELSE 0 END) as approved_count,
                SUM(CASE WHEN status = 'Rejected' THEN 1 ELSE 0 END) as rejected_count
            FROM leaves
        """)
        stats = cur.fetchone()
        
        # Get leave balance for active staff
        cur.execute("""
            SELECT 
                s.id,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department,
                COALESCE(SUM(CASE 
                    WHEN l.status = 'Approved' AND strftime('%Y', l.start_date) = strftime('%Y', CURRENT_DATE)
                    THEN l.days_requested ELSE 0 END), 0) as days_taken,
                20 - COALESCE(SUM(CASE 
                    WHEN l.status = 'Approved' AND strftime('%Y', l.start_date) = strftime('%Y', CURRENT_DATE)
                    THEN l.days_requested ELSE 0 END), 0) as days_remaining
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN leaves l ON s.id = l.staff_id AND l.status = 'Approved'
            WHERE s.status = 'Active'
            GROUP BY s.id, s.first_name, s.last_name, s.position, d.name
            ORDER BY days_remaining ASC
            LIMIT 10
        """)
        
        leave_balances = cur.fetchall()
        
        # Get leave types distribution
        cur.execute("""
            SELECT 
                leave_type,
                COUNT(*) as count,
                SUM(days_requested) as total_days
            FROM leaves
            WHERE status = 'Approved' 
                AND strftime('%Y', start_date) = strftime('%Y', CURRENT_DATE)
            GROUP BY leave_type
            ORDER BY count DESC
        """)
        
        leave_types = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error fetching leave data: {e}")
        pending_leaves = []
        approved_leaves = []
        rejected_leaves = []
        stats = (0, 0, 0, 0)
        leave_balances = []
        leave_types = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "leave_management.html",
        pending_leaves=pending_leaves,
        approved_leaves=approved_leaves,
        rejected_leaves=rejected_leaves,
        total_requests=stats[0],
        pending_count=stats[1],
        approved_count=stats[2],
        rejected_count=stats[3],
        leave_balances=leave_balances,
        leave_types=leave_types,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )

@app.route("/hr/leave/request", methods=["GET", "POST"])
def request_leave():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        staff_id = request.form.get("staff_id")
        leave_type = request.form.get("leave_type")
        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        reason = request.form.get("reason")
        
        # Calculate days requested
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
        days_requested = (end - start).days + 1
        
        if days_requested <= 0:
            flash("End date must be after start date", "danger")
            return redirect(url_for("request_leave"))
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Check for overlapping leaves
            cur.execute("""
                SELECT id FROM leaves 
                WHERE staff_id = ? 
                AND status IN ('Pending', 'Approved')
                AND NOT (end_date < ? OR start_date > ?)
            """, (staff_id, start_date, end_date))
            
            if cur.fetchone():
                flash("Staff already has a leave request for this period", "warning")
                return redirect(url_for("request_leave"))
            
            # Insert leave request
            cur.execute("""
                INSERT INTO leaves (
                    staff_id, leave_type, start_date, end_date, 
                    days_requested, reason, status, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, 'Pending', CURRENT_TIMESTAMP)
            """, (staff_id, leave_type, start_date, end_date, days_requested, reason))
            
            leave_id = cur.lastrowid
            conn.commit()
            
            flash(f"Leave request #{leave_id} submitted successfully!", "success")
            return redirect(url_for("leave_management"))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating leave request: {e}")
            flash(f"Error creating leave request: {str(e)}", "danger")
        
        finally:
            cur.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get active staff for dropdown
        cur.execute("""
            SELECT s.id, s.first_name, s.last_name, s.position, d.name as department
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.status = 'Active'
            ORDER BY s.first_name, s.last_name
        """)
        staff_list = cur.fetchall()
        
        # Get leave types
        leave_types = [
            'Annual Leave',
            'Sick Leave',
            'Maternity Leave',
            'Paternity Leave',
            'Bereavement Leave',
            'Study Leave',
            'Unpaid Leave',
            'Compensatory Off',
            'Emergency Leave'
        ]
        
    except Exception as e:
        app.logger.error(f"Error loading leave form data: {e}")
        staff_list = []
        leave_types = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "request_leave.html",
        staff_list=staff_list,
        leave_types=leave_types,
        today=date.today().strftime("%Y-%m-%d"),
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    )

@app.route("/hr/leave/approve/<int:leave_id>", methods=["POST"])
def approve_leave(leave_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            UPDATE leaves 
            SET status = 'Approved', 
                approved_by = ?, 
                approved_at = CURRENT_TIMESTAMP
            WHERE id = ? AND status = 'Pending'
        """, (session["hr_user_id"], leave_id))
        
        if cur.rowcount > 0:
            conn.commit()
            return jsonify({"success": True, "message": "Leave approved successfully"})
        else:
            return jsonify({"success": False, "message": "Leave request not found or already processed"}), 404
            
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error approving leave: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()

@app.route("/hr/leave/reject/<int:leave_id>", methods=["POST"])
def reject_leave(leave_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    rejection_reason = data.get("rejection_reason", "")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            UPDATE leaves 
            SET status = 'Rejected', 
                approved_by = ?, 
                approved_at = CURRENT_TIMESTAMP,
                reason = reason || ' [Rejected: ' || ? || ']'
            WHERE id = ? AND status = 'Pending'
        """, (session["hr_user_id"], rejection_reason, leave_id))
        
        if cur.rowcount > 0:
            conn.commit()
            return jsonify({"success": True, "message": "Leave rejected"})
        else:
            return jsonify({"success": False, "message": "Leave request not found or already processed"}), 404
            
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error rejecting leave: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()
@app.route("/hr/departments/analytics")
def department_analytics():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Staff distribution by department
        cur.execute("""
            SELECT 
                d.name,
                COUNT(s.id) as total_staff,
                SUM(CASE WHEN s.status = 'Active' THEN 1 ELSE 0 END) as active_staff,
                SUM(CASE WHEN s.employment_type = 'Full-Time' THEN 1 ELSE 0 END) as full_time,
                SUM(CASE WHEN s.employment_type = 'Part-Time' THEN 1 ELSE 0 END) as part_time,
                SUM(CASE WHEN s.employment_type = 'Contract' THEN 1 ELSE 0 END) as contract
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            WHERE d.status = 'Active'
            GROUP BY d.name
            ORDER BY total_staff DESC
        """)
        staff_distribution = cur.fetchall()
        
        # Attendance rates by department (last 30 days)
        cur.execute("""
            WITH dept_attendance AS (
                SELECT 
                    d.id,
                    d.name,
                    COUNT(a.id) as total_attendance,
                    SUM(CASE WHEN a.status IN ('Present', 'Late') THEN 1 ELSE 0 END) as present_count,
                    COUNT(DISTINCT a.date) as working_days
                FROM departments d
                JOIN staff s ON d.id = s.department_id
                LEFT JOIN attendance a ON s.id = a.staff_id 
                    AND a.date >= date('now', '-30 days')
                WHERE d.status = 'Active'
                GROUP BY d.id, d.name
            )
            SELECT 
                name,
                total_attendance,
                present_count,
                working_days,
                CASE 
                    WHEN working_days > 0 AND staff_count > 0
                    THEN ROUND((present_count * 1.0 / (working_days * staff_count)) * 100, 1)
                    ELSE 0 
                END as attendance_rate
            FROM dept_attendance
            JOIN (
                SELECT department_id, COUNT(*) as staff_count 
                FROM staff 
                WHERE status = 'Active'
                GROUP BY department_id
            ) staff_counts ON dept_attendance.id = staff_counts.department_id
            ORDER BY attendance_rate DESC
        """)
        attendance_rates = cur.fetchall()
        
        # Leave statistics by department
        cur.execute("""
            SELECT 
                d.name,
                COUNT(l.id) as total_leaves,
                SUM(CASE WHEN l.status = 'Approved' THEN 1 ELSE 0 END) as approved_leaves,
                SUM(CASE WHEN l.status = 'Pending' THEN 1 ELSE 0 END) as pending_leaves,
                COALESCE(SUM(l.days_requested), 0) as total_days
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            LEFT JOIN leaves l ON s.id = l.staff_id
                AND strftime('%Y', l.start_date) = strftime('%Y', 'now')
            WHERE d.status = 'Active'
            GROUP BY d.name
            ORDER BY total_leaves DESC
        """)
        leave_stats = cur.fetchall()
        
        # Schedule coverage by department
        cur.execute("""
            SELECT 
                d.name,
                COUNT(DISTINCT sch.id) as total_shifts,
                COUNT(DISTINCT sch.staff_id) as staff_scheduled,
                COUNT(DISTINCT sch.schedule_date) as days_covered
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            LEFT JOIN schedules sch ON s.id = sch.staff_id
                AND sch.schedule_date >= date('now')
            WHERE d.status = 'Active'
            GROUP BY d.name
            ORDER BY total_shifts DESC
        """)
        schedule_coverage = cur.fetchall()
        
        # Salary distribution by department
        cur.execute("""
            SELECT 
                d.name,
                ROUND(AVG(s.salary), 2) as avg_salary,
                MIN(s.salary) as min_salary,
                MAX(s.salary) as max_salary,
                SUM(s.salary) as total_salary
            FROM departments d
            JOIN staff s ON d.id = s.department_id
            WHERE d.status = 'Active' AND s.status = 'Active'
            GROUP BY d.name
            ORDER BY avg_salary DESC
        """)
        salary_stats = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error generating department analytics: {e}")
        staff_distribution = []
        attendance_rates = []
        leave_stats = []
        schedule_coverage = []
        salary_stats = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "department_analytics.html",
        staff_distribution=staff_distribution,
        attendance_rates=attendance_rates,
        leave_stats=leave_stats,
        schedule_coverage=schedule_coverage,
        salary_stats=salary_stats,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )
    
# ==================== ROUTES: ATTENDANCE RECORD MODULE ====================

@app.route("/hr/attendance-record")
def attendance_record():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get filter parameters
    department_id = request.args.get("department_id", "")
    staff_id = request.args.get("staff_id", "")
    date_filter = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    status_filter = request.args.get("status", "")
    
    # Parse date
    try:
        selected_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
    except:
        selected_date = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get attendance records for selected date with filters
        query = """
            SELECT 
                a.id,
                a.staff_id,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department,
                a.date,
                a.check_in,
                a.check_out,
                a.status,
                a.remarks,
                a.created_at,
                s.staff_id as employee_id
            FROM attendance a
            JOIN staff s ON a.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE a.date = ?
        """
        params = [selected_date.isoformat()]
        
        if department_id:
            query += " AND s.department_id = ?"
            params.append(department_id)
        
        if staff_id:
            query += " AND a.staff_id = ?"
            params.append(staff_id)
        
        if status_filter:
            query += " AND a.status = ?"
            params.append(status_filter)
        
        query += " ORDER BY s.first_name, s.last_name"
        
        cur.execute(query, params)
        attendance_records = cur.fetchall()
        
        # Get all active staff for attendance marking
        cur.execute("""
            SELECT s.id, s.first_name, s.last_name, s.position, d.name as department
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.status = 'Active'
            ORDER BY s.first_name, s.last_name
        """)
        all_staff = cur.fetchall()
        
        # Get departments for filter
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get staff for filter
        cur.execute("""
            SELECT id, first_name, last_name 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        # Calculate statistics for selected date
        cur.execute("""
            SELECT 
                COUNT(*) as total_present,
                SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) as present,
                SUM(CASE WHEN status = 'Late' THEN 1 ELSE 0 END) as late,
                SUM(CASE WHEN status = 'Absent' THEN 1 ELSE 0 END) as absent,
                SUM(CASE WHEN status = 'Half Day' THEN 1 ELSE 0 END) as half_day,
                SUM(CASE WHEN status = 'Holiday' THEN 1 ELSE 0 END) as holiday
            FROM attendance
            WHERE date = ?
        """, (selected_date.isoformat(),))
        
        stats = cur.fetchone()
        
        # Get total active staff count
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active'")
        total_active = cur.fetchone()[0] or 0
        
        # Get recent attendance history (last 7 days)
        week_ago = (selected_date - timedelta(days=7)).isoformat()
        cur.execute("""
            SELECT 
                date,
                COUNT(*) as total,
                SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) as present,
                SUM(CASE WHEN status = 'Late' THEN 1 ELSE 0 END) as late,
                SUM(CASE WHEN status = 'Absent' THEN 1 ELSE 0 END) as absent
            FROM attendance
            WHERE date BETWEEN ? AND ?
            GROUP BY date
            ORDER BY date DESC
        """, (week_ago, selected_date.isoformat()))
        
        history = cur.fetchall()
        
        # Get staff without attendance for today
        if selected_date == date.today():
            marked_staff_ids = [r[1] for r in attendance_records]
            unmarked_staff = [s for s in all_staff if s[0] not in marked_staff_ids]
        else:
            unmarked_staff = []
        
    except Exception as e:
        app.logger.error(f"Error fetching attendance data: {e}")
        attendance_records = []
        all_staff = []
        departments = []
        staff_list = []
        stats = (0, 0, 0, 0, 0, 0)
        total_active = 0
        history = []
        unmarked_staff = []
    
    finally:
        cur.close()
        conn.close()
    
    # Format attendance records for template
    formatted_records = []
    for record in attendance_records:
        check_in = record[7]
        check_out = record[8]
        
        # Calculate hours worked if both check-in and check-out exist
        hours_worked = None
        if check_in and check_out:
            # Parse time strings if needed
            if isinstance(check_in, str):
                check_in_time = datetime.strptime(check_in, '%H:%M:%S').time()
                check_out_time = datetime.strptime(check_out, '%H:%M:%S').time()
            else:
                check_in_time = check_in
                check_out_time = check_out
            
            check_in_dt = datetime.combine(selected_date, check_in_time)
            check_out_dt = datetime.combine(selected_date, check_out_time)
            if check_out_dt < check_in_dt:  # Overnight shift
                check_out_dt += timedelta(days=1)
            hours_worked = (check_out_dt - check_in_dt).total_seconds() / 3600
        
        formatted_records.append({
            'id': record[0],
            'staff_id': record[1],
            'first_name': record[2],
            'last_name': record[3],
            'staff_name': f"{record[2]} {record[3]}",
            'position': record[4],
            'department': record[5] or 'N/A',
            'date': record[6],
            'check_in': record[7],
            'check_out': record[8],
            'status': record[9],
            'remarks': record[10],
            'created_at': record[11],
            'employee_id': record[12],
            'hours_worked': round(hours_worked, 1) if hours_worked else None
        })
    
    return render_template(
        "attendance_record.html",
        attendance_records=formatted_records,
        all_staff=all_staff,
        departments=departments,
        staff_list=staff_list,
        selected_date=selected_date,
        selected_department=department_id,
        selected_staff=staff_id,
        selected_status=status_filter,
        total_present=stats[1] or 0,
        total_late=stats[2] or 0,
        total_absent=stats[3] or 0,
        total_half_day=stats[4] or 0,
        total_holiday=stats[5] or 0,
        total_marked=(stats[1] or 0) + (stats[2] or 0) + (stats[3] or 0) + (stats[4] or 0) + (stats[5] or 0),
        total_active=total_active,
        history=history,
        unmarked_staff=unmarked_staff,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )

@app.route("/hr/attendance/mark", methods=["POST"])
def mark_attendance():
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    staff_id = data.get("staff_id")
    attendance_date = data.get("date")
    check_in = data.get("check_in")
    check_out = data.get("check_out")
    status = data.get("status")
    remarks = data.get("remarks", "")
    
    if not all([staff_id, attendance_date, status]):
        return jsonify({"success": False, "message": "Missing required fields"}), 400
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Check if attendance already exists for this staff on this date
        cur.execute("""
            SELECT id FROM attendance 
            WHERE staff_id = ? AND date = ?
        """, (staff_id, attendance_date))
        
        existing = cur.fetchone()
        
        if existing:
            # Update existing attendance
            cur.execute("""
                UPDATE attendance 
                SET check_in = ?,
                    check_out = ?,
                    status = ?,
                    remarks = ?,
                    recorded_by = ?
                WHERE id = ?
            """, (check_in, check_out, status, remarks, session["hr_user_id"], existing[0]))
        else:
            # Insert new attendance
            cur.execute("""
                INSERT INTO attendance (
                    staff_id, date, check_in, check_out, status, remarks, recorded_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (staff_id, attendance_date, check_in, check_out, status, remarks, session["hr_user_id"]))
        
        conn.commit()
        
        return jsonify({
            "success": True, 
            "message": "Attendance marked successfully"
        })
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error marking attendance: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()
        


@app.route("/hr/scheduling/reports")
def schedule_reports():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get report parameters
    report_type = request.args.get("report_type", "monthly")
    month = request.args.get("month", date.today().month, type=int)
    year = request.args.get("year", date.today().year, type=int)
    department_id = request.args.get("department_id", "")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    report_data = []
    total_shifts = 0
    total_hours = 0
    unique_staff = 0
    
    try:
        # Build report query based on report type
        if report_type == "monthly":
            start_date = date(int(year), int(month), 1)
            if int(month) == 12:
                end_date = date(int(year) + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(int(year), int(month) + 1, 1) - timedelta(days=1)
            
            query = """
                SELECT 
                    COALESCE(d.name, 'No Department') as department,
                    st.first_name || ' ' || st.last_name as staff_name,
                    COUNT(*) as total_shifts,
                    COALESCE(SUM((julianday(sch.end_time) - julianday(sch.start_time)) * 24), 0) as total_hours,
                    COUNT(DISTINCT sch.schedule_date) as days_scheduled,
                    st.position,
                    st.id as staff_id
                FROM schedules sch
                JOIN staff st ON sch.staff_id = st.id
                LEFT JOIN departments d ON st.department_id = d.id
                WHERE sch.schedule_date BETWEEN ? AND ?
            """
            params = [start_date.isoformat(), end_date.isoformat()]
            
            if department_id:
                query += " AND st.department_id = ?"
                params.append(department_id)
            
            query += """
                GROUP BY d.name, st.first_name, st.last_name, st.position, st.id
                ORDER BY d.name, staff_name
            """
            
            cur.execute(query, params)
            rows = cur.fetchall()
            
            # Format the report data
            for row in rows:
                report_data.append({
                    'department': row[0],
                    'staff_name': row[1],
                    'total_shifts': row[2],
                    'total_hours': float(row[3]) if row[3] else 0,
                    'days_scheduled': row[4],
                    'position': row[5],
                    'staff_id': row[6]
                })
                total_shifts += row[2]
                total_hours += float(row[3]) if row[3] else 0
            
            unique_staff = len(report_data)
            
        elif report_type == "weekly":
            # Get current week
            today = date.today()
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            
            # Allow custom week if specified
            if request.args.get("week_start"):
                try:
                    start_date = datetime.strptime(request.args.get("week_start"), "%Y-%m-%d").date()
                    end_date = start_date + timedelta(days=6)
                except:
                    pass
            
            query = """
                SELECT 
                    sch.schedule_date,
                    COALESCE(d.name, 'No Department') as department,
                    st.first_name || ' ' || st.last_name as staff_name,
                    sch.shift_type,
                    sch.start_time,
                    sch.end_time,
                    (julianday(sch.end_time) - julianday(sch.start_time)) * 24 as hours,
                    st.position,
                    sch.id as schedule_id
                FROM schedules sch
                JOIN staff st ON sch.staff_id = st.id
                LEFT JOIN departments d ON st.department_id = d.id
                WHERE sch.schedule_date BETWEEN ? AND ?
            """
            params = [start_date.isoformat(), end_date.isoformat()]
            
            if department_id:
                query += " AND st.department_id = ?"
                params.append(department_id)
            
            query += " ORDER BY sch.schedule_date, d.name, sch.start_time"
            
            cur.execute(query, params)
            rows = cur.fetchall()
            
            # Group by date for the report
            daily_data = {}
            for row in rows:
                date_str = row[0] if isinstance(row[0], str) else row[0].strftime("%Y-%m-%d")
                if date_str not in daily_data:
                    daily_data[date_str] = {
                        'date': row[0],
                        'shifts': [],
                        'total_hours': 0,
                        'staff_count': 0
                    }
                
                hours = float(row[6]) if row[6] else 0
                daily_data[date_str]['shifts'].append({
                    'department': row[1],
                    'staff_name': row[2],
                    'shift_type': row[3],
                    'start_time': row[4],
                    'end_time': row[5],
                    'hours': hours,
                    'position': row[7],
                    'schedule_id': row[8]
                })
                daily_data[date_str]['total_hours'] += hours
                daily_data[date_str]['staff_count'] += 1
                
                total_hours += hours
                total_shifts += 1
            
            report_data = list(daily_data.values())
            unique_staff = len(set([s['staff_name'] for day in report_data for s in day['shifts']]))
        
        # Get departments for filter dropdown
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get months and years for filter
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        
        # Calculate summary statistics
        avg_hours_per_shift = total_hours / total_shifts if total_shifts > 0 else 0
        
    except Exception as e:
        app.logger.error(f"Error generating schedule report: {e}")
        report_data = []
        departments = []
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        total_shifts = 0
        total_hours = 0
        unique_staff = 0
        avg_hours_per_shift = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "schedule_reports.html",
        report_data=report_data,
        report_type=report_type,
        departments=departments,
        months=months,
        years=years,
        selected_month=int(month),
        selected_year=int(year),
        selected_department=department_id,
        total_shifts=total_shifts,
        total_hours=round(total_hours, 1),
        unique_staff=unique_staff,
        avg_hours_per_shift=round(avg_hours_per_shift, 1),
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )
    
@app.route("/hr/scheduling/shift-swap", methods=["GET", "POST"])
def shift_swap():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == "POST":
        action = request.form.get("action")
        
        if action == "request":
            # Request shift swap
            from_staff_id = request.form.get("from_staff_id")
            to_staff_id = request.form.get("to_staff_id")
            schedule_id = request.form.get("schedule_id")
            reason = request.form.get("reason")
            
            try:
                # Get schedule details
                cur.execute("""
                    SELECT schedule_date, start_time, end_time, shift_type 
                    FROM schedules 
                    WHERE id = ?
                """, (schedule_id,))
                schedule = cur.fetchone()
                
                if not schedule:
                    flash("Schedule not found", "danger")
                    return redirect(url_for("shift_swap"))
                
                # Check if target staff is available on that date
                cur.execute("""
                    SELECT id FROM schedules 
                    WHERE staff_id = ? AND schedule_date = ?
                """, (to_staff_id, schedule[0]))
                
                if cur.fetchone():
                    flash("Selected staff already has a schedule on this date", "warning")
                    return redirect(url_for("shift_swap"))
                
                # Create shift swap request
                cur.execute("""
                    INSERT INTO shift_swap_requests (
                        schedule_id, from_staff_id, to_staff_id, 
                        reason, status, requested_by, requested_at
                    ) VALUES (?, ?, ?, ?, 'Pending', ?, CURRENT_TIMESTAMP)
                """, (schedule_id, from_staff_id, to_staff_id, reason, session["hr_user_id"]))
                
                conn.commit()
                flash("Shift swap request submitted successfully!", "success")
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error creating shift swap request: {e}")
                flash(f"Error creating swap request: {str(e)}", "danger")
            
            return redirect(url_for("shift_swap"))
        
        elif action == "approve":
            # Approve shift swap
            swap_id = request.form.get("swap_id")
            
            try:
                # Get swap request details
                cur.execute("""
                    SELECT schedule_id, from_staff_id, to_staff_id 
                    FROM shift_swap_requests 
                    WHERE id = ? AND status = 'Pending'
                """, (swap_id,))
                
                swap_request = cur.fetchone()
                if not swap_request:
                    flash("Swap request not found or already processed", "warning")
                    return redirect(url_for("shift_swap"))
                
                # Update schedule with new staff
                cur.execute("""
                    UPDATE schedules 
                    SET staff_id = ?, 
                        notes = COALESCE(notes, '') || ' [Shift swapped from staff ID ' || ? || ']'
                    WHERE id = ?
                """, (swap_request[2], swap_request[1], swap_request[0]))
                
                # Update swap request status
                cur.execute("""
                    UPDATE shift_swap_requests 
                    SET status = 'Approved', 
                        approved_by = ?, 
                        approved_at = CURRENT_TIMESTAMP 
                    WHERE id = ?
                """, (session["hr_user_id"], swap_id))
                
                conn.commit()
                flash("Shift swap request approved successfully!", "success")
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error approving shift swap: {e}")
                flash(f"Error approving swap: {str(e)}", "danger")
            
            return redirect(url_for("shift_swap"))
        
        elif action == "reject":
            # Reject shift swap
            swap_id = request.form.get("swap_id")
            rejection_reason = request.form.get("rejection_reason")
            
            try:
                cur.execute("""
                    UPDATE shift_swap_requests 
                    SET status = 'Rejected', 
                        rejection_reason = ?,
                        reviewed_by = ?, 
                        reviewed_at = CURRENT_TIMESTAMP 
                    WHERE id = ?
                """, (rejection_reason, session["hr_user_id"], swap_id))
                
                conn.commit()
                flash("Shift swap request rejected", "info")
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error rejecting shift swap: {e}")
                flash(f"Error rejecting swap: {str(e)}", "danger")
            
            return redirect(url_for("shift_swap"))
    
    # GET request - show shift swap page
    try:
        # Get all shift swap requests with details
        cur.execute("""
            SELECT 
                ssr.id,
                ssr.schedule_id,
                ssr.from_staff_id,
                ssr.to_staff_id,
                ssr.reason,
                ssr.status,
                ssr.requested_at,
                ssr.approved_at,
                ssr.rejection_reason,
                -- From staff details
                s1.first_name as from_first_name,
                s1.last_name as from_last_name,
                s1.position as from_position,
                -- To staff details
                s2.first_name as to_first_name,
                s2.last_name as to_last_name,
                s2.position as to_position,
                -- Schedule details
                sch.schedule_date,
                sch.start_time,
                sch.end_time,
                sch.shift_type,
                -- Requester details
                hu.username as requested_by_username
            FROM shift_swap_requests ssr
            JOIN schedules sch ON ssr.schedule_id = sch.id
            JOIN staff s1 ON ssr.from_staff_id = s1.id
            JOIN staff s2 ON ssr.to_staff_id = s2.id
            LEFT JOIN hr_users hu ON ssr.requested_by = hu.id
            ORDER BY 
                CASE 
                    WHEN ssr.status = 'Pending' THEN 1
                    WHEN ssr.status = 'Approved' THEN 2
                    ELSE 3
                END,
                ssr.requested_at DESC
        """)
        
        all_swaps = cur.fetchall()
        
        # Separate pending and history swaps
        pending_swaps = []
        swap_history = []
        
        for swap in all_swaps:
            swap_dict = {
                'id': swap[0],
                'schedule_id': swap[1],
                'from_staff_id': swap[2],
                'to_staff_id': swap[3],
                'reason': swap[4] or 'No reason provided',
                'status': swap[5],
                'requested_at': swap[6],
                'approved_at': swap[7],
                'rejection_reason': swap[8],
                'from_staff_name': f"{swap[9]} {swap[10]}",
                'from_staff_position': swap[11],
                'to_staff_name': f"{swap[12]} {swap[13]}",
                'to_staff_position': swap[14],
                'schedule_date': swap[15],
                'start_time': swap[16],
                'end_time': swap[17],
                'shift_type': swap[18] or 'Regular',
                'requested_by': swap[19]
            }
            
            if swap_dict['status'] == 'Pending':
                pending_swaps.append(swap_dict)
            else:
                swap_history.append(swap_dict)
        
        # Get upcoming schedules for swap requests (next 30 days)
        thirty_days_later = (date.today() + timedelta(days=30)).isoformat()
        cur.execute("""
            SELECT 
                sch.id,
                sch.schedule_date,
                sch.start_time,
                sch.end_time,
                sch.shift_type,
                st.first_name,
                st.last_name,
                st.position,
                d.name as department_name
            FROM schedules sch
            JOIN staff st ON sch.staff_id = st.id
            LEFT JOIN departments d ON st.department_id = d.id
            WHERE sch.schedule_date BETWEEN ? AND ?
            ORDER BY sch.schedule_date, sch.start_time
            LIMIT 50
        """, (date.today().isoformat(), thirty_days_later))
        
        upcoming_schedules = cur.fetchall()
        formatted_schedules = []
        for sched in upcoming_schedules:
            formatted_schedules.append({
                'id': sched[0],
                'date': sched[1],
                'start_time': sched[2],
                'end_time': sched[3],
                'shift_type': sched[4],
                'staff_name': f"{sched[5]} {sched[6]}",
                'position': sched[7],
                'department': sched[8] or 'N/A'
            })
        
        # Get all active staff for swap requests
        cur.execute("""
            SELECT id, first_name, last_name, position, 
                   COALESCE(d.name, 'No Department') as department
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.status = 'Active' 
            ORDER BY first_name, last_name
        """)
        
        staff_list = cur.fetchall()
        formatted_staff = []
        for staff in staff_list:
            formatted_staff.append({
                'id': staff[0],
                'name': f"{staff[1]} {staff[2]}",
                'position': staff[3],
                'department': staff[4]
            })
        
        # Get statistics
        cur.execute("SELECT COUNT(*) FROM shift_swap_requests WHERE status = 'Pending'")
        pending_count = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM shift_swap_requests WHERE status = 'Approved' AND approved_at >= date('now', '-30 days')")
        approved_30days = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error loading shift swap data: {e}")
        pending_swaps = []
        swap_history = []
        formatted_schedules = []
        formatted_staff = []
        pending_count = 0
        approved_30days = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "shift_swap.html",
        pending_swaps=pending_swaps,
        swap_history=swap_history,
        upcoming_schedules=formatted_schedules,
        staff_list=formatted_staff,
        pending_count=pending_count,
        approved_30days=approved_30days,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )


@app.route("/hr/scheduling/check-availability", methods=["POST"])
def check_availability():
    """AJAX endpoint to check staff availability"""
    if "hr_user_id" not in session:
        return jsonify({"available": False, "message": "Unauthorized"}), 401
    
    data = request.json
    staff_id = data.get("staff_id")
    schedule_date = data.get("schedule_date")
    
    if not all([staff_id, schedule_date]):
        return jsonify({"available": False, "message": "Missing parameters"}), 400
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Check if staff has schedule on that date
        cur.execute("""
            SELECT id, shift_type, start_time, end_time 
            FROM schedules 
            WHERE staff_id = ? AND schedule_date = ?
        """, (staff_id, schedule_date))
        
        existing = cur.fetchone()
        available = existing is None
        
        if available:
            return jsonify({
                "available": True,
                "message": "Staff is available on this date"
            })
        else:
            return jsonify({
                "available": False,
                "message": f"Staff already scheduled for {existing[1]} shift ({existing[2]} - {existing[3]})"
            })
        
    except Exception as e:
        app.logger.error(f"Error checking availability: {e}")
        return jsonify({"available": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()
        
# -------------------- ROUTES: LABORATORY MODULE --------------------

@app.route('/laboratory/login', methods=['GET', 'POST'])
def laboratory_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return render_template("laboratory_login.html")

        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, username, password, full_name, role FROM lab_users WHERE username=? AND is_active=1",
            (username,)
        )
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if user and check_password_hash(user[2], password):
            session['lab_user_id'] = user[0]
            session['lab_username'] = user[1]
            session['lab_full_name'] = user[3]
            session['lab_role'] = user[4]
            return redirect(url_for('laboratory_dashboard'))
        else:
            flash("Invalid username or password", "danger")

    return render_template("laboratory_login.html")

@app.route('/laboratory/dashboard')
def laboratory_dashboard():
    if 'lab_user_id' not in session:
        return redirect(url_for('laboratory_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get statistics
        cur.execute("SELECT COUNT(*) FROM lab_requests WHERE status = 'Pending'")
        pending_tests = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM lab_requests WHERE status = 'In Progress'")
        in_progress = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM lab_requests WHERE status = 'Completed' AND DATE(completed_date) = DATE('now')")
        completed_today = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM lab_tests WHERE is_active = 1")
        total_tests = cur.fetchone()[0] or 0
        
        # Get recent requests
        cur.execute("""
            SELECT id, request_no, patient_name, test_names, status, priority, requested_date
            FROM lab_requests
            ORDER BY created_at DESC
            LIMIT 10
        """)
        recent_requests = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error loading lab dashboard: {e}")
        pending_tests = in_progress = completed_today = total_tests = 0
        recent_requests = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "laboratory_dashboard.html",
        pending_tests=pending_tests,
        in_progress=in_progress,
        completed_today=completed_today,
        total_tests=total_tests,
        recent_requests=recent_requests,
        lab_user=session.get('lab_full_name')
    )

@app.route('/laboratory/logout')
def laboratory_logout():
    session.pop('lab_user_id', None)
    session.pop('lab_username', None)
    session.pop('lab_full_name', None)
    session.pop('lab_role', None)
    flash("Logged out successfully", "success")
    return redirect(url_for('laboratory_login'))

@app.route('/laboratory/tests')
def lab_tests():
    if 'lab_user_id' not in session:
        return redirect(url_for('laboratory_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT id, test_id, test_name, category, price, description, 
               normal_range, turnaround_time, is_active
        FROM lab_tests
        ORDER BY test_name
    """)
    tests = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("lab_tests.html", tests=tests)

@app.route('/laboratory/tests/add', methods=['GET', 'POST'])
def add_lab_test():
    if 'lab_user_id' not in session:
        return redirect(url_for('laboratory_login'))
    
    if request.method == 'POST':
        test_id = request.form.get('test_id')
        test_name = request.form.get('test_name')
        category = request.form.get('category')
        price = float(request.form.get('price', 0))
        description = request.form.get('description')
        preparation_instructions = request.form.get('preparation_instructions')
        normal_range = request.form.get('normal_range')
        turnaround_time = request.form.get('turnaround_time')
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            cur.execute("""
                INSERT INTO lab_tests (test_id, test_name, category, price, description,
                                       preparation_instructions, normal_range, turnaround_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (test_id, test_name, category, price, description, 
                  preparation_instructions, normal_range, turnaround_time))
            conn.commit()
            flash(f"Test '{test_name}' added successfully!", "success")
            return redirect(url_for('lab_tests'))
        except Exception as e:
            conn.rollback()
            flash(f"Error adding test: {str(e)}", "danger")
        finally:
            cur.close()
            conn.close()
    
    return render_template("add_lab_test.html")

@app.route('/laboratory/requests')
def lab_requests():
    if 'lab_user_id' not in session:
        return redirect(url_for('laboratory_login'))
    
    status_filter = request.args.get('status', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    query = """
        SELECT id, request_no, patient_name, patient_id, test_names, 
               grand_total as total_amount, status, priority, requested_date
        FROM lab_requests
    """
    params = []
    
    if status_filter:
        query += " WHERE status = ?"
        params.append(status_filter)
    
    query += " ORDER BY requested_date DESC"
    
    cur.execute(query, params)
    lab_requests_list = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("lab_requests.html", requests=lab_requests_list, status_filter=status_filter)

@app.route('/laboratory/request/new', methods=['GET', 'POST'])
def new_lab_request():
    if 'lab_user_id' not in session:
        return redirect(url_for('laboratory_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == 'POST':
        patient_name = request.form.get('patient_name')
        patient_id = request.form.get('patient_id')
        patient_age = request.form.get('patient_age')
        patient_gender = request.form.get('patient_gender')
        doctor_name = request.form.get('doctor_name')
        test_ids = request.form.getlist('test_ids')
        custom_prices_json = request.form.get('custom_prices', '{}')
        discount_amount = float(request.form.get('discount_amount', 0))
        priority = request.form.get('priority', 'Normal')
        notes = request.form.get('notes')
        
        if not patient_name:
            flash("Patient name is required", "danger")
            return redirect(url_for('new_lab_request'))
        
        if not test_ids:
            flash("Please add at least one test", "danger")
            return redirect(url_for('new_lab_request'))
        
        # Parse custom prices
        import json
        custom_prices = json.loads(custom_prices_json)
        
        # Get test details
        placeholders = ','.join(['?' for _ in test_ids])
        cur.execute(f"""
            SELECT id, test_name, price FROM lab_tests 
            WHERE id IN ({placeholders}) AND is_active = 1
        """, test_ids)
        tests = cur.fetchall()
        
        if not tests:
            flash("Selected tests not found", "danger")
            return redirect(url_for('new_lab_request'))
        
        # Use custom prices if provided, otherwise use default prices
        test_names = []
        total_amount = 0
        for test in tests:
            test_id_str = str(test[0])
            test_names.append(test[1])
            price = custom_prices.get(test_id_str, test[2])
            total_amount += float(price)
        
        # Apply discount
        grand_total = total_amount - discount_amount
        if grand_total < 0:
            grand_total = 0
        
        # Generate unique request number
        request_no = f"LAB-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        
        try:
            cur.execute("""
                INSERT INTO lab_requests (
                    request_no, patient_name, patient_id, patient_age, patient_gender,
                    doctor_name, test_ids, test_names, total_amount, discount, grand_total,
                    priority, notes, requested_date, created_by, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending')
            """, (request_no, patient_name, patient_id, patient_age, patient_gender,
                  doctor_name, ','.join(test_ids), ','.join(test_names), total_amount,
                  discount_amount, grand_total, priority, notes, 
                  date.today().isoformat(), session['lab_user_id']))
            
            conn.commit()
            flash(f"Lab request {request_no} created successfully! Total: ₦{grand_total:,.2f}", "success")
            return redirect(url_for('lab_requests'))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating request: {e}")
            flash(f"Error creating request: {str(e)}", "danger")
    
    # GET request - show form
    cur.execute("SELECT id, test_name, price, category FROM lab_tests WHERE is_active = 1 ORDER BY category, test_name")
    tests = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("new_lab_request.html", tests=tests)

@app.route('/laboratory/request/<int:request_id>')
def view_lab_request(request_id):
    if 'lab_user_id' not in session:
        return redirect(url_for('laboratory_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("SELECT * FROM lab_requests WHERE id = ?", (request_id,))
    lab_request = cur.fetchone()
    
    if lab_request:
        # Get results if any
        cur.execute("SELECT * FROM lab_results WHERE request_id = ?", (request_id,))
        results = cur.fetchall()
    else:
        results = []
    
    cur.close()
    conn.close()
    
    if not lab_request:
        flash("Lab request not found", "danger")
        return redirect(url_for('lab_requests'))
    
    return render_template("view_lab_request.html", request=lab_request, results=results)

@app.route('/laboratory/request/<int:request_id>/process', methods=['GET', 'POST'])
def process_lab_request(request_id):
    if 'lab_user_id' not in session:
        return redirect(url_for('laboratory_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == 'POST':
        # Update status to In Progress
        cur.execute("""
            UPDATE lab_requests 
            SET status = 'In Progress', collected_date = ?, collected_by = ?
            WHERE id = ?
        """, (date.today().isoformat(), session['lab_full_name'], request_id))
        conn.commit()
        flash("Sample collected and test is now in progress", "success")
        return redirect(url_for('enter_lab_results', request_id=request_id))
    
    # GET request - fetch the lab request
    cur.execute("SELECT * FROM lab_requests WHERE id = ?", (request_id,))
    lab_request = cur.fetchone()
    cur.close()
    conn.close()
    
    if not lab_request:
        flash("Lab request not found", "danger")
        return redirect(url_for('lab_requests'))
    
    return render_template("process_lab_request.html", request=lab_request)

@app.route('/laboratory/request/<int:request_id>/results', methods=['GET', 'POST'])
def enter_lab_results(request_id):
    if 'lab_user_id' not in session:
        return redirect(url_for('laboratory_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == 'POST':
        # Save results for each test
        for key, value in request.form.items():
            if key.startswith('result_'):
                test_id = key.split('_')[1]
                result_value = value
                interpretation = request.form.get(f'interpretation_{test_id}')
                remarks = request.form.get(f'remarks_{test_id}')
                
                cur.execute("""
                    INSERT OR REPLACE INTO lab_results (
                        request_id, test_id, test_name, result_value, 
                        interpretation, remarks, performed_by, performed_date
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (request_id, test_id, 
                      request.form.get(f'test_name_{test_id}'),
                      result_value, interpretation, remarks,
                      session['lab_full_name'], date.today().isoformat()))
        
        # Update request status to Completed
        cur.execute("""
            UPDATE lab_requests 
            SET status = 'Completed', completed_date = ?, completed_by = ?
            WHERE id = ?
        """, (date.today().isoformat(), session['lab_full_name'], request_id))
        
        conn.commit()
        flash("Results saved and request marked as completed!", "success")
        return redirect(url_for('lab_requests'))
    
    # GET request - fetch the lab request
    cur.execute("SELECT * FROM lab_requests WHERE id = ?", (request_id,))
    lab_request = cur.fetchone()
    
    if not lab_request:
        flash("Lab request not found", "danger")
        return redirect(url_for('lab_requests'))
    
    # Parse test IDs and names
    test_ids = lab_request[7].split(',') if lab_request[7] else []
    test_names = lab_request[8].split(',') if lab_request[8] else []
    
    # Get existing results
    cur.execute("SELECT * FROM lab_results WHERE request_id = ?", (request_id,))
    existing_results = {}
    for row in cur.fetchall():
        existing_results[str(row[2])] = row
    
    cur.close()
    conn.close()
    
    return render_template("enter_lab_results.html", 
                         request=lab_request, 
                         test_ids=test_ids, 
                         test_names=test_names,
                         existing_results=existing_results)

@app.route('/laboratory/reports')
def lab_reports():
    if 'lab_user_id' not in session:
        return redirect(url_for('laboratory_login'))
    
    report_type = request.args.get('type', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    month = request.args.get('month', date.today().month)
    year = request.args.get('year', date.today().year)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        if report_type == 'daily':
            if not start_date:
                start_date = date.today().isoformat()
            
            # Get stats
            cur.execute("""
                SELECT COUNT(*), COALESCE(SUM(grand_total), 0)
                FROM lab_requests 
                WHERE DATE(requested_date) = ? AND status = 'Completed'
            """, (start_date,))
            stats = cur.fetchone()
            
            # Get requests
            cur.execute("""
                SELECT id, request_no, patient_name, test_names, grand_total as amount, completed_date
                FROM lab_requests 
                WHERE DATE(requested_date) = ? AND status = 'Completed'
                ORDER BY completed_date DESC
            """, (start_date,))
            lab_requests_list = cur.fetchall()
            
        else:  # monthly
            # Get stats
            cur.execute("""
                SELECT COUNT(*), COALESCE(SUM(grand_total), 0)
                FROM lab_requests 
                WHERE strftime('%Y-%m', requested_date) = ? AND status = 'Completed'
            """, (f"{year}-{int(month):02d}",))
            stats = cur.fetchone()
            
            # Get requests
            cur.execute("""
                SELECT id, request_no, patient_name, test_names, grand_total as amount, requested_date
                FROM lab_requests 
                WHERE strftime('%Y-%m', requested_date) = ? AND status = 'Completed'
                ORDER BY requested_date DESC
            """, (f"{year}-{int(month):02d}",))
            lab_requests_list = cur.fetchall()
        
        # Convert to list of dictionaries for easier template access
        requests_data = []
        for req in lab_requests_list:
            requests_data.append({
                'id': req[0],
                'request_no': req[1],
                'patient_name': req[2],
                'test_names': req[3],
                'amount': req[4],
                'date': req[5]
            })
        
    except Exception as e:
        app.logger.error(f"Error generating lab reports: {e}")
        stats = (0, 0)
        requests_data = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("lab_reports.html", 
                         report_type=report_type,
                         stats=stats,
                         requests=requests_data,
                         start_date=start_date,
                         month=month,
                         year=year)
def create_default_hr_data():
    """Insert default HR data into SQLite tables."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    # Default password
    default_password = 'hr@admin123'
    hashed_password = generate_password_hash(default_password)
    
    # Insert or update default HR users
    try:
        # Check if hr_admin exists
        cursor.execute("SELECT id, password FROM hr_users WHERE username = 'hr_admin'")
        admin = cursor.fetchone()
        
        if admin:
            # Check if password needs updating
            try:
                if not check_password_hash(admin[1], default_password):
                    cursor.execute("UPDATE hr_users SET password = ? WHERE username = 'hr_admin'", (hashed_password,))
                    print("Updated hr_admin password")
            except (ValueError, TypeError):
                cursor.execute("UPDATE hr_users SET password = ? WHERE username = 'hr_admin'", (hashed_password,))
                print("Updated hr_admin password (was not hashed)")
        else:
            cursor.execute("""
                INSERT INTO hr_users (username, password, full_name, email, role, is_active) 
                VALUES (?, ?, ?, ?, ?, ?)
            """, ('hr_admin', hashed_password, 'HR Administrator', 'admin@hospital.com', 'HR Manager', 1))
            print("Created hr_admin user")
        
        # Check if hr_staff exists
        cursor.execute("SELECT id, password FROM hr_users WHERE username = 'hr_staff'")
        staff = cursor.fetchone()
        
        if staff:
            try:
                if not check_password_hash(staff[1], default_password):
                    cursor.execute("UPDATE hr_users SET password = ? WHERE username = 'hr_staff'", (hashed_password,))
                    print("Updated hr_staff password")
            except (ValueError, TypeError):
                cursor.execute("UPDATE hr_users SET password = ? WHERE username = 'hr_staff'", (hashed_password,))
                print("Updated hr_staff password (was not hashed)")
        else:
            cursor.execute("""
                INSERT INTO hr_users (username, password, full_name, email, role, is_active) 
                VALUES (?, ?, ?, ?, ?, ?)
            """, ('hr_staff', hashed_password, 'HR Staff', 'staff@hospital.com', 'HR Officer', 1))
            print("Created hr_staff user")
            
    except Exception as e:
        app.logger.error(f"Error inserting HR users: {e}")
        print(f"HR users error: {e}")
    
    # Insert sample departments
    departments = [
        ('Administration', 'ADMIN', 'Hospital Administration and Management', 'Dr. John Smith'),
        ('Medical', 'MED', 'Medical Services Department', 'Dr. Sarah Johnson'),
        ('Nursing', 'NURS', 'Nursing Services', 'Mrs. Grace Williams'),
        ('Pharmacy', 'PHARM', 'Pharmacy Department', 'Mr. Michael Brown'),
        ('Laboratory', 'LAB', 'Laboratory Services', 'Dr. David Miller'),
        ('Radiology', 'RAD', 'Radiology Department', 'Dr. Lisa Davis'),
        ('Finance', 'FIN', 'Finance and Billing Department', 'Mr. Robert Wilson'),
        ('Human Resources', 'HR', 'Human Resources Department', 'Ms. Patricia Taylor'),
        ('Maintenance', 'MAINT', 'Facility Maintenance', 'Mr. Thomas Anderson'),
        ('Security', 'SEC', 'Hospital Security', 'Mr. Richard Clark')
    ]
    
    for dept in departments:
        try:
            cursor.execute("""
                INSERT OR IGNORE INTO departments (name, code, description, head_of_dept, status) 
                VALUES (?, ?, ?, ?, 'Active')
            """, dept)
        except Exception as e:
            app.logger.error(f"Error inserting department {dept[0]}: {e}")
    
    # Get admin department ID for sample staff
    cursor.execute("SELECT id FROM departments WHERE code = 'ADMIN' AND status = 'Active' LIMIT 1;")
    admin_dept = cursor.fetchone()
    
    # Insert sample staff if departments exist
    if admin_dept:
        admin_dept_id = admin_dept[0]
        sample_staff = [
            ('EMP001', 'John', 'Doe', admin_dept_id, 'Hospital Administrator', 'Full-Time', 
             'john.doe@hospital.com', '08012345678', '2022-01-15', 850000.00, 'Jane Doe - 08087654321', '123 Admin Street, Enugu'),
            ('EMP002', 'Sarah', 'Johnson', admin_dept_id, 'Senior Doctor', 'Full-Time', 
             'sarah.j@hospital.com', '08023456789', '2021-03-20', 1200000.00, 'Mark Johnson - 08098765432', '456 Medical Road, Enugu'),
            ('EMP003', 'Michael', 'Brown', admin_dept_id, 'Chief Pharmacist', 'Full-Time', 
             'michael.b@hospital.com', '08034567890', '2020-06-10', 950000.00, 'Emily Brown - 08076543210', '789 Pharmacy Lane, Enugu'),
            ('EMP004', 'Grace', 'Williams', admin_dept_id, 'Head Nurse', 'Full-Time', 
             'grace.w@hospital.com', '08045678901', '2019-08-05', 750000.00, 'James Williams - 08065432109', '321 Nursing Avenue, Enugu'),
            ('EMP005', 'David', 'Miller', admin_dept_id, 'Lab Technician', 'Full-Time', 
             'david.m@hospital.com', '08056789012', '2022-11-30', 650000.00, 'Sarah Miller - 08054321098', '654 Lab Street, Enugu')
        ]
        
        for staff in sample_staff:
            try:
                cursor.execute("""
                    INSERT OR IGNORE INTO staff (
                        staff_id, first_name, last_name, department_id, position, 
                        employment_type, email, phone, hire_date, salary, 
                        emergency_contact, address, status
                    ) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active')
                """, staff)
            except Exception as e:
                app.logger.error(f"Error inserting staff {staff[0]}: {e}")
        
        print(f"Inserted {len(sample_staff)} sample staff records")
    
    conn.commit()
    
    # Verify the data was inserted correctly
    try:
        cursor.execute("SELECT COUNT(*) FROM hr_users")
        user_count = cursor.fetchone()[0]
        print(f"Total HR users in database: {user_count}")
        
        cursor.execute("SELECT COUNT(*) FROM departments")
        dept_count = cursor.fetchone()[0]
        print(f"Total departments in database: {dept_count}")
        
        cursor.execute("SELECT COUNT(*) FROM staff")
        staff_count = cursor.fetchone()[0]
        print(f"Total staff in database: {staff_count}")
        
    except Exception as e:
        app.logger.error(f"Error verifying data: {e}")
    
    cursor.close()
    conn.close()
    print("HR data initialization completed")
        
def create_sample_lab_tests():
    """Insert sample lab tests into database."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    sample_tests = [
        ('CBC-001', 'Complete Blood Count', 'Hematology', 5000.00, 'Complete blood count test', 'No special preparation', 'Various', '2 hours'),
        ('LFT-001', 'Liver Function Test', 'Biochemistry', 8000.00, 'Liver function test panel', 'Fasting required (8 hours)', 'Various', '4 hours'),
        ('RFT-001', 'Renal Function Test', 'Biochemistry', 7000.00, 'Kidney function test', 'Fasting required (8 hours)', 'Various', '4 hours'),
        ('LIP-001', 'Lipid Profile', 'Biochemistry', 6000.00, 'Cholesterol and lipids test', 'Fasting required (12 hours)', 'Various', '4 hours'),
        ('U/A-001', 'Urinalysis', 'Urinalysis', 3000.00, 'Urine analysis', 'Clean catch mid-stream urine', 'Various', '1 hour'),
        ('MP-001', 'Malaria Parasite', 'Microbiology', 2000.00, 'Malaria blood test', 'No special preparation', 'Negative', '1 hour'),
        ('W/F-001', 'Widal Test', 'Microbiology', 4000.00, 'Typhoid fever test', 'No special preparation', 'Negative', '3 hours'),
        ('HBsAg-001', 'Hepatitis B Surface Antigen', 'Immunology', 4500.00, 'Hepatitis B screening', 'No special preparation', 'Negative', '3 hours'),
        ('HIV-001', 'HIV Test', 'Immunology', 3500.00, 'HIV screening test', 'No special preparation', 'Negative', '2 hours'),
        ('HBA1C-001', 'HbA1c', 'Biochemistry', 5500.00, 'Diabetes monitoring test', 'No special preparation', '4-5.6%', '3 hours'),
    ]
    
    for test in sample_tests:
        try:
            cursor.execute("""
                INSERT OR IGNORE INTO lab_tests 
                (test_id, test_name, category, price, description, preparation_instructions, normal_range, turnaround_time, is_active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
            """, test)
        except Exception as e:
            app.logger.error(f"Error inserting test {test[0]}: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()
    print("Sample lab tests added successfully!")

# Call this function after create_tables() in your main block

def update_lab_requests_table():
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Check if grand_total column exists, if not add it
        cur.execute("PRAGMA table_info(lab_requests)")
        columns = [col[1] for col in cur.fetchall()]
        
        if 'grand_total' not in columns:
            cur.execute("ALTER TABLE lab_requests ADD COLUMN grand_total DECIMAL(10, 2) DEFAULT 0")
            print("Added grand_total column")
        
        if 'discount' not in columns:
            cur.execute("ALTER TABLE lab_requests ADD COLUMN discount DECIMAL(10, 2) DEFAULT 0")
            print("Added discount column")
        
        if 'amount_paid' not in columns:
            cur.execute("ALTER TABLE lab_requests ADD COLUMN amount_paid DECIMAL(10, 2) DEFAULT 0")
            print("Added amount_paid column")
        
        if 'balance' not in columns:
            cur.execute("ALTER TABLE lab_requests ADD COLUMN balance DECIMAL(10, 2) DEFAULT 0")
            print("Added balance column")
        
        # Update existing records - set grand_total = total_amount for old records
        if 'grand_total' in columns and 'total_amount' in columns:
            cur.execute("UPDATE lab_requests SET grand_total = total_amount WHERE grand_total IS NULL OR grand_total = 0")
            print("Updated grand_total for existing records")
        
        conn.commit()
        print("Lab requests table updated successfully!")
        
    except Exception as e:
        print(f"Error updating table: {e}")
        conn.rollback()
    finally:
        cur.close()
        conn.close()

# Call this function after create_tables() in your main block
# if __name__ == "__main__":
#     create_tables()
#     create_default_users()
#     create_default_hr_data()
#     create_sample_lab_tests()
#     update_lab_requests_table()  # Add this line
#     app.run(debug=True)
# @app.route('/laboratory/tests/<int:test_id>/view')
# def view_lab_test(test_id):
#     if 'lab_user_id' not in session:
#         return jsonify({"success": False, "message": "Unauthorized"}), 401
    
#     conn = get_db_connection()
#     cur = conn.cursor()
    
#     cur.execute("SELECT * FROM lab_tests WHERE id = ?", (test_id,))
#     test = cur.fetchone()
#     cur.close()
#     conn.close()
    
#     if test:
#         return jsonify({
#             "success": True,
#             "test": {
#                 "id": test[0],
#                 "test_id": test[1],
#                 "test_name": test[2],
#                 "category": test[3],
#                 "price": test[4],
#                 "description": test[5],
#                 "preparation_instructions": test[6],
#                 "normal_range": test[7],
#                 "turnaround_time": test[8],
#                 "is_active": test[9]
#             }
#         })
#     else:
#         return jsonify({"success": False, "message": "Test not found"}), 404

# @app.route('/laboratory/tests/<int:test_id>/edit', methods=['POST'])
# def edit_lab_test(test_id):
#     if 'lab_user_id' not in session:
#         return jsonify({"success": False, "message": "Unauthorized"}), 401
    
#     data = request.get_json()
    
#     conn = get_db_connection()
#     cur = conn.cursor()
    
#     try:
#         cur.execute("""
#             UPDATE lab_tests 
#             SET test_name = ?,
#                 category = ?,
#                 price = ?,
#                 description = ?,
#                 preparation_instructions = ?,
#                 normal_range = ?,
#                 turnaround_time = ?
#             WHERE id = ?
#         """, (data['test_name'], data['category'], data['price'], 
#               data['description'], data['preparation_instructions'],
#               data['normal_range'], data['turnaround_time'], test_id))
        
#         conn.commit()
#         return jsonify({"success": True, "message": "Test updated successfully"})
        
#     except Exception as e:
#         conn.rollback()
#         return jsonify({"success": False, "message": str(e)}), 500
#     finally:
#         cur.close()
#         conn.close()

# @app.route('/laboratory/tests/<int:test_id>/toggle-status', methods=['POST'])
# def toggle_lab_test_status(test_id):
#     if 'lab_user_id' not in session:
#         return jsonify({"success": False, "message": "Unauthorized"}), 401
    
#     data = request.get_json()
#     is_active = data.get('is_active', 0)
    
#     conn = get_db_connection()
#     cur = conn.cursor()
    
#     try:
#         cur.execute("UPDATE lab_tests SET is_active = ? WHERE id = ?", (is_active, test_id))
#         conn.commit()
#         return jsonify({"success": True, "message": "Status updated successfully"})
        
#     except Exception as e:
#         conn.rollback()
#         return jsonify({"success": False, "message": str(e)}), 500
#     finally:
#         cur.close()
#         conn.close()
        
        
@app.route('/laboratory/tests/<int:test_id>/view')
def view_lab_test(test_id):
    if 'lab_user_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("SELECT * FROM lab_tests WHERE id = ?", (test_id,))
        test = cur.fetchone()
        cur.close()
        conn.close()
        
        if test:
            # Convert to dictionary for JSON response
            return jsonify({
                "success": True,
                "test": {
                    "id": test[0],
                    "test_id": test[1],
                    "test_name": test[2],
                    "category": test[3],
                    "price": test[4],
                    "description": test[5] or '',
                    "preparation_instructions": test[6] or '',
                    "normal_range": test[7] or '',
                    "turnaround_time": test[8] or '',
                    "is_active": test[9]
                }
            })
        else:
            return jsonify({"success": False, "message": "Test not found"}), 404
    except Exception as e:
        app.logger.error(f"Error viewing test: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/laboratory/tests/<int:test_id>/edit', methods=['POST'])
def edit_lab_test(test_id):
    if 'lab_user_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    try:
        data = request.get_json()
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            UPDATE lab_tests 
            SET test_name = ?,
                category = ?,
                price = ?,
                description = ?,
                preparation_instructions = ?,
                normal_range = ?,
                turnaround_time = ?
            WHERE id = ?
        """, (data.get('test_name'), data.get('category'), data.get('price'), 
              data.get('description', ''), data.get('preparation_instructions', ''),
              data.get('normal_range', ''), data.get('turnaround_time', ''), test_id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Test updated successfully"})
        
    except Exception as e:
        app.logger.error(f"Error editing test: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/laboratory/tests/<int:test_id>/toggle-status', methods=['POST'])
def toggle_lab_test_status(test_id):
    if 'lab_user_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    try:
        data = request.get_json()
        is_active = data.get('is_active', 0)
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("UPDATE lab_tests SET is_active = ? WHERE id = ?", (is_active, test_id))
        conn.commit()
        cur.close()
        conn.close()
        
        status_text = "activated" if is_active == 1 else "deactivated"
        return jsonify({"success": True, "message": f"Test {status_text} successfully"})
        
    except Exception as e:
        app.logger.error(f"Error toggling test status: {e}")
        return jsonify({"success": False, "message": str(e)}), 500   
    
# -------------------- ROUTES: PATIENT SERVICES MODULE --------------------

# 

# @app.route('/patient_services/dashboard')
# def patient_services_dashboard():
#     if 'patient_user_id' not in session:
#         return redirect(url_for('patient_services_login'))
    
#     conn = get_db_connection()
#     cur = conn.cursor()
    
#     try:
#         # Today's statistics
#         today = date.today().isoformat()
        
#         cur.execute("SELECT COUNT(*) FROM patients WHERE registration_date = ?", (today,))
#         today_registrations = cur.fetchone()[0] or 0
        
#         cur.execute("SELECT COUNT(*) FROM appointments WHERE appointment_date = ? AND status = 'Scheduled'", (today,))
#         today_appointments = cur.fetchone()[0] or 0
        
#         cur.execute("SELECT COUNT(*) FROM queue WHERE status = 'Waiting'")
#         waiting_queue = cur.fetchone()[0] or 0
        
#         cur.execute("SELECT COUNT(*) FROM patient_services WHERE DATE(check_in_time) = ? AND status = 'Active'", (today,))
#         active_services = cur.fetchone()[0] or 0
        
#         # Current queue
#         cur.execute("""
#             SELECT id, token_no, patient_name, service_type, priority, queue_position, created_at
#             FROM queue 
#             WHERE status = 'Waiting'
#             ORDER BY queue_position ASC
#             LIMIT 10
#         """)
#         current_queue = cur.fetchall()
        
#     except Exception as e:
#         app.logger.error(f"Error loading patient dashboard: {e}")
#         today_registrations = today_appointments = waiting_queue = active_services = 0
#         current_queue = []
    
#     finally:
#         cur.close()
#         conn.close()
    
#     return render_template("patient_services/dashboard.html",
#                          today_registrations=today_registrations,
#                          today_appointments=today_appointments,
#                          waiting_queue=waiting_queue,
#                          active_services=active_services,
#                          current_queue=current_queue,
#                          user_name=session.get('patient_full_name'))

# @app.route('/patient_services/logout')
# def patient_services_logout():
#     session.pop('patient_user_id', None)
#     session.pop('patient_username', None)
#     session.pop('patient_full_name', None)
#     session.pop('patient_role', None)
#     flash("Logged out successfully", "success")
#     return redirect(url_for('patient_services_login'))

# ==================== PATIENT REGISTRATION ====================

@app.route('/patient_services/patients/register', methods=['GET', 'POST'])
def register_patient():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    if request.method == 'POST':
        # Get form data
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        date_of_birth = request.form.get('date_of_birth')
        gender = request.form.get('gender')
        blood_group = request.form.get('blood_group')
        genotype = request.form.get('genotype')
        phone = request.form.get('phone')
        email = request.form.get('email')
        address = request.form.get('address')
        emergency_contact_name = request.form.get('emergency_contact_name')
        emergency_contact_phone = request.form.get('emergency_contact_phone')
        occupation = request.form.get('occupation')
        marital_status = request.form.get('marital_status')
        nationality = request.form.get('nationality')
        religion = request.form.get('religion')
        next_of_kin = request.form.get('next_of_kin')
        next_of_kin_phone = request.form.get('next_of_kin_phone')
        notes = request.form.get('notes')
        
        # Validate required fields
        if not all([first_name, last_name, date_of_birth, gender]):
            flash("Please fill in all required fields", "danger")
            return redirect(url_for('register_patient'))
        
        # Generate unique IDs
        import random
        patient_id = f"PT-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        hospital_no = f"HN-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        
        # Generate sequential card number
        card_number = generate_card_number()
        
        # Handle empty values - convert empty strings to None
        blood_group = blood_group if blood_group and blood_group != '' else None
        genotype = genotype if genotype and genotype != '' else None
        phone = phone if phone and phone != '' else None
        email = email if email and email != '' else None
        address = address if address and address != '' else None
        emergency_contact_name = emergency_contact_name if emergency_contact_name and emergency_contact_name != '' else None
        emergency_contact_phone = emergency_contact_phone if emergency_contact_phone and emergency_contact_phone != '' else None
        occupation = occupation if occupation and occupation != '' else None
        marital_status = marital_status if marital_status and marital_status != '' else None
        nationality = nationality if nationality and nationality != '' else 'Nigerian'
        religion = religion if religion and religion != '' else None
        next_of_kin = next_of_kin if next_of_kin and next_of_kin != '' else None
        next_of_kin_phone = next_of_kin_phone if next_of_kin_phone and next_of_kin_phone != '' else None
        notes = notes if notes and notes != '' else None
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            cur.execute("""
                INSERT INTO patients (
                    patient_id, hospital_no, card_number, first_name, last_name, date_of_birth, gender,
                    blood_group, genotype, phone, email, address, emergency_contact_name,
                    emergency_contact_phone, occupation, marital_status, nationality, religion,
                    next_of_kin, next_of_kin_phone, registration_date, registration_time, notes, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (patient_id, hospital_no, card_number, first_name, last_name, date_of_birth, gender,
                  blood_group, genotype, phone, email, address, emergency_contact_name,
                  emergency_contact_phone, occupation, marital_status, nationality, religion,
                  next_of_kin, next_of_kin_phone, date.today().isoformat(), 
                  datetime.now().strftime('%H:%M:%S'), notes, session['patient_user_id']))
            
            conn.commit()
            patient_db_id = cur.lastrowid
            
            flash(f"Patient {first_name} {last_name} registered successfully! Card Number: {card_number}", "success")
            return redirect(url_for('view_patient', patient_id=patient_db_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error registering patient: {e}")
            flash(f"Error registering patient: {str(e)}", "danger")
            return redirect(url_for('register_patient'))
        finally:
            cur.close()
            conn.close()
    
    # GET request - show form
    # Import date here to avoid conflict with variable name
    from datetime import date as date_today
    current_date = date_today.today()
    max_date = current_date.isoformat()
    
    return render_template("patient_services/register_patient.html", 
                         current_date=current_date,
                         max_date=max_date)


@app.route('/patient_services/patients/search')
def search_patients():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    search_term = request.args.get('q', '')
    show_deleted = request.args.get('show_deleted', '0')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Build the query with status filter
    query = """
        SELECT 
            p.id, 
            p.patient_id, 
            p.hospital_no, 
            p.card_number, 
            p.first_name, 
            p.last_name, 
            p.phone, 
            p.registration_date,
            p.status,
            p.blood_group,
            p.date_of_birth,
            p.gender,
            p.email,
            p.address
        FROM patients p
        WHERE 1=1
    """
    params = []
    
    # Filter by deleted status
    if show_deleted != '1':
        query += " AND p.status != 'Deleted'"
    
    # Search term filter
    if search_term:
        query += """ AND (p.first_name LIKE ? OR p.last_name LIKE ? 
                   OR p.patient_id LIKE ? OR p.phone LIKE ? 
                   OR p.hospital_no LIKE ? OR p.card_number LIKE ?)"""
        search_pattern = f'%{search_term}%'
        params.extend([search_pattern, search_pattern, search_pattern, 
                      search_pattern, search_pattern, search_pattern])
    
    query += " ORDER BY p.first_name, p.last_name LIMIT 100"
    
    cur.execute(query, params)
    patients = cur.fetchall()
    cur.close()
    conn.close()
    
    # Convert to list of dicts for easier template access
    patients_list = []
    for patient in patients:
        patients_list.append({
            'id': patient[0],
            'patient_id': patient[1],
            'hospital_no': patient[2],
            'card_number': patient[3],
            'first_name': patient[4],
            'last_name': patient[5],
            'phone': patient[6],
            'registration_date': patient[7],
            'status': patient[8] if len(patient) > 8 else 'Active',
            'blood_group': patient[9] if len(patient) > 9 else None,
            'date_of_birth': patient[10] if len(patient) > 10 else None,
            'gender': patient[11] if len(patient) > 11 else None,
            'email': patient[12] if len(patient) > 12 else None,
            'address': patient[13] if len(patient) > 13 else None
        })
    
    return render_template("patient_services/search_patients.html",
                         patients=patients_list,
                         search_term=search_term,
                         show_deleted=show_deleted)
    
    
@app.route('/api/patient/<int:patient_id>/card-number')
def api_patient_card_number(patient_id):
    """API endpoint to get patient card number"""
    if 'patient_user_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT card_number FROM patients WHERE id = ?", (patient_id,))
    result = cur.fetchone()
    cur.close()
    conn.close()
    
    if result:
        return jsonify({"card_number": result[0]})
    else:
        return jsonify({"card_number": None})
    
    
        
    
@app.route('/patient_services/patients/<int:patient_id>')
def view_patient(patient_id):
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("SELECT * FROM patients WHERE id = ?", (patient_id,))
    patient = cur.fetchone()
    
    # Get patient's appointments
    cur.execute("""
        SELECT * FROM appointments 
        WHERE patient_id = ? 
        ORDER BY appointment_date DESC, appointment_time DESC
        LIMIT 10
    """, (patient_id,))
    appointments = cur.fetchall()
    
    # Get patient's queue history
    cur.execute("""
        SELECT * FROM queue 
        WHERE patient_id = ? 
        ORDER BY created_at DESC
        LIMIT 10
    """, (patient_id,))
    queue_history = cur.fetchall()
    
    cur.close()
    conn.close()
    
    if not patient:
        flash("Patient not found", "danger")
        return redirect(url_for('search_patients'))
    
    return render_template("patient_services/view_patient.html", 
                         patient=patient, 
                         appointments=appointments,
                         queue_history=queue_history)

@app.route('/patient_services/patients/<int:patient_id>/edit', methods=['GET', 'POST'])
def edit_patient(patient_id):
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == 'POST':
        # Update patient data
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        date_of_birth = request.form.get('date_of_birth')
        gender = request.form.get('gender')
        blood_group = request.form.get('blood_group')
        genotype = request.form.get('genotype')
        phone = request.form.get('phone')
        email = request.form.get('email')
        address = request.form.get('address')
        emergency_contact_name = request.form.get('emergency_contact_name')
        emergency_contact_phone = request.form.get('emergency_contact_phone')
        occupation = request.form.get('occupation')
        marital_status = request.form.get('marital_status')
        nationality = request.form.get('nationality')
        religion = request.form.get('religion')
        next_of_kin = request.form.get('next_of_kin')
        next_of_kin_phone = request.form.get('next_of_kin_phone')
        notes = request.form.get('notes')
        
        try:
            cur.execute("""
                UPDATE patients SET
                    first_name = ?, last_name = ?, date_of_birth = ?, gender = ?,
                    blood_group = ?, genotype = ?, phone = ?, email = ?, address = ?,
                    emergency_contact_name = ?, emergency_contact_phone = ?, occupation = ?,
                    marital_status = ?, nationality = ?, religion = ?, next_of_kin = ?,
                    next_of_kin_phone = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (first_name, last_name, date_of_birth, gender, blood_group, genotype,
                  phone, email, address, emergency_contact_name, emergency_contact_phone,
                  occupation, marital_status, nationality, religion, next_of_kin,
                  next_of_kin_phone, notes, patient_id))
            
            conn.commit()
            flash("Patient information updated successfully!", "success")
            return redirect(url_for('view_patient', patient_id=patient_id))
            
        except Exception as e:
            conn.rollback()
            flash(f"Error updating patient: {str(e)}", "danger")
    
    # GET request - show form
    cur.execute("SELECT * FROM patients WHERE id = ?", (patient_id,))
    patient = cur.fetchone()
    cur.close()
    conn.close()
    
    if not patient:
        flash("Patient not found", "danger")
        return redirect(url_for('search_patients'))
    
    return render_template("patient_services/edit_patient.html", patient=patient)

# ==================== APPOINTMENTS ====================

@app.route('/patient_services/appointments/book', methods=['GET', 'POST'])
def book_appointment():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    if request.method == 'POST':
        patient_id = request.form.get('patient_id')
        doctor_name = request.form.get('doctor_name')
        department = request.form.get('department')
        appointment_date = request.form.get('appointment_date')
        appointment_time = request.form.get('appointment_time')
        purpose = request.form.get('purpose')
        priority = request.form.get('priority', 'Normal')
        notes = request.form.get('notes')
        
        # Get patient name
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT first_name, last_name FROM patients WHERE id = ?", (patient_id,))
        patient = cur.fetchone()
        
        if not patient:
            flash("Patient not found", "danger")
            return redirect(url_for('book_appointment'))
        
        patient_name = f"{patient[0]} {patient[1]}"
        appointment_no = f"APT-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        
        try:
            cur.execute("""
                INSERT INTO appointments (
                    appointment_no, patient_id, patient_name, doctor_name, department,
                    appointment_date, appointment_time, purpose, priority, notes, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (appointment_no, patient_id, patient_name, doctor_name, department,
                  appointment_date, appointment_time, purpose, priority, notes, session['patient_user_id']))
            
            conn.commit()
            flash(f"Appointment booked successfully! No: {appointment_no}", "success")
            return redirect(url_for('view_appointments'))
            
        except Exception as e:
            conn.rollback()
            flash(f"Error booking appointment: {str(e)}", "danger")
        finally:
            cur.close()
            conn.close()
    
    return render_template("patient_services/book_appointment.html")

@app.route('/patient_services/appointments')
def view_appointments():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    date_filter = request.args.get('date', date.today().isoformat())
    status_filter = request.args.get('status', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    query = """
        SELECT * FROM appointments 
        WHERE appointment_date = ?
    """
    params = [date_filter]
    
    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)
    
    query += " ORDER BY appointment_time"
    
    cur.execute(query, params)
    appointments = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("patient_services/appointments.html", 
                         appointments=appointments, 
                         date_filter=date_filter,
                         status_filter=status_filter)

@app.route('/patient_services/appointments/<int:appointment_id>/cancel', methods=['POST'])
def cancel_appointment(appointment_id):
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("UPDATE appointments SET status = 'Cancelled' WHERE id = ?", (appointment_id,))
        conn.commit()
        flash("Appointment cancelled successfully", "success")
    except Exception as e:
        flash(f"Error cancelling appointment: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('view_appointments'))

# ==================== QUEUE MANAGEMENT ====================

@app.route('/patient_services/queue/generate', methods=['POST'])
def generate_queue_token():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    patient_id = request.form.get('patient_id')
    service_type = request.form.get('service_type')
    priority = request.form.get('priority', 'Normal')
    notes = request.form.get('notes')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get patient name
    cur.execute("SELECT first_name, last_name FROM patients WHERE id = ?", (patient_id,))
    patient = cur.fetchone()
    
    if not patient:
        flash("Patient not found", "danger")
        return redirect(url_for('search_patients'))
    
    patient_name = f"{patient[0]} {patient[1]}"
    
    # Get next queue position
    cur.execute("SELECT COUNT(*) FROM queue WHERE status = 'Waiting'")
    queue_count = cur.fetchone()[0] or 0
    queue_position = queue_count + 1
    
    token_no = f"TK-{datetime.now().strftime('%Y%m%d')}-{queue_position:03d}"
    
    try:
        cur.execute("""
            INSERT INTO queue (
                token_no, patient_id, patient_name, service_type, priority,
                queue_position, status, notes, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, 'Waiting', ?, CURRENT_TIMESTAMP)
        """, (token_no, patient_id, patient_name, service_type, priority, queue_position, notes))
        
        conn.commit()
        flash(f"Queue token generated: {token_no} (Position: {queue_position})", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error generating token: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('view_queue'))

@app.route('/patient_services/queue')
def view_queue():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT * FROM queue 
        WHERE status = 'Waiting'
        ORDER BY 
            CASE priority WHEN 'High' THEN 1 WHEN 'Normal' THEN 2 ELSE 3 END,
            queue_position ASC
    """)
    waiting_queue = cur.fetchall()
    
    cur.execute("""
        SELECT * FROM queue 
        WHERE status IN ('Called', 'In Progress')
        ORDER BY called_at DESC
        LIMIT 10
    """)
    active_calls = cur.fetchall()
    
    cur.close()
    conn.close()
    
    return render_template("patient_services/queue.html", 
                         waiting_queue=waiting_queue,
                         active_calls=active_calls)

@app.route('/patient_services/queue/<int:queue_id>/call', methods=['POST'])
def call_patient(queue_id):
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            UPDATE queue 
            SET status = 'Called', 
                called_at = CURRENT_TIMESTAMP, 
                called_by = ?
            WHERE id = ?
        """, (session['patient_full_name'], queue_id))
        conn.commit()
        flash("Patient called successfully", "success")
    except Exception as e:
        flash(f"Error calling patient: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('view_queue'))

# ==================== PATIENT SERVICES ====================

@app.route('/patient_services/checkin', methods=['GET', 'POST'])
def patient_checkin():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    if request.method == 'POST':
        patient_id = request.form.get('patient_id')
        service_type = request.form.get('service_type')
        is_emergency = request.form.get('is_emergency', '0')
        ward_requested = request.form.get('ward_requested', '0')
        ward_type = request.form.get('ward_type') if ward_requested == '1' else None
        notes = request.form.get('notes')
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Get patient name
        cur.execute("SELECT first_name, last_name FROM patients WHERE id = ?", (patient_id,))
        patient = cur.fetchone()
        
        if not patient:
            flash("Patient not found", "danger")
            return redirect(url_for('patient_checkin'))
        
        patient_name = f"{patient[0]} {patient[1]}"
        service_no = f"SVC-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        
        try:
            cur.execute("""
                INSERT INTO patient_services (
                    service_no, patient_id, patient_name, service_type, check_in_time,
                    is_emergency, ward_requested, ward_type, notes, created_by
                ) VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP, ?, ?, ?, ?, ?)
            """, (service_no, patient_id, patient_name, service_type, 
                  is_emergency, ward_requested, ward_type, notes, session['patient_user_id']))
            
            conn.commit()
            flash(f"Patient checked in successfully! Service No: {service_no}", "success")
            
            # Also generate queue token
            if service_type:
                # Get next queue position
                cur.execute("SELECT COUNT(*) FROM queue WHERE status = 'Waiting'")
                queue_count = cur.fetchone()[0] or 0
                queue_position = queue_count + 1
                token_no = f"TK-{datetime.now().strftime('%Y%m%d')}-{queue_position:03d}"
                
                cur.execute("""
                    INSERT INTO queue (
                        token_no, patient_id, patient_name, service_type, priority,
                        queue_position, status, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, 'Waiting', CURRENT_TIMESTAMP)
                """, (token_no, patient_id, patient_name, service_type, 
                      'High' if is_emergency == '1' else 'Normal', queue_position))
                conn.commit()
                flash(f"Queue token generated: {token_no} (Position: {queue_position})", "success")
            
        except Exception as e:
            conn.rollback()
            flash(f"Error during check-in: {str(e)}", "danger")
        finally:
            cur.close()
            conn.close()
        
        return redirect(url_for('active_services'))
    
    return render_template("patient_services/checkin.html")

@app.route('/patient_services/services/active')
def active_services():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT * FROM patient_services 
        WHERE status = 'Active'
        ORDER BY check_in_time DESC
    """)
    services = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("patient_services/active_services.html", services=services)

# ==================== REPORTS ====================

@app.route('/patient_services/reports/daily')
def daily_registration_report():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    # Fix: Don't use 'date' as variable name since it conflicts with the date module
    from datetime import date as date_today
    
    report_date_str = request.args.get('date', date_today.today().isoformat())
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT COUNT(*)
        FROM patients 
        WHERE registration_date = ?
    """, (report_date_str,))
    total_registrations = cur.fetchone()[0] or 0
    
    cur.execute("""
        SELECT * FROM patients 
        WHERE registration_date = ?
        ORDER BY registration_time DESC
    """, (report_date_str,))
    patients = cur.fetchall()
    
    cur.close()
    conn.close()
    
    return render_template("patient_services/daily_report.html", 
                         date=report_date_str, 
                         total_registrations=total_registrations,
                         patients=patients)
        
 
 # -------------------- ROUTES: PATIENT SERVICES MODULE --------------------


@app.route('/patient_services/login', methods=['GET', 'POST'])
def patient_services_login():
    if 'patient_user_id' in session:
        return redirect(url_for('patient_services_dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM patient_users WHERE username = ? AND is_active = 1", (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user['password'], password):
            session['patient_user_id'] = user['id']
            session['patient_username'] = user['username']
            session['patient_full_name'] = user['full_name']
            session['patient_role'] = user['role']
            flash("Login successful!", "success")
            return redirect(url_for('patient_services_dashboard'))
        else:
            flash("Invalid username or password.", "danger")

    return render_template("patient_services/login.html")

@app.route('/patient_services/dashboard')
def patient_services_dashboard():
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    return render_template("patient_services/dashboard.html",
                         today_registrations=0,
                         today_appointments=0,
                         waiting_queue=0,
                         active_services=0,
                         current_queue=[],
                         user_name=session.get('patient_full_name', 'User'))

@app.route('/patient_services/logout')
def patient_services_logout():
    session.pop('patient_user_id', None)
    session.pop('patient_username', None)
    session.pop('patient_full_name', None)
    session.pop('patient_role', None)
    flash("Logged out successfully", "success")
    return redirect(url_for('patient_services_login'))   

def generate_card_number():
    """Generate sequential card number in format JTSSH-XXXXXX"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Check if card_number column exists
        cur.execute("PRAGMA table_info(patients)")
        columns = [col[1] for col in cur.fetchall()]
        
        if 'card_number' not in columns:
            # Column doesn't exist yet, add it without UNIQUE constraint
            cur.execute("ALTER TABLE patients ADD COLUMN card_number VARCHAR(50)")
            conn.commit()
            print("Added card_number column")
        
        # Get the highest card number
        cur.execute("""
            SELECT card_number FROM patients 
            WHERE card_number IS NOT NULL AND card_number LIKE 'JTSSH-%'
            ORDER BY id DESC 
            LIMIT 1
        """)
        
        last_card = cur.fetchone()
        
        if last_card and last_card[0]:
            # Extract the number part (e.g., from 'JTSSH-000001' get '000001')
            try:
                last_num = int(last_card[0].split('-')[1])
                next_num = last_num + 1
            except (ValueError, IndexError):
                next_num = 1
        else:
            # Count existing patients to determine next number
            cur.execute("SELECT COUNT(*) FROM patients")
            count = cur.fetchone()[0] or 0
            next_num = count + 1
        
        # Format with leading zeros (6 digits)
        return f"JTSSH-{next_num:06d}"
        
    except Exception as e:
        print(f"Error generating card number: {e}")
        # Fallback: use timestamp-based number
        import time
        return f"JTSSH-{int(time.time()) % 1000000:06d}"
    finally:
        cur.close()
        conn.close()

def add_card_number_column_and_backfill():
    """Add card_number column to patients table and backfill existing records"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Check if column exists
        cur.execute("PRAGMA table_info(patients)")
        columns = [col[1] for col in cur.fetchall()]
        
        if 'card_number' not in columns:
            # Add the column
            cur.execute("ALTER TABLE patients ADD COLUMN card_number VARCHAR(50) UNIQUE")
            print("Added card_number column")
            
            # Backfill existing patients with sequential card numbers
            cur.execute("SELECT id FROM patients ORDER BY id")
            patients = cur.fetchall()
            
            for index, patient in enumerate(patients, start=1):
                card_number = f"JTSSH-{index:06d}"
                cur.execute("UPDATE patients SET card_number = ? WHERE id = ?", (card_number, patient[0]))
            
            conn.commit()
            print(f"Backfilled {len(patients)} patients with card numbers")
        else:
            print("card_number column already exists")
            
    except Exception as e:
        print(f"Error updating table: {e}")
        conn.rollback()
    finally:
        cur.close()
        conn.close()
        
def add_card_number_column():
    """Add card_number column to patients table if it doesn't exist"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Check if column exists
        cur.execute("PRAGMA table_info(patients)")
        columns = [col[1] for col in cur.fetchall()]
        
        if 'card_number' not in columns:
            # Add the column
            cur.execute("ALTER TABLE patients ADD COLUMN card_number VARCHAR(50) UNIQUE")
            print("✓ Added card_number column to patients table")
            
            # Backfill existing patients with sequential card numbers
            cur.execute("SELECT id FROM patients ORDER BY id")
            patients = cur.fetchall()
            
            for index, patient in enumerate(patients, start=1):
                card_number = f"JTSSH-{index:06d}"
                cur.execute("UPDATE patients SET card_number = ? WHERE id = ?", (card_number, patient[0]))
            
            conn.commit()
            print(f"✓ Backfilled {len(patients)} patients with card numbers")
        else:
            print("✓ card_number column already exists")
            
    except Exception as e:
        print(f"Error adding card_number column: {e}")
        conn.rollback()
    finally:
        cur.close()
        conn.close()


# ==================== DELETE PATIENT (WITH SOFT DELETE) ====================

@app.route('/patient_services/patients/<int:patient_id>/delete', methods=['POST'])
def delete_patient(patient_id):
    """Delete a patient (soft delete by setting status to 'Deleted')"""
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    # Optional: Check if user has permission to delete (e.g., only admins or managers)
    # You can uncomment this if you want to restrict deletion to specific roles
    # if session.get('patient_role') not in ['Manager', 'Admin']:
    #     flash("You don't have permission to delete patients", "danger")
    #     return redirect(url_for('search_patients'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # First, get patient info for confirmation message
        cur.execute("SELECT first_name, last_name, patient_id FROM patients WHERE id = ?", (patient_id,))
        patient = cur.fetchone()
        
        if not patient:
            flash("Patient not found", "danger")
            return redirect(url_for('search_patients'))
        
        # Check if patient has any active appointments or services
        cur.execute("""
            SELECT COUNT(*) FROM appointments 
            WHERE patient_id = ? AND status IN ('Scheduled', 'Confirmed')
        """, (patient_id,))
        active_appointments = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COUNT(*) FROM patient_services 
            WHERE patient_id = ? AND status = 'Active'
        """, (patient_id,))
        active_services = cur.fetchone()[0] or 0
        
        # Option 1: Soft delete - just mark as deleted (Recommended)
        cur.execute("""
            UPDATE patients 
            SET status = 'Deleted', 
                updated_at = CURRENT_TIMESTAMP 
            WHERE id = ?
        """, (patient_id,))
        
        # Option 2: Hard delete - permanently remove (Uncomment if you want this instead)
        # cur.execute("DELETE FROM patients WHERE id = ?", (patient_id,))
        
        conn.commit()
        
        patient_name = f"{patient[0]} {patient[1]}"
        
        # Show warning if there were active appointments/services
        warning_msg = ""
        if active_appointments > 0 or active_services > 0:
            warning_msg = f" Note: This patient had {active_appointments} active appointment(s) and {active_services} active service(s)."
        
        flash(f"Patient {patient_name} (ID: {patient[2]}) has been deleted successfully.{warning_msg}", "success")
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error deleting patient {patient_id}: {e}")
        flash(f"Error deleting patient: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    
    # Redirect back to the referring page or search page
    referrer = request.referrer
    if referrer and 'view_patient' in referrer:
        return redirect(url_for('search_patients'))
    return redirect(request.referrer or url_for('search_patients'))


@app.route('/patient_services/patients/<int:patient_id>/restore', methods=['POST'])
def restore_patient(patient_id):
    """Restore a soft-deleted patient"""
    if 'patient_user_id' not in session:
        return redirect(url_for('patient_services_login'))
    
    # Optional: Check permission
    # if session.get('patient_role') not in ['Manager', 'Admin']:
    #     flash("You don't have permission to restore patients", "danger")
    #     return redirect(url_for('search_patients'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            UPDATE patients 
            SET status = 'Active', 
                updated_at = CURRENT_TIMESTAMP 
            WHERE id = ? AND status = 'Deleted'
        """, (patient_id,))
        
        if cur.rowcount > 0:
            conn.commit()
            flash("Patient restored successfully", "success")
        else:
            flash("Patient not found or not deleted", "warning")
            
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error restoring patient {patient_id}: {e}")
        flash(f"Error restoring patient: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    
    return redirect(request.referrer or url_for('search_patients'))

@app.route('/api/next-card-number')
def api_next_card_number():
    """API endpoint to get the next card number"""
    if 'patient_user_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    next_card = generate_card_number()
    return jsonify({"card_number": next_card})


# ==================== CLINICAL/DOCTOR MODULE ====================

@app.route('/clinical/login', methods=['GET', 'POST'])
def clinical_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Simple hardcoded check for testing
        if username == 'doctor1' and password == 'doctor123':
            session['clinical_user_id'] = 1
            session['clinical_username'] = 'doctor1'
            session['clinical_name'] = 'Dr. John Smith'
            session['clinical_specialization'] = 'General Medicine'
            flash("Welcome, Dr. John Smith!", "success")
            return redirect(url_for('clinical_dashboard'))
        elif username == 'dr.sarah' and password == 'sarah123':
            session['clinical_user_id'] = 2
            session['clinical_username'] = 'dr.sarah'
            session['clinical_name'] = 'Dr. Sarah Johnson'
            session['clinical_specialization'] = 'Cardiology'
            flash("Welcome, Dr. Sarah Johnson!", "success")
            return redirect(url_for('clinical_dashboard'))
        else:
            # Check database as fallback
            conn = get_db_connection()
            if conn:
                cur = conn.cursor()
                cur.execute("SELECT id, username, full_name, specialization FROM clinical_users WHERE username=? AND password=?", (username, password))
                user = cur.fetchone()
                cur.close()
                conn.close()
                
                if user:
                    session['clinical_user_id'] = user[0]
                    session['clinical_username'] = user[1]
                    session['clinical_name'] = user[2]
                    session['clinical_specialization'] = user[3] if len(user) > 3 else 'General Medicine'
                    flash(f"Welcome, {session['clinical_name']}!", "success")
                    return redirect(url_for('clinical_dashboard'))
            
            flash("Invalid username or password. Try: doctor1 / doctor123", "danger")
    
    return render_template("clinical/login.html")

@app.route('/clinical/dashboard')
def clinical_dashboard():
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    doctor_name = session.get('clinical_name', 'Doctor')
    doctor_id = session['clinical_user_id']
    today = date.today().isoformat()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get statistics - using doctor_name instead of doctor_id since appointments table doesn't have doctor_id
        cur.execute("SELECT COUNT(*) FROM appointments WHERE doctor_name LIKE ? AND appointment_date = ?", (f'%{doctor_name}%', today))
        appointments_today = cur.fetchone()[0] or 0
        
        # For consultations - these tables might not exist yet, so use try/except
        active_consultations = 0
        completed_today = 0
        try:
            cur.execute("SELECT COUNT(*) FROM consultations WHERE doctor_id=? AND status='In Progress'", (doctor_id,))
            active_consultations = cur.fetchone()[0] or 0
        except:
            pass
        
        try:
            cur.execute("SELECT COUNT(*) FROM consultations WHERE doctor_id=? AND DATE(completed_at)=?", (doctor_id, today))
            completed_today = cur.fetchone()[0] or 0
        except:
            pass
        
        # Get queue - doctor_queue table might not exist yet
        waiting_patients = 0
        queue = []
        try:
            cur.execute("SELECT COUNT(*) FROM doctor_queue WHERE doctor_id=? AND status='Waiting'", (doctor_id,))
            waiting_patients = cur.fetchone()[0] or 0
            
            # Get queue
            cur.execute("""
                SELECT id, queue_no as token_no, patient_name, priority, consultation_type, 
                       strftime('%s', 'now') - strftime('%s', created_at) as wait_seconds
                FROM doctor_queue 
                WHERE doctor_id=? AND status='Waiting'
                ORDER BY 
                    CASE priority WHEN 'Emergency' THEN 1 WHEN 'High' THEN 2 WHEN 'Normal' THEN 3 ELSE 4 END,
                    created_at ASC
                LIMIT 10
            """, (doctor_id,))
            
            for row in cur.fetchall():
                queue.append({
                    'id': row[0],
                    'token_no': row[1],
                    'patient_name': row[2],
                    'priority': row[3] or 'Normal',
                    'consultation_type': row[4] or 'General',
                    'wait_time': round(row[5] / 60) if row[5] else 0
                })
        except Exception as e:
            app.logger.warning(f"Queue table not ready: {e}")
        
    except Exception as e:
        app.logger.error(f"Error loading clinical dashboard: {e}")
        appointments_today = 0
        active_consultations = 0
        completed_today = 0
        waiting_patients = 0
        queue = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("clinical/dashboard.html",
                         doctor_name=doctor_name,
                         today_date=date.today().strftime("%A, %B %d, %Y"),
                         appointments_today=appointments_today,
                         active_consultations=active_consultations,
                         completed_today=completed_today,
                         waiting_patients=waiting_patients,
                         pending_lab_results=0,
                         pending_imaging=0,
                         followup_patients=0,
                         pending_referrals=0,
                         queue=queue)    
    

@app.route('/clinical/logout')
def clinical_logout():
    session.pop('clinical_user_id', None)
    session.pop('clinical_username', None)
    session.pop('clinical_name', None)
    session.pop('clinical_specialization', None)
    flash("Logged out successfully", "success")
    return redirect(url_for('clinical_login'))

def create_default_clinical_user():
    """Create default clinical/doctor user if none exists."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    # Check if clinical_users table exists
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='clinical_users'")
    if not cursor.fetchone():
        print("clinical_users table doesn't exist yet - will be created by create_tables()")
        cursor.close()
        conn.close()
        return
    
    # Check if any clinical users exist
    cursor.execute("SELECT COUNT(*) FROM clinical_users")
    count = cursor.fetchone()[0]
    
    if count == 0:
        # Create default doctor
        default_password = generate_password_hash('doctor123')
        cursor.execute("""
            INSERT INTO clinical_users (username, password, full_name, email, specialization, license_number, role, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, ('doctor1', default_password, 'Dr. John Smith', 'doctor@hospital.com', 'General Medicine', 'LIC-001', 'Doctor', 1))
        
        # Add a second test doctor
        cursor.execute("""
            INSERT INTO clinical_users (username, password, full_name, email, specialization, license_number, role, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, ('dr.sarah', generate_password_hash('sarah123'), 'Dr. Sarah Johnson', 'sarah.j@hospital.com', 'Cardiology', 'LIC-002', 'Doctor', 1))
        
        conn.commit()
        print("Created default clinical users: doctor1/doctor123, dr.sarah/sarah123")
    else:
        print(f"Clinical users already exist: {count} user(s)")
        
        # Optional: List existing users for debugging
        cursor.execute("SELECT id, username, full_name, specialization FROM clinical_users WHERE is_active = 1")
        users = cursor.fetchall()
        for user in users:
            print(f"  - {user[1]}: {user[2]} ({user[3]})")
    
    cursor.close()
    conn.close()
    
@app.route('/clinical/debug-users')
def debug_clinical_users():
    if 'clinical_user_id' not in session:
        # Allow access for debugging
        pass
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, username, password[:50] as password_preview, full_name, specialization, is_active FROM clinical_users")
    users = cur.fetchall()
    cur.close()
    conn.close()
    
    result = "<h3>Clinical Users:</h3>"
    for user in users:
        result += f"<p>ID: {user[0]}, Username: {user[1]}, Password Hash: {user[2]}..., Name: {user[3]}, Specialization: {user[4]}, Active: {user[5]}</p>"
    
    # Add test info
    result += "<hr>"
    result += "<h3>Test Login Credentials:</h3>"
    result += "<p><strong>Username:</strong> doctor1<br><strong>Password:</strong> doctor123</p>"
    result += "<p><strong>Username:</strong> dr.sarah<br><strong>Password:</strong> sarah123</p>"
    
    return result

@app.route('/clinical/check-db')
def clinical_check_db():
    conn = get_db_connection()
    if not conn:
        return "Database connection failed"
    
    cur = conn.cursor()
    
    # Check table structure
    cur.execute("PRAGMA table_info(clinical_users)")
    columns = cur.fetchall()
    
    result = "<h3>clinical_users table structure:</h3>"
    result += "<table border='1'><tr><th>Column</th><th>Type</th><th>Nullable</th></tr>"
    for col in columns:
        result += f"<tr><td>{col[1]}</td><td>{col[2]}</td><td>{'YES' if col[3] else 'NO'}</td></tr>"
    result += "</table>"
    
    # Get users
    cur.execute("SELECT id, username, full_name, specialization, is_active FROM clinical_users")
    users = cur.fetchall()
    
    result += "<h3>Current Users:</h3>"
    result += "<table border='1'><tr><th>ID</th><th>Username</th><th>Full Name</th><th>Specialization</th><th>Active</th></tr>"
    for user in users:
        result += f"<tr><td>{user[0]}</td><td>{user[1]}</td><td>{user[2]}</td><td>{user[3]}</td><td>{user[4]}</td></tr>"
    result += "</table>"
    
    # Test password verification for doctor1
    cur.execute("SELECT password FROM clinical_users WHERE username = 'doctor1'")
    pw_row = cur.fetchone()
    if pw_row:
        from werkzeug.security import check_password_hash
        stored_pw = pw_row[0]
        test_password = 'doctor123'
        
        result += "<h3>Password Test:</h3>"
        result += f"<p>Testing password '{test_password}' for user 'doctor1'</p>"
        
        try:
            if check_password_hash(stored_pw, test_password):
                result += "<p style='color:green'>✓ Password verification SUCCESSFUL!</p>"
            else:
                result += "<p style='color:red'>✗ Password verification FAILED!</p>"
                result += f"<p>Stored password hash: {stored_pw[:50]}...</p>"
        except Exception as e:
            result += f"<p style='color:red'>Error checking password: {e}</p>"
    
    cur.close()
    conn.close()
    
    return result



@app.route('/clinical/debug-schema')
def debug_schema():
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    result = "<h3>Database Schema Check</h3>"
    
    # Check appointments table
    cur.execute("PRAGMA table_info(appointments)")
    columns = cur.fetchall()
    result += "<h4>appointments table columns:</h4>"
    result += "<ul>"
    for col in columns:
        result += f"<li>{col[1]} ({col[2]})</li>"
    result += "</ul>"
    
    # Check if consultations table exists
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='consultations'")
    if cur.fetchone():
        cur.execute("PRAGMA table_info(consultations)")
        cols = cur.fetchall()
        result += "<h4>consultations table columns:</h4>"
        result += "<ul>"
        for col in cols:
            result += f"<li>{col[1]} ({col[2]})</li>"
        result += "</ul>"
    else:
        result += "<p>consultations table does not exist yet</p>"
    
    # Check if doctor_queue table exists
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='doctor_queue'")
    if cur.fetchone():
        cur.execute("PRAGMA table_info(doctor_queue)")
        cols = cur.fetchall()
        result += "<h4>doctor_queue table columns:</h4>"
        result += "<ul>"
        for col in cols:
            result += f"<li>{col[1]} ({col[2]})</li>"
        result += "</ul>"
    else:
        result += "<p>doctor_queue table does not exist yet</p>"
    
    cur.close()
    conn.close()
    
    return result

@app.route('/clinical/patients/search')
def clinical_patient_search():
    """Search patients for clinical use"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    search_term = request.args.get('q', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    query = """
        SELECT id, patient_id, card_number, first_name, last_name, 
               date_of_birth, gender, phone, registration_date
        FROM patients 
        WHERE status = 'Active'
    """
    params = []
    
    if search_term:
        query += """ AND (first_name LIKE ? OR last_name LIKE ? 
                   OR patient_id LIKE ? OR phone LIKE ? 
                   OR hospital_no LIKE ? OR card_number LIKE ?)"""
        search_pattern = f'%{search_term}%'
        params = [search_pattern] * 6
    
    query += " ORDER BY first_name, last_name LIMIT 50"
    
    cur.execute(query, params)
    patients = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("clinical/patient_search.html", 
                         patients=patients, 
                         search_term=search_term)
    

@app.route('/clinical/consultation/start/<int:patient_id>', methods=['GET', 'POST'])
def start_consultation(patient_id):
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get patient details
    cur.execute("SELECT * FROM patients WHERE id = ?", (patient_id,))
    patient = cur.fetchone()
    
    if not patient:
        flash("Patient not found", "danger")
        return redirect(url_for('clinical_patient_search'))
    
    if request.method == 'POST':
        # Save consultation and update queue status
        import uuid
        consultation_no = f"CON-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        
        chief_complaint = request.form.get('chief_complaint')
        history = request.form.get('history')
        past_history = request.form.get('past_history')
        examination = request.form.get('examination')
        diagnosis = request.form.get('diagnosis')
        treatment = request.form.get('treatment')
        follow_up = request.form.get('follow_up')
        notes = request.form.get('notes')
        
        try:
            cur.execute("""
                INSERT INTO consultations (
                    consultation_no, patient_id, patient_name, doctor_id, doctor_name,
                    consultation_date, consultation_time, status,
                    chief_complaint, history_of_present_illness, past_medical_history,
                    physical_exam, diagnosis, treatment_plan, follow_up_date, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, 'Completed', ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                consultation_no, patient_id, f"{patient[4]} {patient[5]}",
                session['clinical_user_id'], session['clinical_name'],
                date.today().isoformat(), datetime.now().strftime('%H:%M:%S'),
                chief_complaint, history, past_history, examination,
                diagnosis, treatment, follow_up if follow_up else None, notes
            ))
            
            # Update queue status if patient is in queue
            cur.execute("""
                UPDATE doctor_queue 
                SET status = 'Completed', completed_at = CURRENT_TIMESTAMP 
                WHERE patient_id = ? AND doctor_id = ? AND status IN ('Waiting', 'Called', 'In Progress')
            """, (patient_id, session['clinical_user_id']))
            
            conn.commit()
            flash(f"Consultation saved successfully! No: {consultation_no}", "success")
            return redirect(url_for('clinical_dashboard'))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error saving consultation: {e}")
            flash(f"Error saving consultation: {str(e)}", "danger")
    
    cur.close()
    conn.close()
    
    return render_template("clinical/start_consultation.html", patient=patient)


# ==================== CLINICAL ROUTES (Placeholders for now) ====================

@app.route('/clinical/appointments')
def clinical_appointments():
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    flash("Appointments feature coming soon!", "info")
    return redirect(url_for('clinical_dashboard'))

@app.route('/clinical/active-consultations')
def clinical_active_consultations():
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    flash("Active consultations feature coming soon!", "info")
    return redirect(url_for('clinical_dashboard'))

@app.route('/clinical/consultation-history')
def clinical_consultation_history():
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    flash("Consultation history feature coming soon!", "info")
    return redirect(url_for('clinical_dashboard'))


@app.route('/clinical/lab-results')
def clinical_lab_results():
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    flash("Lab results feature coming soon!", "info")
    return redirect(url_for('clinical_dashboard'))

@app.route('/clinical/imaging-reports')
def clinical_imaging_reports():
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    flash("Imaging reports feature coming soon!", "info")
    return redirect(url_for('clinical_dashboard'))

@app.route('/clinical/followups')
def clinical_followups():
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    flash("Follow-ups feature coming soon!", "info")
    return redirect(url_for('clinical_dashboard'))

@app.route('/clinical/referrals-inbox')
def clinical_referrals_inbox():
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    flash("Referrals feature coming soon!", "info")
    return redirect(url_for('clinical_dashboard'))

@app.route('/clinical/queue')
def clinical_queue():
    """View doctor's patient queue"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    doctor_id = session['clinical_user_id']
    doctor_name = session.get('clinical_name', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get all patients in queue for this doctor
    cur.execute("""
        SELECT 
            dq.id,
            dq.queue_no,
            dq.patient_id,
            dq.patient_name,
            dq.priority,
            dq.status,
            dq.consultation_type,
            dq.wait_time_minutes,
            dq.token_no,
            dq.called_at,
            dq.created_at,
            p.card_number,
            p.phone,
            p.date_of_birth,
            strftime('%s', 'now') - strftime('%s', dq.created_at) as wait_seconds
        FROM doctor_queue dq
        LEFT JOIN patients p ON dq.patient_id = p.id
        WHERE dq.doctor_id = ? 
        ORDER BY 
            CASE dq.priority 
                WHEN 'Emergency' THEN 1 
                WHEN 'High' THEN 2 
                WHEN 'Normal' THEN 3 
                ELSE 4 
            END,
            dq.created_at ASC
    """, (doctor_id,))
    
    queue_items = cur.fetchall()
    
    # Process queue items
    waiting_queue = []
    in_progress_queue = []
    completed_queue = []
    
    for item in queue_items:
        wait_minutes = round(item[14] / 60) if item[14] else item[8] or 0
        
        queue_item = {
            'id': item[0],
            'queue_no': item[1],
            'patient_id': item[2],
            'patient_name': item[3],
            'priority': item[4] or 'Normal',
            'status': item[5] or 'Waiting',
            'consultation_type': item[6] or 'General',
            'wait_time': wait_minutes,
            'token_no': item[8],
            'called_at': item[9],
            'created_at': item[10],
            'card_number': item[11],
            'phone': item[12],
            'date_of_birth': item[13]
        }
        
        if queue_item['status'] == 'Waiting':
            waiting_queue.append(queue_item)
        elif queue_item['status'] == 'In Progress':
            in_progress_queue.append(queue_item)
        elif queue_item['status'] == 'Completed':
            completed_queue.append(queue_item)
    
    # Get statistics
    total_waiting = len(waiting_queue)
    total_in_progress = len(in_progress_queue)
    avg_wait_time = 0
    if total_waiting > 0:
        total_seconds = sum([item['wait_time'] for item in waiting_queue])
        avg_wait_time = round(total_seconds / total_waiting)
    
    cur.close()
    conn.close()
    
    return render_template("clinical/queue.html",
                         waiting_queue=waiting_queue,
                         in_progress_queue=in_progress_queue,
                         completed_queue=completed_queue,
                         total_waiting=total_waiting,
                         total_in_progress=total_in_progress,
                         avg_wait_time=avg_wait_time,
                         doctor_name=doctor_name)
    
@app.route('/clinical/call-patient', methods=['POST'])
def clinical_call_patient():
    """Call a patient from the queue"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    queue_id = request.form.get('queue_id')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            UPDATE doctor_queue 
            SET status = 'Called', 
                called_at = CURRENT_TIMESTAMP 
            WHERE id = ? AND doctor_id = ?
        """, (queue_id, session['clinical_user_id']))
        
        conn.commit()
        flash("Patient has been called!", "success")
    except Exception as e:
        app.logger.error(f"Error calling patient: {e}")
        flash("Error calling patient", "danger")
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('clinical_queue'))

@app.route('/clinical/remove-from-queue', methods=['POST'])
def clinical_remove_from_queue():
    """Remove a patient from the queue"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    queue_id = request.form.get('queue_id')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            DELETE FROM doctor_queue 
            WHERE id = ? AND doctor_id = ?
        """, (queue_id, session['clinical_user_id']))
        
        conn.commit()
        flash("Patient removed from queue", "success")
    except Exception as e:
        app.logger.error(f"Error removing patient: {e}")
        flash("Error removing patient", "danger")
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('clinical_queue'))

@app.route('/clinical/write-prescription', methods=['GET', 'POST'])
def clinical_write_prescription():
    """Write prescription for a patient"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    from datetime import date as date_today
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get available drugs from pharmacy
    cur.execute("""
        SELECT id, name, strength, unit_price, stock_quantity
        FROM drugs 
        WHERE stock_quantity > 0
        ORDER BY name ASC
        LIMIT 50
    """)
    drugs = cur.fetchall()
    
    if request.method == 'POST':
        patient_name = request.form.get('patient_name')
        patient_id_input = request.form.get('patient_id')
        drug_name = request.form.get('drug_name')
        strength = request.form.get('strength')
        dosage = request.form.get('dosage')
        frequency = request.form.get('frequency')
        duration = request.form.get('duration')
        quantity = request.form.get('quantity')
        instructions = request.form.get('instructions')
        refills = request.form.get('refills', 0)
        is_controlled = request.form.get('is_controlled', '0')
        
        # Validate required fields
        if not patient_name:
            flash("Patient name is required", "danger")
            cur.close()
            conn.close()
            return redirect(url_for('clinical_write_prescription'))
        
        if not drug_name:
            flash("Drug name is required", "danger")
            cur.close()
            conn.close()
            return redirect(url_for('clinical_write_prescription'))
        
        # Handle patient_id - if provided and exists, use it; otherwise create new patient or set to None
        actual_patient_id = None
        
        if patient_id_input and patient_id_input != '':
            # Check if the patient ID exists
            cur.execute("SELECT id FROM patients WHERE id = ?", (patient_id_input,))
            if cur.fetchone():
                actual_patient_id = patient_id_input
            else:
                # Try to find patient by name
                name_parts = patient_name.strip().split(' ', 1)
                first_name = name_parts[0]
                last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                cur.execute("SELECT id FROM patients WHERE first_name = ? AND last_name = ?", (first_name, last_name))
                existing = cur.fetchone()
                if existing:
                    actual_patient_id = existing[0]
        else:
            # Try to find patient by name
            name_parts = patient_name.strip().split(' ', 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            
            cur.execute("SELECT id FROM patients WHERE first_name = ? AND last_name = ?", (first_name, last_name))
            existing = cur.fetchone()
            if existing:
                actual_patient_id = existing[0]
        
        prescription_no = f"RX-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
        prescribed_date = date_today.today().isoformat()
        prescribed_time = datetime.now().strftime('%H:%M:%S')
        expires_date = (date_today.today() + timedelta(days=90)).isoformat()
        
        try:
            # Insert prescription (patient_id can be NULL for now)
            cur.execute("""
                INSERT INTO prescriptions (
                    prescription_no, patient_id, patient_name, doctor_id, doctor_name,
                    drug_name, strength, dosage, frequency, duration, quantity,
                    instructions, refills, is_controlled, status,
                    prescribed_date, prescribed_time, expires_date
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active', ?, ?, ?)
            """, (
                prescription_no, actual_patient_id, patient_name, session['clinical_user_id'],
                session['clinical_name'], drug_name, strength, dosage, frequency,
                duration, quantity, instructions, refills, is_controlled,
                prescribed_date, prescribed_time, expires_date
            ))
            
            conn.commit()
            prescription_id = cur.lastrowid
            flash(f"Prescription {prescription_no} written successfully!", "success")
            cur.close()
            conn.close()
            return redirect(url_for('clinical_prescription_view', prescription_id=prescription_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error writing prescription: {e}")
            flash(f"Error writing prescription: {str(e)}", "danger")
    
    cur.close()
    conn.close()
    
    return render_template("clinical/write_prescription.html",
                         drugs=drugs,
                         today_date=date_today.today().strftime("%A, %B %d, %Y"))        
    
@app.route('/clinical/prescription/<int:prescription_id>')
def clinical_prescription_view(prescription_id):
    """View a specific prescription"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT * FROM prescriptions WHERE id = ? AND doctor_id = ?
    """, (prescription_id, session['clinical_user_id']))
    
    prescription = cur.fetchone()
    cur.close()
    conn.close()
    
    if not prescription:
        flash("Prescription not found", "danger")
        return redirect(url_for('clinical_dashboard'))
    
    return render_template("clinical/prescription_view.html", prescription=prescription)

@app.route('/clinical/prescriptions')
def clinical_prescriptions():
    """List all prescriptions by the doctor"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT id, prescription_no, patient_name, drug_name, dosage, frequency,
               duration, status, prescribed_date
        FROM prescriptions 
        WHERE doctor_id = ?
        ORDER BY prescribed_date DESC
        LIMIT 50
    """, (session['clinical_user_id'],))
    
    prescriptions = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("clinical/prescriptions_list.html", prescriptions=prescriptions)

@app.route('/clinical/prescription/<int:prescription_id>/print')
def clinical_prescription_print(prescription_id):
    """Print a prescription"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT p.*, 
               pat.first_name, pat.last_name, pat.date_of_birth, pat.gender
        FROM prescriptions p
        JOIN patients pat ON p.patient_id = pat.id
        WHERE p.id = ? AND p.doctor_id = ?
    """, (prescription_id, session['clinical_user_id']))
    
    prescription = cur.fetchone()
    cur.close()
    conn.close()
    
    if not prescription:
        flash("Prescription not found", "danger")
        return redirect(url_for('clinical_dashboard'))
    
    return render_template("clinical/prescription_print.html", prescription=prescription)

@app.route('/api/clinical/patients/search')
def api_clinical_patient_search():
    """API endpoint to search patients for prescription form"""
    if 'clinical_user_id' not in session:
        return jsonify([])
    
    search_term = request.args.get('q', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT id, patient_id, card_number, first_name, last_name, phone
        FROM patients 
        WHERE status = 'Active'
        AND (first_name LIKE ? OR last_name LIKE ? OR patient_id LIKE ? OR card_number LIKE ? OR phone LIKE ?)
        LIMIT 15
    """, (f'%{search_term}%', f'%{search_term}%', f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'))
    
    patients = []
    for row in cur.fetchall():
        patients.append({
            'id': row[0],
            'patient_id': row[1],
            'card_number': row[2],
            'first_name': row[3],
            'last_name': row[4],
            'name': f"{row[3]} {row[4]}",
            'phone': row[5]
        })
    
    cur.close()
    conn.close()
    
    return jsonify(patients)


@app.route('/api/clinical/prescriptions/list')
def api_clinical_prescriptions_list():
    """API endpoint to get prescriptions for the doctor"""
    if 'clinical_user_id' not in session:
        return jsonify([])
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT 
            id, prescription_no, patient_name, drug_name, dosage, frequency,
            duration, quantity, status, prescribed_date, expires_date,
            instructions, refills, is_controlled
        FROM prescriptions 
        WHERE doctor_id = ?
        ORDER BY prescribed_date DESC
        LIMIT 200
    """, (session['clinical_user_id'],))
    
    prescriptions = []
    for row in cur.fetchall():
        prescriptions.append({
            'id': row[0],
            'prescription_no': row[1],
            'patient_name': row[2],
            'drug_name': row[3],
            'dosage': row[4],
            'frequency': row[5],
            'duration': row[6],
            'quantity': row[7],
            'status': row[8],
            'prescribed_date': row[9],
            'expires_date': row[10],
            'instructions': row[11],
            'refills': row[12],
            'is_controlled': row[13]
        })
    
    cur.close()
    conn.close()
    
    return jsonify(prescriptions)

@app.route('/api/clinical/prescriptions/delete/<int:prescription_id>', methods=['DELETE'])
def api_clinical_prescription_delete(prescription_id):
    """Delete a prescription"""
    if 'clinical_user_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # First check if prescription belongs to this doctor
        cur.execute("SELECT id FROM prescriptions WHERE id = ? AND doctor_id = ?", 
                   (prescription_id, session['clinical_user_id']))
        if not cur.fetchone():
            return jsonify({"success": False, "message": "Prescription not found or access denied"}), 404
        
        # Delete the prescription
        cur.execute("DELETE FROM prescriptions WHERE id = ?", (prescription_id,))
        conn.commit()
        
        return jsonify({"success": True, "message": "Prescription deleted successfully"})
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error deleting prescription: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()
        
@app.route('/clinical/order-lab', methods=['GET', 'POST'])
def clinical_order_lab():
    """Order laboratory tests for a patient"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    from datetime import date as date_today
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get available lab tests (without price)
    cur.execute("""
        SELECT id, test_id, test_name, category, turnaround_time, normal_range
        FROM lab_tests 
        WHERE is_active = 1
        ORDER BY category, test_name
    """)
    lab_tests = cur.fetchall()
    
    # Group tests by category
    tests_by_category = {}
    for test in lab_tests:
        category = test[3] or 'Other'
        if category not in tests_by_category:
            tests_by_category[category] = []
        tests_by_category[category].append({
            'id': test[0],
            'test_id': test[1],
            'name': test[2],
            'turnaround': test[4],
            'normal_range': test[5]
        })
    
    if request.method == 'POST':
        patient_name = request.form.get('patient_name')
        patient_id_input = request.form.get('patient_id')
        selected_tests = request.form.getlist('selected_tests')
        manual_tests_str = request.form.get('manual_tests', '')
        priority = request.form.get('priority', 'Normal')
        clinical_notes = request.form.get('clinical_notes')
        diagnosis = request.form.get('diagnosis')
        
        if not patient_name:
            flash("Patient name is required", "danger")
            cur.close()
            conn.close()
            return redirect(url_for('clinical_order_lab'))
        
        # Parse manual tests
        manual_tests = [t.strip() for t in manual_tests_str.split('|') if t.strip()]
        
        if not selected_tests and not manual_tests:
            flash("Please select at least one test", "danger")
            cur.close()
            conn.close()
            return redirect(url_for('clinical_order_lab'))
        
        # Get test details for selected tests
        test_ids = []
        test_names = []
        
        if selected_tests:
            placeholders = ','.join(['?' for _ in selected_tests])
            cur.execute(f"""
                SELECT id, test_name
                FROM lab_tests 
                WHERE id IN ({placeholders}) AND is_active = 1
            """, selected_tests)
            tests = cur.fetchall()
            for test in tests:
                test_ids.append(str(test[0]))
                test_names.append(test[1])
        
        # Add manual tests
        for manual_test in manual_tests:
            test_names.append(f"[CUSTOM] {manual_test}")
        
        # Generate unique request number
        request_no = f"LAB-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        
        # Handle patient_id - try to find existing patient
        actual_patient_id = None
        if patient_id_input and patient_id_input != '':
            actual_patient_id = patient_id_input
        else:
            # Try to find patient by name
            name_parts = patient_name.strip().split(' ', 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            cur.execute("SELECT id FROM patients WHERE first_name = ? AND last_name = ?", (first_name, last_name))
            existing = cur.fetchone()
            if existing:
                actual_patient_id = existing[0]
        
        try:
            cur.execute("""
                INSERT INTO lab_requests (
                    request_no, patient_name, patient_id, doctor_name,
                    test_ids, test_names, total_amount, discount, grand_total,
                    priority, notes, requested_date, created_by, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending')
            """, (
                request_no, patient_name, actual_patient_id, session['clinical_name'],
                ','.join(test_ids), ','.join(test_names), 0, 0, 0,
                priority, f"Clinical Notes: {clinical_notes}\nDiagnosis: {diagnosis}" if clinical_notes or diagnosis else None,
                date_today.today().isoformat(), session['clinical_user_id']
            ))
            
            conn.commit()
            lab_request_id = cur.lastrowid
            flash(f"Lab request {request_no} submitted successfully! Tests: {len(test_names)}", "success")
            cur.close()
            conn.close()
            return redirect(url_for('clinical_lab_request_view', request_id=lab_request_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating lab request: {e}")
            flash(f"Error creating lab request: {str(e)}", "danger")
    
    cur.close()
    conn.close()
    
    return render_template("clinical/order_lab.html",
                         tests_by_category=tests_by_category,
                         today_date=date_today.today().strftime("%A, %B %d, %Y"))
    
        
@app.route('/clinical/lab-requests')
def clinical_lab_requests():
    """View all lab requests by the doctor"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT id, request_no, patient_name, test_names, total_amount, 
               grand_total, status, priority, requested_date
        FROM lab_requests 
        WHERE doctor_name = ?
        ORDER BY requested_date DESC
        LIMIT 100
    """, (session['clinical_name'],))
    
    requests = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("clinical/lab_requests_list.html", requests=requests)

@app.route('/clinical/lab-request/<int:request_id>')
def clinical_lab_request_view(request_id):
    """View a specific lab request"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT * FROM lab_requests 
        WHERE id = ? AND doctor_name = ?
    """, (request_id, session['clinical_name']))
    
    lab_request = cur.fetchone()
    cur.close()
    conn.close()
    
    if not lab_request:
        flash("Lab request not found", "danger")
        return redirect(url_for('clinical_lab_requests'))
    
    return render_template("clinical/lab_request_view.html", request=lab_request)

@app.route('/clinical/notes', methods=['GET', 'POST'])
def clinical_notes():
    """Manage clinical notes for patients"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    from datetime import date as date_today
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'save_note':
            patient_name = request.form.get('patient_name')
            patient_id_input = request.form.get('patient_id')
            note_type = request.form.get('note_type')
            subjective = request.form.get('subjective')
            objective = request.form.get('objective')
            assessment = request.form.get('assessment')
            plan = request.form.get('plan')
            is_private = request.form.get('is_private', '0')
            
            if not patient_name or not subjective:
                flash("Patient name and subjective findings are required", "danger")
                return redirect(url_for('clinical_notes'))
            
            # Generate note number
            note_no = f"NOTE-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
            
            # Handle patient_id
            actual_patient_id = None
            if patient_id_input and patient_id_input != '':
                actual_patient_id = patient_id_input
            else:
                # Try to find patient by name
                name_parts = patient_name.strip().split(' ', 1)
                first_name = name_parts[0]
                last_name = name_parts[1] if len(name_parts) > 1 else ''
                cur.execute("SELECT id FROM patients WHERE first_name = ? AND last_name = ?", (first_name, last_name))
                existing = cur.fetchone()
                if existing:
                    actual_patient_id = existing[0]
            
            try:
                cur.execute("""
                    INSERT INTO clinical_notes (
                        note_no, patient_id, patient_name, doctor_id, doctor_name,
                        note_type, subjective, objective, assessment, plan,
                        note_date, note_time, status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Final')
                """, (
                    note_no, actual_patient_id, patient_name, session['clinical_user_id'],
                    session['clinical_name'], note_type, subjective, objective, assessment, plan,
                    date_today.today().isoformat(), datetime.now().strftime('%H:%M:%S')
                ))
                
                conn.commit()
                flash(f"Clinical note saved successfully! Note No: {note_no}", "success")
                return redirect(url_for('clinical_note_view', note_id=cur.lastrowid))
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error saving clinical note: {e}")
                flash(f"Error saving note: {str(e)}", "danger")
        
        elif action == 'update_note':
            note_id = request.form.get('note_id')
            subjective = request.form.get('subjective')
            objective = request.form.get('objective')
            assessment = request.form.get('assessment')
            plan = request.form.get('plan')
            
            try:
                cur.execute("""
                    UPDATE clinical_notes 
                    SET subjective = ?, objective = ?, assessment = ?, plan = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ? AND doctor_id = ?
                """, (subjective, objective, assessment, plan, note_id, session['clinical_user_id']))
                
                conn.commit()
                flash("Clinical note updated successfully!", "success")
                return redirect(url_for('clinical_note_view', note_id=note_id))
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error updating clinical note: {e}")
                flash(f"Error updating note: {str(e)}", "danger")
        
        elif action == 'delete_note':
            note_id = request.form.get('note_id')
            
            try:
                cur.execute("DELETE FROM clinical_notes WHERE id = ? AND doctor_id = ?", 
                           (note_id, session['clinical_user_id']))
                conn.commit()
                flash("Clinical note deleted successfully!", "success")
                return redirect(url_for('clinical_notes'))
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error deleting clinical note: {e}")
                flash(f"Error deleting note: {str(e)}", "danger")
    
    # GET request - get all notes for this doctor
    cur.execute("""
        SELECT id, note_no, patient_name, note_type, note_date, 
               subjective, assessment, created_at
        FROM clinical_notes 
        WHERE doctor_id = ?
        ORDER BY note_date DESC, created_at DESC
        LIMIT 100
    """, (session['clinical_user_id'],))
    
    notes = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("clinical/notes_list.html", 
                         notes=notes,
                         today_date=date_today.today().strftime("%A, %B %d, %Y"))

@app.route('/clinical/notes/new')
def clinical_note_new():
    """Create a new clinical note"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    from datetime import date as date_today
    
    return render_template("clinical/note_form.html",
                         today_date=date_today.today().strftime("%A, %B %d, %Y"),
                         note=None)

@app.route('/clinical/notes/<int:note_id>')
def clinical_note_view(note_id):
    """View a specific clinical note"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT * FROM clinical_notes 
        WHERE id = ? AND doctor_id = ?
    """, (note_id, session['clinical_user_id']))
    
    note = cur.fetchone()
    cur.close()
    conn.close()
    
    if not note:
        flash("Clinical note not found", "danger")
        return redirect(url_for('clinical_notes'))
    
    return render_template("clinical/note_view.html", note=note)

@app.route('/clinical/notes/<int:note_id>/edit')
def clinical_note_edit(note_id):
    """Edit a clinical note"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    from datetime import date as date_today
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT * FROM clinical_notes 
        WHERE id = ? AND doctor_id = ?
    """, (note_id, session['clinical_user_id']))
    
    note = cur.fetchone()
    cur.close()
    conn.close()
    
    if not note:
        flash("Clinical note not found", "danger")
        return redirect(url_for('clinical_notes'))
    
    return render_template("clinical/note_form.html",
                         today_date=date_today.today().strftime("%A, %B %d, %Y"),
                         note=note)

@app.route('/api/clinical/notes/search')
def api_clinical_notes_search():
    """API endpoint to search clinical notes"""
    if 'clinical_user_id' not in session:
        return jsonify([])
    
    search_term = request.args.get('q', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT id, note_no, patient_name, note_type, note_date, 
               subjective, assessment
        FROM clinical_notes 
        WHERE doctor_id = ?
        AND (patient_name LIKE ? OR subjective LIKE ? OR assessment LIKE ?)
        ORDER BY note_date DESC
        LIMIT 50
    """, (session['clinical_user_id'], f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'))
    
    notes = []
    for row in cur.fetchall():
        notes.append({
            'id': row[0],
            'note_no': row[1],
            'patient_name': row[2],
            'note_type': row[3],
            'note_date': row[4],
            'subjective': row[5][:100] if row[5] else '',
            'assessment': row[6][:100] if row[6] else ''
        })
    
    cur.close()
    conn.close()
    
    return jsonify(notes)

@app.route('/clinical/patient-notes/<int:patient_id>')
def clinical_patient_notes(patient_id):
    """View all notes for a specific patient"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get patient info
    cur.execute("SELECT first_name, last_name, patient_id FROM patients WHERE id = ?", (patient_id,))
    patient = cur.fetchone()
    
    if not patient:
        flash("Patient not found", "danger")
        return redirect(url_for('clinical_notes'))
    
    # Get patient notes
    cur.execute("""
        SELECT id, note_no, note_type, note_date, subjective, assessment, plan
        FROM clinical_notes 
        WHERE patient_id = ? AND doctor_id = ?
        ORDER BY note_date DESC
    """, (patient_id, session['clinical_user_id']))
    
    notes = cur.fetchall()
    cur.close()
    conn.close()
    
    patient_name = f"{patient[0]} {patient[1]}"
    
    return render_template("clinical/patient_notes.html", 
                         notes=notes, 
                         patient=patient,
                         patient_name=patient_name)
    
    
@app.route('/clinical/order-imaging', methods=['GET', 'POST'])
def clinical_order_imaging():
    """Order imaging/radiology studies for a patient"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    from datetime import date as date_today
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Predefined imaging types
    imaging_types = [
        'X-Ray', 'CT Scan', 'MRI', 'Ultrasound', 'Mammogram',
        'Fluoroscopy', 'PET Scan', 'Nuclear Medicine', 'DEXA Scan',
        'Angiography', 'Doppler Study', 'Echocardiogram'
    ]
    
    # Body parts list
    body_parts = [
        'Head', 'Neck', 'Chest', 'Abdomen', 'Pelvis', 'Spine',
        'Upper Extremity', 'Lower Extremity', 'Knee', 'Shoulder',
        'Hip', 'Ankle', 'Wrist', 'Elbow', 'Hand', 'Foot'
    ]
    
    if request.method == 'POST':
        patient_name = request.form.get('patient_name')
        patient_id_input = request.form.get('patient_id')
        imaging_type = request.form.get('imaging_type')
        custom_imaging = request.form.get('custom_imaging', '').strip()
        body_part = request.form.get('body_part')
        custom_body_part = request.form.get('custom_body_part', '').strip()
        urgency = request.form.get('urgency', 'Routine')
        clinical_indication = request.form.get('clinical_indication')
        notes = request.form.get('notes')
        
        # Determine final imaging type and body part
        final_imaging_type = custom_imaging if custom_imaging else imaging_type
        final_body_part = custom_body_part if custom_body_part else body_part
        
        if not patient_name:
            flash("Patient name is required", "danger")
            return redirect(url_for('clinical_order_imaging'))
        
        if not final_imaging_type:
            flash("Please select or enter an imaging type", "danger")
            return redirect(url_for('clinical_order_imaging'))
        
        if not final_body_part:
            flash("Please select or enter a body part", "danger")
            return redirect(url_for('clinical_order_imaging'))
        
        # Generate order number
        order_no = f"IMG-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
        
        # Handle patient_id
        actual_patient_id = None
        if patient_id_input and patient_id_input != '':
            actual_patient_id = patient_id_input
        else:
            name_parts = patient_name.strip().split(' ', 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            cur.execute("SELECT id FROM patients WHERE first_name = ? AND last_name = ?", (first_name, last_name))
            existing = cur.fetchone()
            if existing:
                actual_patient_id = existing[0]
        
        try:
            cur.execute("""
                INSERT INTO imaging_orders (
                    order_no, patient_id, patient_name, doctor_id, doctor_name,
                    imaging_type, body_part, urgency, clinical_indication, status,
                    ordered_date, ordered_time, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending', ?, ?, ?)
            """, (
                order_no, actual_patient_id, patient_name, session['clinical_user_id'],
                session['clinical_name'], final_imaging_type, final_body_part, urgency,
                clinical_indication, date_today.today().isoformat(),
                datetime.now().strftime('%H:%M:%S'), notes
            ))
            
            conn.commit()
            order_id = cur.lastrowid
            flash(f"Imaging order {order_no} submitted successfully!", "success")
            return redirect(url_for('clinical_imaging_view', order_id=order_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating imaging order: {e}")
            flash(f"Error creating order: {str(e)}", "danger")
    
    cur.close()
    conn.close()
    
    return render_template("clinical/order_imaging.html",
                         imaging_types=imaging_types,
                         body_parts=body_parts,
                         today_date=date_today.today().strftime("%A, %B %d, %Y"))

@app.route('/clinical/imaging-requests')
def clinical_imaging_requests():
    """List all imaging orders by the doctor"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT id, order_no, patient_name, imaging_type, body_part, urgency, status, ordered_date
        FROM imaging_orders 
        WHERE doctor_id = ?
        ORDER BY ordered_date DESC
        LIMIT 100
    """, (session['clinical_user_id'],))
    
    orders = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template("clinical/imaging_requests.html", orders=orders)

@app.route('/clinical/imaging/<int:order_id>')
def clinical_imaging_view(order_id):
    """View a specific imaging order"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT * FROM imaging_orders 
        WHERE id = ? AND doctor_id = ?
    """, (order_id, session['clinical_user_id']))
    
    order = cur.fetchone()
    cur.close()
    conn.close()
    
    if not order:
        flash("Imaging order not found", "danger")
        return redirect(url_for('clinical_imaging_requests'))
    
    return render_template("clinical/imaging_view.html", order=order)

@app.route('/clinical/imaging/<int:order_id>/print')
def clinical_imaging_print(order_id):
    """Print an imaging order"""
    if 'clinical_user_id' not in session:
        return redirect(url_for('clinical_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT * FROM imaging_orders 
        WHERE id = ? AND doctor_id = ?
    """, (order_id, session['clinical_user_id']))
    
    order = cur.fetchone()
    cur.close()
    conn.close()
    
    if not order:
        flash("Imaging order not found", "danger")
        return redirect(url_for('clinical_imaging_requests'))
    
    return render_template("clinical/imaging_print.html", order=order)


#============================Radiography Route ==============================================
@app.route('/radiology/login', methods=['GET', 'POST'])
def radiology_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, username, password, full_name, role FROM radiology_users WHERE username=? AND is_active=1", (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()
        if user and check_password_hash(user[2], password):
            session['radiology_user_id'] = user[0]
            session['radiology_username'] = user[1]
            session['radiology_full_name'] = user[3]
            session['radiology_role'] = user[4]
            return redirect(url_for('radiology_dashboard'))
        else:
            flash("Invalid credentials", "danger")
    return render_template("radiology/login.html")

@app.route('/radiology/dashboard')
def radiology_dashboard():
    if 'radiology_user_id' not in session:
        return redirect(url_for('radiology_login'))
    conn = get_db_connection()
    cur = conn.cursor()
    # Count pending, in progress, completed orders
    cur.execute("SELECT COUNT(*) FROM imaging_orders WHERE status='Pending'")
    pending = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM imaging_orders WHERE status='In Progress'")
    in_progress = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM imaging_orders WHERE status='Completed' AND DATE(performed_date)=DATE('now')")
    completed_today = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM imaging_orders WHERE DATE(ordered_date)=DATE('now')")
    ordered_today = cur.fetchone()[0] or 0
    # Recent orders
    cur.execute("SELECT id, order_no, patient_name, imaging_type, body_part, urgency, status, ordered_date FROM imaging_orders ORDER BY ordered_date DESC LIMIT 10")
    recent_orders = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("radiology/dashboard.html",
                         pending=pending,
                         in_progress=in_progress,
                         completed_today=completed_today,
                         ordered_today=ordered_today,
                         recent_orders=recent_orders,
                         user_name=session.get('radiology_full_name'))
    
    
@app.route('/radiology/orders')
def radiology_orders():
    if 'radiology_user_id' not in session:
        return redirect(url_for('radiology_login'))
    status_filter = request.args.get('status', '')
    conn = get_db_connection()
    cur = conn.cursor()
    query = "SELECT * FROM imaging_orders WHERE 1=1"
    params = []
    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)
    query += " ORDER BY urgency DESC, ordered_date DESC"
    cur.execute(query, params)
    orders = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("radiology/orders.html", orders=orders, status_filter=status_filter)


@app.route('/radiology/order/<int:order_id>', methods=['GET', 'POST'])
def radiology_order_detail(order_id):
    if 'radiology_user_id' not in session:
        return redirect(url_for('radiology_login'))
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        status = request.form.get('status')
        report = request.form.get('report')
        performed_date = request.form.get('performed_date') if request.form.get('performed_date') else None
        # Update imaging order
        cur.execute("""
            UPDATE imaging_orders
            SET status=?, report=?, performed_date=?, performed_by=?, report_date=CURRENT_DATE
            WHERE id=?
        """, (status, report, performed_date, session['radiology_full_name'], order_id))
        conn.commit()
        flash("Order updated successfully", "success")
        return redirect(url_for('radiology_orders'))
    # GET: fetch order details
    cur.execute("SELECT * FROM imaging_orders WHERE id=?", (order_id,))
    order = cur.fetchone()
    cur.close()
    conn.close()
    if not order:
        flash("Order not found", "danger")
        return redirect(url_for('radiology_orders'))
    return render_template("radiology/order_detail.html", order=order)


@app.route('/radiology/reports')
def radiology_reports():
    if 'radiology_user_id' not in session:
        return redirect(url_for('radiology_login'))
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, order_no, patient_name, imaging_type, report_date, report FROM imaging_orders WHERE status='Completed' ORDER BY report_date DESC LIMIT 50")
    reports = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("radiology/reports.html", reports=reports)


@app.route('/radiology/logout')
def radiology_logout():
    session.pop('radiology_user_id', None)
    session.pop('radiology_username', None)
    session.pop('radiology_full_name', None)
    session.pop('radiology_role', None)
    flash("Logged out", "success")
    return redirect(url_for('radiology_login'))


def create_default_radiology_user():
    """Create default radiology user if none exists."""
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()
    # Check if radiology_users table exists
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='radiology_users'")
    if not cursor.fetchone():
        cursor.close()
        conn.close()
        return
    cursor.execute("SELECT COUNT(*) FROM radiology_users")
    count = cursor.fetchone()[0]
    if count == 0:
        from werkzeug.security import generate_password_hash
        hashed_pw = generate_password_hash('radio123')
        cursor.execute("""
            INSERT INTO radiology_users (username, password, full_name, role, is_active)
            VALUES (?, ?, ?, ?, 1)
        """, ('radiologist1', hashed_pw, 'Radiologist', 'Radiologist'))
        conn.commit()
        print("Created default radiology user: radiologist1 / radio123")
    cursor.close()
    conn.close()
    
def update_imaging_orders_table():
    """Add missing columns to imaging_orders table"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Check if table exists
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='imaging_orders'")
    if cur.fetchone():
        # Get existing columns
        cur.execute("PRAGMA table_info(imaging_orders)")
        existing_columns = [col[1] for col in cur.fetchall()]
        
        # Add missing columns
        if 'report' not in existing_columns:
            cur.execute("ALTER TABLE imaging_orders ADD COLUMN report TEXT")
            print("Added 'report' column to imaging_orders")
        
        if 'performed_date' not in existing_columns:
            cur.execute("ALTER TABLE imaging_orders ADD COLUMN performed_date DATE")
            print("Added 'performed_date' column to imaging_orders")
        
        if 'performed_by' not in existing_columns:
            cur.execute("ALTER TABLE imaging_orders ADD COLUMN performed_by VARCHAR(100)")
            print("Added 'performed_by' column to imaging_orders")
        
        if 'report_date' not in existing_columns:
            cur.execute("ALTER TABLE imaging_orders ADD COLUMN report_date DATE")
            print("Added 'report_date' column to imaging_orders")
        
        if 'report_approved_by' not in existing_columns:
            cur.execute("ALTER TABLE imaging_orders ADD COLUMN report_approved_by VARCHAR(100)")
            print("Added 'report_approved_by' column to imaging_orders")
        
        if 'images_stored' not in existing_columns:
            cur.execute("ALTER TABLE imaging_orders ADD COLUMN images_stored TEXT")
            print("Added 'images_stored' column to imaging_orders")
        
        conn.commit()
    else:
        print("imaging_orders table does not exist yet - will be created by create_tables()")
    
    cur.close()
    conn.close()


@app.route('/fix-imaging-table')
def fix_imaging_table():
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("ALTER TABLE imaging_orders ADD COLUMN report TEXT")
    except: pass
    try:
        cur.execute("ALTER TABLE imaging_orders ADD COLUMN performed_date DATE")
    except: pass
    try:
        cur.execute("ALTER TABLE imaging_orders ADD COLUMN performed_by VARCHAR(100)")
    except: pass
    try:
        cur.execute("ALTER TABLE imaging_orders ADD COLUMN report_date DATE")
    except: pass
    try:
        cur.execute("ALTER TABLE imaging_orders ADD COLUMN report_approved_by VARCHAR(100)")
    except: pass
    try:
        cur.execute("ALTER TABLE imaging_orders ADD COLUMN images_stored TEXT")
    except: pass
    conn.commit()
    cur.close()
    conn.close()
    return "Table updated successfully"

def add_missing_imaging_columns():
    conn = get_db_connection()
    cur = conn.cursor()
    for col in ['report', 'performed_date', 'performed_by', 'report_date']:
        try:
            cur.execute(f"ALTER TABLE imaging_orders ADD COLUMN {col} TEXT")
        except: pass
    conn.commit()
    cur.close()
    conn.close()
    
    
    
# ==================== MANAGEMENT MODULE ====================

@app.route('/management/login', methods=['GET', 'POST'])
def management_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, username, password, full_name, role FROM management_users WHERE username=? AND is_active=1", (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()
        if user and check_password_hash(user[2], password):
            session['management_user_id'] = user[0]
            session['management_username'] = user[1]
            session['management_full_name'] = user[3]
            session['management_role'] = user[4]
            return redirect(url_for('management_dashboard'))
        else:
            flash("Invalid credentials", "danger")
    return render_template("management/login.html")

@app.route('/management/dashboard')
def management_dashboard():
    if 'management_user_id' not in session:
        return redirect(url_for('management_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Existing metrics
    cur.execute("SELECT COUNT(*) FROM patients WHERE status='Active'")
    total_patients = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM appointments WHERE appointment_date = date('now')")
    today_appointments = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM staff WHERE status='Active'")
    total_staff = cur.fetchone()[0] or 0
    cur.execute("SELECT COALESCE(SUM(amount_paid), 0) FROM payments WHERE payment_date = date('now')")
    today_revenue = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM consultations WHERE status='Completed' AND DATE(completed_at) = date('now')")
    today_consultations = cur.fetchone()[0] or 0
    
    # Additional metrics
    cur.execute("SELECT COUNT(*) FROM appointments WHERE appointment_date > date('now')")
    upcoming_appointments = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM patients WHERE registration_date = date('now')")
    new_patients_today = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM lab_requests WHERE status='Pending'")
    pending_lab_requests = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM imaging_orders WHERE status='Pending'")
    pending_imaging = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM prescriptions WHERE status='Active' AND expires_date >= date('now')")
    active_prescriptions = cur.fetchone()[0] or 0
    
    # Revenue trend last 7 days for chart
    revenue_labels = []
    revenue_data = []
    today = date.today()
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        cur.execute("SELECT COALESCE(SUM(amount_paid), 0) FROM payments WHERE payment_date = ?", (d.isoformat(),))
        revenue_data.append(float(cur.fetchone()[0]))
        revenue_labels.append(d.strftime("%a"))
    
    # Recent payments (last 5)
    cur.execute("""
        SELECT id, patient_name, amount_paid, payment_method, payment_date 
        FROM payments 
        ORDER BY payment_date DESC, id DESC 
        LIMIT 5
    """)
    recent_payments = cur.fetchall()
    
    cur.close()
    conn.close()
    
    return render_template("management/dashboard.html",
                         total_patients=total_patients,
                         today_appointments=today_appointments,
                         total_staff=total_staff,
                         today_revenue=today_revenue,
                         today_consultations=today_consultations,
                         upcoming_appointments=upcoming_appointments,
                         new_patients_today=new_patients_today,
                         pending_lab_requests=pending_lab_requests,
                         pending_imaging=pending_imaging,
                         active_prescriptions=active_prescriptions,
                         revenue_labels=revenue_labels,
                         revenue_data=revenue_data,
                         recent_payments=recent_payments,
                         user_name=session.get('management_full_name'))
    
    
@app.route('/management/financial-reports')
def management_financial_reports():
    if 'management_user_id' not in session:
        return redirect(url_for('management_login'))
    
    period = request.args.get('period', 'monthly')  # daily, weekly, monthly
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Base data: payment methods (always shown)
    cur.execute("""
        SELECT payment_method, COALESCE(SUM(amount_paid), 0) as total
        FROM payments
        GROUP BY payment_method
        ORDER BY total DESC
    """)
    payment_methods = cur.fetchall()
    
    # Revenue data based on period
    labels = []
    revenue_data = []
    
    today = date.today()
    
    if period == 'daily':
        # Last 7 days
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            date_str = d.isoformat()
            cur.execute("SELECT COALESCE(SUM(amount_paid), 0) FROM payments WHERE payment_date = ?", (date_str,))
            total = cur.fetchone()[0]
            labels.append(d.strftime("%a, %d %b"))
            revenue_data.append(float(total))
    elif period == 'weekly':
        # Last 6 weeks (starting from current week's Monday)
        # Find Monday of current week
        start_of_week = today - timedelta(days=today.weekday())
        for i in range(5, -1, -1):
            week_start = start_of_week - timedelta(weeks=i)
            week_end = week_start + timedelta(days=6)
            # Use SQLite strftime to extract week number? Simpler: sum between dates
            cur.execute("""
                SELECT COALESCE(SUM(amount_paid), 0) FROM payments 
                WHERE payment_date BETWEEN ? AND ?
            """, (week_start.isoformat(), week_end.isoformat()))
            total = cur.fetchone()[0]
            labels.append(f"Week {week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m')}")
            revenue_data.append(float(total))
    else:  # monthly
        # Last 12 months
        for i in range(11, -1, -1):
            month = today.month - i
            year = today.year
            if month <= 0:
                month += 12
                year -= 1
            month_str = f"{year}-{month:02d}"
            cur.execute("""
                SELECT COALESCE(SUM(amount_paid), 0) FROM payments 
                WHERE strftime('%Y-%m', payment_date) = ?
            """, (month_str,))
            total = cur.fetchone()[0]
            labels.append(f"{year}-{month:02d}")
            revenue_data.append(float(total))
    
    cur.close()
    conn.close()
    
    return render_template("management/financial_reports.html",
                         payment_methods=payment_methods,
                         labels=labels,
                         revenue_data=revenue_data,
                         period=period)
    
    
@app.route('/management/clinical-reports')
def management_clinical_reports():
    if 'management_user_id' not in session:
        return redirect(url_for('management_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Top diagnoses (from consultations)
    cur.execute("""
        SELECT diagnosis, COUNT(*) as count
        FROM consultations
        WHERE diagnosis IS NOT NULL AND diagnosis != ''
        GROUP BY diagnosis
        ORDER BY count DESC
        LIMIT 10
    """)
    top_diagnoses = cur.fetchall()
    
    # Monthly consultations
    cur.execute("""
        SELECT strftime('%Y-%m', consultation_date) as month, COUNT(*) as count
        FROM consultations
        WHERE consultation_date >= date('now', '-12 months')
        GROUP BY month
        ORDER BY month DESC
    """)
    consultation_trend = cur.fetchall()
    
    cur.close()
    conn.close()
    
    return render_template("management/clinical_reports.html",
                         top_diagnoses=top_diagnoses,
                         consultation_trend=consultation_trend)

@app.route('/management/operational-reports')
def management_operational_reports():
    if 'management_user_id' not in session:
        return redirect(url_for('management_login'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Bed occupancy (simplified - can be extended with ward beds)
    # For now, show active patient services
    cur.execute("SELECT COUNT(*) FROM patient_services WHERE status='Active'")
    active_services = cur.fetchone()[0] or 0
    # Assume total beds = 100 (configurable)
    total_beds = 100
    occupancy_rate = (active_services / total_beds * 100) if total_beds > 0 else 0
    
    # Average wait times (from queue)
    cur.execute("SELECT AVG(waiting_time) FROM queue WHERE waiting_time IS NOT NULL")
    avg_wait = cur.fetchone()[0] or 0
    
    cur.close()
    conn.close()
    
    return render_template("management/operational_reports.html",
                         occupancy_rate=round(occupancy_rate, 1),
                         active_patients=active_services,
                         total_beds=total_beds,
                         avg_wait_time=round(avg_wait, 1))

@app.route('/management/export-data')
def management_export_data():
    if 'management_user_id' not in session:
        return redirect(url_for('management_login'))
    
    import csv
    from io import StringIO
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, patient_id, first_name, last_name, phone, registration_date FROM patients ORDER BY registration_date DESC")
    data = cur.fetchall()
    cur.close()
    conn.close()
    
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Patient ID', 'First Name', 'Last Name', 'Phone', 'Registration Date'])
    for row in data:
        writer.writerow(row)
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=patients_export.csv'
    response.headers['Content-type'] = 'text/csv'
    return response


@app.route('/management/logout')
def management_logout():
    session.pop('management_user_id', None)
    session.pop('management_username', None)
    session.pop('management_full_name', None)
    session.pop('management_role', None)
    flash("Logged out successfully", "success")
    return redirect(url_for('management_login'))

def create_default_users():
    """Create default users for all modules."""
    default_users = {
        "pharmacists": ("pharmacist1", "pharma123"),
        "billing_users": ("billing1", "billing123"),
        "lab_users": ("labtech1", "lab123"),
        "patient_users": ("reception1", "reception123"),
        "clinical_users": ("doctor1", "doctor123"),
        "radiology_users": ("radiologist1", "radio123"),
        "management_users": ("manager1", "manager123")
    }

    conn = get_db_connection()
    if not conn:
        return

    cursor = conn.cursor()
    for table, (username, password) in default_users.items():
        hashed_pw = generate_password_hash(password)
        try:
            if table == "lab_users":
                cursor.execute(f"""
                    INSERT OR REPLACE INTO {table} (username, password, full_name, role)
                    VALUES (?, ?, ?, ?)
                """, (username, hashed_pw, 'Lab Technician', 'Lab Technician'))
            elif table == "patient_users":
                cursor.execute(f"""
                    INSERT OR REPLACE INTO {table} (username, password, full_name, role, is_active)
                    VALUES (?, ?, ?, ?, 1)
                """, (username, hashed_pw, 'Receptionist', 'Receptionist'))
            elif table == "clinical_users":
                cursor.execute(f"""
                    INSERT OR REPLACE INTO {table} (username, password, full_name, specialization, is_active)
                    VALUES (?, ?, ?, ?, 1)
                """, (username, hashed_pw, 'Dr. John Smith', 'General Medicine'))
            elif table == "radiology_users":
                cursor.execute(f"""
                    INSERT OR REPLACE INTO {table} (username, password, full_name, role, is_active)
                    VALUES (?, ?, ?, ?, 1)
                """, (username, hashed_pw, 'Radiologist', 'Radiologist'))
            elif table == "management_users":
                cursor.execute(f"""
                    INSERT OR REPLACE INTO {table} (username, password, full_name, role, is_active)
                    VALUES (?, ?, ?, ?, 1)
                """, (username, hashed_pw, 'Hospital Manager', 'Manager'))
            else:
                cursor.execute(f"""
                    INSERT OR REPLACE INTO {table} (username, password)
                    VALUES (?, ?)
                """, (username, hashed_pw))
        except Exception as e:
            app.logger.error(f"Error creating default user {username}: {e}")
    conn.commit()
    cursor.close()
    conn.close()

# Call this function after create_tables() in your main block
if __name__ == "__main__":
    create_tables()
    add_card_number_column()
    create_default_users()
    create_default_hr_data()
    create_sample_lab_tests()
    create_default_clinical_user()
    update_lab_requests_table()
    create_default_radiology_user()
    update_imaging_orders_table() 
    update_imaging_orders_table() 
    update_prescriptions_table()
    add_card_number_column_and_backfill()
    app.run(debug=True)