# -------------------- IMPORTS --------------------
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import psycopg2
import os
from datetime import datetime, date, timedelta
from calendar import month_name
import uuid
import json
import io
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# -------------------- FLASK APP SETUP --------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "super_secret_key_change_later")

# -------------------- DATABASE CONFIGURATION --------------------
DATABASE_URL = os.environ.get(
    "DATABASE_URL", "postgresql://flask_user:Olarewaju1.@localhost:5432/hospital_db"
)

def get_db_connection():
    """Establish database connection with error handling."""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        app.logger.error(f"Database connection error: {e}")
        return None

# -------------------- DATABASE INITIALIZATION --------------------
def create_tables():
    """Create all necessary tables if they don't exist."""
    queries = {
        "pharmacists": """
            CREATE TABLE IF NOT EXISTS pharmacists (
                id SERIAL PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                is_active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "drugs": """
            CREATE TABLE IF NOT EXISTS drugs (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) NOT NULL,
                strength VARCHAR(50) NOT NULL,
                unit_price DECIMAL(10, 2) NOT NULL,
                stock_quantity INT NOT NULL,
                expiry_date DATE NOT NULL,
                low_stock_threshold INT DEFAULT 20,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "drug_sales": """
            CREATE TABLE IF NOT EXISTS drug_sales (
                id SERIAL PRIMARY KEY,
                receipt_no VARCHAR(50) UNIQUE NOT NULL,
                patient_name VARCHAR(100),
                patient_id VARCHAR(50),
                items JSONB NOT NULL,
                subtotal DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0.00,
                tax DECIMAL(10, 2) DEFAULT 0.00,
                grand_total DECIMAL(10, 2) NOT NULL,
                pharmacist VARCHAR(50) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "receipts": """
            CREATE TABLE IF NOT EXISTS receipts (
                id SERIAL PRIMARY KEY,
                patient_name VARCHAR(100),
                patient_id VARCHAR(50),
                subtotal DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0.00,
                tax DECIMAL(10, 2) DEFAULT 0.00,
                total_amount DECIMAL(10, 2) NOT NULL,
                grand_total DECIMAL(10, 2) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "receipt_items": """
            CREATE TABLE IF NOT EXISTS receipt_items (
                id SERIAL PRIMARY KEY,
                receipt_id INT NOT NULL REFERENCES receipts(id),
                drug_name VARCHAR(100) NOT NULL,
                strength VARCHAR(50) NOT NULL,
                quantity INT NOT NULL,
                unit_price DECIMAL(10, 2) NOT NULL
            );
        """,
        "stock_movements": """
            CREATE TABLE IF NOT EXISTS stock_movements (
                id SERIAL PRIMARY KEY,
                drug_id INT NOT NULL REFERENCES drugs(id),
                movement_type VARCHAR(20) NOT NULL,
                quantity INT NOT NULL,
                user_id INT NOT NULL,
                note TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "billing_users": """
            CREATE TABLE IF NOT EXISTS billing_users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "billing_invoice": """
            CREATE TABLE IF NOT EXISTS billing_invoice (
                id SERIAL PRIMARY KEY,
                patient_name VARCHAR(100) NOT NULL,
                service_type VARCHAR(100) NOT NULL,
                amount DECIMAL(10, 2) NOT NULL,
                status VARCHAR(20) DEFAULT 'UNPAID',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "billing_receipt": """
            CREATE TABLE IF NOT EXISTS billing_receipt (
                id SERIAL PRIMARY KEY,
                invoice_id INT NOT NULL REFERENCES billing_invoice(id),
                amount_paid DECIMAL(10, 2) NOT NULL,
                payment_method VARCHAR(50) NOT NULL,
                received_by VARCHAR(50) NOT NULL,
                payment_date TIMESTAMP NOT NULL
            );
        """,
        "payments": """
            CREATE TABLE IF NOT EXISTS payments (
                id SERIAL PRIMARY KEY,
                patient_name VARCHAR(100) NOT NULL,
                service_type VARCHAR(100) NOT NULL,
                subtotal DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0.00,
                tax DECIMAL(10, 2) DEFAULT 0.00,
                grand_total DECIMAL(10, 2) NOT NULL,
                amount_paid DECIMAL(10, 2) NOT NULL,
                balance DECIMAL(10, 2) NOT NULL,
                payment_method VARCHAR(50) NOT NULL,
                status VARCHAR(20) NOT NULL,
                payment_date DATE NOT NULL,
                recorded_by INT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        "users": """
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """
        
    }

    conn = get_db_connection()
    if not conn:
        return

    cursor = conn.cursor()
    for table, query in queries.items():
        try:
            cursor.execute(query)
        except Exception as e:
            app.logger.error(f"Error creating table {table}: {e}")
    conn.commit()
    cursor.close()
    conn.close()

def create_default_users():
    """Create default users for pharmacy and billing modules."""
    default_users = {
        "pharmacists": ("pharmacist1", "pharma123"),
        "billing_users": ("billing1", "billing123")
    }

    conn = get_db_connection()
    if not conn:
        return

    cursor = conn.cursor()
    for table, (username, password) in default_users.items():
        hashed_pw = generate_password_hash(password)
        try:
            cursor.execute(f"""
                INSERT INTO {table} (username, password)
                VALUES (%s, %s)
                ON CONFLICT (username) DO UPDATE SET password = EXCLUDED.password
            """, (username, hashed_pw))
        except Exception as e:
            app.logger.error(f"Error creating default user {username}: {e}")
    conn.commit()
    cursor.close()
    conn.close()
    
def create_hr_tables():
    """Create HR-related tables in PostgreSQL."""
    conn = get_db_connection()
    if not conn:
        app.logger.error("Cannot connect to database for HR table creation")
        return

    cursor = conn.cursor()
    
    # Enable UUID extension if needed
    try:
        cursor.execute("CREATE EXTENSION IF NOT EXISTS \"uuid-ossp\";")
    except Exception as e:
        app.logger.warning(f"Could not enable uuid-ossp extension: {e}")
    
    # HR Users Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS hr_users (
            id SERIAL PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            full_name VARCHAR(100) NOT NULL,
            email VARCHAR(100),
            role VARCHAR(50) DEFAULT 'HR Staff',
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Departments Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS departments (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            code VARCHAR(20) UNIQUE NOT NULL,
            description TEXT,
            head_of_dept VARCHAR(100),
            status VARCHAR(20) DEFAULT 'Active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Staff Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS staff (
            id SERIAL PRIMARY KEY,
            staff_id VARCHAR(50) UNIQUE NOT NULL,
            first_name VARCHAR(100) NOT NULL,
            last_name VARCHAR(100) NOT NULL,
            department_id INTEGER REFERENCES departments(id),
            position VARCHAR(100) NOT NULL,
            employment_type VARCHAR(50),
            email VARCHAR(100),
            phone VARCHAR(20),
            hire_date DATE NOT NULL,
            salary DECIMAL(12, 2),
            status VARCHAR(20) DEFAULT 'Active',
            emergency_contact VARCHAR(100),
            address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Attendance Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            date DATE NOT NULL,
            check_in TIME,
            check_out TIME,
            status VARCHAR(20),
            remarks TEXT,
            recorded_by INTEGER REFERENCES hr_users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Leaves Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leaves (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            leave_type VARCHAR(50) NOT NULL,
            start_date DATE NOT NULL,
            end_date DATE NOT NULL,
            days_requested INTEGER NOT NULL,
            reason TEXT,
            status VARCHAR(20) DEFAULT 'Pending',
            approved_by INTEGER REFERENCES hr_users(id),
            approved_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Schedules Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS schedules (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            schedule_date DATE NOT NULL,
            shift_type VARCHAR(50),
            start_time TIME NOT NULL,
            end_time TIME NOT NULL,
            location VARCHAR(100),
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Payroll Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            pay_period VARCHAR(50),
            basic_salary DECIMAL(12, 2),
            allowances DECIMAL(12, 2),
            deductions DECIMAL(12, 2),
            net_salary DECIMAL(12, 2),
            status VARCHAR(20) DEFAULT 'Pending',
            payment_date DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Documents Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS documents (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            document_type VARCHAR(50),
            document_name VARCHAR(255),
            file_path VARCHAR(500),
            uploaded_by INTEGER REFERENCES hr_users(id),
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # ADD THIS NEW TABLE FOR SHIFT SWAP REQUESTS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS shift_swap_requests (
            id SERIAL PRIMARY KEY,
            schedule_id INTEGER NOT NULL REFERENCES schedules(id) ON DELETE CASCADE,
            from_staff_id INTEGER NOT NULL REFERENCES staff(id) ON DELETE CASCADE,
            to_staff_id INTEGER NOT NULL REFERENCES staff(id) ON DELETE CASCADE,
            reason TEXT,
            status VARCHAR(20) DEFAULT 'Pending',
            requested_by INTEGER REFERENCES hr_users(id),
            requested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            approved_by INTEGER REFERENCES hr_users(id),
            approved_at TIMESTAMP,
            reviewed_by INTEGER REFERENCES hr_users(id),
            reviewed_at TIMESTAMP,
            rejection_reason TEXT,
            notes TEXT
        );
    """)
    
    # Create indexes for better performance
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_staff_department ON staff(department_id);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_staff_date ON attendance(staff_id, date);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_staff_status ON leaves(staff_id, status);",
        "CREATE INDEX IF NOT EXISTS idx_schedules_staff_date ON schedules(staff_id, schedule_date);",
        "CREATE INDEX IF NOT EXISTS idx_payroll_staff_period ON payroll(staff_id, pay_period);",
        "CREATE INDEX IF NOT EXISTS idx_staff_status ON staff(status);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(date);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_status ON leaves(status);",
        "CREATE INDEX IF NOT EXISTS idx_shift_swap_status ON shift_swap_requests(status);",  # ADD THIS INDEX
        "CREATE INDEX IF NOT EXISTS idx_shift_swap_schedule ON shift_swap_requests(schedule_id);"  # ADD THIS INDEX
    ]
    
    for index_query in indexes:
        try:
            cursor.execute(index_query)
        except Exception as e:
            app.logger.warning(f"Could not create index: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()
    
    # Insert default data
    create_default_hr_data()

# -------------------- HELPER FUNCTIONS --------------------
def format_currency(amount):
    """Format amount as Nigerian Naira currency."""
    return f"₦{amount:,.2f}"


def build_stock_snapshot(rows, today):
    stock = []
    for r in rows:
        expiry_date = r[5]
        quantity = r[3]
        threshold = r[6] or 20

        if expiry_date:
            days_left = (expiry_date - today).days
            status = "EXPIRED" if days_left < 0 else "EXPIRING_SOON" if days_left <= 30 else "VALID"
        else:
            days_left = None
            status = "UNKNOWN"

        stock.append({
            "id": r[0], "name": r[1], "strength": r[2],
            "quantity": quantity, "unit_price": r[4],
            "expiry_date": expiry_date, "days_left": days_left,
            "status": status, "low_stock_threshold": threshold,
            "total_value": quantity * r[4]
        })
    return stock

def apply_stock_filter(stock, filter_type):
    if filter_type == "expired":
        return [d for d in stock if d["status"] == "EXPIRED"]
    elif filter_type in ("expiring", "expiring_soon"):
        return [d for d in stock if d["status"] == "EXPIRING_SOON"]
    elif filter_type in ("low", "low_stock"):
        return [d for d in stock if d["quantity"] <= d["low_stock_threshold"]]
    return stock

# -------------------- ROUTES: LANDING & MODULES --------------------
@app.route('/')
def landing_page():
    hospital_name = "John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    modules = [
        "System Admin", "Patient Services", "Clinical Services",
        "Pharmacy", "Laboratory", "Radiology", "Billing and Revenue",
        "Human Resources", "Management and Reports"
    ]
    return render_template("dashboard.html", hospital_name=hospital_name, modules=modules)

@app.route('/<module_name>')
def module_placeholder(module_name):
    display_name = module_name.replace('_', ' ').title()
    if module_name.lower() == "pharmacy":
        return redirect(url_for('pharmacy_login'))
    return render_template("module_placeholder.html", module_name=display_name)

# -------------------- ROUTES: PHARMACY MODULE --------------------
@app.route('/pharmacy/login', methods=['GET', 'POST'])
def pharmacy_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return render_template("pharmacy_login.html")

        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, username, password FROM pharmacists WHERE username=%s AND is_active=TRUE",
            (username,)
        )
        pharmacist = cursor.fetchone()
        cursor.close()
        conn.close()

        if pharmacist and check_password_hash(pharmacist[2], password):
            session['pharmacist_id'] = pharmacist[0]
            session['pharmacist_username'] = pharmacist[1]
            return redirect(url_for('pharmacy_dashboard'))
        else:
            flash("Invalid username or password", "danger")

    return render_template("pharmacy_login.html")

@app.route('/pharmacy/dashboard')
def pharmacy_dashboard():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))
    return render_template(
        "pharmacy_dashboard.html",
        pharmacist_name=session.get('pharmacist_username')
    )

@app.route('/pharmacy/logout')
def pharmacy_logout():
    session.clear()
    return redirect(url_for('pharmacy_login'))

@app.route('/pharmacy/drug_sales')
def drug_sales():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))
    return render_template(
        "drug_sales_dashboard.html",
        pharmacist_name=session.get('pharmacist_username'),
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    )

@app.route('/pharmacy/add-stock', methods=['GET', 'POST'])
def add_stock():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    if request.method == 'POST':
        drug_name = request.form.get('drug_name', '').strip()
        strength = request.form.get('strength', '').strip()
        unit_price = request.form.get('unit_price', '').strip()
        quantity = request.form.get('quantity', '').strip()
        expiry_date = request.form.get('expiry_date', '').strip()

        if not all([drug_name, strength, unit_price, quantity, expiry_date]):
            flash("All fields including expiry date are required.", "danger")
            return redirect(url_for('add_stock'))

        try:
            unit_price = float(unit_price)
            quantity = int(quantity)
            expiry_date_obj = datetime.strptime(expiry_date, "%Y-%m-%d").date()
        except ValueError:
            flash("Invalid input data.", "danger")
            return redirect(url_for('add_stock'))

        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('add_stock'))

        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, stock_quantity FROM drugs
            WHERE name = %s AND strength = %s AND expiry_date = %s
        """, (drug_name, strength, expiry_date_obj))

        existing = cursor.fetchone()

        if existing:
            cursor.execute("""
                UPDATE drugs
                SET stock_quantity = stock_quantity + %s,
                    unit_price = %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (quantity, unit_price, existing[0]))
        else:
            cursor.execute("""
                INSERT INTO drugs (name, strength, unit_price, stock_quantity, expiry_date)
                VALUES (%s, %s, %s, %s, %s)
            """, (drug_name, strength, unit_price, quantity, expiry_date_obj))

        conn.commit()
        cursor.close()
        conn.close()

        flash("Stock added successfully.", "success")
        return redirect(url_for('add_stock'))

    return render_template("add_stock.html")

@app.route('/api/drugs')
def api_drugs():
    if 'pharmacist_id' not in session:
        return jsonify([])

    search = request.args.get('q', '').strip()
    conn = get_db_connection()
    if not conn:
        return jsonify([])

    cur = conn.cursor()
    query = """
        SELECT id, name, strength, unit_price, stock_quantity
        FROM drugs
        WHERE stock_quantity > 0
    """
    params = ()
    
    if search:
        query += " AND LOWER(name) LIKE LOWER(%s)"
        params = (f"{search}%",)
    
    query += " ORDER BY name ASC"
    cur.execute(query, params)
    
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify([{
        "id": r[0], "name": r[1], "strength": r[2],
        "unit_price": float(r[3]), "stock_quantity": r[4]
    } for r in rows])

@app.route('/pharmacy/receipt', methods=['POST'])
def pharmacy_receipt():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    data = request.json
    receipt_no = f"RX-{uuid.uuid4().hex[:8].upper()}"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO drug_sales
        (receipt_no, items, subtotal, discount, tax, grand_total, pharmacist)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        receipt_no,
        json.dumps(data["items"]),
        data["subtotal"],
        data["discount"],
        data["tax"],
        data["grand_total"],
        session.get('pharmacist_username')
    ))

    conn.commit()
    cur.close()
    conn.close()

    data["receipt_no"] = receipt_no
    return render_template(
        "receipt.html",
        receipt=data,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        pharmacist_name=session.get('pharmacist_username')
    )

@app.route('/pharmacy/save-patient', methods=['POST'])
def save_patient_info():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    receipt_no = request.form['receipt_no']
    patient_name = request.form['patient_name']
    patient_id = request.form.get('patient_id')

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE drug_sales
        SET patient_name=%s, patient_id=%s
        WHERE receipt_no=%s
    """, (patient_name, patient_id, receipt_no))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for('reprint_receipt', receipt_no=receipt_no))

@app.route('/pharmacy/receipt/<receipt_no>')
def reprint_receipt(receipt_no):
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT receipt_no, patient_name, patient_id, items,
               subtotal, discount, tax, grand_total, pharmacist, created_at
        FROM drug_sales
        WHERE receipt_no = %s
    """, (receipt_no,))

    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        flash("Receipt not found", "danger")
        return redirect(url_for('pharmacy_dashboard'))

    receipt = {
        "receipt_no": row[0], "patient_name": row[1], "patient_id": row[2],
        "items": json.loads(row[3]) if row[3] else [],
        "subtotal": float(row[4]), "discount": float(row[5]),
        "tax": float(row[6]), "grand_total": float(row[7]),
        "pharmacist": row[8], "date": row[9].strftime("%Y-%m-%d %H:%M:%S")
    }

    return render_template("receipt.html", receipt=receipt, hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

@app.route("/pharmacy/confirm-payment", methods=["POST"])
def confirm_payment():
    if "pharmacist_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json()
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO receipts (
                patient_name, patient_id, subtotal, discount, tax,
                total_amount, grand_total, created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id;
        """, (
            data.get("patient_name"), data.get("patient_id"),
            data["subtotal"], data["discount"], data["tax"],
            data["grand_total"], data["grand_total"], datetime.now()
        ))

        receipt_id = cur.fetchone()[0]

        for item in data["items"]:
            cur.execute("""
                INSERT INTO receipt_items (
                    receipt_id, drug_name, strength, quantity, unit_price
                )
                VALUES (%s, %s, %s, %s, %s);
            """, (
                receipt_id, item["drug_name"], item["strength"],
                item["quantity"], item["unit_price"]
            ))

            cur.execute("""
                UPDATE drugs
                SET stock_quantity = stock_quantity - %s,
                    updated_at = NOW()
                WHERE name = %s AND strength = %s;
            """, (
                item["quantity"], item["drug_name"], item["strength"]
            ))

        conn.commit()
        return jsonify({"success": True, "receipt_id": receipt_id})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        cur.close()
        conn.close()

@app.route("/pharmacy/stock-report")
def stock_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    filter_type = request.args.get("filter", "all")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, strength, stock_quantity, unit_price, 
               expiry_date, low_stock_threshold
        FROM drugs
        ORDER BY expiry_date ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    stock = build_stock_snapshot(rows, date.today())
    stock = apply_stock_filter(stock, filter_type)

    return render_template(
        "stock_report.html",
        stock=stock,
        current_filter=filter_type,
        expired_count=sum(1 for d in stock if d["status"] == "EXPIRED"),
        expiring_soon_count=sum(1 for d in stock if d["status"] == "EXPIRING_SOON"),
        low_stock_count=sum(1 for d in stock if d["quantity"] <= d["low_stock_threshold"]),
        total_stock_value=sum(d["total_value"] for d in stock)
    )

@app.route("/pharmacy/stock-report/export")
def export_stock_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    filter_type = request.args.get("filter", "all")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, strength, stock_quantity, unit_price, 
               expiry_date, low_stock_threshold
        FROM drugs
        ORDER BY expiry_date ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    stock = build_stock_snapshot(rows, date.today())
    stock = apply_stock_filter(stock, filter_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    headers = [
        "Drug Name", "Strength", "Quantity", "Unit Price (₦)",
        "Expiry Date", "Days Left", "Status", "Total Value (₦)", "Low Stock Threshold"
    ]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    fills = {
        "EXPIRED": PatternFill("solid", fgColor="FF9999"),
        "EXPIRING_SOON": PatternFill("solid", fgColor="FFFF99"),
        "LOW": PatternFill("solid", fgColor="ADD8E6")
    }

    for item in stock:
        ws.append([
            item["name"], item["strength"], item["quantity"],
            float(item["unit_price"]), item["expiry_date"],
            item["days_left"], item["status"],
            float(item["total_value"]), item["low_stock_threshold"]
        ])

        row_idx = ws.max_row
        if item["status"] in fills:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fills[item["status"]]
        elif item["quantity"] <= item["low_stock_threshold"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fills["LOW"]

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=f"pharmacy_stock_report_{date.today()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/pharmacy/stock-movements')
def stock_movements():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT sm.id, d.name, d.strength, sm.movement_type, 
               sm.quantity, u.username, sm.created_at, sm.note
        FROM stock_movements sm
        JOIN drugs d ON sm.drug_id = d.id
        JOIN users u ON sm.user_id = u.id
        ORDER BY sm.created_at DESC
    """)
    movements = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('stock_movements.html', movements=movements)

@app.route("/pharmacy/revenue-report", methods=["GET", "POST"])
def revenue_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    report_type = request.form.get("period", "daily")
    selected_day = request.form.get("day")
    selected_month = request.form.get("month")
    selected_year = request.form.get("year")
    today = date.today()

    if report_type == "daily":
        start_date = end_date = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today
    elif report_type == "weekly":
        d = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today
        start_date = d - timedelta(days=d.weekday())
        end_date = start_date + timedelta(days=6)
    elif report_type == "monthly":
        month = int(selected_month) if selected_month else today.month
        year = int(selected_year) if selected_year else today.year
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    else:
        flash("Invalid report period", "danger")
        return redirect(url_for("pharmacy_dashboard"))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, patient_name, patient_id, grand_total, created_at
        FROM receipts
        WHERE DATE(created_at) BETWEEN %s AND %s
        ORDER BY created_at ASC;
    """, (start_date, end_date))
    sales = cur.fetchall()
    cur.close()
    conn.close()

    total_revenue = sum(float(s[3]) if s[3] is not None else 0.0 for s in sales)
    months = [(i, month_name[i]) for i in range(1, 13)]
    years = range(2024, today.year + 1)

    return render_template(
        "revenue_report.html",
        sales=sales,
        total_revenue=total_revenue,
        period=report_type,
        start_date=start_date,
        end_date=end_date,
        selected_day=selected_day,
        selected_month=int(selected_month) if selected_month else today.month,
        selected_year=int(selected_year) if selected_year else today.year,
        months=months,
        years=years
    )

@app.route("/receipt/<int:receipt_id>")
def receipt(receipt_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM receipts WHERE id = %s;", (receipt_id,))
    receipt = cur.fetchone()

    cur.execute("""
        SELECT drug_name, strength, quantity, unit_price
        FROM receipt_items
        WHERE receipt_id = %s;
    """, (receipt_id,))
    items = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "receipt.html",
        receipt=receipt,
        items=items
    )

@app.route("/pharmacy/receipt/<int:receipt_id>")
def view_receipt(receipt_id):
    if "pharmacist_id" not in session:
        return redirect(url_for("pharmacy_login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, patient_name, patient_id, subtotal, discount, tax, grand_total, created_at
        FROM receipts
        WHERE id = %s
    """, (receipt_id,))

    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        flash("Receipt not found", "danger")
        return redirect(url_for("pharmacy_dashboard"))

    receipt = {
        "id": row[0],
        "patient_name": row[1],
        "patient_id": row[2],
        "subtotal": float(row[3]),
        "discount": float(row[4]),
        "tax": float(row[5]),
        "grand_total": float(row[6]),
        "date": row[7].strftime("%Y-%m-%d %H:%M:%S")
    }

    cur.execute("""
        SELECT drug_name, strength, quantity, unit_price
        FROM receipt_items
        WHERE receipt_id = %s
    """, (receipt_id,))

    items_rows = cur.fetchall()
    items = []
    for i in items_rows:
        items.append({
            "drug_name": i[0],
            "strength": i[1],
            "quantity": i[2],
            "unit_price": float(i[3])
        })

    cur.close()
    conn.close()

    return render_template(
        "receipt.html",
        receipt=receipt,
        items=items,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    )

# -------------------- ROUTES: BILLING MODULE --------------------
@app.route("/billing/login", methods=["GET", "POST"])
def billing_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, password
            FROM billing_users
            WHERE username = %s
        """, (username,))

        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user[1], password):
            session["billing_user_id"] = user[0]
            session["billing_username"] = username
            return redirect(url_for("billing_dashboard"))
        else:
            flash("Invalid login credentials", "danger")

    return render_template("billing_login.html")

# REPLACE the old billing_dashboard function with this one:

@app.route("/billing/dashboard")
def billing_dashboard():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))

    # No longer need to fetch invoices
    return render_template("billing_dashboard.html")

@app.route("/billing/logout")
def billing_logout():
    session.pop("billing_user_id", None)
    session.pop("billing_username", None)
    flash("Logged out successfully", "success")
    return redirect(url_for("billing_login"))



@app.route("/billing/confirm-payment", methods=["POST"])
def billing_confirm_payment():
    if "billing_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    patient_name = request.form.get("patient_name")
    service_type = request.form.get("service_type")
    receipt_date = request.form.get("receipt_date")
    payment_method = request.form.get("payment_method")
    amount_paid = float(request.form.get("amount_paid", 0))
    vat_percent = float(request.form.get("vat", 0))
    discount = float(request.form.get("discount", 0))

    subtotal = amount_paid
    vat_amount = (subtotal * vat_percent) / 100
    grand_total = subtotal + vat_amount - discount
    balance = 0 if amount_paid >= grand_total else grand_total - amount_paid
    status = "Paid" if balance <= 0 else "Partial"

    payment_date = datetime.strptime(receipt_date, "%Y-%m-%d").date()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO payments (
                patient_name, service_type, subtotal, discount, tax,
                grand_total, amount_paid, balance, payment_method,
                status, payment_date, recorded_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id;
        """, (
            patient_name, service_type, subtotal, discount, vat_amount,
            grand_total, amount_paid, balance, payment_method,
            status, payment_date, session["billing_user_id"]
        ))

        payment_id = cur.fetchone()[0]
        conn.commit()
        flash(f"Payment recorded successfully. Receipt No: {payment_id}", "success")

        # Redirect to a receipt page
        return redirect(url_for("view_payment_receipt", payment_id=payment_id))

    except Exception as e:
        conn.rollback()
        flash(f"Payment error: {e}", "danger")
        return redirect(url_for("accept_payment_page"))

    finally:
        cur.close()
        conn.close()



@app.route("/billing/receipt/<int:payment_id>")
def billing_receipt(payment_id):
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT id, patient_name, service_type, subtotal, discount, 
                   tax, grand_total, amount_paid, balance, payment_method, 
                   status, payment_date, created_at
            FROM payments
            WHERE id = %s
        """, (payment_id,))

        row = cur.fetchone()
        if not row:
            flash("Receipt not found", "danger")
            return redirect(url_for("billing_dashboard"))

        receipt = {
            "id": row[0],
            "patient_name": row[1],
            "service_type": row[2],
            "subtotal": float(row[3]),
            "discount": float(row[4]),
            "tax": float(row[5]),
            "grand_total": float(row[6]),
            "amount_paid": float(row[7]),
            "balance": float(row[8]),
            "payment_method": row[9],
            "status": row[10],
            "payment_date": row[11].strftime("%Y-%m-%d"),
            "created_at": row[12].strftime("%Y-%m-%d %H:%M:%S")
        }

        cur.close()
        conn.close()

        return render_template(
            "billing_receipt.html",
            receipt=receipt,
            hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
            user_name=session.get("billing_username")
        )

    except Exception as e:
        cur.close()
        conn.close()
        flash(f"Error retrieving receipt: {e}", "danger")
        return redirect(url_for("billing_dashboard"))

# 


@app.route("/billing/accept-payment", methods=["GET"])
def accept_payment_page():
    return render_template("accept_payment.html")


@app.route("/billing/receipt/<int:payment_id>")
def billing_receipt_page(payment_id):
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))

    conn = get_db_connection()
    cur = conn.cursor()

    # Fetch payment info
    cur.execute("""
        SELECT id, patient_name, service_type, subtotal, discount, tax,
               grand_total, amount_paid, balance, payment_method, payment_date
        FROM payments
        WHERE id = %s
    """, (payment_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        flash("Receipt not found", "danger")
        return redirect(url_for("billing_dashboard"))

    receipt = {
        "id": row[0],
        "patient_name": row[1],
        "service_type": row[2],
        "subtotal": float(row[3]),
        "discount": float(row[4]),
        "tax": float(row[5]),
        "grand_total": float(row[6]),
        "amount_paid": float(row[7]),
        "balance": float(row[8]),
        "payment_method": row[9],
        "date": row[10].strftime("%Y-%m-%d %H:%M:%S")
    }

    return render_template("billing_receipt.html", receipt=receipt, hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

@app.route("/billing/receipt/<int:payment_id>")
def view_payment_receipt(payment_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM payments WHERE id=%s", (payment_id,))
    payment = cur.fetchone()
    cur.close()
    conn.close()

    if not payment:
        flash("Payment not found", "danger")
        return redirect(url_for("billing_dashboard"))

    return render_template("payment_receipt.html", payment=payment)


# -------------------- ROUTES: BILLING MODULE - PAYMENT HISTORY --------------------

@app.route("/billing/payment-history")
def payment_history():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))

    # Get filter parameters
    patient_name = request.args.get("patient_name", "").strip()
    service_type = request.args.get("service_type", "")
    payment_method = request.args.get("payment_method", "")
    status = request.args.get("status", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    
    # Pagination
    page = request.args.get("page", 1, type=int)
    per_page = 20
    
    # Build query
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Base query
    query = "SELECT * FROM payments WHERE 1=1"
    count_query = "SELECT COUNT(*) FROM payments WHERE 1=1"
    params = []
    
    # Apply filters
    if patient_name:
        query += " AND LOWER(patient_name) LIKE LOWER(%s)"
        count_query += " AND LOWER(patient_name) LIKE LOWER(%s)"
        params.append(f"%{patient_name}%")
    
    if service_type:
        query += " AND service_type = %s"
        count_query += " AND service_type = %s"
        params.append(service_type)
    
    if payment_method:
        query += " AND payment_method = %s"
        count_query += " AND payment_method = %s"
        params.append(payment_method)
    
    if status:
        query += " AND status = %s"
        count_query += " AND status = %s"
        params.append(status)
    
    if start_date:
        query += " AND payment_date >= %s"
        count_query += " AND payment_date >= %s"
        params.append(start_date)
    
    if end_date:
        query += " AND payment_date <= %s"
        count_query += " AND payment_date <= %s"
        params.append(end_date)
    
    # Get total count
    cur.execute(count_query, params)
    total_items = cur.fetchone()[0]
    
    # Apply ordering and pagination
    query += " ORDER BY created_at DESC LIMIT %s OFFSET %s"
    offset = (page - 1) * per_page
    params.extend([per_page, offset])
    
    # Execute main query
    cur.execute(query, params)
    payments = cur.fetchall()
    
    # Get unique service types for dropdown
    cur.execute("SELECT DISTINCT service_type FROM payments WHERE service_type IS NOT NULL ORDER BY service_type")
    service_types = [row[0] for row in cur.fetchall()]
    
    # Calculate total amount
    total_amount = 0
    formatted_payments = []
    for payment in payments:
        payment_dict = {
            "id": payment[0],
            "patient_name": payment[1],
            "service_type": payment[2],
            "subtotal": float(payment[3]),
            "discount": float(payment[4]),
            "tax": float(payment[5]),
            "grand_total": float(payment[6]),
            "amount_paid": float(payment[7]),
            "balance": float(payment[8]),
            "payment_method": payment[9],
            "status": payment[10],
            "payment_date": payment[11],
            "created_at": payment[13]
        }
        formatted_payments.append(payment_dict)
        total_amount += payment_dict["amount_paid"]
    
    cur.close()
    conn.close()
    
    # Calculate pagination
    total_pages = (total_items + per_page - 1) // per_page
    
    return render_template(
        "billing_payment_history.html",
        payments=formatted_payments,
        service_types=service_types,
        total_items=total_items,
        total_amount=total_amount,
        page=page,
        total_pages=total_pages,
        current_filters=request.args
    )


@app.route("/billing/payment-history/export")
def export_payment_history():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    # Get filter parameters (same as payment_history)
    patient_name = request.args.get("patient_name", "").strip()
    service_type = request.args.get("service_type", "")
    payment_method = request.args.get("payment_method", "")
    status = request.args.get("status", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Build query without pagination
    query = "SELECT * FROM payments WHERE 1=1"
    params = []
    
    if patient_name:
        query += " AND LOWER(patient_name) LIKE LOWER(%s)"
        params.append(f"%{patient_name}%")
    
    if service_type:
        query += " AND service_type = %s"
        params.append(service_type)
    
    if payment_method:
        query += " AND payment_method = %s"
        params.append(payment_method)
    
    if status:
        query += " AND status = %s"
        params.append(status)
    
    if start_date:
        query += " AND payment_date >= %s"
        params.append(start_date)
    
    if end_date:
        query += " AND payment_date <= %s"
        params.append(end_date)
    
    query += " ORDER BY created_at DESC"
    cur.execute(query, params)
    payments = cur.fetchall()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment History"
    
    # Add headers
    headers = [
        "Receipt No", "Patient Name", "Service Type", 
        "Subtotal (₦)", "Discount (₦)", "Tax (₦)", "Grand Total (₦)",
        "Amount Paid (₦)", "Balance (₦)", "Payment Method",
        "Status", "Payment Date", "Created At", "Recorded By"
    ]
    ws.append(headers)
    
    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for payment in payments:
        ws.append([
            payment[0],  # id
            payment[1],  # patient_name
            payment[2],  # service_type
            float(payment[3]),  # subtotal
            float(payment[4]),  # discount
            float(payment[5]),  # tax
            float(payment[6]),  # grand_total
            float(payment[7]),  # amount_paid
            float(payment[8]),  # balance
            payment[9],  # payment_method
            payment[10],  # status
            payment[11],  # payment_date
            payment[13],  # created_at
            payment[12]   # recorded_by
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary row
    ws.append([])
    ws.append(["SUMMARY", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    
    if payments:
        total_amount = sum(float(p[7]) for p in payments)
        total_balance = sum(float(p[8]) for p in payments)
        
        summary_headers = ["Total Payments", "Total Amount", "Total Balance"]
        summary_values = [len(payments), total_amount, total_balance]
        
        for i, (header, value) in enumerate(zip(summary_headers, summary_values)):
            ws.append([header, value])
        
        # Style summary
        summary_row = ws.max_row - len(summary_headers) + 1
        for i in range(len(summary_headers)):
            ws.cell(row=summary_row + i, column=1).font = Font(bold=True)
    
    # Save to BytesIO
    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    cur.close()
    conn.close()
    
    # Generate filename
    filename = f"payment_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -------------------- ROUTES: TODAY'S COLLECTION --------------------

@app.route("/billing/todays-collection")
def todays_collection():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    today = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get today's payments
    cur.execute("""
        SELECT id, patient_name, service_type, amount_paid, 
               payment_method, status, created_at
        FROM payments 
        WHERE DATE(payment_date) = %s
        ORDER BY created_at DESC
    """, (today,))
    
    today_payments = cur.fetchall()
    
    # Calculate totals by payment method
    payment_methods_data = {
        'Cash': {'amount': 0, 'count': 0},
        'Card': {'amount': 0, 'count': 0},
        'Transfer': {'amount': 0, 'count': 0},
        'POS': {'amount': 0, 'count': 0},
        'Insurance': {'amount': 0, 'count': 0},
        'Other': {'amount': 0, 'count': 0}
    }
    
    # Process payments
    total_transactions = len(today_payments)
    grand_total = 0
    amounts = []
    
    recent_transactions = []
    for payment in today_payments:
        amount_paid = float(payment[3])
        payment_method = payment[4]
        
        # Add to grand total
        grand_total += amount_paid
        amounts.append(amount_paid)
        
        # Add to payment method totals
        if payment_method in payment_methods_data:
            payment_methods_data[payment_method]['amount'] += amount_paid
            payment_methods_data[payment_method]['count'] += 1
        else:
            payment_methods_data['Other']['amount'] += amount_paid
            payment_methods_data['Other']['count'] += 1
        
        # Prepare recent transactions data
        recent_transactions.append({
            'id': payment[0],
            'patient_name': payment[1],
            'service_type': payment[2],
            'amount_paid': amount_paid,
            'payment_method': payment_method,
            'status': payment[5],
            'created_at': payment[6]
        })
    
    # Calculate additional statistics
    average_transaction = grand_total / total_transactions if total_transactions > 0 else 0
    highest_transaction = max(amounts) if amounts else 0
    lowest_transaction = min(amounts) if amounts else 0
    
    # Calculate totals for time periods
    morning_total = 0  # 6AM - 12PM
    afternoon_total = 0  # 12PM - 4PM
    evening_total = 0  # 4PM - 10PM
    
    for payment in today_payments:
        created_at = payment[6]
        if created_at:
            hour = created_at.hour
            amount = float(payment[3])
            
            if 6 <= hour < 12:
                morning_total += amount
            elif 12 <= hour < 16:
                afternoon_total += amount
            elif 16 <= hour < 22:
                evening_total += amount
    
    # Prepare payment methods for template
    payment_methods = []
    for method_name, data in payment_methods_data.items():
        if data['count'] > 0:  # Only include methods with transactions
            percentage = (data['amount'] / grand_total * 100) if grand_total > 0 else 0
            payment_methods.append({
                'name': method_name,
                'amount': data['amount'],
                'count': data['count'],
                'percentage': round(percentage, 1)
            })
    
    # Set daily target (you can make this configurable)
    daily_target = 500000.00  # ₦500,000 daily target
    
    cur.close()
    conn.close()
    
    # Format date for display
    today_date = today.strftime("%A, %B %d, %Y")
    
    return render_template(
        "todays_collection.html",
        today_date=today_date,
        grand_total=grand_total,
        cash_total=payment_methods_data['Cash']['amount'],
        card_total=payment_methods_data['Card']['amount'],
        transfer_total=payment_methods_data['Transfer']['amount'],
        pos_total=payment_methods_data['POS']['amount'],
        insurance_total=payment_methods_data['Insurance']['amount'],
        other_total=payment_methods_data['Other']['amount'],
        payment_methods=payment_methods,
        recent_transactions=recent_transactions,
        total_transactions=total_transactions,
        average_transaction=average_transaction,
        highest_transaction=highest_transaction,
        lowest_transaction=lowest_transaction,
        morning_total=morning_total,
        afternoon_total=afternoon_total,
        evening_total=evening_total,
        daily_target=daily_target
    )


@app.route("/billing/todays-collection/export")
def export_todays_collection():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    today = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get today's payments
    cur.execute("""
        SELECT id, patient_name, service_type, subtotal, discount, tax,
               grand_total, amount_paid, balance, payment_method, 
               status, payment_date, created_at
        FROM payments 
        WHERE DATE(payment_date) = %s
        ORDER BY created_at DESC
    """, (today,))
    
    today_payments = cur.fetchall()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Today's Collection - {today}"
    
    # Add title
    ws.append([f"Today's Collection Report - {today.strftime('%B %d, %Y')}"])
    ws.append([])
    
    # Add summary section
    ws.append(["SUMMARY"])
    ws.append([])
    
    # Calculate totals by payment method
    payment_methods_data = {
        'Cash': {'amount': 0, 'count': 0},
        'Card': {'amount': 0, 'count': 0},
        'Transfer': {'amount': 0, 'count': 0},
        'POS': {'amount': 0, 'count': 0},
        'Insurance': {'amount': 0, 'count': 0},
        'Other': {'amount': 0, 'count': 0}
    }
    
    grand_total = 0
    total_transactions = len(today_payments)
    
    for payment in today_payments:
        amount_paid = float(payment[7])
        payment_method = payment[9]
        grand_total += amount_paid
        
        if payment_method in payment_methods_data:
            payment_methods_data[payment_method]['amount'] += amount_paid
            payment_methods_data[payment_method]['count'] += 1
        else:
            payment_methods_data['Other']['amount'] += amount_paid
            payment_methods_data['Other']['count'] += 1
    
    # Write summary
    ws.append(["Total Transactions:", total_transactions])
    ws.append(["Grand Total:", grand_total])
    ws.append([])
    ws.append(["Payment Method Breakdown"])
    ws.append(["Method", "Count", "Amount", "Percentage"])
    
    for method_name, data in payment_methods_data.items():
        if data['count'] > 0:
            percentage = (data['amount'] / grand_total * 100) if grand_total > 0 else 0
            ws.append([
                method_name,
                data['count'],
                data['amount'],
                f"{percentage:.1f}%"
            ])
    
    ws.append([])
    ws.append([])
    
    # Add detailed transactions
    ws.append(["DETAILED TRANSACTIONS"])
    ws.append([])
    
    headers = [
        "Receipt No", "Patient Name", "Service Type", 
        "Subtotal", "Discount", "Tax", "Grand Total",
        "Amount Paid", "Balance", "Payment Method",
        "Status", "Payment Date", "Time"
    ]
    ws.append(headers)
    
    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))
    
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Add data rows
    for payment in today_payments:
        ws.append([
            payment[0],  # id
            payment[1],  # patient_name
            payment[2],  # service_type
            float(payment[3]),  # subtotal
            float(payment[4]),  # discount
            float(payment[5]),  # tax
            float(payment[6]),  # grand_total
            float(payment[7]),  # amount_paid
            float(payment[8]),  # balance
            payment[9],  # payment_method
            payment[10],  # status
            payment[11].strftime('%Y-%m-%d'),  # payment_date
            payment[12].strftime('%H:%M:%S') if payment[12] else ''  # created_at time
        ])
    
    # Apply borders to data rows
    for row in ws.iter_rows(min_row=ws.max_row - len(today_payments) + 1, max_row=ws.max_row):
        for cell in row:
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    cur.close()
    conn.close()
    
    # Save to BytesIO
    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # Generate filename
    filename = f"todays_collection_{today.strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Custom filter for currency formatting
@app.template_filter('currency')
def currency_filter(amount):
    """Format amount as Nigerian Naira currency."""
    if amount is None:
        return "₦0.00"
    return f"₦{float(amount):,.2f}"

# -------------------- HR DATABASE TABLES --------------------
def create_hr_tables():
    """Create HR-related tables in PostgreSQL."""
    conn = get_db_connection()
    if not conn:
        app.logger.error("Cannot connect to database for HR table creation")
        return

    cursor = conn.cursor()
    
    # Enable UUID extension if needed
    try:
        cursor.execute("CREATE EXTENSION IF NOT EXISTS \"uuid-ossp\";")
    except Exception as e:
        app.logger.warning(f"Could not enable uuid-ossp extension: {e}")
    
    # HR Users Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS hr_users (
            id SERIAL PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            full_name VARCHAR(100) NOT NULL,
            email VARCHAR(100),
            role VARCHAR(50) DEFAULT 'HR Staff',
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Departments Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS departments (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            code VARCHAR(20) UNIQUE NOT NULL,
            description TEXT,
            head_of_dept VARCHAR(100),
            status VARCHAR(20) DEFAULT 'Active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Staff Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS staff (
            id SERIAL PRIMARY KEY,
            staff_id VARCHAR(50) UNIQUE NOT NULL,
            first_name VARCHAR(100) NOT NULL,
            last_name VARCHAR(100) NOT NULL,
            department_id INTEGER REFERENCES departments(id),
            position VARCHAR(100) NOT NULL,
            employment_type VARCHAR(50),
            email VARCHAR(100),
            phone VARCHAR(20),
            hire_date DATE NOT NULL,
            salary DECIMAL(12, 2),
            status VARCHAR(20) DEFAULT 'Active',
            emergency_contact VARCHAR(100),
            address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Attendance Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            date DATE NOT NULL,
            check_in TIME,
            check_out TIME,
            status VARCHAR(20),
            remarks TEXT,
            recorded_by INTEGER REFERENCES hr_users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Leaves Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leaves (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            leave_type VARCHAR(50) NOT NULL,
            start_date DATE NOT NULL,
            end_date DATE NOT NULL,
            days_requested INTEGER NOT NULL,
            reason TEXT,
            status VARCHAR(20) DEFAULT 'Pending',
            approved_by INTEGER REFERENCES hr_users(id),
            approved_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Schedules Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS schedules (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            schedule_date DATE NOT NULL,
            shift_type VARCHAR(50),
            start_time TIME NOT NULL,
            end_time TIME NOT NULL,
            location VARCHAR(100),
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Payroll Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            pay_period VARCHAR(50),
            basic_salary DECIMAL(12, 2),
            allowances DECIMAL(12, 2),
            deductions DECIMAL(12, 2),
            net_salary DECIMAL(12, 2),
            status VARCHAR(20) DEFAULT 'Pending',
            payment_date DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Documents Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS documents (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER REFERENCES staff(id),
            document_type VARCHAR(50),
            document_name VARCHAR(255),
            file_path VARCHAR(500),
            uploaded_by INTEGER REFERENCES hr_users(id),
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Create indexes for better performance
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_staff_department ON staff(department_id);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_staff_date ON attendance(staff_id, date);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_staff_status ON leaves(staff_id, status);",
        "CREATE INDEX IF NOT EXISTS idx_schedules_staff_date ON schedules(staff_id, schedule_date);",
        "CREATE INDEX IF NOT EXISTS idx_payroll_staff_period ON payroll(staff_id, pay_period);",
        "CREATE INDEX IF NOT EXISTS idx_staff_status ON staff(status);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(date);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_status ON leaves(status);"
    ]
    
    for index_query in indexes:
        try:
            cursor.execute(index_query)
        except Exception as e:
            app.logger.warning(f"Could not create index: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()
    
    # Insert default data
    create_default_hr_data()

def create_default_hr_data():
    """Insert default HR data into PostgreSQL tables."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    # Default hashed password for 'hr@admin123'
    hashed_password = generate_password_hash('hr@admin123')
    
    # Insert default HR users
    try:
        cursor.execute("""
            INSERT INTO hr_users (username, password, full_name, email, role) 
            VALUES 
                (%s, %s, %s, %s, %s),
                (%s, %s, %s, %s, %s)
            ON CONFLICT (username) DO NOTHING;
        """, (
            'hr_admin', hashed_password, 'HR Administrator', 'admin@hospital.com', 'HR Manager',
            'hr_staff', hashed_password, 'HR Staff', 'staff@hospital.com', 'HR Officer'
        ))
    except Exception as e:
        app.logger.error(f"Error inserting HR users: {e}")
    
    # Insert sample departments
    departments = [
        ('Administration', 'ADMIN', 'Hospital Administration and Management', 'Dr. John Smith'),
        ('Medical', 'MED', 'Medical Services Department', 'Dr. Sarah Johnson'),
        ('Nursing', 'NURS', 'Nursing Services', 'Mrs. Grace Williams'),
        ('Pharmacy', 'PHARM', 'Pharmacy Department', 'Mr. Michael Brown'),
        ('Laboratory', 'LAB', 'Laboratory Services', 'Dr. David Miller'),
        ('Radiology', 'RAD', 'Radiology Department', 'Dr. Lisa Davis'),
        ('Finance', 'FIN', 'Finance and Billing Department', 'Mr. Robert Wilson'),
        ('Human Resources', 'HR', 'Human Resources Department', 'Ms. Patricia Taylor'),
        ('Maintenance', 'MAINT', 'Facility Maintenance', 'Mr. Thomas Anderson'),
        ('Security', 'SEC', 'Hospital Security', 'Mr. Richard Clark')
    ]
    
    for dept in departments:
        try:
            cursor.execute("""
                INSERT INTO departments (name, code, description, head_of_dept) 
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (code) DO NOTHING;
            """, dept)
        except Exception as e:
            app.logger.error(f"Error inserting department {dept[0]}: {e}")
    
    # Get admin department ID for sample staff
    cursor.execute("SELECT id FROM departments WHERE code = 'ADMIN' LIMIT 1;")
    admin_dept = cursor.fetchone()
    
    # Insert sample staff if departments exist
    if admin_dept:
        sample_staff = [
            ('EMP001', 'John', 'Doe', admin_dept[0], 'Hospital Administrator', 'Full-Time', 
             'john.doe@hospital.com', '08012345678', '2022-01-15', 850000.00, 'Jane Doe - 08087654321'),
            ('EMP002', 'Sarah', 'Johnson', admin_dept[0], 'Senior Doctor', 'Full-Time', 
             'sarah.j@hospital.com', '08023456789', '2021-03-20', 1200000.00, 'Mark Johnson - 08098765432'),
            ('EMP003', 'Michael', 'Brown', admin_dept[0], 'Chief Pharmacist', 'Full-Time', 
             'michael.b@hospital.com', '08034567890', '2020-06-10', 950000.00, 'Emily Brown - 08076543210'),
            ('EMP004', 'Grace', 'Williams', admin_dept[0], 'Head Nurse', 'Full-Time', 
             'grace.w@hospital.com', '08045678901', '2019-08-05', 750000.00, 'James Williams - 08065432109'),
            ('EMP005', 'David', 'Miller', admin_dept[0], 'Lab Technician', 'Full-Time', 
             'david.m@hospital.com', '08056789012', '2022-11-30', 650000.00, 'Sarah Miller - 08054321098')
        ]
        
        for staff in sample_staff:
            try:
                cursor.execute("""
                    INSERT INTO staff (staff_id, first_name, last_name, department_id, position, 
                                      employment_type, email, phone, hire_date, salary, emergency_contact) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (staff_id) DO NOTHING;
                """, staff)
            except Exception as e:
                app.logger.error(f"Error inserting staff {staff[0]}: {e}")
    
    # Get HR admin ID for recording
    cursor.execute("SELECT id FROM hr_users WHERE username = 'hr_admin' LIMIT 1;")
    hr_admin = cursor.fetchone()
    
    # Get staff IDs for sample data
    cursor.execute("SELECT id, staff_id FROM staff ORDER BY id LIMIT 5;")
    staff_members = cursor.fetchall()
    
    if hr_admin and staff_members:
        hr_admin_id = hr_admin[0]
        today = date.today()
        
        # Insert sample attendance for today
        for i, staff in enumerate(staff_members[:3]):  # First 3 staff
            check_in = '08:00:00' if i != 1 else '08:30:00'  # Make second staff late
            status = 'Present' if i != 1 else 'Late'
            
            try:
                cursor.execute("""
                    INSERT INTO attendance (staff_id, date, check_in, check_out, status, recorded_by)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING;
                """, (staff[0], today, check_in, '16:00:00', status, hr_admin_id))
            except Exception as e:
                app.logger.error(f"Error inserting attendance for {staff[1]}: {e}")
        
        # Insert sample leave requests
        try:
            cursor.execute("""
                INSERT INTO leaves (staff_id, leave_type, start_date, end_date, days_requested, reason, status)
                SELECT 
                    id,
                    'Annual Leave',
                    %s + INTERVAL '5 days',
                    %s + INTERVAL '12 days',
                    8,
                    'Family vacation',
                    'Pending'
                FROM staff WHERE staff_id = 'EMP004'
                UNION ALL
                SELECT 
                    id,
                    'Sick Leave',
                    %s - INTERVAL '2 days',
                    %s + INTERVAL '2 days',
                    5,
                    'Medical treatment',
                    'Approved'
                FROM staff WHERE staff_id = 'EMP005'
                ON CONFLICT DO NOTHING;
            """, (today, today, today, today))
        except Exception as e:
            app.logger.error(f"Error inserting leaves: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()

# -------------------- ROUTES: HR MODULE --------------------
@app.route("/hr/login", methods=["GET", "POST"])
def hr_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return render_template("hr_login.html", 
                                 hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Enugu State")

        cur = conn.cursor()

        try:
            cur.execute("""
                SELECT id, username, password, full_name, role
                FROM hr_users
                WHERE username = %s AND is_active = TRUE
            """, (username,))

            user = cur.fetchone()
            cur.close()
            conn.close()

            if user:
                stored_hash = user[2]
                
                # SIMPLE CHECK for our known bcrypt hash
                known_hash = "$2b$12$LQv3c1yqBWVHxkd0LsZcdeJN8L7Fmm8Zz3qG9XwFk8kC1YdV6n4Oq"
                
                if stored_hash == known_hash and password == "hr@admin123":
                    session["hr_user_id"] = user[0]
                    session["hr_username"] = user[1]
                    session["hr_full_name"] = user[3]
                    session["hr_role"] = user[4]
                    flash(f"Welcome, {user[3]}!", "success")
                    return redirect(url_for("hr_dashboard"))
                else:
                    flash("Invalid password. Use: hr@admin123", "danger")
            else:
                flash("Invalid username. Use: hr_admin or hr_staff", "danger")

        except Exception as e:
            app.logger.error(f"HR login error: {e}")
            flash("Login error. Please try again.", "danger")

    return render_template("hr_login.html", 
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")    


# @app.route("/hr/scheduling")
# def scheduling():
#     if "hr_user_id" not in session:
#         return redirect(url_for("hr_login"))
#     return render_template("module_placeholder.html", 
#                          module_name="Scheduling",
#                          description="Create and manage work schedules and shifts")

# 



# @app.route("/hr/departments")
# def departments():
#     if "hr_user_id" not in session:
#         return redirect(url_for("hr_login"))
    
#     conn = get_db_connection()
#     cur = conn.cursor()
    
#     try:
#         # Get all departments with staff count
#         cur.execute("""
#             SELECT d.*, 
#                    COUNT(s.id) as staff_count
#             FROM departments d
#             LEFT JOIN staff s ON d.id = s.department_id AND s.status = 'Active'
#             GROUP BY d.id
#             ORDER BY d.name
#         """)
#         dept_list = cur.fetchall()
        
#     except Exception as e:
#         app.logger.error(f"Error fetching departments: {e}")
#         dept_list = []
    
#     finally:
#         cur.close()
#         conn.close()
    
#     return render_template("hr_departments.html", 
#                          module_name="Departments",
#                          description="Manage hospital departments and reporting structure",
#                          departments=dept_list)

@app.route("/hr/reports")
def hr_reports():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    return render_template("module_placeholder.html", 
                         module_name="HR Reports",
                         description="Generate HR analytics and reports")

# # -------------------- QUICK ACTION PLACEHOLDERS --------------------
# @app.route("/hr/add-staff")
# def add_staff():
#     if "hr_user_id" not in session:
#         return redirect(url_for("hr_login"))
#     flash("Feature coming soon: Add New Staff", "info")
#     return redirect(url_for("hr_dashboard"))





@app.route("/hr/generate-payroll")
def generate_payroll():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    flash("Feature coming soon: Generate Payroll", "info")
    return redirect(url_for("hr_dashboard"))

@app.route("/hr/upload-documents")
def upload_documents():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    flash("Feature coming soon: Upload Documents", "info")
    return redirect(url_for("hr_dashboard"))

@app.route("/hr/send-notifications")
def send_notifications():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    flash("Feature coming soon: Send Notifications", "info")
    return redirect(url_for("hr_dashboard"))

@app.route("/hr/dashboard")
def hr_dashboard():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get HR statistics
    conn = get_db_connection()
    if not conn:
        flash("Database connection error", "danger")
        return redirect(url_for("hr_login"))
    
    cur = conn.cursor()
    
    try:
        # Total staff
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active'")
        total_staff = cur.fetchone()[0] or 0
        
        # Staff active today (attendance)
        today = date.today()
        cur.execute("""
            SELECT COUNT(DISTINCT staff_id) 
            FROM attendance 
            WHERE date = %s AND status IN ('Present', 'Late')
        """, (today,))
        active_staff = cur.fetchone()[0] or 0
        
        # Staff on leave today
        cur.execute("""
            SELECT COUNT(*) 
            FROM leaves 
            WHERE %s BETWEEN start_date AND end_date 
            AND status = 'Approved'
        """, (today,))
        on_leave = cur.fetchone()[0] or 0
        
        # Departments count
        cur.execute("SELECT COUNT(*) FROM departments WHERE status = 'Active'")
        departments_count = cur.fetchone()[0] or 0
        
        # Pending leave requests
        cur.execute("SELECT COUNT(*) FROM leaves WHERE status = 'Pending'")
        pending_leave = cur.fetchone()[0] or 0
        
        # Upcoming shifts (next 7 days)
        next_week = today + timedelta(days=7)
        cur.execute("""
            SELECT COUNT(*) 
            FROM schedules 
            WHERE schedule_date BETWEEN %s AND %s
        """, (today, next_week))
        upcoming_shifts = cur.fetchone()[0] or 0
        
        # Pending updates (staff with missing info)
        cur.execute("""
            SELECT COUNT(*) 
            FROM staff 
            WHERE emergency_contact IS NULL OR address IS NULL
        """)
        pending_updates = cur.fetchone()[0] or 0
        
        # Late arrivals today
        cur.execute("""
            SELECT COUNT(*) 
            FROM attendance 
            WHERE date = %s AND status = 'Late'
        """, (today,))
        late_arrivals = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error fetching HR stats: {e}")
        # Set default values on error
        total_staff = active_staff = on_leave = departments_count = 0
        pending_leave = upcoming_shifts = pending_updates = late_arrivals = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "hr_dashboard.html",
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        total_staff=total_staff,
        active_staff=active_staff,
        on_leave=on_leave,
        departments_count=departments_count,
        pending_leave=pending_leave,
        upcoming_shifts=upcoming_shifts,
        pending_updates=pending_updates,
        late_arrivals=late_arrivals,
        current_year=date.today().year
    )
    
@app.route("/hr/logout")
def hr_logout():
    session.pop("hr_user_id", None)
    session.pop("hr_username", None)
    session.pop("hr_full_name", None)
    session.pop("hr_role", None)
    flash("Logged out successfully", "success")
    return redirect(url_for("hr_login"))

@app.route("/hr/staff-management")
def staff_management():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    if not conn:
        flash("Database connection error", "danger")
        return redirect(url_for("hr_login"))
    
    cur = conn.cursor()
    
    try:
        # Get staff list with department names
        cur.execute("""
            SELECT s.id, s.staff_id, s.first_name, s.last_name, 
                   s.position, s.employment_type, s.email, s.phone,
                   s.hire_date, s.salary, s.status, s.emergency_contact,
                   s.address, d.name as department_name
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            ORDER BY s.id DESC
            LIMIT 100
        """)
        staff_list = cur.fetchall()
        
        # Get statistics
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active'")
        total_staff = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active' AND employment_type = 'Full-Time'")
        active_staff = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM staff WHERE employment_type = 'Contract'")
        on_contract = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(DISTINCT department_id) FROM staff")
        departments_count = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error fetching staff data: {e}")
        staff_list = []
        total_staff = active_staff = on_contract = departments_count = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("staff_management.html", 
                         module_name="Staff Management",
                         description="Manage staff profiles, positions, and employment details",
                         staff_list=staff_list,
                         total_staff=total_staff,
                         active_staff=active_staff,
                         on_contract=on_contract,
                         departments_count=departments_count,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
                         current_year=date.today().year)    
# View Staff Details
@app.route("/hr/staff/<int:staff_id>")
def view_staff(staff_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get staff details with department
        cur.execute("""
            SELECT s.*, d.name as department_name, d.code as department_code
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.id = %s
        """, (staff_id,))
        
        staff = cur.fetchone()
        
        if not staff:
            flash("Staff member not found", "danger")
            return redirect(url_for("staff_management"))
        
        # Convert to dictionary for easier template access
        staff_dict = {
            'id': staff[0],
            'staff_id': staff[1],
            'first_name': staff[2],
            'last_name': staff[3],
            'department_id': staff[4],
            'position': staff[5],
            'employment_type': staff[6],
            'email': staff[7],
            'phone': staff[8],
            'hire_date': staff[9],
            'salary': float(staff[10]) if staff[10] else 0,
            'status': staff[11],
            'emergency_contact': staff[12],
            'address': staff[13],
            'department_name': staff[14],
            'department_code': staff[15]
        }
        
        # Calculate employment duration
        from datetime import date
        today = date.today()
        
        # Ensure hire_date is a date object
        hire_date = staff[9]
        if isinstance(hire_date, str):
            from datetime import datetime
            hire_date = datetime.strptime(hire_date, '%Y-%m-%d').date()
        
        years = today.year - hire_date.year
        months = today.month - hire_date.month
        
        if months < 0:
            years -= 1
            months += 12
        
        employment_duration = f"{years} year(s), {months} month(s)"
        
    except Exception as e:
        app.logger.error(f"Error fetching staff details: {e}")
        flash("Error loading staff details", "danger")
        return redirect(url_for("staff_management"))
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("view_staff.html", 
                         staff=staff_dict,
                         today=today,
                         employment_duration=employment_duration,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")
# Add New Staff
@app.route("/hr/staff/add", methods=["GET", "POST"])
def add_staff():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        # Get form data
        staff_id = request.form.get('staff_id')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        department_id = request.form.get('department_id')
        position = request.form.get('position')
        employment_type = request.form.get('employment_type')
        email = request.form.get('email')
        phone = request.form.get('phone')
        hire_date = request.form.get('hire_date')
        salary = request.form.get('salary')
        emergency_contact = request.form.get('emergency_contact')
        address = request.form.get('address')
        
        # Validate required fields
        if not all([staff_id, first_name, last_name, department_id, position, hire_date]):
            flash("Please fill in all required fields", "danger")
            return redirect(url_for("add_staff"))
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return redirect(url_for("add_staff"))
        
        cur = conn.cursor()
        
        try:
            # Check if staff ID already exists
            cur.execute("SELECT id FROM staff WHERE staff_id = %s", (staff_id,))
            if cur.fetchone():
                flash(f"Staff ID '{staff_id}' already exists. Please use a different ID.", "danger")
                return redirect(url_for("add_staff"))
            
            # Convert salary to decimal or set to 0
            try:
                salary_decimal = float(salary) if salary else 0.00
            except ValueError:
                salary_decimal = 0.00
            
            # Insert new staff
            cur.execute("""
                INSERT INTO staff (
                    staff_id, first_name, last_name, department_id, 
                    position, employment_type, email, phone, 
                    hire_date, salary, emergency_contact, address, status
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Active')
            """, (
                staff_id, first_name, last_name, department_id,
                position, employment_type, email, phone,
                hire_date, salary_decimal, emergency_contact, address
            ))
            
            conn.commit()
            flash(f"Staff member {first_name} {last_name} (ID: {staff_id}) added successfully!", "success")
            
            # Redirect to staff management or view the new staff
            cur.execute("SELECT id FROM staff WHERE staff_id = %s", (staff_id,))
            new_staff_id = cur.fetchone()[0]
            return redirect(url_for("view_staff", staff_id=new_staff_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error adding staff: {e}")
            flash(f"Error adding staff: {str(e)}", "danger")
            return redirect(url_for("add_staff"))
            
        finally:
            cur.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    if not conn:
        flash("Database connection error", "danger")
        return redirect(url_for("staff_management"))
    
    cur = conn.cursor()
    
    try:
        # Get departments for dropdown
        cur.execute("SELECT id, name, code FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get next staff ID suggestion
        cur.execute("""
            SELECT MAX(staff_id) FROM staff 
            WHERE staff_id ~ '^EMP[0-9]+$'
        """)
        last_staff_id = cur.fetchone()[0]
        
        if last_staff_id:
            # Extract number and increment
            import re
            match = re.search(r'EMP(\d+)', last_staff_id)
            if match:
                next_num = int(match.group(1)) + 1
                suggested_id = f"EMP{next_num:03d}"
            else:
                suggested_id = "EMP001"
        else:
            suggested_id = "EMP001"
            
        # Get current date for hire date default
        today = date.today().strftime("%Y-%m-%d")
        
    except Exception as e:
        app.logger.error(f"Error loading form data: {e}")
        departments = []
        suggested_id = "EMP001"
        today = date.today().strftime("%Y-%m-%d")
        
    finally:
        cur.close()
        conn.close()
    
    return render_template("add_staff.html",
                         departments=departments,
                         suggested_id=suggested_id,
                         today=today,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

# Edit Staff
@app.route("/hr/staff/edit/<int:staff_id>", methods=["GET", "POST"])
def edit_staff(staff_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == "POST":
        # Update staff
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        department_id = request.form.get('department_id')
        position = request.form.get('position')
        employment_type = request.form.get('employment_type')
        email = request.form.get('email')
        phone = request.form.get('phone')
        salary = request.form.get('salary')
        status = request.form.get('status')
        emergency_contact = request.form.get('emergency_contact')
        address = request.form.get('address')
        
        try:
            cur.execute("""
                UPDATE staff SET
                    first_name = %s,
                    last_name = %s,
                    department_id = %s,
                    position = %s,
                    employment_type = %s,
                    email = %s,
                    phone = %s,
                    salary = %s,
                    status = %s,
                    emergency_contact = %s,
                    address = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (
                first_name, last_name, department_id,
                position, employment_type, email, phone,
                salary, status, emergency_contact, address,
                staff_id
            ))
            
            conn.commit()
            flash("Staff details updated successfully!", "success")
            return redirect(url_for("view_staff", staff_id=staff_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error updating staff: {e}")
            flash("Error updating staff details", "danger")
    
    # GET request - load staff data
    try:
        cur.execute("SELECT * FROM staff WHERE id = %s", (staff_id,))
        staff = cur.fetchone()
        
        if not staff:
            flash("Staff member not found", "danger")
            return redirect(url_for("staff_management"))
        
        # Get departments
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error loading staff for edit: {e}")
        flash("Error loading staff details", "danger")
        return redirect(url_for("staff_management"))
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("edit_staff.html",
                         staff=staff,
                         departments=departments,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")


# ==================== ROUTES: SCHEDULING MODULE ====================

@app.route("/hr/scheduling")
def scheduling():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get current month schedules
        current_month = date.today().replace(day=1)
        if current_month.month == 12:
            next_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            next_month = current_month.replace(month=current_month.month + 1)
        
        cur.execute("""
            SELECT s.*, st.first_name, st.last_name, st.position, d.name as department_name
            FROM schedules s
            JOIN staff st ON s.staff_id = st.id
            LEFT JOIN departments d ON st.department_id = d.id
            WHERE s.schedule_date >= %s AND s.schedule_date < %s
            ORDER BY s.schedule_date, s.start_time
        """, (current_month, next_month))
        
        schedules = cur.fetchall()
        
        # Get statistics
        cur.execute("SELECT COUNT(*) FROM schedules WHERE schedule_date >= CURRENT_DATE")
        upcoming_shifts = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COUNT(DISTINCT staff_id) 
            FROM schedules 
            WHERE schedule_date >= CURRENT_DATE
        """)
        staff_scheduled = cur.fetchone()[0] or 0
        
        # Get departments for filter
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get staff for filter
        cur.execute("""
            SELECT id, first_name, last_name, position 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error fetching scheduling data: {e}")
        schedules = []
        upcoming_shifts = 0
        staff_scheduled = 0
        departments = []
        staff_list = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("scheduling_dashboard.html",
                         schedules=schedules,
                         upcoming_shifts=upcoming_shifts,
                         staff_scheduled=staff_scheduled,
                         departments=departments,
                         staff_list=staff_list,
                         current_month=current_month.strftime("%B %Y"),
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")
    
    
@app.route("/hr/scheduling/create", methods=["GET", "POST"])
def create_schedule():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        staff_id = request.form.get("staff_id")
        schedule_date = request.form.get("schedule_date")
        shift_type = request.form.get("shift_type")
        start_time = request.form.get("start_time")
        end_time = request.form.get("end_time")
        location = request.form.get("location")
        notes = request.form.get("notes")
        
        if not all([staff_id, schedule_date, start_time, end_time]):
            flash("Please fill in all required fields", "danger")
            return redirect(url_for("create_schedule"))
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Check for existing schedule for same staff on same date
            cur.execute("""
                SELECT id FROM schedules 
                WHERE staff_id = %s AND schedule_date = %s
            """, (staff_id, schedule_date))
            
            if cur.fetchone():
                flash("This staff already has a schedule for this date", "warning")
                return redirect(url_for("create_schedule"))
            
            # Insert new schedule
            cur.execute("""
                INSERT INTO schedules (
                    staff_id, schedule_date, shift_type, 
                    start_time, end_time, location, notes
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (staff_id, schedule_date, shift_type, start_time, end_time, location, notes))
            
            conn.commit()
            flash("Schedule created successfully!", "success")
            return redirect(url_for("scheduling"))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating schedule: {e}")
            flash(f"Error creating schedule: {str(e)}", "danger")
            return redirect(url_for("create_schedule"))
            
        finally:
            cur.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get active staff
        cur.execute("""
            SELECT id, first_name, last_name, position, 
                   (SELECT name FROM departments WHERE id = staff.department_id) as department
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        # Get default tomorrow's date
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        
    except Exception as e:
        app.logger.error(f"Error loading schedule form data: {e}")
        staff_list = []
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        
    finally:
        cur.close()
        conn.close()
    
    return render_template("create_schedule.html",
                         staff_list=staff_list,
                         tomorrow=tomorrow,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")

@app.route("/hr/scheduling/roster")
def view_roster():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get filter parameters
    department_id = request.args.get("department_id", "")
    staff_id = request.args.get("staff_id", "")
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    
    # Default to current week if no dates specified
    if not start_date:
        today = date.today()
        start_date = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
    
    if not end_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = (start_date_obj + timedelta(days=6)).strftime("%Y-%m-%d")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Build query with filters
        query = """
            SELECT 
                s.id as schedule_id,
                s.schedule_date,
                s.shift_type,
                s.start_time,
                s.end_time,
                s.location,
                s.notes,
                st.id as staff_id,
                st.first_name,
                st.last_name,
                st.position,
                d.name as department_name,
                d.id as department_id
            FROM schedules s
            JOIN staff st ON s.staff_id = st.id
            LEFT JOIN departments d ON st.department_id = d.id
            WHERE s.schedule_date BETWEEN %s AND %s
        """
        params = [start_date, end_date]
        
        if department_id:
            query += " AND st.department_id = %s"
            params.append(department_id)
        
        if staff_id:
            query += " AND s.staff_id = %s"
            params.append(staff_id)
        
        query += " ORDER BY s.schedule_date, d.name, st.first_name, s.start_time"
        
        cur.execute(query, params)
        schedules = cur.fetchall()
        
        # Get departments for filter dropdown
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get staff for filter dropdown
        cur.execute("""
            SELECT id, first_name, last_name 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        # Group schedules by date for calendar view - FIXED: Use tuple indices
        schedule_dict = {}
        for schedule in schedules:
            schedule_date = schedule[1].strftime("%Y-%m-%d")
            if schedule_date not in schedule_dict:
                schedule_dict[schedule_date] = []
            
            schedule_dict[schedule_date].append({
                'id': schedule[0],
                'date': schedule[1],
                'shift_type': schedule[2],
                'start_time': schedule[3],
                'end_time': schedule[4],
                'location': schedule[5],
                'notes': schedule[6],
                'staff_id': schedule[7],
                'first_name': schedule[8],
                'last_name': schedule[9],
                'position': schedule[10],
                'department': schedule[11]
            })
        
        # Calculate statistics
        total_shifts = len(schedules)
        unique_staff = len(set([s[7] for s in schedules]))
        unique_departments = len(set([s[11] for s in schedules if s[11]]))
        
    except Exception as e:
        app.logger.error(f"Error fetching roster: {e}")
        schedules = []
        departments = []
        staff_list = []
        schedule_dict = {}
        total_shifts = 0
        unique_staff = 0
        unique_departments = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("view_roster.html",
                         schedules=schedules,
                         schedule_dict=schedule_dict,
                         departments=departments,
                         staff_list=staff_list,
                         start_date=start_date,
                         end_date=end_date,
                         selected_department=department_id,
                         selected_staff=staff_id,
                         total_shifts=total_shifts,
                         unique_staff=unique_staff,
                         unique_departments=unique_departments,
                         hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")
@app.route("/hr/scheduling/shift-swap", methods=["GET", "POST"])
def shift_swap():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == "POST":
        action = request.form.get("action")
        
        if action == "request":
            # Request shift swap
            from_staff_id = request.form.get("from_staff_id")
            to_staff_id = request.form.get("to_staff_id")
            schedule_id = request.form.get("schedule_id")
            reason = request.form.get("reason")
            
            try:
                # Get schedule details
                cur.execute("""
                    SELECT schedule_date, start_time, end_time, shift_type 
                    FROM schedules 
                    WHERE id = %s
                """, (schedule_id,))
                schedule = cur.fetchone()
                
                if not schedule:
                    flash("Schedule not found", "danger")
                    return redirect(url_for("shift_swap"))
                
                # Check if target staff is available on that date
                cur.execute("""
                    SELECT id FROM schedules 
                    WHERE staff_id = %s AND schedule_date = %s
                """, (to_staff_id, schedule[0]))
                
                if cur.fetchone():
                    flash("Selected staff already has a schedule on this date", "warning")
                    return redirect(url_for("shift_swap"))
                
                # Create shift swap request
                cur.execute("""
                    INSERT INTO shift_swap_requests (
                        schedule_id, from_staff_id, to_staff_id, 
                        reason, status, requested_by, requested_at
                    ) VALUES (%s, %s, %s, %s, 'Pending', %s, NOW())
                    RETURNING id
                """, (schedule_id, from_staff_id, to_staff_id, reason, session["hr_user_id"]))
                
                swap_id = cur.fetchone()[0]
                conn.commit()
                flash(f"Shift swap request #{swap_id} submitted successfully!", "success")
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error creating shift swap request: {e}")
                flash(f"Error creating swap request: {str(e)}", "danger")
            
            return redirect(url_for("shift_swap"))
        
        elif action == "approve":
            # Approve shift swap
            swap_id = request.form.get("swap_id")
            
            try:
                # Get swap request details
                cur.execute("""
                    SELECT schedule_id, from_staff_id, to_staff_id 
                    FROM shift_swap_requests 
                    WHERE id = %s AND status = 'Pending'
                """, (swap_id,))
                
                swap_request = cur.fetchone()
                if not swap_request:
                    flash("Swap request not found or already processed", "warning")
                    return redirect(url_for("shift_swap"))
                
                # Update schedule with new staff
                cur.execute("""
                    UPDATE schedules 
                    SET staff_id = %s, 
                        notes = COALESCE(notes, '') || ' [Shift swapped from staff ID ' || %s || ']'
                    WHERE id = %s
                """, (swap_request[2], swap_request[1], swap_request[0]))
                
                # Update swap request status
                cur.execute("""
                    UPDATE shift_swap_requests 
                    SET status = 'Approved', 
                        approved_by = %s, 
                        approved_at = NOW() 
                    WHERE id = %s
                """, (session["hr_user_id"], swap_id))
                
                conn.commit()
                flash(f"Shift swap request #{swap_id} approved successfully!", "success")
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error approving shift swap: {e}")
                flash(f"Error approving swap: {str(e)}", "danger")
            
            return redirect(url_for("shift_swap"))
        
        elif action == "reject":
            # Reject shift swap
            swap_id = request.form.get("swap_id")
            rejection_reason = request.form.get("rejection_reason")
            
            try:
                cur.execute("""
                    UPDATE shift_swap_requests 
                    SET status = 'Rejected', 
                        rejection_reason = %s,
                        reviewed_by = %s, 
                        reviewed_at = NOW() 
                    WHERE id = %s
                """, (rejection_reason, session["hr_user_id"], swap_id))
                
                conn.commit()
                flash(f"Shift swap request #{swap_id} rejected", "info")
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error rejecting shift swap: {e}")
                flash(f"Error rejecting swap: {str(e)}", "danger")
            
            return redirect(url_for("shift_swap"))
    
    # GET request - show shift swap page
    try:
        # Get all shift swap requests with details
        cur.execute("""
            SELECT 
                ssr.id,
                ssr.schedule_id,
                ssr.from_staff_id,
                ssr.to_staff_id,
                ssr.reason,
                ssr.status,
                ssr.requested_at,
                ssr.approved_at,
                ssr.rejection_reason,
                -- From staff details
                s1.first_name as from_first_name,
                s1.last_name as from_last_name,
                s1.position as from_position,
                -- To staff details
                s2.first_name as to_first_name,
                s2.last_name as to_last_name,
                s2.position as to_position,
                -- Schedule details
                sch.schedule_date,
                sch.start_time,
                sch.end_time,
                sch.shift_type,
                -- Requester details
                hu.username as requested_by_username
            FROM shift_swap_requests ssr
            JOIN schedules sch ON ssr.schedule_id = sch.id
            JOIN staff s1 ON ssr.from_staff_id = s1.id
            JOIN staff s2 ON ssr.to_staff_id = s2.id
            LEFT JOIN hr_users hu ON ssr.requested_by = hu.id
            ORDER BY 
                CASE 
                    WHEN ssr.status = 'Pending' THEN 1
                    WHEN ssr.status = 'Approved' THEN 2
                    ELSE 3
                END,
                ssr.requested_at DESC
        """)
        
        all_swaps = cur.fetchall()
        
        # Separate pending and history swaps
        pending_swaps = []
        swap_history = []
        
        for swap in all_swaps:
            swap_dict = {
                'id': swap[0],
                'schedule_id': swap[1],
                'from_staff_id': swap[2],
                'to_staff_id': swap[3],
                'reason': swap[4],
                'status': swap[5],
                'requested_at': swap[6],
                'approved_at': swap[7],
                'rejection_reason': swap[8],
                'from_staff_name': f"{swap[9]} {swap[10]}",
                'from_staff_position': swap[11],
                'to_staff_name': f"{swap[12]} {swap[13]}",
                'to_staff_position': swap[14],
                'schedule_date': swap[15],
                'start_time': swap[16],
                'end_time': swap[17],
                'shift_type': swap[18],
                'requested_by': swap[19]
            }
            
            if swap_dict['status'] == 'Pending':
                pending_swaps.append(swap_dict)
            else:
                swap_history.append(swap_dict)
        
        # Get upcoming schedules for swap requests
        cur.execute("""
            SELECT 
                sch.id,
                sch.schedule_date,
                sch.start_time,
                sch.end_time,
                sch.shift_type,
                st.first_name,
                st.last_name,
                st.position,
                d.name as department_name
            FROM schedules sch
            JOIN staff st ON sch.staff_id = st.id
            LEFT JOIN departments d ON st.department_id = d.id
            WHERE sch.schedule_date >= CURRENT_DATE
            ORDER BY sch.schedule_date, sch.start_time
            LIMIT 50
        """)
        
        upcoming_schedules = cur.fetchall()
        formatted_schedules = []
        for sched in upcoming_schedules:
            formatted_schedules.append({
                'id': sched[0],
                'date': sched[1],
                'start_time': sched[2],
                'end_time': sched[3],
                'shift_type': sched[4],
                'staff_name': f"{sched[5]} {sched[6]}",
                'position': sched[7],
                'department': sched[8]
            })
        
        # Get all active staff for swap requests
        cur.execute("""
            SELECT id, first_name, last_name, position, 
                   (SELECT name FROM departments WHERE id = staff.department_id) as department
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        
        staff_list = cur.fetchall()
        formatted_staff = []
        for staff in staff_list:
            formatted_staff.append({
                'id': staff[0],
                'name': f"{staff[1]} {staff[2]}",
                'position': staff[3],
                'department': staff[4]
            })
        
        # Get statistics
        cur.execute("SELECT COUNT(*) FROM shift_swap_requests WHERE status = 'Pending'")
        pending_count = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM shift_swap_requests WHERE status = 'Approved' AND approved_at >= NOW() - INTERVAL '30 days'")
        approved_30days = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error loading shift swap data: {e}")
        pending_swaps = []
        swap_history = []
        formatted_schedules = []
        formatted_staff = []
        pending_count = 0
        approved_30days = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "shift_swap.html",
        pending_swaps=pending_swaps,
        swap_history=swap_history,
        upcoming_schedules=formatted_schedules,
        staff_list=formatted_staff,
        pending_count=pending_count,
        approved_30days=approved_30days,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )
    
    
@app.route("/hr/scheduling/reports")
def schedule_reports():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get report parameters
    report_type = request.args.get("report_type", "monthly")
    month = request.args.get("month", date.today().month)
    year = request.args.get("year", date.today().year)
    department_id = request.args.get("department_id", "")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    report_data = []
    total_shifts = 0
    total_hours = 0
    unique_staff = 0
    
    try:
        # Build report query based on report type
        if report_type == "monthly":
            start_date = date(int(year), int(month), 1)
            if int(month) == 12:
                end_date = date(int(year) + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(int(year), int(month) + 1, 1) - timedelta(days=1)
            
            query = """
                SELECT 
                    COALESCE(d.name, 'No Department') as department,
                    st.first_name || ' ' || st.last_name as staff_name,
                    COUNT(*) as total_shifts,
                    COALESCE(SUM(EXTRACT(EPOCH FROM (end_time - start_time))/3600), 0) as total_hours,
                    COUNT(DISTINCT sch.schedule_date) as days_scheduled,
                    st.position,
                    st.id as staff_id
                FROM schedules sch
                JOIN staff st ON sch.staff_id = st.id
                LEFT JOIN departments d ON st.department_id = d.id
                WHERE sch.schedule_date BETWEEN %s AND %s
            """
            params = [start_date, end_date]
            
            if department_id:
                query += " AND st.department_id = %s"
                params.append(department_id)
            
            query += """
                GROUP BY d.name, st.first_name, st.last_name, st.position, st.id
                ORDER BY d.name, staff_name
            """
            
            cur.execute(query, params)
            rows = cur.fetchall()
            
            # Format the report data
            for row in rows:
                report_data.append({
                    'department': row[0],
                    'staff_name': row[1],
                    'total_shifts': row[2],
                    'total_hours': float(row[3]) if row[3] else 0,
                    'days_scheduled': row[4],
                    'position': row[5],
                    'staff_id': row[6]
                })
                total_shifts += row[2]
                total_hours += float(row[3]) if row[3] else 0
            
            unique_staff = len(report_data)
            
        elif report_type == "weekly":
            # Get current week
            today = date.today()
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            
            # Allow custom week if specified
            if request.args.get("week_start"):
                try:
                    start_date = datetime.strptime(request.args.get("week_start"), "%Y-%m-%d").date()
                    end_date = start_date + timedelta(days=6)
                except:
                    pass
            
            query = """
                SELECT 
                    sch.schedule_date,
                    COALESCE(d.name, 'No Department') as department,
                    st.first_name || ' ' || st.last_name as staff_name,
                    sch.shift_type,
                    sch.start_time,
                    sch.end_time,
                    EXTRACT(EPOCH FROM (sch.end_time - sch.start_time))/3600 as hours,
                    st.position,
                    sch.id as schedule_id
                FROM schedules sch
                JOIN staff st ON sch.staff_id = st.id
                LEFT JOIN departments d ON st.department_id = d.id
                WHERE sch.schedule_date BETWEEN %s AND %s
            """
            params = [start_date, end_date]
            
            if department_id:
                query += " AND st.department_id = %s"
                params.append(department_id)
            
            query += " ORDER BY sch.schedule_date, d.name, sch.start_time"
            
            cur.execute(query, params)
            rows = cur.fetchall()
            
            # Group by date for the report
            daily_data = {}
            for row in rows:
                date_str = row[0].strftime("%Y-%m-%d")
                if date_str not in daily_data:
                    daily_data[date_str] = {
                        'date': row[0],
                        'shifts': [],
                        'total_hours': 0,
                        'staff_count': 0
                    }
                
                hours = float(row[6]) if row[6] else 0
                daily_data[date_str]['shifts'].append({
                    'department': row[1],
                    'staff_name': row[2],
                    'shift_type': row[3],
                    'start_time': row[4],
                    'end_time': row[5],
                    'hours': hours,
                    'position': row[7],
                    'schedule_id': row[8]
                })
                daily_data[date_str]['total_hours'] += hours
                daily_data[date_str]['staff_count'] += 1
                
                total_hours += hours
                total_shifts += 1
            
            report_data = list(daily_data.values())
            unique_staff = len(set([s['staff_name'] for day in report_data for s in day['shifts']]))
            
        elif report_type == "coverage":
            start_date = date(int(year), int(month), 1)
            if int(month) == 12:
                end_date = date(int(year) + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(int(year), int(month) + 1, 1) - timedelta(days=1)
            
            query = """
                SELECT 
                    sch.schedule_date,
                    COUNT(DISTINCT sch.staff_id) as staff_count,
                    COUNT(*) as total_shifts,
                    STRING_AGG(DISTINCT st.first_name || ' ' || st.last_name, ', ') as staff_names,
                    SUM(EXTRACT(EPOCH FROM (sch.end_time - sch.start_time))/3600) as daily_hours,
                    COUNT(DISTINCT d.id) as departments_covered
                FROM schedules sch
                JOIN staff st ON sch.staff_id = st.id
                LEFT JOIN departments d ON st.department_id = d.id
                WHERE sch.schedule_date BETWEEN %s AND %s
            """
            params = [start_date, end_date]
            
            if department_id:
                query += " AND st.department_id = %s"
                params.append(department_id)
            
            query += """
                GROUP BY sch.schedule_date
                ORDER BY sch.schedule_date
            """
            
            cur.execute(query, params)
            rows = cur.fetchall()
            
            for row in rows:
                report_data.append({
                    'date': row[0],
                    'staff_count': row[1],
                    'total_shifts': row[2],
                    'staff_names': row[3][:100] + '...' if row[3] and len(row[3]) > 100 else row[3],
                    'daily_hours': float(row[4]) if row[4] else 0,
                    'departments_covered': row[5]
                })
                total_shifts += row[2]
                total_hours += float(row[4]) if row[4] else 0
            
            unique_staff = len(set([name for row in report_data for name in (row['staff_names'] or '').split(', ') if name]))
        
        # Get departments for filter dropdown
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get months and years for filter
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        
        # Calculate summary statistics
        avg_hours_per_shift = total_hours / total_shifts if total_shifts > 0 else 0
        avg_staff_per_day = total_shifts / len(report_data) if report_data and report_type == "coverage" else 0
        
    except Exception as e:
        app.logger.error(f"Error generating schedule report: {e}")
        report_data = []
        departments = []
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        total_shifts = 0
        total_hours = 0
        unique_staff = 0
        avg_hours_per_shift = 0
        avg_staff_per_day = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "schedule_reports.html",
        report_data=report_data,
        report_type=report_type,
        departments=departments,
        months=months,
        years=years,
        selected_month=int(month),
        selected_year=int(year),
        selected_department=department_id,
        total_shifts=total_shifts,
        total_hours=total_hours,
        unique_staff=unique_staff,
        avg_hours_per_shift=avg_hours_per_shift,
        avg_staff_per_day=avg_staff_per_day,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )

# Route to delete schedule
@app.route("/hr/scheduling/delete/<int:schedule_id>", methods=["POST"])
def delete_schedule(schedule_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("DELETE FROM schedules WHERE id = %s", (schedule_id,))
        conn.commit()
        return jsonify({"success": True, "message": "Schedule deleted successfully"})
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error deleting schedule: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()
        
# ==================== ROUTES: SCHEDULING MODULE ====================

# @app.route("/hr/scheduling")
# def scheduling():
#     if "hr_user_id" not in session:
#         return redirect(url_for("hr_login"))
    
#     conn = get_db_connection()
#     cur = conn.cursor()
    
#     try:
#         # Get current month schedules
#         current_month = date.today().replace(day=1)
#         if current_month.month == 12:
#             next_month = current_month.replace(year=current_month.year + 1, month=1)
#         else:
#             next_month = current_month.replace(month=current_month.month + 1)
        
#         cur.execute("""
#             SELECT s.*, st.first_name, st.last_name, st.position, d.name as department_name
#             FROM schedules s
#             JOIN staff st ON s.staff_id = st.id
#             LEFT JOIN departments d ON st.department_id = d.id
#             WHERE s.schedule_date >= %s AND s.schedule_date < %s
#             ORDER BY s.schedule_date, s.start_time
#         """, (current_month, next_month))
        
#         schedules = cur.fetchall()
        
#         # Get statistics
#         cur.execute("SELECT COUNT(*) FROM schedules WHERE schedule_date >= CURRENT_DATE")
#         upcoming_shifts = cur.fetchone()[0] or 0
        
#         cur.execute("""
#             SELECT COUNT(DISTINCT staff_id) 
#             FROM schedules 
#             WHERE schedule_date >= CURRENT_DATE
#         """)
#         staff_scheduled = cur.fetchone()[0] or 0
        
#         # Get departments for filter
#         cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
#         departments = cur.fetchall()
        
#         # Get staff for filter
#         cur.execute("""
#             SELECT id, first_name, last_name, position 
#             FROM staff 
#             WHERE status = 'Active' 
#             ORDER BY first_name, last_name
#         """)
#         staff_list = cur.fetchall()
        
#     except Exception as e:
#         app.logger.error(f"Error fetching scheduling data: {e}")
#         schedules = []
#         upcoming_shifts = 0
#         staff_scheduled = 0
#         departments = []
#         staff_list = []
    
#     finally:
#         cur.close()
#         conn.close()
    
#     return render_template("scheduling_dashboard.html",
#                          schedules=schedules,
#                          upcoming_shifts=upcoming_shifts,
#                          staff_scheduled=staff_scheduled,
#                          departments=departments,
#                          staff_list=staff_list,
#                          current_month=current_month.strftime("%B %Y"),
#                          hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State")









@app.route("/hr/scheduling/check-availability", methods=["POST"])
def check_availability():
    """AJAX endpoint to check staff availability"""
    if "hr_user_id" not in session:
        return jsonify({"available": False, "message": "Unauthorized"}), 401
    
    data = request.json
    staff_id = data.get("staff_id")
    schedule_date = data.get("schedule_date")
    start_time = data.get("start_time")
    end_time = data.get("end_time")
    
    if not all([staff_id, schedule_date, start_time, end_time]):
        return jsonify({"available": False, "message": "Missing parameters"}), 400
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT id FROM schedules 
            WHERE staff_id = %s 
            AND schedule_date = %s
            AND NOT (%s <= start_time OR %s >= end_time)
        """, (staff_id, schedule_date, end_time, start_time))
        
        conflict = cur.fetchone()
        available = conflict is None
        
        return jsonify({
            "available": available,
            "message": "Available" if available else "Staff has a conflicting schedule"
        })
        
    except Exception as e:
        app.logger.error(f"Error checking availability: {e}")
        return jsonify({"available": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()
        
# ==================== ROUTES: LEAVE MANAGEMENT MODULE ====================

@app.route("/hr/leave-management")
def leave_management():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get all leave requests with staff details
        cur.execute("""
            SELECT 
                l.id,
                l.leave_type,
                l.start_date,
                l.end_date,
                l.days_requested,
                l.reason,
                l.status,
                l.created_at,
                l.approved_at,
                l.approved_by,
                s.id as staff_id,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department,
                s.staff_id as employee_id,
                hu.username as approved_by_name
            FROM leaves l
            JOIN staff s ON l.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN hr_users hu ON l.approved_by = hu.id
            ORDER BY 
                CASE 
                    WHEN l.status = 'Pending' THEN 1
                    WHEN l.status = 'Approved' THEN 2
                    ELSE 3
                END,
                l.start_date DESC
        """)
        
        leaves = cur.fetchall()
        
        # Separate pending and history leaves
        pending_leaves = []
        approved_leaves = []
        rejected_leaves = []
        
        for leave in leaves:
            leave_dict = {
                'id': leave[0],
                'leave_type': leave[1],
                'start_date': leave[2],
                'end_date': leave[3],
                'days_requested': leave[4],
                'reason': leave[5],
                'status': leave[6],
                'created_at': leave[7],
                'approved_at': leave[8],
                'approved_by': leave[9],
                'staff_id': leave[10],
                'first_name': leave[11],
                'last_name': leave[12],
                'staff_name': f"{leave[11]} {leave[12]}",
                'position': leave[13],
                'department': leave[14] or 'N/A',
                'employee_id': leave[15],
                'approved_by_name': leave[16]
            }
            
            if leave_dict['status'] == 'Pending':
                pending_leaves.append(leave_dict)
            elif leave_dict['status'] == 'Approved':
                approved_leaves.append(leave_dict)
            elif leave_dict['status'] == 'Rejected':
                rejected_leaves.append(leave_dict)
        
        # Get leave statistics
        cur.execute("""
            SELECT 
                COUNT(*) as total_requests,
                SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending_count,
                SUM(CASE WHEN status = 'Approved' THEN 1 ELSE 0 END) as approved_count,
                SUM(CASE WHEN status = 'Rejected' THEN 1 ELSE 0 END) as rejected_count
            FROM leaves
        """)
        stats = cur.fetchone()
        
        # Get leave balance for active staff
        cur.execute("""
            SELECT 
                s.id,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department,
                COALESCE(SUM(CASE 
                    WHEN l.status = 'Approved' AND EXTRACT(YEAR FROM l.start_date) = EXTRACT(YEAR FROM CURRENT_DATE)
                    THEN l.days_requested ELSE 0 END), 0) as days_taken,
                20 - COALESCE(SUM(CASE 
                    WHEN l.status = 'Approved' AND EXTRACT(YEAR FROM l.start_date) = EXTRACT(YEAR FROM CURRENT_DATE)
                    THEN l.days_requested ELSE 0 END), 0) as days_remaining
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN leaves l ON s.id = l.staff_id AND l.status = 'Approved'
            WHERE s.status = 'Active'
            GROUP BY s.id, s.first_name, s.last_name, s.position, d.name
            ORDER BY days_remaining ASC
            LIMIT 10
        """)
        
        leave_balances = cur.fetchall()
        
        # Get leave types distribution
        cur.execute("""
            SELECT 
                leave_type,
                COUNT(*) as count,
                SUM(days_requested) as total_days
            FROM leaves
            WHERE status = 'Approved' 
                AND EXTRACT(YEAR FROM start_date) = EXTRACT(YEAR FROM CURRENT_DATE)
            GROUP BY leave_type
            ORDER BY count DESC
        """)
        
        leave_types = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error fetching leave data: {e}")
        pending_leaves = []
        approved_leaves = []
        rejected_leaves = []
        stats = (0, 0, 0, 0)
        leave_balances = []
        leave_types = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "leave_management.html",
        pending_leaves=pending_leaves,
        approved_leaves=approved_leaves,
        rejected_leaves=rejected_leaves,
        total_requests=stats[0],
        pending_count=stats[1],
        approved_count=stats[2],
        rejected_count=stats[3],
        leave_balances=leave_balances,
        leave_types=leave_types,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )


@app.route("/hr/leave/request", methods=["GET", "POST"])
def request_leave():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        staff_id = request.form.get("staff_id")
        leave_type = request.form.get("leave_type")
        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        reason = request.form.get("reason")
        
        # Calculate days requested
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
        days_requested = (end - start).days + 1
        
        if days_requested <= 0:
            flash("End date must be after start date", "danger")
            return redirect(url_for("request_leave"))
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Check for overlapping leaves
            cur.execute("""
                SELECT id FROM leaves 
                WHERE staff_id = %s 
                AND status IN ('Pending', 'Approved')
                AND NOT (end_date < %s OR start_date > %s)
            """, (staff_id, start_date, end_date))
            
            if cur.fetchone():
                flash("Staff already has a leave request for this period", "warning")
                return redirect(url_for("request_leave"))
            
            # Insert leave request
            cur.execute("""
                INSERT INTO leaves (
                    staff_id, leave_type, start_date, end_date, 
                    days_requested, reason, status, created_at
                ) VALUES (%s, %s, %s, %s, %s, %s, 'Pending', NOW())
                RETURNING id
            """, (staff_id, leave_type, start_date, end_date, days_requested, reason))
            
            leave_id = cur.fetchone()[0]
            conn.commit()
            
            flash(f"Leave request #{leave_id} submitted successfully!", "success")
            return redirect(url_for("leave_management"))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating leave request: {e}")
            flash(f"Error creating leave request: {str(e)}", "danger")
        
        finally:
            cur.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get active staff for dropdown
        cur.execute("""
            SELECT s.id, s.first_name, s.last_name, s.position, d.name as department
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.status = 'Active'
            ORDER BY s.first_name, s.last_name
        """)
        staff_list = cur.fetchall()
        
        # Get leave types
        leave_types = [
            'Annual Leave',
            'Sick Leave',
            'Maternity Leave',
            'Paternity Leave',
            'Bereavement Leave',
            'Study Leave',
            'Unpaid Leave',
            'Compensatory Off',
            'Emergency Leave'
        ]
        
    except Exception as e:
        app.logger.error(f"Error loading leave form data: {e}")
        staff_list = []
        leave_types = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "request_leave.html",
        staff_list=staff_list,
        leave_types=leave_types,
        today=date.today().strftime("%Y-%m-%d"),
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    )


@app.route("/hr/leave/approve/<int:leave_id>", methods=["POST"])
def approve_leave(leave_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            UPDATE leaves 
            SET status = 'Approved', 
                approved_by = %s, 
                approved_at = NOW() 
            WHERE id = %s AND status = 'Pending'
            RETURNING id
        """, (session["hr_user_id"], leave_id))
        
        if cur.fetchone():
            conn.commit()
            return jsonify({"success": True, "message": "Leave approved successfully"})
        else:
            return jsonify({"success": False, "message": "Leave request not found or already processed"}), 404
            
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error approving leave: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()


@app.route("/hr/leave/reject/<int:leave_id>", methods=["POST"])
def reject_leave(leave_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    rejection_reason = data.get("rejection_reason", "")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            UPDATE leaves 
            SET status = 'Rejected', 
                approved_by = %s, 
                approved_at = NOW(),
                reason = CONCAT(reason, ' [Rejected: ', %s, ']')
            WHERE id = %s AND status = 'Pending'
            RETURNING id
        """, (session["hr_user_id"], rejection_reason, leave_id))
        
        if cur.fetchone():
            conn.commit()
            return jsonify({"success": True, "message": "Leave rejected"})
        else:
            return jsonify({"success": False, "message": "Leave request not found or already processed"}), 404
            
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error rejecting leave: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()


@app.route("/hr/leave/balance/<int:staff_id>")
def get_leave_balance(staff_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Calculate leave balance (assuming 20 days annual leave)
        cur.execute("""
            SELECT 
                COALESCE(SUM(CASE 
                    WHEN status = 'Approved' AND EXTRACT(YEAR FROM start_date) = EXTRACT(YEAR FROM CURRENT_DATE)
                    THEN days_requested ELSE 0 END), 0) as days_taken
            FROM leaves
            WHERE staff_id = %s
        """, (staff_id,))
        
        days_taken = cur.fetchone()[0] or 0
        days_remaining = 20 - days_taken
        
        # Get upcoming leaves
        cur.execute("""
            SELECT leave_type, start_date, end_date, days_requested, status
            FROM leaves
            WHERE staff_id = %s AND start_date >= CURRENT_DATE
            ORDER BY start_date
            LIMIT 5
        """, (staff_id,))
        
        upcoming = cur.fetchall()
        
        return jsonify({
            "success": True,
            "days_taken": days_taken,
            "days_remaining": max(0, days_remaining),
            "upcoming": [{
                "type": u[0],
                "start": u[1].strftime("%Y-%m-%d"),
                "end": u[2].strftime("%Y-%m-%d"),
                "days": u[3],
                "status": u[4]
            } for u in upcoming]
        })
        
    except Exception as e:
        app.logger.error(f"Error fetching leave balance: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()


@app.route("/hr/leave/export")
def export_leave_report():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get all leave data for export
        cur.execute("""
            SELECT 
                l.id,
                s.staff_id as employee_id,
                s.first_name || ' ' || s.last_name as staff_name,
                d.name as department,
                s.position,
                l.leave_type,
                l.start_date,
                l.end_date,
                l.days_requested,
                l.reason,
                l.status,
                l.created_at,
                l.approved_at,
                hu.username as approved_by
            FROM leaves l
            JOIN staff s ON l.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN hr_users hu ON l.approved_by = hu.id
            ORDER BY l.created_at DESC
        """)
        
        leaves = cur.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leave Report"
        
        # Add headers
        headers = [
            "Leave ID", "Employee ID", "Staff Name", "Department", "Position",
            "Leave Type", "Start Date", "End Date", "Days Requested",
            "Reason", "Status", "Requested Date", "Approved Date", "Approved By"
        ]
        ws.append(headers)
        
        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for leave in leaves:
            ws.append([
                leave[0], leave[1], leave[2], leave[3], leave[4],
                leave[5], leave[6], leave[7], leave[8],
                leave[9], leave[10], leave[11], leave[12] or '', leave[13] or ''
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        # Generate filename
        filename = f"leave_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error exporting leave report: {e}")
        flash(f"Error exporting report: {str(e)}", "danger")
        return redirect(url_for("leave_management"))
        
    finally:
        cur.close()
        conn.close()


@app.route("/hr/leave/calendar")
def leave_calendar():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    month = request.args.get("month", date.today().month, type=int)
    year = request.args.get("year", date.today().year, type=int)
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT 
                l.start_date,
                l.end_date,
                l.leave_type,
                l.status,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department
            FROM leaves l
            JOIN staff s ON l.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE l.status IN ('Approved', 'Pending')
                AND (l.start_date BETWEEN %s AND %s OR l.end_date BETWEEN %s AND %s)
            ORDER BY l.start_date
        """, (start_date, end_date, start_date, end_date))
        
        leaves = cur.fetchall()
        
        # Group leaves by date for calendar
        calendar_data = {}
        for leave in leaves:
            current = max(leave[0], start_date)
            end = min(leave[1], end_date)
            
            while current <= end:
                date_str = current.strftime("%Y-%m-%d")
                if date_str not in calendar_data:
                    calendar_data[date_str] = []
                
                calendar_data[date_str].append({
                    'type': leave[2],
                    'status': leave[3],
                    'staff_name': f"{leave[4]} {leave[5]}",
                    'position': leave[6],
                    'department': leave[7] or 'N/A'
                })
                current += timedelta(days=1)
        
        # Get months for navigation
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        
    except Exception as e:
        app.logger.error(f"Error generating leave calendar: {e}")
        calendar_data = {}
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "leave_calendar.html",
        calendar_data=calendar_data,
        month=month,
        year=year,
        months=months,
        years=years,
        start_date=start_date,
        end_date=end_date,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )

 # ==================== ROUTES: ATTENDANCE RECORD MODULE ====================

@app.route("/hr/attendance-record")
def attendance_record():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get filter parameters
    department_id = request.args.get("department_id", "")
    staff_id = request.args.get("staff_id", "")
    date_filter = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    status_filter = request.args.get("status", "")
    
    # Parse date
    try:
        selected_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
    except:
        selected_date = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get attendance records for selected date with filters
        query = """
            SELECT 
                a.id,
                a.staff_id,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department,
                a.date,
                a.check_in,
                a.check_out,
                a.status,
                a.remarks,
                a.created_at,
                s.staff_id as employee_id
            FROM attendance a
            JOIN staff s ON a.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE a.date = %s
        """
        params = [selected_date]
        
        if department_id:
            query += " AND s.department_id = %s"
            params.append(department_id)
        
        if staff_id:
            query += " AND a.staff_id = %s"
            params.append(staff_id)
        
        if status_filter:
            query += " AND a.status = %s"
            params.append(status_filter)
        
        query += " ORDER BY s.first_name, s.last_name"
        
        cur.execute(query, params)
        attendance_records = cur.fetchall()
        
        # Get all active staff for attendance marking
        cur.execute("""
            SELECT s.id, s.first_name, s.last_name, s.position, d.name as department
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.status = 'Active'
            ORDER BY s.first_name, s.last_name
        """)
        all_staff = cur.fetchall()
        
        # Get departments for filter
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get staff for filter
        cur.execute("""
            SELECT id, first_name, last_name 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        # Calculate statistics for selected date
        cur.execute("""
            SELECT 
                COUNT(*) as total_present,
                SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) as present,
                SUM(CASE WHEN status = 'Late' THEN 1 ELSE 0 END) as late,
                SUM(CASE WHEN status = 'Absent' THEN 1 ELSE 0 END) as absent,
                SUM(CASE WHEN status = 'Half Day' THEN 1 ELSE 0 END) as half_day,
                SUM(CASE WHEN status = 'Holiday' THEN 1 ELSE 0 END) as holiday
            FROM attendance
            WHERE date = %s
        """, (selected_date,))
        
        stats = cur.fetchone()
        
        # Get total active staff count
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active'")
        total_active = cur.fetchone()[0] or 0
        
        # Get recent attendance history (last 7 days)
        week_ago = selected_date - timedelta(days=7)
        cur.execute("""
            SELECT 
                date,
                COUNT(*) as total,
                SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) as present,
                SUM(CASE WHEN status = 'Late' THEN 1 ELSE 0 END) as late,
                SUM(CASE WHEN status = 'Absent' THEN 1 ELSE 0 END) as absent
            FROM attendance
            WHERE date BETWEEN %s AND %s
            GROUP BY date
            ORDER BY date DESC
        """, (week_ago, selected_date))
        
        history = cur.fetchall()
        
        # Get staff without attendance for today
        if selected_date == date.today():
            marked_staff_ids = [r[1] for r in attendance_records]
            unmarked_staff = [s for s in all_staff if s[0] not in marked_staff_ids]
        else:
            unmarked_staff = []
        
    except Exception as e:
        app.logger.error(f"Error fetching attendance data: {e}")
        attendance_records = []
        all_staff = []
        departments = []
        staff_list = []
        stats = (0, 0, 0, 0, 0, 0)
        total_active = 0
        history = []
        unmarked_staff = []
    
    finally:
        cur.close()
        conn.close()
    
    # Format attendance records for template
    formatted_records = []
    for record in attendance_records:
        check_in = record[7]
        check_out = record[8]
        
        # Calculate hours worked if both check-in and check-out exist
        hours_worked = None
        if check_in and check_out:
            # Convert time to datetime for calculation
            check_in_dt = datetime.combine(selected_date, check_in)
            check_out_dt = datetime.combine(selected_date, check_out)
            if check_out_dt < check_in_dt:  # Overnight shift
                check_out_dt += timedelta(days=1)
            hours_worked = (check_out_dt - check_in_dt).total_seconds() / 3600
        
        formatted_records.append({
            'id': record[0],
            'staff_id': record[1],
            'first_name': record[2],
            'last_name': record[3],
            'staff_name': f"{record[2]} {record[3]}",
            'position': record[4],
            'department': record[5] or 'N/A',
            'date': record[6],
            'check_in': check_in.strftime("%H:%M") if check_in else None,
            'check_out': check_out.strftime("%H:%M") if check_out else None,
            'status': record[9],
            'remarks': record[10],
            'created_at': record[11],
            'employee_id': record[12],
            'hours_worked': round(hours_worked, 1) if hours_worked else None
        })
    
    return render_template(
        "attendance_record.html",
        attendance_records=formatted_records,
        all_staff=all_staff,
        departments=departments,
        staff_list=staff_list,
        selected_date=selected_date,
        selected_department=department_id,
        selected_staff=staff_id,
        selected_status=status_filter,
        total_present=stats[1] or 0,
        total_late=stats[2] or 0,
        total_absent=stats[3] or 0,
        total_half_day=stats[4] or 0,
        total_holiday=stats[5] or 0,
        total_marked=(stats[1] or 0) + (stats[2] or 0) + (stats[3] or 0) + (stats[4] or 0) + (stats[5] or 0),
        total_active=total_active,
        history=history,
        unmarked_staff=unmarked_staff,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )


@app.route("/hr/attendance/mark", methods=["POST"])
def mark_attendance():
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    staff_id = data.get("staff_id")
    attendance_date = data.get("date")
    check_in = data.get("check_in")
    check_out = data.get("check_out")
    status = data.get("status")
    remarks = data.get("remarks", "")
    
    if not all([staff_id, attendance_date, status]):
        return jsonify({"success": False, "message": "Missing required fields"}), 400
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Check if attendance already exists for this staff on this date
        cur.execute("""
            SELECT id FROM attendance 
            WHERE staff_id = %s AND date = %s
        """, (staff_id, attendance_date))
        
        existing = cur.fetchone()
        
        if existing:
            # Update existing attendance
            cur.execute("""
                UPDATE attendance 
                SET check_in = %s,
                    check_out = %s,
                    status = %s,
                    remarks = %s,
                    recorded_by = %s
                WHERE id = %s
                RETURNING id
            """, (check_in, check_out, status, remarks, session["hr_user_id"], existing[0]))
        else:
            # Insert new attendance
            cur.execute("""
                INSERT INTO attendance (
                    staff_id, date, check_in, check_out, status, remarks, recorded_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (staff_id, attendance_date, check_in, check_out, status, remarks, session["hr_user_id"]))
        
        attendance_id = cur.fetchone()[0]
        conn.commit()
        
        return jsonify({
            "success": True, 
            "message": "Attendance marked successfully",
            "attendance_id": attendance_id
        })
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error marking attendance: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()


@app.route("/hr/attendance/bulk-mark", methods=["POST"])
def bulk_mark_attendance():
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    attendance_date = data.get("date")
    status = data.get("status")
    staff_ids = data.get("staff_ids", [])
    
    if not all([attendance_date, status]) or not staff_ids:
        return jsonify({"success": False, "message": "Missing required fields"}), 400
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        marked_count = 0
        for staff_id in staff_ids:
            # Check if attendance already exists
            cur.execute("""
                SELECT id FROM attendance 
                WHERE staff_id = %s AND date = %s
            """, (staff_id, attendance_date))
            
            existing = cur.fetchone()
            
            if existing:
                # Update existing
                cur.execute("""
                    UPDATE attendance 
                    SET status = %s,
                        recorded_by = %s
                    WHERE id = %s
                """, (status, session["hr_user_id"], existing[0]))
            else:
                # Insert new
                cur.execute("""
                    INSERT INTO attendance (
                        staff_id, date, status, recorded_by
                    ) VALUES (%s, %s, %s, %s)
                """, (staff_id, attendance_date, status, session["hr_user_id"]))
            
            marked_count += 1
        
        conn.commit()
        
        return jsonify({
            "success": True, 
            "message": f"Bulk attendance marked for {marked_count} staff members"
        })
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error bulk marking attendance: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()


@app.route("/hr/attendance/report")
def attendance_report():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get report parameters
    report_type = request.args.get("type", "daily")
    month = request.args.get("month", date.today().month, type=int)
    year = request.args.get("year", date.today().year, type=int)
    department_id = request.args.get("department_id", "")
    staff_id = request.args.get("staff_id", "")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        if report_type == "daily":
            # Daily report for selected date
            report_date = request.args.get("date", date.today().strftime("%Y-%m-%d"))
            start_date = end_date = datetime.strptime(report_date, "%Y-%m-%d").date()
            
        elif report_type == "weekly":
            # Weekly report
            week_start = request.args.get("week_start")
            if week_start:
                start_date = datetime.strptime(week_start, "%Y-%m-%d").date()
            else:
                today = date.today()
                start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            
        else:  # monthly
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # Build query
        query = """
            SELECT 
                a.date,
                s.first_name || ' ' || s.last_name as staff_name,
                s.position,
                d.name as department,
                a.check_in,
                a.check_out,
                a.status,
                a.remarks,
                EXTRACT(EPOCH FROM (a.check_out - a.check_in))/3600 as hours_worked
            FROM attendance a
            JOIN staff s ON a.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE a.date BETWEEN %s AND %s
        """
        params = [start_date, end_date]
        
        if department_id:
            query += " AND s.department_id = %s"
            params.append(department_id)
        
        if staff_id:
            query += " AND a.staff_id = %s"
            params.append(staff_id)
        
        query += " ORDER BY a.date DESC, d.name, s.first_name"
        
        cur.execute(query, params)
        report_data = cur.fetchall()
        
        # Calculate summary statistics
        cur.execute("""
            SELECT 
                COUNT(*) as total_records,
                SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) as present,
                SUM(CASE WHEN status = 'Late' THEN 1 ELSE 0 END) as late,
                SUM(CASE WHEN status = 'Absent' THEN 1 ELSE 0 END) as absent,
                SUM(CASE WHEN status = 'Half Day' THEN 1 ELSE 0 END) as half_day,
                SUM(CASE WHEN status = 'Holiday' THEN 1 ELSE 0 END) as holiday
            FROM attendance
            WHERE date BETWEEN %s AND %s
        """, (start_date, end_date))
        
        summary = cur.fetchone()
        
        # Get departments for filter
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        # Get staff for filter
        cur.execute("""
            SELECT id, first_name, last_name 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        # Get months for dropdown
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        
    except Exception as e:
        app.logger.error(f"Error generating attendance report: {e}")
        report_data = []
        summary = (0, 0, 0, 0, 0, 0)
        departments = []
        staff_list = []
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        start_date = date.today()
        end_date = date.today()
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "attendance_report.html",
        report_data=report_data,
        summary=summary,
        report_type=report_type,
        departments=departments,
        staff_list=staff_list,
        months=months,
        years=years,
        start_date=start_date,
        end_date=end_date,
        selected_month=month,
        selected_year=year,
        selected_department=department_id,
        selected_staff=staff_id,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )


@app.route("/hr/attendance/export")
def export_attendance():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get filter parameters (same as attendance_record)
    department_id = request.args.get("department_id", "")
    staff_id = request.args.get("staff_id", "")
    start_date = request.args.get("start_date", (date.today() - timedelta(days=30)).strftime("%Y-%m-%d"))
    end_date = request.args.get("end_date", date.today().strftime("%Y-%m-%d"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Build query
        query = """
            SELECT 
                a.date,
                s.staff_id as employee_id,
                s.first_name || ' ' || s.last_name as staff_name,
                s.position,
                d.name as department,
                a.check_in,
                a.check_out,
                a.status,
                a.remarks,
                EXTRACT(EPOCH FROM (a.check_out - a.check_in))/3600 as hours_worked,
                a.created_at
            FROM attendance a
            JOIN staff s ON a.staff_id = s.id
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE a.date BETWEEN %s AND %s
        """
        params = [start_date, end_date]
        
        if department_id:
            query += " AND s.department_id = %s"
            params.append(department_id)
        
        if staff_id:
            query += " AND a.staff_id = %s"
            params.append(staff_id)
        
        query += " ORDER BY a.date DESC, d.name, s.first_name"
        
        cur.execute(query, params)
        attendance_data = cur.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Add headers
        headers = [
            "Date", "Employee ID", "Staff Name", "Position", "Department",
            "Check In", "Check Out", "Status", "Hours Worked", "Remarks", "Recorded At"
        ]
        ws.append(headers)
        
        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for record in attendance_data:
            ws.append([
                record[0].strftime("%Y-%m-%d"),
                record[1],
                record[2],
                record[3],
                record[4] or 'N/A',
                record[5].strftime("%H:%M") if record[5] else '',
                record[6].strftime("%H:%M") if record[6] else '',
                record[7],
                round(record[8], 1) if record[8] else '',
                record[9] or '',
                record[10].strftime("%Y-%m-%d %H:%M") if record[10] else ''
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        # Generate filename
        filename = f"attendance_report_{start_date}_to_{end_date}.xlsx"
        
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error exporting attendance: {e}")
        flash(f"Error exporting report: {str(e)}", "danger")
        return redirect(url_for("attendance_record"))
        
    finally:
        cur.close()
        conn.close()


@app.route("/hr/attendance/summary")
def attendance_summary():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    # Get parameters
    period = request.args.get("period", "month")
    month = request.args.get("month", date.today().month, type=int)
    year = request.args.get("year", date.today().year, type=int)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        if period == "month":
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
        else:  # year
            start_date = date(year, 1, 1)
            end_date = date(year, 12, 31)
        
        # Get attendance summary by staff
        cur.execute("""
            SELECT 
                s.id,
                s.first_name || ' ' || s.last_name as staff_name,
                s.position,
                d.name as department,
                COUNT(CASE WHEN a.status = 'Present' THEN 1 END) as present_days,
                COUNT(CASE WHEN a.status = 'Late' THEN 1 END) as late_days,
                COUNT(CASE WHEN a.status = 'Absent' THEN 1 END) as absent_days,
                COUNT(CASE WHEN a.status = 'Half Day' THEN 1 END) as half_days,
                COUNT(CASE WHEN a.status = 'Holiday' THEN 1 END) as holiday_days,
                COUNT(a.id) as total_days,
                ROUND(AVG(EXTRACT(EPOCH FROM (a.check_out - a.check_in))/3600), 1) as avg_hours
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            LEFT JOIN attendance a ON s.id = a.staff_id 
                AND a.date BETWEEN %s AND %s
            WHERE s.status = 'Active'
            GROUP BY s.id, s.first_name, s.last_name, s.position, d.name
            ORDER BY d.name, s.first_name
        """, (start_date, end_date))
        
        staff_summary = cur.fetchall()
        
        # Get daily summary
        cur.execute("""
            SELECT 
                a.date,
                COUNT(DISTINCT a.staff_id) as staff_present,
                COUNT(CASE WHEN a.status = 'Present' THEN 1 END) as present,
                COUNT(CASE WHEN a.status = 'Late' THEN 1 END) as late,
                COUNT(CASE WHEN a.status = 'Absent' THEN 1 END) as absent,
                COUNT(CASE WHEN a.status = 'Half Day' THEN 1 END) as half_day
            FROM attendance a
            WHERE a.date BETWEEN %s AND %s
            GROUP BY a.date
            ORDER BY a.date
        """, (start_date, end_date))
        
        daily_summary = cur.fetchall()
        
        # Get months for dropdown
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
        
    except Exception as e:
        app.logger.error(f"Error generating attendance summary: {e}")
        staff_summary = []
        daily_summary = []
        months = [(i, month_name[i]) for i in range(1, 13)]
        years = range(date.today().year - 2, date.today().year + 2)
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "attendance_summary.html",
        staff_summary=staff_summary,
        daily_summary=daily_summary,
        period=period,
        months=months,
        years=years,
        selected_month=month,
        selected_year=year,
        start_date=start_date,
        end_date=end_date,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )


@app.route("/hr/attendance/delete/<int:attendance_id>", methods=["POST"])
def delete_attendance(attendance_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("DELETE FROM attendance WHERE id = %s RETURNING id", (attendance_id,))
        deleted = cur.fetchone()
        conn.commit()
        
        if deleted:
            return jsonify({"success": True, "message": "Attendance record deleted"})
        else:
            return jsonify({"success": False, "message": "Record not found"}), 404
            
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error deleting attendance: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()  

# ==================== ROUTES: DEPARTMENT MANAGEMENT MODULE ====================

@app.route("/hr/departments")
def departments():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get all departments with detailed statistics
        cur.execute("""
            SELECT 
                d.id,
                d.name,
                d.code,
                d.description,
                d.head_of_dept,
                d.status,
                d.created_at,
                COUNT(DISTINCT s.id) as staff_count,
                COUNT(DISTINCT CASE WHEN s.status = 'Active' THEN s.id END) as active_staff,
                COUNT(DISTINCT sch.id) as total_shifts,
                COALESCE(SUM(CASE 
                    WHEN a.status IN ('Present', 'Late') THEN 1 
                    ELSE 0 
                END), 0) as attendance_count
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            LEFT JOIN schedules sch ON s.id = sch.staff_id 
                AND sch.schedule_date >= CURRENT_DATE
            LEFT JOIN attendance a ON s.id = a.staff_id 
                AND a.date = CURRENT_DATE
            GROUP BY d.id, d.name, d.code, d.description, d.head_of_dept, d.status, d.created_at
            ORDER BY d.name
        """)
        
        departments_list = cur.fetchall()
        
        # Get staff count by department for chart
        cur.execute("""
            SELECT 
                d.name,
                COUNT(s.id) as staff_count
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            WHERE d.status = 'Active'
            GROUP BY d.name
            ORDER BY staff_count DESC
        """)
        chart_data = cur.fetchall()
        
        # Get department heads list
        cur.execute("""
            SELECT DISTINCT head_of_dept 
            FROM departments 
            WHERE head_of_dept IS NOT NULL 
            ORDER BY head_of_dept
        """)
        dept_heads = cur.fetchall()
        
        # Statistics
        cur.execute("SELECT COUNT(*) FROM departments WHERE status = 'Active'")
        active_depts = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM staff WHERE department_id IS NOT NULL")
        staff_with_dept = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COALESCE(AVG(staff_count), 0)
            FROM (
                SELECT COUNT(*) as staff_count 
                FROM staff 
                WHERE department_id IS NOT NULL 
                GROUP BY department_id
            ) as dept_stats
        """)
        avg_staff_per_dept = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error fetching departments: {e}")
        departments_list = []
        chart_data = []
        dept_heads = []
        active_depts = 0
        staff_with_dept = 0
        avg_staff_per_dept = 0
    
    finally:
        cur.close()
        conn.close()
    
    # Format departments for template
    formatted_depts = []
    for dept in departments_list:
        formatted_depts.append({
            'id': dept[0],
            'name': dept[1],
            'code': dept[2],
            'description': dept[3] or 'No description',
            'head_of_dept': dept[4] or 'Not assigned',
            'status': dept[5],
            'created_at': dept[6],
            'staff_count': dept[7] or 0,
            'active_staff': dept[8] or 0,
            'total_shifts': dept[9] or 0,
            'attendance_count': dept[10] or 0,
            'attendance_rate': round((dept[10] / dept[8] * 100) if dept[8] > 0 else 0, 1)
        })
    
    return render_template(
        "departments.html",
        departments=formatted_depts,
        chart_data=chart_data,
        dept_heads=dept_heads,
        active_depts=active_depts,
        staff_with_dept=staff_with_dept,
        avg_staff_per_dept=round(avg_staff_per_dept, 1),
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )


@app.route("/hr/departments/add", methods=["GET", "POST"])
def add_department():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        name = request.form.get("name")
        code = request.form.get("code")
        description = request.form.get("description")
        head_of_dept = request.form.get("head_of_dept")
        status = request.form.get("status", "Active")
        
        # Validate required fields
        if not all([name, code]):
            flash("Department name and code are required", "danger")
            return redirect(url_for("add_department"))
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Check if code already exists
            cur.execute("SELECT id FROM departments WHERE code = %s", (code,))
            if cur.fetchone():
                flash(f"Department code '{code}' already exists", "danger")
                return redirect(url_for("add_department"))
            
            # Insert new department
            cur.execute("""
                INSERT INTO departments (name, code, description, head_of_dept, status)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
            """, (name, code.upper(), description, head_of_dept, status))
            
            dept_id = cur.fetchone()[0]
            conn.commit()
            
            flash(f"Department '{name}' added successfully!", "success")
            return redirect(url_for("view_department", department_id=dept_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error adding department: {e}")
            flash(f"Error adding department: {str(e)}", "danger")
            
        finally:
            cur.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get potential department heads (staff members)
        cur.execute("""
            SELECT id, first_name, last_name, position 
            FROM staff 
            WHERE status = 'Active'
            ORDER BY first_name, last_name
        """)
        potential_heads = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error loading form data: {e}")
        potential_heads = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "add_department.html",
        potential_heads=potential_heads,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
        )


@app.route("/hr/departments/<int:department_id>")
def view_department(department_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get department details
        cur.execute("""
            SELECT 
                d.id,
                d.name,
                d.code,
                d.description,
                d.head_of_dept,
                d.status,
                d.created_at,
                COUNT(DISTINCT s.id) as total_staff,
                COUNT(DISTINCT CASE WHEN s.status = 'Active' THEN s.id END) as active_staff,
                COUNT(DISTINCT sch.id) as upcoming_shifts,
                COUNT(DISTINCT l.id) as pending_leaves
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            LEFT JOIN schedules sch ON s.id = sch.staff_id 
                AND sch.schedule_date >= CURRENT_DATE
            LEFT JOIN leaves l ON s.id = l.staff_id 
                AND l.status = 'Pending'
            WHERE d.id = %s
            GROUP BY d.id, d.name, d.code, d.description, d.head_of_dept, d.status, d.created_at
        """, (department_id,))
        
        department = cur.fetchone()
        
        if not department:
            flash("Department not found", "danger")
            return redirect(url_for("departments"))
        
        # Get staff in this department
        cur.execute("""
            SELECT 
                s.id,
                s.staff_id,
                s.first_name,
                s.last_name,
                s.position,
                s.email,
                s.phone,
                s.status,
                s.hire_date,
                s.salary,
                COUNT(DISTINCT a.id) as attendance_count,
                COUNT(DISTINCT sch.id) as shift_count
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id 
                AND a.date >= CURRENT_DATE - INTERVAL '30 days'
            LEFT JOIN schedules sch ON s.id = sch.staff_id 
                AND sch.schedule_date >= CURRENT_DATE
            WHERE s.department_id = %s
            GROUP BY s.id, s.staff_id, s.first_name, s.last_name, s.position, 
                     s.email, s.phone, s.status, s.hire_date, s.salary
            ORDER BY s.first_name, s.last_name
        """, (department_id,))
        
        staff_list = cur.fetchall()
        
        # Get recent attendance for department (last 7 days)
        cur.execute("""
            SELECT 
                a.date,
                COUNT(CASE WHEN a.status = 'Present' THEN 1 END) as present,
                COUNT(CASE WHEN a.status = 'Late' THEN 1 END) as late,
                COUNT(CASE WHEN a.status = 'Absent' THEN 1 END) as absent,
                COUNT(CASE WHEN a.status = 'Half Day' THEN 1 END) as half_day
            FROM attendance a
            JOIN staff s ON a.staff_id = s.id
            WHERE s.department_id = %s
                AND a.date >= CURRENT_DATE - INTERVAL '7 days'
            GROUP BY a.date
            ORDER BY a.date DESC
        """, (department_id,))
        
        weekly_attendance = cur.fetchall()
        
        # Get upcoming schedules (next 7 days)
        cur.execute("""
            SELECT 
                sch.id,
                sch.schedule_date,
                sch.start_time,
                sch.end_time,
                sch.shift_type,
                s.first_name,
                s.last_name,
                s.position
            FROM schedules sch
            JOIN staff s ON sch.staff_id = s.id
            WHERE s.department_id = %s
                AND sch.schedule_date >= CURRENT_DATE
                AND sch.schedule_date <= CURRENT_DATE + INTERVAL '7 days'
            ORDER BY sch.schedule_date, sch.start_time
            LIMIT 10
        """, (department_id,))
        
        upcoming_schedules = cur.fetchall()
        
        # Get pending leave requests
        cur.execute("""
            SELECT 
                l.id,
                l.leave_type,
                l.start_date,
                l.end_date,
                l.days_requested,
                l.status,
                s.first_name,
                s.last_name
            FROM leaves l
            JOIN staff s ON l.staff_id = s.id
            WHERE s.department_id = %s
                AND l.status = 'Pending'
            ORDER BY l.start_date
            LIMIT 10
        """, (department_id,))
        
        leave_requests = cur.fetchall()
        
        # Get department statistics
        cur.execute("""
            SELECT 
                COALESCE(AVG(s.salary), 0) as avg_salary,
                COUNT(DISTINCT CASE WHEN s.gender = 'Male' THEN s.id END) as male_count,
                COUNT(DISTINCT CASE WHEN s.gender = 'Female' THEN s.id END) as female_count,
                MIN(s.hire_date) as oldest_employee
            FROM staff s
            WHERE s.department_id = %s
        """, (department_id,))
        
        dept_stats = cur.fetchone()
        
    except Exception as e:
        app.logger.error(f"Error fetching department details: {e}")
        flash(f"Error loading department details: {str(e)}", "danger")
        return redirect(url_for("departments"))
    
    finally:
        cur.close()
        conn.close()
    
    # Format department data
    dept_dict = {
        'id': department[0],
        'name': department[1],
        'code': department[2],
        'description': department[3] or 'No description provided',
        'head_of_dept': department[4] or 'Not assigned',
        'status': department[5],
        'created_at': department[6],
        'total_staff': department[7] or 0,
        'active_staff': department[8] or 0,
        'upcoming_shifts': department[9] or 0,
        'pending_leaves': department[10] or 0
    }
    
    # Format staff list
    formatted_staff = []
    for staff in staff_list:
        formatted_staff.append({
            'id': staff[0],
            'staff_id': staff[1],
            'first_name': staff[2],
            'last_name': staff[3],
            'name': f"{staff[2]} {staff[3]}",
            'position': staff[4] or 'N/A',
            'email': staff[5] or 'N/A',
            'phone': staff[6] or 'N/A',
            'status': staff[7],
            'hire_date': staff[8],
            'salary': float(staff[9]) if staff[9] else 0,
            'attendance_count': staff[10] or 0,
            'shift_count': staff[11] or 0
        })
    
    # Format weekly attendance
    formatted_attendance = []
    for att in weekly_attendance:
        formatted_attendance.append({
            'date': att[0],
            'present': att[1] or 0,
            'late': att[2] or 0,
            'absent': att[3] or 0,
            'half_day': att[4] or 0,
            'total': (att[1] or 0) + (att[2] or 0) + (att[3] or 0) + (att[4] or 0)
        })
    
    # Format upcoming schedules
    formatted_schedules = []
    for sch in upcoming_schedules:
        formatted_schedules.append({
            'id': sch[0],
            'date': sch[1],
            'start_time': sch[2],
            'end_time': sch[3],
            'shift_type': sch[4] or 'Regular',
            'staff_name': f"{sch[5]} {sch[6]}",
            'position': sch[7]
        })
    
    # Format leave requests
    formatted_leaves = []
    for leave in leave_requests:
        formatted_leaves.append({
            'id': leave[0],
            'type': leave[1],
            'start_date': leave[2],
            'end_date': leave[3],
            'days': leave[4],
            'status': leave[5],
            'staff_name': f"{leave[6]} {leave[7]}"
        })
    
    return render_template(
        "view_department.html",
        department=dept_dict,
        staff_list=formatted_staff,
        weekly_attendance=formatted_attendance,
        upcoming_schedules=formatted_schedules,
        leave_requests=formatted_leaves,
        avg_salary=float(dept_stats[0]) if dept_stats else 0,
        male_count=dept_stats[1] or 0 if dept_stats else 0,
        female_count=dept_stats[2] or 0 if dept_stats else 0,
        oldest_employee=dept_stats[3] if dept_stats else None,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )


@app.route("/hr/departments/edit/<int:department_id>", methods=["GET", "POST"])
def edit_department(department_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == "POST":
        name = request.form.get("name")
        code = request.form.get("code")
        description = request.form.get("description")
        head_of_dept = request.form.get("head_of_dept")
        status = request.form.get("status")
        
        if not all([name, code]):
            flash("Department name and code are required", "danger")
            return redirect(url_for("edit_department", department_id=department_id))
        
        try:
            # Check if code exists for other departments
            cur.execute("""
                SELECT id FROM departments 
                WHERE code = %s AND id != %s
            """, (code.upper(), department_id))
            
            if cur.fetchone():
                flash(f"Department code '{code}' already exists", "danger")
                return redirect(url_for("edit_department", department_id=department_id))
            
            # Update department
            cur.execute("""
                UPDATE departments 
                SET name = %s,
                    code = %s,
                    description = %s,
                    head_of_dept = %s,
                    status = %s
                WHERE id = %s
                RETURNING id
            """, (name, code.upper(), description, head_of_dept, status, department_id))
            
            updated = cur.fetchone()
            conn.commit()
            
            if updated:
                flash("Department updated successfully!", "success")
                return redirect(url_for("view_department", department_id=department_id))
            else:
                flash("Department not found", "danger")
                return redirect(url_for("departments"))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error updating department: {e}")
            flash(f"Error updating department: {str(e)}", "danger")
    
    # GET request - load department data
    try:
        cur.execute("SELECT * FROM departments WHERE id = %s", (department_id,))
        department = cur.fetchone()
        
        if not department:
            flash("Department not found", "danger")
            return redirect(url_for("departments"))
        
        # Get potential department heads (active staff)
        cur.execute("""
            SELECT 
                s.id,
                s.first_name,
                s.last_name,
                s.position,
                d.name as department_name
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.status = 'Active'
            ORDER BY s.first_name, s.last_name
        """)
        potential_heads = cur.fetchall()
        
        # Format potential heads for template
        formatted_heads = []
        for head in potential_heads:
            formatted_heads.append({
                'id': head[0],
                'name': f"{head[1]} {head[2]}",
                'position': head[3],
                'department': head[4] or 'No Department',
                'full_title': f"{head[1]} {head[2]} - {head[3]} ({head[4] or 'No Dept'})"
            })
        
    except Exception as e:
        app.logger.error(f"Error loading department for edit: {e}")
        flash("Error loading department details", "danger")
        return redirect(url_for("departments"))
    
    finally:
        cur.close()
        conn.close()
    
    # Format department data
    dept_dict = {
        'id': department[0],
        'name': department[1],
        'code': department[2],
        'description': department[3] or '',
        'head_of_dept': department[4] or '',
        'status': department[5],
        'created_at': department[6]
    }
    
    return render_template(
        "edit_department.html",
        department=dept_dict,
        potential_heads=formatted_heads,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )


@app.route("/hr/departments/delete/<int:department_id>", methods=["POST"])
def delete_department(department_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Check if department has staff
        cur.execute("SELECT COUNT(*) FROM staff WHERE department_id = %s", (department_id,))
        staff_count = cur.fetchone()[0]
        
        if staff_count > 0:
            return jsonify({
                "success": False, 
                "message": f"Cannot delete department with {staff_count} staff members. Reassign staff first."
            }), 400
        
        # Delete department
        cur.execute("DELETE FROM departments WHERE id = %s RETURNING id", (department_id,))
        deleted = cur.fetchone()
        conn.commit()
        
        if deleted:
            return jsonify({"success": True, "message": "Department deleted successfully"})
        else:
            return jsonify({"success": False, "message": "Department not found"}), 404
            
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error deleting department: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()


@app.route("/hr/departments/assign-head", methods=["POST"])
def assign_department_head():
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    department_id = data.get("department_id")
    staff_id = data.get("staff_id")
    
    if not all([department_id, staff_id]):
        return jsonify({"success": False, "message": "Missing required fields"}), 400
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get staff name
        cur.execute("""
            SELECT first_name, last_name, position 
            FROM staff 
            WHERE id = %s
        """, (staff_id,))
        
        staff = cur.fetchone()
        if not staff:
            return jsonify({"success": False, "message": "Staff not found"}), 404
        
        head_name = f"{staff[0]} {staff[1]} - {staff[2]}"
        
        # Update department head
        cur.execute("""
            UPDATE departments 
            SET head_of_dept = %s 
            WHERE id = %s
            RETURNING id
        """, (head_name, department_id))
        
        updated = cur.fetchone()
        conn.commit()
        
        if updated:
            return jsonify({
                "success": True, 
                "message": "Department head assigned successfully",
                "head_name": head_name
            })
        else:
            return jsonify({"success": False, "message": "Department not found"}), 404
            
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error assigning department head: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close()


@app.route("/hr/departments/analytics")
def department_analytics():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Staff distribution by department
        cur.execute("""
            SELECT 
                d.name,
                COUNT(s.id) as total_staff,
                SUM(CASE WHEN s.status = 'Active' THEN 1 ELSE 0 END) as active_staff,
                SUM(CASE WHEN s.employment_type = 'Full-Time' THEN 1 ELSE 0 END) as full_time,
                SUM(CASE WHEN s.employment_type = 'Part-Time' THEN 1 ELSE 0 END) as part_time,
                SUM(CASE WHEN s.employment_type = 'Contract' THEN 1 ELSE 0 END) as contract
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            WHERE d.status = 'Active'
            GROUP BY d.name
            ORDER BY total_staff DESC
        """)
        staff_distribution = cur.fetchall()
        
        # Attendance rates by department (last 30 days)
        cur.execute("""
            WITH dept_attendance AS (
                SELECT 
                    d.id,
                    d.name,
                    COUNT(a.id) as total_attendance,
                    SUM(CASE WHEN a.status IN ('Present', 'Late') THEN 1 ELSE 0 END) as present_count,
                    COUNT(DISTINCT a.date) as working_days
                FROM departments d
                JOIN staff s ON d.id = s.department_id
                LEFT JOIN attendance a ON s.id = a.staff_id 
                    AND a.date >= CURRENT_DATE - INTERVAL '30 days'
                WHERE d.status = 'Active'
                GROUP BY d.id, d.name
            )
            SELECT 
                name,
                total_attendance,
                present_count,
                working_days,
                CASE 
                    WHEN working_days > 0 
                    THEN ROUND((present_count::decimal / (working_days * COUNT_STAFF)) * 100, 1)
                    ELSE 0 
                END as attendance_rate
            FROM dept_attendance
            CROSS JOIN LATERAL (
                SELECT COUNT(*) as COUNT_STAFF 
                FROM staff 
                WHERE department_id = dept_attendance.id
            ) staff_count
            ORDER BY attendance_rate DESC
        """)
        attendance_rates = cur.fetchall()
        
        # Leave statistics by department
        cur.execute("""
            SELECT 
                d.name,
                COUNT(l.id) as total_leaves,
                SUM(CASE WHEN l.status = 'Approved' THEN 1 ELSE 0 END) as approved_leaves,
                SUM(CASE WHEN l.status = 'Pending' THEN 1 ELSE 0 END) as pending_leaves,
                SUM(l.days_requested) as total_days
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            LEFT JOIN leaves l ON s.id = l.staff_id
                AND EXTRACT(YEAR FROM l.start_date) = EXTRACT(YEAR FROM CURRENT_DATE)
            WHERE d.status = 'Active'
            GROUP BY d.name
            ORDER BY total_leaves DESC
        """)
        leave_stats = cur.fetchall()
        
        # Schedule coverage by department
        cur.execute("""
            SELECT 
                d.name,
                COUNT(DISTINCT sch.id) as total_shifts,
                COUNT(DISTINCT sch.staff_id) as staff_scheduled,
                COUNT(DISTINCT sch.schedule_date) as days_covered
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            LEFT JOIN schedules sch ON s.id = sch.staff_id
                AND sch.schedule_date >= CURRENT_DATE
            WHERE d.status = 'Active'
            GROUP BY d.name
            ORDER BY total_shifts DESC
        """)
        schedule_coverage = cur.fetchall()
        
        # Salary distribution by department
        cur.execute("""
            SELECT 
                d.name,
                ROUND(AVG(s.salary), 2) as avg_salary,
                MIN(s.salary) as min_salary,
                MAX(s.salary) as max_salary,
                SUM(s.salary) as total_salary
            FROM departments d
            JOIN staff s ON d.id = s.department_id
            WHERE d.status = 'Active' AND s.status = 'Active'
            GROUP BY d.name
            ORDER BY avg_salary DESC
        """)
        salary_stats = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error generating department analytics: {e}")
        staff_distribution = []
        attendance_rates = []
        leave_stats = []
        schedule_coverage = []
        salary_stats = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "department_analytics.html",
        staff_distribution=staff_distribution,
        attendance_rates=attendance_rates,
        leave_stats=leave_stats,
        schedule_coverage=schedule_coverage,
        salary_stats=salary_stats,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        current_year=date.today().year
    )


@app.route("/hr/departments/export")
def export_departments():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get all departments with statistics
        cur.execute("""
            SELECT 
                d.name,
                d.code,
                d.head_of_dept,
                d.status,
                d.created_at,
                COUNT(s.id) as total_staff,
                SUM(CASE WHEN s.status = 'Active' THEN 1 ELSE 0 END) as active_staff,
                COALESCE(AVG(s.salary), 0) as avg_salary
            FROM departments d
            LEFT JOIN staff s ON d.id = s.department_id
            GROUP BY d.id, d.name, d.code, d.head_of_dept, d.status, d.created_at
            ORDER BY d.name
        """)
        
        departments = cur.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Departments Report"
        
        # Add headers
        headers = [
            "Department Name", "Code", "Head of Department", "Status",
            "Created Date", "Total Staff", "Active Staff", "Average Salary (₦)"
        ]
        ws.append(headers)
        
        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for dept in departments:
            ws.append([
                dept[0],
                dept[1],
                dept[2] or 'Not assigned',
                dept[3],
                dept[4].strftime("%Y-%m-%d") if dept[4] else '',
                dept[5] or 0,
                dept[6] or 0,
                round(dept[7] or 0, 2)
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        # Generate filename
        filename = f"departments_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error exporting departments: {e}")
        flash(f"Error exporting report: {str(e)}", "danger")
        return redirect(url_for("departments"))
        
    finally:
        cur.close()
        conn.close()


@app.route("/api/departments/staff/<int:department_id>")
def get_department_staff(department_id):
    if "hr_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT 
                s.id,
                s.staff_id,
                s.first_name || ' ' || s.last_name as name,
                s.position,
                s.status
            FROM staff s
            WHERE s.department_id = %s AND s.status = 'Active'
            ORDER BY s.first_name, s.last_name
        """, (department_id,))
        
        staff = cur.fetchall()
        
        return jsonify({
            "success": True,
            "staff": [{
                "id": s[0],
                "staff_id": s[1],
                "name": s[2],
                "position": s[3],
                "status": s[4]
            } for s in staff]
        })
        
    except Exception as e:
        app.logger.error(f"Error fetching department staff: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cur.close()
        conn.close() 
    
# -------------------- RUN APP --------------------
if __name__ == "__main__":
    create_tables()  # Your existing tables
    create_default_users()  # Your existing default users
    create_hr_tables()  # Add this line for HR tables
    app.run(debug=True)