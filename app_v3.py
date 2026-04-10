# -------------------- IMPORTS --------------------
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import psycopg2
import os
import json
from datetime import datetime, date, timedelta
from calendar import month_name
import uuid
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# -------------------- FLASK APP SETUP --------------------
app = Flask(__name__)
app.secret_key = "super_secret_key_change_later"  # change for production

# -------------------- DATABASE SETUP --------------------
DATABASE_URL = "postgresql://flask_user:Olarewaju1.@localhost:5432/hospital_db"

def get_db_connection():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        print(f"Database connection error: {e}")
        return None

# -------------------- DATABASE INITIALIZATION --------------------
def create_pharmacists_table():
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS pharmacists (
            id SERIAL PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    cursor.close()
    conn.close()

def create_default_pharmacist():
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()
    hashed_password = generate_password_hash("pharma123")
    cursor.execute("""
        INSERT INTO pharmacists (username, password)
        VALUES (%s, %s)
        ON CONFLICT (username)
        DO UPDATE SET password = EXCLUDED.password
    """, ("pharmacist1", hashed_password))
    conn.commit()
    cursor.close()
    conn.close()

def create_default_billing_user():
    conn = get_db_connection()
    cur = conn.cursor()
    hashed_password = generate_password_hash("billing123")
    cur.execute("""
        INSERT INTO billing_users (username, password)
        VALUES (%s, %s)
        ON CONFLICT (username)
        DO NOTHING
    """, ("billing1", hashed_password))
    conn.commit()
    cur.close()
    conn.close()

# Initialize tables & default users on startup
create_pharmacists_table()
create_default_pharmacist()
create_default_billing_user()

# -------------------- LANDING PAGE / DASHBOARD --------------------
@app.route('/')
def landing_page():
    hospital_name = "John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    modules = [
        "System Admin",
        "Patient Services",
        "Clinical Services",
        "Pharmacy",
        "Laboratory",
        "Radiology",
        "Billing and Revenue",
        "Human Resources",
        "Management and Reports"
    ]
    return render_template("dashboard.html", hospital_name=hospital_name, modules=modules)

# -------------------- MODULE ROUTES --------------------
@app.route('/<module_name>')
def module_placeholder(module_name):
    display_name = module_name.replace('_', ' ').title()
    if module_name.lower() == "pharmacy":
        return redirect(url_for('pharmacy_login'))
    elif module_name.lower() == "billing":
        return redirect(url_for('billing_login'))
    return render_template("module_placeholder.html", module_name=display_name)

# -------------------- PHARMACY LOGIN --------------------
@app.route('/pharmacy/login', methods=['GET', 'POST'])
def pharmacy_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return render_template("pharmacy_login.html")

        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, username, password FROM pharmacists WHERE username=%s AND is_active=TRUE",
            (username,)
        )
        pharmacist = cursor.fetchone()
        cursor.close()
        conn.close()

        if pharmacist and check_password_hash(pharmacist[2], password):
            session['pharmacist_id'] = pharmacist[0]
            session['pharmacist_username'] = pharmacist[1]
            return redirect(url_for('pharmacy_dashboard'))
        else:
            flash("Invalid username or password", "danger")

    return render_template("pharmacy_login.html")

# -------------------- PHARMACY DASHBOARD --------------------
@app.route('/pharmacy/dashboard')
def pharmacy_dashboard():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))
    return render_template("pharmacy_dashboard.html", pharmacist_name=session.get('pharmacist_username'))

# -------------------- PHARMACY LOGOUT --------------------
@app.route('/pharmacy/logout')
def pharmacy_logout():
    session.clear()
    return redirect(url_for('pharmacy_login'))

# -------------------- PHARMACY STOCK, SALES, RECEIPTS, API --------------------
# All previous pharmacy endpoints (add_stock, drug_sales, api_drugs, pharmacy_receipt, save_patient_info, reprint_receipt, confirm_payment, stock_report, exports, stock_movements, monthly_valuation, revenue_report)
# (Include every single function from your previous code for pharmacy)
# [Omitted here for brevity but in the final file, include everything exactly as you have]
# # -------------------- DRUG SALES DASHBOARD --------------------# Pharmacy Dashboard

@app.route('/pharmacy/drug_sales')
def drug_sales():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    return render_template(
        "drug_sales_dashboard.html",
        pharmacist_name=session.get('pharmacist_username'),
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    )


# -------------------- ADD or RESTOCK --------------------

@app.route('/pharmacy/add-stock', methods=['GET', 'POST'])
def add_stock():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    if request.method == 'POST':
        drug_name = request.form.get('drug_name', '').strip()
        strength = request.form.get('strength', '').strip()
        unit_price = request.form.get('unit_price', '').strip()
        quantity = request.form.get('quantity', '').strip()
        expiry_date = request.form.get('expiry_date', '').strip()  # New field

        # Validation
        if not all([drug_name, strength, unit_price, quantity, expiry_date]):
            flash("All fields including expiry date are required.", "danger")
            return redirect(url_for('add_stock'))

        try:
            unit_price = float(unit_price)
            quantity = int(quantity)
        except ValueError:
            flash("Invalid price or quantity.", "danger")
            return redirect(url_for('add_stock'))

        # Validate expiry date format
        try:
            expiry_date_obj = datetime.strptime(expiry_date, "%Y-%m-%d").date()
        except ValueError:
            flash("Invalid expiry date format.", "danger")
            return redirect(url_for('add_stock'))

        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('add_stock'))

        cursor = conn.cursor()

        # Check if the drug + strength + expiry already exists
        cursor.execute("""
            SELECT id, stock_quantity
            FROM drugs
            WHERE name = %s AND strength = %s AND expiry_date = %s
        """, (drug_name, strength, expiry_date_obj))

        existing = cursor.fetchone()

        if existing:
            # Increase stock if the same batch exists
            cursor.execute("""
                UPDATE drugs
                SET stock_quantity = stock_quantity + %s,
                    unit_price = %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (quantity, unit_price, existing[0]))
        else:
            # Insert new drug batch
            cursor.execute("""
                INSERT INTO drugs (name, strength, unit_price, stock_quantity, expiry_date)
                VALUES (%s, %s, %s, %s, %s)
            """, (drug_name, strength, unit_price, quantity, expiry_date_obj))

        conn.commit()
        cursor.close()
        conn.close()

        flash("Stock added successfully.", "success")
        return redirect(url_for('add_stock'))

    return render_template("add_stock.html")



from flask import jsonify

@app.route('/api/drugs')
def api_drugs():
    if 'pharmacist_id' not in session:
        return jsonify([])

    search = request.args.get('q', '').strip()

    conn = get_db_connection()
    cur = conn.cursor()

    if search:
        cur.execute("""
            SELECT id, name, strength, unit_price, stock_quantity
            FROM drugs
            WHERE stock_quantity > 0
              AND LOWER(name) LIKE LOWER(%s)
            ORDER BY name ASC
        """, (f"{search}%",))
    else:
        cur.execute("""
            SELECT id, name, strength, unit_price, stock_quantity
            FROM drugs
            WHERE stock_quantity > 0
            ORDER BY name ASC
        """)

    rows = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify([
        {
            "id": r[0],
            "name": r[1],
            "strength": r[2],
            "unit_price": float(r[3]),
            "stock_quantity": r[4]
        }
        for r in rows
    ])

import uuid
from flask import jsonify

@app.route('/pharmacy/receipt', methods=['POST'])
def pharmacy_receipt():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    data = request.json
    receipt_no = f"RX-{uuid.uuid4().hex[:8].upper()}"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO drug_sales
        (receipt_no, items, subtotal, discount, tax, grand_total, pharmacist)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        receipt_no,
        json.dumps(data["items"]),
        data["subtotal"],
        data["discount"],
        data["tax"],
        data["grand_total"],
        session.get('pharmacist_username')
    ))

    conn.commit()
    cur.close()
    conn.close()

    data["receipt_no"] = receipt_no

    return render_template(
        "receipt.html",
        receipt=data,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State",
        pharmacist_name=session.get('pharmacist_username')
    )

@app.route('/pharmacy/save-patient', methods=['POST'])
def save_patient_info():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    receipt_no = request.form['receipt_no']
    patient_name = request.form['patient_name']
    patient_id = request.form.get('patient_id')

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE drug_sales
        SET patient_name=%s, patient_id=%s
        WHERE receipt_no=%s
    """, (patient_name, patient_id, receipt_no))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for('reprint_receipt', receipt_no=receipt_no))


import json

@app.route('/pharmacy/receipt/<receipt_no>')
def reprint_receipt(receipt_no):
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    conn = get_db_connection()
    cur = conn.cursor()

    # Fetch receipt header
    cur.execute("""
        SELECT receipt_no, patient_name, patient_id, items,
               subtotal, discount, tax, grand_total, pharmacist, created_at
        FROM drug_sales
        WHERE receipt_no = %s
    """, (receipt_no,))

    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        flash("Receipt not found", "danger")
        return redirect(url_for('pharmacy_dashboard'))

    # Convert row tuple to dict and deserialize items
    receipt = {
        "receipt_no": row[0],
        "patient_name": row[1],
        "patient_id": row[2],
        "items": json.loads(row[3]) if row[3] else [],  # <-- deserialize JSON string to list
        "subtotal": float(row[4]),
        "discount": float(row[5]),
        "tax": float(row[6]),
        "grand_total": float(row[7]),
        "pharmacist": row[8],
        "date": row[9].strftime("%Y-%m-%d %H:%M:%S")
    }

    return render_template(
        "receipt.html",
        receipt=receipt,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    )




@app.route("/receipt_search", methods=["GET", "POST"])
def receipt_search():
    receipts = []

    if request.method == "POST":
        query = request.form.get("query")

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, patient_name, patient_id, created_at, grand_total
            FROM receipts
            WHERE patient_name ILIKE %s
               OR patient_id ILIKE %s
            ORDER BY created_at DESC;
        """, (f"%{query}%", f"%{query}%"))

        receipts = cur.fetchall()
        cur.close()
        conn.close()

    return render_template("receipt_search.html", receipts=receipts)


# @app.route("/pharmacy/confirm-payment", methods=["POST"])
# def confirm_payment():
#     data = request.get_json()

#     conn = get_db_connection()
#     cur = conn.cursor()

#     cur.execute("""
#         INSERT INTO receipts (
#             patient_name,
#             patient_id,
#             subtotal,
#             discount,
#             tax,
#             total_amount,
#             grand_total
#         )
#         VALUES (%s, %s, %s, %s, %s, %s, %s)
#         RETURNING id;
#     """, (
#         data["patient_name"],
#         data["patient_id"],
#         data["subtotal"],
#         data["discount"],
#         data["tax"],
#         data["grand_total"],
#         data["grand_total"]
#     ))

#     receipt_id = cur.fetchone()[0]

#     for item in data["items"]:
#         cur.execute("""
#             INSERT INTO receipt_items (
#                 receipt_id,
#                 drug_name,
#                 strength,
#                 quantity,
#                 unit_price
#             )
#             VALUES (%s, %s, %s, %s, %s);
#         """, (
#             receipt_id,
#             item["drug_name"],
#             item["strength"],
#             item["quantity"],
#             item["unit_price"]
#         ))

#         cur.execute("""
#             UPDATE drugs
#             SET stock_quantity = stock_quantity - %s
#             WHERE name = %s AND strength = %s;
#         """, (
#             item["quantity"],
#             item["drug_name"],
#             item["strength"]
#         ))

#     conn.commit()
#     cur.close()
#     conn.close()

#     return jsonify({"success": True, "receipt_id": receipt_id})

from flask import jsonify
from datetime import datetime

@app.route("/pharmacy/confirm-payment", methods=["POST"])
def confirm_payment():
    if "pharmacist_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # ---------------- INSERT RECEIPT (POSTGRES RETURNING) ----------------
        cur.execute("""
            INSERT INTO receipts (
                patient_name,
                patient_id,
                subtotal,
                discount,
                tax,
                total_amount,
                grand_total,
                created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id;
        """, (
            data.get("patient_name"),
            data.get("patient_id"),
            data["subtotal"],
            data["discount"],
            data["tax"],
            data["grand_total"],
            data["grand_total"],
            datetime.now()
        ))

        receipt_id = cur.fetchone()[0]  # ✅ SAFE

        # ---------------- INSERT RECEIPT ITEMS + UPDATE STOCK ----------------
        for item in data["items"]:
            # Save sold item
            cur.execute("""
                INSERT INTO receipt_items (
                    receipt_id,
                    drug_name,
                    strength,
                    quantity,
                    unit_price
                )
                VALUES (%s, %s, %s, %s, %s);
            """, (
                receipt_id,
                item["drug_name"],
                item["strength"],
                item["quantity"],
                item["unit_price"]
            ))

            # Reduce stock
            cur.execute("""
                UPDATE drugs
                SET stock_quantity = stock_quantity - %s,
                    updated_at = NOW()
                WHERE name = %s AND strength = %s;
            """, (
                item["quantity"],
                item["drug_name"],
                item["strength"]
            ))

        conn.commit()

    except Exception as e:
        conn.rollback()
        cur.close()
        conn.close()
        return jsonify({"success": False, "error": str(e)}), 500

    cur.close()
    conn.close()

    return jsonify({
        "success": True,
        "receipt_id": receipt_id
    })



@app.route("/receipt/<int:receipt_id>")
def receipt(receipt_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM receipts WHERE id = %s;", (receipt_id,))
    receipt = cur.fetchone()

    cur.execute("""
        SELECT drug_name, strength, quantity, unit_price
        FROM receipt_items
        WHERE receipt_id = %s;
    """, (receipt_id,))
    items = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "receipt.html",
        receipt=receipt,
        items=items
    )

@app.route("/pharmacy/receipt/<int:receipt_id>")
def view_receipt(receipt_id):
    if "pharmacist_id" not in session:
        return redirect(url_for("pharmacy_login"))

    conn = get_db_connection()
    cur = conn.cursor()

    # Fetch receipt header
    cur.execute("""
        SELECT id, patient_name, patient_id, subtotal, discount, tax, grand_total, created_at
        FROM receipts
        WHERE id = %s
    """, (receipt_id,))

    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        flash("Receipt not found", "danger")
        return redirect(url_for("pharmacy_dashboard"))

    # Convert to dict
    receipt = {
        "id": row[0],
        "patient_name": row[1],
        "patient_id": row[2],
        "subtotal": float(row[3]),
        "discount": float(row[4]),
        "tax": float(row[5]),
        "grand_total": float(row[6]),
        "date": row[7].strftime("%Y-%m-%d %H:%M:%S")
    }

    # Fetch receipt items
    cur.execute("""
        SELECT drug_name, strength, quantity, unit_price
        FROM receipt_items
        WHERE receipt_id = %s
    """, (receipt_id,))

    items_rows = cur.fetchall()
    items = []
    for i in items_rows:
        items.append({
            "drug_name": i[0],
            "strength": i[1],
            "quantity": i[2],
            "unit_price": float(i[3])
        })

    cur.close()
    conn.close()

    return render_template(
        "receipt.html",
        receipt=receipt,
        items=items,
        hospital_name="John-Theresa Spiritan Specialist Hospital, Uvuru, Nsukka, Enugu State"
    )


def build_stock_snapshot(rows, today):
    stock = []

    for r in rows:
        expiry_date = r[5]
        quantity = r[3]
        threshold = r[6] if r[6] else 20

        if expiry_date:
            days_left = (expiry_date - today).days
            if days_left < 0:
                status = "EXPIRED"
            elif days_left <= 30:
                status = "EXPIRING_SOON"
            else:
                status = "VALID"
        else:
            days_left = None
            status = "UNKNOWN"

        stock.append({
            "id": r[0],
            "name": r[1],
            "strength": r[2],
            "quantity": quantity,
            "unit_price": r[4],
            "expiry_date": expiry_date,
            "days_left": days_left,
            "status": status,
            "low_stock_threshold": threshold,
            "total_value": quantity * r[4]
        })

    return stock

def apply_stock_filter(stock, filter_type):
    if filter_type == "expired":
        return [d for d in stock if d["status"] == "EXPIRED"]
    elif filter_type in ("expiring", "expiring_soon"):
        return [d for d in stock if d["status"] == "EXPIRING_SOON"]
    elif filter_type in ("low", "low_stock"):
        return [d for d in stock if d["quantity"] <= d["low_stock_threshold"]]
    return stock



@app.route("/pharmacy/stock-report")
def stock_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    filter_type = request.args.get("filter", "all")

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            id,
            name,
            strength,
            stock_quantity,
            unit_price,
            expiry_date,
            low_stock_threshold
        FROM drugs
        ORDER BY expiry_date ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    stock = build_stock_snapshot(rows, date.today())
    stock = apply_stock_filter(stock, filter_type)

    return render_template(
        "stock_report.html",
        stock=stock,
        current_filter=filter_type,
        expired_count=sum(1 for d in stock if d["status"] == "EXPIRED"),
        expiring_soon_count=sum(1 for d in stock if d["status"] == "EXPIRING_SOON"),
        low_stock_count=sum(1 for d in stock if d["quantity"] <= d["low_stock_threshold"]),
        total_stock_value=sum(d["total_value"] for d in stock)
    )




from openpyxl import Workbook
from openpyxl.styles import Font
from flask import send_file
from datetime import date
import io

from flask import request
from openpyxl.styles import PatternFill

@app.route("/pharmacy/stock-report/export")
def export_stock_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    filter_type = request.args.get("filter", "all")

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            id,
            name,
            strength,
            stock_quantity,
            unit_price,
            expiry_date,
            low_stock_threshold
        FROM drugs
        ORDER BY expiry_date ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    stock = build_stock_snapshot(rows, date.today())
    stock = apply_stock_filter(stock, filter_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    headers = [
        "Drug Name", "Strength", "Quantity", "Unit Price (₦)",
        "Expiry Date", "Days Left", "Status", "Total Value (₦)", "Low Stock Threshold"
    ]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    fills = {
        "EXPIRED": PatternFill("solid", fgColor="FF9999"),
        "EXPIRING_SOON": PatternFill("solid", fgColor="FFFF99"),
        "LOW": PatternFill("solid", fgColor="ADD8E6")
    }

    for item in stock:
        ws.append([
            item["name"],
            item["strength"],
            item["quantity"],
            float(item["unit_price"]),
            item["expiry_date"],
            item["days_left"],
            item["status"],
            float(item["total_value"]),
            item["low_stock_threshold"]
        ])

        row_idx = ws.max_row
        if item["status"] in fills:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fills[item["status"]]
        elif item["quantity"] <= item["low_stock_threshold"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fills["LOW"]

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=f"pharmacy_stock_report_{date.today()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




def log_stock_movement(drug_id, movement_type, quantity, user_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO stock_movements (drug_id, movement_type, quantity, user_id)
        VALUES (%s, %s, %s, %s)
    """, (drug_id, movement_type, quantity, user_id))
    conn.commit()
    cur.close()
    conn.close()

@app.route('/pharmacy/stock-movements')
def stock_movements():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT sm.id, d.name, d.strength, sm.movement_type, sm.quantity, u.username, sm.created_at, sm.note
        FROM stock_movements sm
        JOIN drugs d ON sm.drug_id = d.id
        JOIN users u ON sm.user_id = u.id
        ORDER BY sm.created_at DESC
    """)
    movements = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('stock_movements.html', movements=movements)


@app.route("/pharmacy/stock-report/monthly-valuation")
def monthly_valuation():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT DATE_TRUNC('month', created_at) AS month, 
               SUM(quantity * unit_price) AS total_value
        FROM stock_movements sm
        JOIN drugs d ON sm.drug_id = d.id
        GROUP BY month
        ORDER BY month DESC;
    """)
    monthly_data = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('monthly_valuation.html', monthly_data=monthly_data)



from calendar import month_name

from datetime import date, datetime, timedelta
from calendar import month_name

from datetime import date, datetime, timedelta
from flask import render_template, request, redirect, url_for, session, flash

@app.route("/pharmacy/revenue-report", methods=["GET", "POST"])
def revenue_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    # -----------------------------
    # Get form inputs
    # -----------------------------
    report_type = request.form.get("period", "daily")
    selected_day = request.form.get("day")
    selected_month = request.form.get("month")
    selected_year = request.form.get("year")

    today = date.today()

    # -----------------------------
    # Compute date range
    # -----------------------------
    if report_type == "daily":
        start_date = end_date = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today

    elif report_type == "weekly":
        d = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today
        start_date = d - timedelta(days=d.weekday())  # Monday
        end_date = start_date + timedelta(days=6)     # Sunday

    elif report_type == "monthly":
        month = int(selected_month) if selected_month else today.month
        year = int(selected_year) if selected_year else today.year
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    else:
        flash("Invalid report period", "danger")
        return redirect(url_for("pharmacy_dashboard"))

    # -----------------------------
    # Fetch sales from DB
    # -----------------------------
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, patient_name, patient_id, grand_total, created_at
        FROM receipts
        WHERE DATE(created_at) BETWEEN %s AND %s
        ORDER BY created_at ASC;
    """, (start_date, end_date))
    sales = cur.fetchall()
    cur.close()
    conn.close()

    # -----------------------------
    # Compute total revenue safely
    # -----------------------------
    total_revenue = sum(float(s[3]) if s[3] is not None else 0.0 for s in sales)

    # -----------------------------
    # Month and year dropdowns
    # -----------------------------
    months = [(i, month_name[i]) for i in range(1, 13)]
    years = range(2024, today.year + 1)

    return render_template(
        "revenue_report.html",
        sales=sales,
        total_revenue=total_revenue,
        period=report_type,
        start_date=start_date,
        end_date=end_date,
        selected_day=selected_day,
        selected_month=int(selected_month) if selected_month else today.month,
        selected_year=int(selected_year) if selected_year else today.year,
        months=months,
        years=years
    )
# -------------------- BILLING LOGIN/DASHBOARD --------------------
@app.route('/billing/login', methods=['GET', 'POST'])
def billing_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, password
            FROM billing_users
            WHERE username = %s
        """, (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user[1], password):
            session['billing_user_id'] = user[0]
            session['billing_username'] = username
            return redirect(url_for('billing_dashboard'))
        else:
            flash("Invalid login credentials", "danger")

    return render_template("billing_login.html")

# REPLACE the old billing_dashboard function with this one:

@app.route("/billing/dashboard")
def billing_dashboard():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))

    # No longer need to fetch invoices
    return render_template("billing_dashboard.html")

@app.route('/billing/logout')
def billing_logout():
    session.pop('billing_user_id', None)
    session.pop('billing_username', None)
    flash('Logged out successfully', 'success')
    return redirect(url_for('billing_login'))

# # -------------------- BILLING PAYMENT --------------------
# @app.route('/billing/accept-payment/<int:invoice_id>', methods=['POST'])
# def accept_payment(invoice_id):
#     # Implementation same as before
#     pass

@app.route('/billing/confirm-payment', methods=['POST'])
def billing_confirm_payment():
    # Implementation same as before
    pass

# -------------------- RUN APP --------------------
if __name__ == '__main__':
    app.run(debug=True)
